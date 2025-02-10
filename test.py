import oracledb
from flask import Flask, jsonify, request
from flask_cors import CORS  # Import CORS
import logging
from flask import Flask, request, send_file
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure Oracle database connection pool
DB_POOL = oracledb.create_pool(
    user="compiere",
    password="compiere",
    dsn="192.168.1.213/compiere",
    min=2,
    max=10,
    increment=1
)

# Fetch data from Oracle DB
def fetch_filtred_emplacment_from_db():
    try:
        # Fetch emplacement data
        emplacements_data = fetch_emplacment_data()

        # Extract emplacement values
        emplacements = [entry["EMPLACEMENT"] for entry in emplacements_data]

        # Build the emplacement filter dynamically
        emplacement_filter = ""
        if emplacements:
            emplacement_values = "', '".join(emplacements)  # Convert list to SQL format
            emplacement_filter = f"AND ml.value IN ('{emplacement_values}')"

        # Define the SQL query with dynamic emplacement filtering
        query = f"""
        SELECT 
            mati.value AS fournisseur, 
            m.name,  
            SUM(m_storage.qtyonhand) AS qty,
            SUM(M_ATTRIBUTEINSTANCE.valuenumber * m_storage.qtyonhand) AS prix,
            SUM(m_storage.qtyonhand - m_storage.QTYRESERVED) AS qty_dispo, 
            SUM(M_ATTRIBUTEINSTANCE.valuenumber * (m_storage.qtyonhand - m_storage.QTYRESERVED)) AS prix_dispo,
            ml.M_Locator_ID AS locatorid,
            m.m_product_id AS productid,
            1 AS sort_order
        FROM 
            M_ATTRIBUTEINSTANCE
        JOIN 
            m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
        JOIN 
            M_PRODUCT m ON m.M_PRODUCT_id = m_storage.M_PRODUCT_id
        JOIN 
            M_Locator ml ON ml.M_Locator_ID = m_storage.M_Locator_ID
        INNER JOIN 
            m_attributeinstance mati ON m_storage.m_attributesetinstance_id = mati.m_attributesetinstance_id
        WHERE 
            M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
            AND m_storage.qtyonhand > 0
            AND mati.m_attribute_id = 1000508
            AND m_storage.AD_Client_ID = 1000000
            {emplacement_filter}  -- Dynamically added emplacement filter
        GROUP BY 
            m.name, mati.value, m.m_product_id, ml.M_Locator_ID
        ORDER BY 
            fournisseur, name
        """

        # Execute the query
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            data = [dict(zip(columns, row)) for row in rows]

        return data

    except Exception as e:
        logger.error(f"Error fetching data: {e}")
        return {"error": "An error occurred while fetching data."}



def fetch_emplacment_data():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
            SELECT ml.value AS EMPLACEMENT
            FROM M_Locator ml
            JOIN M_Warehouse m ON m.M_WAREHOUSE_ID = ml.M_WAREHOUSE_ID
            WHERE m.ISACTIVE = 'Y'
                AND m.AD_Client_ID = 1000000
                AND ml.ISACTIVE = 'Y'
                AND ml.AD_Client_ID = 1000000
            ORDER BY m.value
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching remise data: {e}")
        return {"error": "An error occurred while fetching emplacement data."}
    
@app.route('/fetch-emplacement-data', methods=['GET'])
def fetch_data():
    data = fetch_emplacment_data()
    return jsonify(data)

@app.route('/fetch-data', methods=['GET'])
def fetch_stock_data():
    data = fetch_filtred_emplacment_from_db()
    return jsonify(data)

# Route to download Excel for bonus data
@app.route('/download-stock-excel', methods=['GET'])
def download_stock_excel():
    stock_value = request.args.get("stock", "STOCK")
    filters = {
        "locatorid": request.args.get("locatorid"),
        "locatorname": request.args.get("locatorname"),  # Capture the name
        "fournisseur": request.args.get("fournisseur"),
        "name": request.args.get("name")
    }

    data = fetch_filtred_emplacment_from_db()

    # Apply filters
    if filters["locatorid"]:
        data = [row for row in data if str(row["LOCATORID"]) == filters["locatorid"]]
    if filters["fournisseur"]:
        data = [row for row in data if filters["fournisseur"].lower() in row["FOURNISSEUR"].lower()]
    if filters["name"]:
        data = [row for row in data if filters["name"].lower() in row["NAME"].lower()]

    # Generate filename with filter values
    today_date = datetime.now().strftime("%d-%m-%Y")
    location_part = f"_{filters['locatorname']}" if filters["locatorname"] else ""
    filter_values = "_".join([f"{key}-{value}" for key, value in filters.items() if value and key != "locatorname"])
    
    filename = f"{stock_value}_{today_date}{location_part}.xlsx" if not filter_values else f"{stock_value}_{today_date}{location_part}_{filter_values}.xlsx"

    return generate_excel_stock(data, filename)




# Helper function to generate Excel file
def generate_excel_stock(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered Data"

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    table = Table(displayName="DataTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    app.run(debug=True, port=5003)