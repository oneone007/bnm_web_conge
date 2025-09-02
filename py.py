

import oracledb
from flask import Flask, jsonify, request, send_file, make_response
from flask_cors import CORS
import logging
import pandas as pd
from io import BytesIO
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.cell.cell import MergedCell
from datetime import datetime
import os
import json

import calendar
import mysql.connector




app = Flask(__name__)
CORS(app)  # Allow your DDNS domains and ports

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure Oracle database connection pool
DB_POOL = oracledb.create_pool(
    user="compiere",
    password="compiere",
    dsn="192.168.1.213/compiere",
    min=2,
    max=10,
    increment=1
)



# Fetch reserved data from Oracle DB
def fetch_reserved_data():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
            SELECT 
                us.name AS OPERATEUR,
                co.documentno AS NDOCUMENT,
                m.name AS PRODUCT,
                co.dateordered AS DATECOMMANDE,
                s.qtyreserved AS TOTALRESERVE,
                cl.qtyreserved AS QTYRESERVE,
                l.name AS NAME,
                CASE
                    WHEN co.docstatus = 'CO' THEN 'prepared'
                    ELSE 'not prepared'
                END AS STATUS
            FROM
                m_storage s
            INNER JOIN c_orderline cl
                ON cl.m_attributesetinstance_id = s.m_attributesetinstance_id
            INNER JOIN c_order co
                ON co.c_order_id = cl.c_order_id
            INNER JOIN m_product m
                ON m.m_product_id = s.m_product_id
            INNER JOIN ad_user us
                ON us.ad_user_id = co.salesrep_id
            JOIN XX_Laboratory l
                ON l.XX_Laboratory_ID = m.XX_Laboratory_ID
            WHERE
                cl.qtyreserved != 0
                AND co.c_doctypetarget_id = 1000539
                AND cl.m_product_id = s.m_product_id
                AND co.ad_org_id = 1000000
                AND s.qtyonhand > 0
                AND s.qtyreserved > 0
            FETCH FIRST 1048575 ROWS ONLY
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching reserved data: {e}")
        return {"error": "An error occurred while fetching reserved data."}

# Fetch remise data from Oracle DB
def fetch_remise_data():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
            WITH LastFournisseur AS (
                SELECT
                    cf.m_product_id,
                    cf.c_bpartner_id,
                    ROW_NUMBER() OVER (
                        PARTITION BY cf.m_product_id
                        ORDER BY cf.dateinvoiced DESC
                    ) AS rn
                FROM
                    xx_ca_fournisseur_facture cf
            )
            SELECT
                mp.name AS product,
                md.name AS reward,
                l.name AS laboratory_name,
                cb.name AS fournisseur,
                CBG.NAME AS Type_Client
            FROM
                m_product mp
                INNER JOIN C_BPartner_Product cbp ON mp.m_product_id = cbp.m_product_id
                INNER JOIN  C_BP_Group CBG  ON CBG.C_BP_Group_ID = CBP.C_BP_Group_ID
                JOIN XX_Laboratory l ON l.XX_Laboratory_ID = mp.XX_Laboratory_ID
                JOIN LastFournisseur lf ON mp.m_product_id = lf.m_product_id AND lf.rn = 1
                JOIN c_bpartner cb ON cb.c_bpartner_id = lf.c_bpartner_id
                INNER JOIN M_DiscountSchema md ON cbp.M_DiscountSchema_id = md.M_DiscountSchema_id
            WHERE
                cbp.C_BPartner_Product_id IS NOT NULL
                AND cbp.m_discountschema_id IS NOT NULL
            ORDER BY
                mp.name
            FETCH FIRST 1048575 ROWS ONLY
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching remise data: {e}")
        return {"error": "An error occurred while fetching remise data."}

# Fetch marge data from Oracle DB

# Fetch marge data from Oracle DB
def fetch_marge_data():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
SELECT
    *
FROM
    (
        SELECT
            source.FOURNISSEUR,
            source.PRODUCT,
            source.P_ACHAT,
            source.P_VENTE,
            source.REM_ACHAT,
            source.REM_VENTE,
            source.BON_ACHAT,
            source.BON_VENTE,
            source.REMISE_AUTO,
            source.BONUS_AUTO,
            source.P_REVIENT,
            source.MARGE,
            source.LABO,
            source.LOT,
            source.QTY,
            source.QTY_DISPO,
            source.GUARANTEEDATE,
            source.PPA,
            source.LOCATION
        FROM
            (
                SELECT
                    DISTINCT fournisseur,
                    product,
                    p_achat,
                    p_vente,
                    ROUND(rem_achat, 2) AS rem_achat,
                    rem_vente,
                    ROUND(bon_achat, 2) AS bon_achat,
                    bon_vente,
                    remise_auto,
                    bonus_auto,
                    ROUND(p_revient, 2) AS p_revient,
                    LEAST(ROUND(marge, 2), 100) AS marge,
                    labo,
                    lot,
                    qty,
                    qty_dispo,
                    guaranteedate,
                    ppa,
                    CASE 
                        WHEN m_locator_id = 1000614 THEN 'Préparation'
                        WHEN m_locator_id = 1001135 THEN 'HANGAR'
                        WHEN m_locator_id = 1001128 THEN 'Dépot_réserve'
                        WHEN m_locator_id = 1001136 THEN 'HANGAR_'
                        WHEN m_locator_id = 1001020 THEN 'Depot_Vente'
                    END AS location
                FROM
                    (
                        SELECT
                            d.*,
                            LEAST(
                                ROUND(
                                    (((ventef - ((ventef * NVL(rma, 0)) / 100))) - p_revient) / p_revient * 100,
                                    2
                                ), 
                                100
                            ) AS marge
                        FROM
                            (
                                SELECT
                                    det.*,
                                    (det.p_achat - ((det.p_achat * det.rem_achat) / 100)) / (1 + (det.bon_achat / 100)) AS p_revient,
                                    (
                                        det.p_vente - ((det.p_vente * NVL(det.rem_vente, 0)) / 100)
                                    ) / (
                                        1 + (
                                            CASE
                                                WHEN det.bna > 0 THEN det.bna
                                                ELSE det.bon_vente
                                            END / 100
                                        )
                                    ) AS ventef
                                FROM
                                    (
                                        SELECT
                                            p.name AS product,
                                            (
                                                SELECT NAME
                                                FROM XX_Laboratory
                                                WHERE XX_Laboratory_id = p.XX_Laboratory_id
                                            ) AS labo,
                                            mst.qtyonhand AS qty,
                                            (mst.qtyonhand - mst.QTYRESERVED) AS qty_dispo,
                                            mst.m_locator_id,
                                            mati.value AS fournisseur,
                                            mats.guaranteedate,
                                            md.name AS remise_auto,
                                            sal.description AS bonus_auto,
                                            md.flatdiscount AS rma,
                                            TO_NUMBER(
                                                CASE
                                                    WHEN REGEXP_LIKE(sal.name, '^[0-9]+$') THEN sal.name
                                                    ELSE NULL
                                                END
                                            ) AS bna,
                                            (
                                                SELECT valuenumber
                                                FROM m_attributeinstance
                                                WHERE m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                  AND m_attribute_id = 1000501
                                            ) AS p_achat,
                                            (
                                                SELECT valuenumber
                                                FROM m_attributeinstance
                                                WHERE m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                  AND m_attribute_id = 1001009
                                            ) AS rem_achat,
                                            (
                                                SELECT valuenumber
                                                FROM m_attributeinstance
                                                WHERE m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                  AND m_attribute_id = 1000808
                                            ) AS bon_achat,
                                            (
                                                SELECT valuenumber
                                                FROM m_attributeinstance
                                                WHERE m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                  AND m_attribute_id = 1000502
                                            ) AS p_vente,
                                            (
                                                SELECT valuenumber
                                                FROM m_attributeinstance
                                                WHERE m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                  AND m_attribute_id = 1001408
                                            ) AS rem_vente,
                                            (
                                                SELECT valuenumber
                                                FROM m_attributeinstance
                                                WHERE m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                  AND m_attribute_id = 1000908
                                            ) AS bon_vente,
                                            (
                                                SELECT lot
                                                FROM m_attributesetinstance
                                                WHERE m_attributesetinstance_id = mst.m_attributesetinstance_id
                                            ) AS lot,
                                            (
                                                SELECT valuenumber
                                                FROM m_attributeinstance
                                                WHERE m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                  AND m_attribute_id = 1000503
                                            ) AS ppa
                                        FROM
                                            m_product p
                                            INNER JOIN m_storage mst ON p.m_product_id = mst.m_product_id
                                            INNER JOIN m_attributeinstance mati ON mst.m_attributesetinstance_id = mati.m_attributesetinstance_id
                                            INNER JOIN m_attributesetinstance mats ON mst.m_attributesetinstance_id = mats.m_attributesetinstance_id
                                            LEFT JOIN (
                                                SELECT *
                                                FROM C_BPartner_Product
                                                WHERE isactive = 'Y'
                                            ) cp ON cp.m_product_id = p.m_product_id
                                            LEFT JOIN M_DiscountSchema md ON cp.M_DiscountSchema_id = md.M_DiscountSchema_id
                                            LEFT JOIN XX_SalesContext sal ON p.XX_SalesContext_ID = sal.XX_SalesContext_ID
                                        WHERE
                                            mati.m_attribute_id = 1000508
                                            AND mst.m_locator_id IN (1001135, 1000614, 1001128, 1001136, 1001020)
                                            AND mst.qtyonhand != 0
                                        ORDER BY
                                            p.name
                                    ) det
                                WHERE
                                    det.rem_achat < 200
                            ) d
                    )
                GROUP BY
                    fournisseur,
                    product,
                    p_achat,
                    p_vente,
                    rem_achat,
                    rem_vente,
                    bon_achat,
                    bon_vente,
                    remise_auto,
                    bonus_auto,
                    p_revient,
                    marge,
                    labo,
                    lot,
                    qty,
                    qty_dispo,
                    guaranteedate,
                    ppa,
                    m_locator_id
                ORDER BY
                    fournisseur
            ) source
    )
WHERE
    ROWNUM <= 1048575
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching marge data: {e}")
        return {"error": "An error occurred while fetching marge data."}

# Fetch bonus data from Oracle DB
def fetch_bonus_data():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
            SELECT DISTINCT 
                mp.name AS product,
                (SELECT description 
                 FROM XX_SalesContext xsc 
                 WHERE mp.XX_SalesContext_id = xsc.XX_SalesContext_id) AS bonus,
                l.name AS laboratory_name,
                cbp.name AS fournisseur
            FROM 
                m_product mp
            JOIN XX_Laboratory l ON l.XX_Laboratory_ID = mp.XX_Laboratory_ID
            JOIN m_storage s ON s.m_product_id = mp.m_product_id
            JOIN M_ATTRIBUTEINSTANCE asi ON s.M_ATTRIBUTESETINSTANCE_ID = asi.M_ATTRIBUTESETINSTANCE_ID
            JOIN c_bpartner cbp ON cbp.c_bpartner_id = ValueNUMBER_of_ASI('Fournisseur', asi.m_attributesetinstance_id)
            WHERE 
                mp.xx_salescontext_id NOT IN (1000000, 1000100)
                AND mp.ad_org_id = 1000000
            ORDER BY 
                mp.name
            FETCH FIRST 1048575 ROWS ONLY
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching bonus data: {e}")
        return {"error": "An error occurred while fetching bonus data."}



# Test database connection
def test_db_connection():
    try:
        with DB_POOL.acquire() as connection:
            logger.info("Database connection successful.")
        return True
    except Exception as e:
        logger.error(f"Database connection failed: {str(e)}")
        return False
    



# Route to fetch reserved data
@app.route('/fetch-reserved-data', methods=['GET'])
def fetch_reserved():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_reserved_data()
    return jsonify(data)

# Route to fetch remise data
@app.route('/fetch-remise-data', methods=['GET'])
def fetch_remise():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_remise_data()
    return jsonify(data)

# Route to fetch marge data
@app.route('/fetch-data', methods=['GET'])
def fetch_marge():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_marge_data()
    return jsonify(data)



# Route to fetch bonus data
@app.route('/fetch-bonus-data', methods=['GET'])
def fetch_bonus():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_bonus_data()
    return jsonify(data)

# Route to download Excel for reserved data
@app.route('/download-reserved-excel', methods=['GET'])
def download_reserved_excel():
    reserve_value = request.args.get("reserve", "RESERVE")
    data = fetch_reserved_data()
    return generate_excel(data, reserve_value)

# Route to download Excel for remise data
@app.route('/download-remise-excel', methods=['GET'])
def download_remise_excel():
    remise_value = request.args.get("remise", "REMISE")
    data = fetch_remise_data()
    return generate_excel(data, remise_value)


# Route to download Excel for bonus data
@app.route('/download-bonus-excel', methods=['GET'])
def download_bonus_excel():
    bonus_value = request.args.get("bonus", "BONUS")
    data = fetch_bonus_data()
    return generate_excel(data, bonus_value)


# Route to download Excel for marge data
@app.route('/download-marge-excel', methods=['GET'])
def download_marge_excel():
    data = fetch_marge_data()  # Fetch all data
    filters = {
        "fournisseur": request.args.get("fournisseur", "").strip(),
        "product": request.args.get("product", "").strip(),
        "marge": request.args.get("marge", "").strip(),
    }

    # Apply filters if provided
    if filters["fournisseur"]:
        data = [row for row in data if filters["fournisseur"].lower() in row["FOURNISSEUR"].lower()]
    if filters["product"]:
        data = [row for row in data if filters["product"].lower() in row["PRODUCT"].lower()]
    if filters["marge"]:
        try:
            marge_value = float(filters["marge"])
            data = [row for row in data if float(row["MARGE"]) >= marge_value]
        except ValueError:
            return jsonify({"error": "Invalid marge value"}), 400

    return generate_excel_marge(data, filters)

#----------------------------
def generate_excel_marge(data, filters):
    if not data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered Data"

    # Apply header styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(df.columns.tolist())
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Append data with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Generate filename dynamically based on filters (include values)
    today_date = datetime.now().strftime("%d-%m-%Y")
    filter_text = "_".join([f"{k}-{v}" for k, v in filters.items() if v])  # Include values
    filename = f"Marge_{filter_text}_{today_date}.xlsx" if filter_text else f"Marge_{today_date}.xlsx"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# Helper function to generate Excel file
def generate_excel(data, value):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "All Data"

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    table = Table(displayName="DataTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"{value}_{today_date}.xlsx"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route('/fetch-stock-data', methods=['GET'])
def fetch_stock_data():
    try:
        fournisseur = request.args.get("fournisseur", None)
        magasin = request.args.get("magasin", None)
        emplacement = request.args.get("emplacement", None)
        name = request.args.get("name", None)

        data = fetch_stock_data_from_db(fournisseur, magasin, emplacement, name)
        return jsonify(data)

    except Exception as e:
        logger.error(f"Error fetching stock data: {e}")
        return jsonify({"error": "Failed to fetch stock data"}), 500


def fetch_stock_data_from_db(fournisseur=None, magasin=None, emplacement=None, name=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
            SELECT 
                mati.value AS fournisseur, 
                m.name,  
                SUM(m_storage.qtyonhand) AS qty,
                SUM(M_ATTRIBUTEINSTANCE.valuenumber * m_storage.qtyonhand) AS prix,
                SUM(m_storage.qtyonhand - m_storage.QTYRESERVED) AS qty_dispo, 
                SUM(M_ATTRIBUTEINSTANCE.valuenumber * (m_storage.qtyonhand - m_storage.QTYRESERVED)) AS prix_dispo,
                ml.M_Locator_ID AS locatorid,
                m.m_product_id AS productid,
                1 AS sort_order,
                CASE 
                    WHEN ml.M_Locator_ID = 1000614 THEN 'Préparation'
                    WHEN ml.M_Locator_ID = 1001135 THEN 'HANGAR'
                    WHEN ml.M_Locator_ID = 1001128 THEN 'Dépot_réserve'
                    WHEN ml.M_Locator_ID = 1001136 THEN 'HANGAR_'
                    WHEN ml.M_Locator_ID = 1001020 THEN 'Depot_Vente'
                    ELSE 'Unknown' 
                END AS place
            FROM 
                M_ATTRIBUTEINSTANCE
            JOIN 
                m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
            JOIN 
                M_PRODUCT m ON m.M_PRODUCT_id = m_storage.M_PRODUCT_id
            JOIN 
                M_Locator ml ON ml.M_Locator_ID = m_storage.M_Locator_ID
            INNER JOIN 
                m_attributeinstance mati ON m_storage.m_attributesetinstance_id = mati.m_attributesetinstance_id
            WHERE 
                M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                AND m_storage.qtyonhand > 0
                AND mati.m_attribute_id = 1000508
                AND m_storage.AD_Client_ID = 1000000
            """

            params = {}

            if fournisseur:
                query += " AND mati.value LIKE :fournisseur || '%'"
                params["fournisseur"] = fournisseur

            if name:
                query += " AND m.name LIKE :name || '%'"
                params["name"] = name

            if magasin:
                query += """
                AND m_storage.M_Locator_ID IN (
                    SELECT M_Locator_ID 
                    FROM M_Locator 
                    WHERE M_Warehouse_ID IN (
                        SELECT M_Warehouse_ID 
                        FROM M_Warehouse 
                        WHERE VALUE LIKE :magasin || '%'
                    )
                )
                """
                params["magasin"] = magasin
            else:
                query += """
                AND m_storage.M_Locator_ID IN (
                    SELECT M_Locator_ID 
                    FROM M_Locator 
                    WHERE M_Warehouse_ID IN (
                        SELECT M_Warehouse_ID 
                        FROM M_Warehouse 
                        WHERE VALUE IN ('HANGAR', '1-Dépôt Principal', '8-Dépot réserve', '88-Dépot Hangar réserve')
                    )
                )
                """

            if emplacement:
                query += """
                AND (
                    (M_Warehouse_ID != 1000000 AND 
                     m_storage.M_Locator_ID IN (
                         SELECT M_Locator_ID 
                         FROM M_Locator 
                         WHERE value LIKE :emplacement || '%'
                     ))
                    OR 
                    (M_Warehouse_ID = 1000000 AND 
                     m.M_Locator_ID IN (
                         SELECT M_Locator_ID 
                         FROM M_Locator 
                         WHERE value LIKE :emplacement || '%'
                     ))
                )
                """
                params["emplacement"] = emplacement

            query += """
            GROUP BY m.name, mati.value, m.m_product_id, ml.M_Locator_ID

            UNION ALL

            SELECT 
                CAST('Total' AS NVARCHAR2(300)) AS fournisseur, 
                CAST('' AS NVARCHAR2(300)) AS name, 
                SUM(m_storage.qtyonhand) AS qty,
                SUM(M_ATTRIBUTEINSTANCE.valuenumber * m_storage.qtyonhand) AS prix,
                SUM(m_storage.qtyonhand - m_storage.QTYRESERVED) AS qty_dispo, 
                SUM(M_ATTRIBUTEINSTANCE.valuenumber * (m_storage.qtyonhand - m_storage.QTYRESERVED)) AS prix_dispo,
                NULL AS locatorid,
                NULL AS productid,
                0 AS sort_order,
                NULL AS place
            FROM 
                M_ATTRIBUTEINSTANCE
            JOIN 
                m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
            JOIN 
                M_PRODUCT m ON m.M_PRODUCT_ID = m_storage.M_PRODUCT_ID
            JOIN 
                M_Locator ml ON ml.M_Locator_ID = m_storage.M_Locator_ID
            INNER JOIN 
                m_attributeinstance mati ON m_storage.m_attributesetinstance_id = mati.m_attributesetinstance_id
            WHERE 
                M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                AND m_storage.qtyonhand > 0
                AND mati.m_attribute_id = 1000508
                AND m_storage.AD_Client_ID = 1000000
            """

            if fournisseur:
                query += " AND mati.value LIKE :fournisseur || '%'"

            if name:
                query += " AND m.name LIKE :name || '%'"

            if magasin:
                query += """
                AND m_storage.M_Locator_ID IN (
                    SELECT M_Locator_ID 
                    FROM M_Locator 
                    WHERE M_Warehouse_ID IN (
                        SELECT M_Warehouse_ID 
                        FROM M_Warehouse 
                        WHERE VALUE LIKE :magasin || '%'
                    )
                )
                """
            else:
                query += """
                AND m_storage.M_Locator_ID IN (
                    SELECT M_Locator_ID 
                    FROM M_Locator 
                    WHERE M_Warehouse_ID IN (
                        SELECT M_Warehouse_ID 
                        FROM M_Warehouse 
                        WHERE VALUE IN ('HANGAR', '1-Dépôt Principal', '8-Dépot réserve', '88-Dépot Hangar réserve')
                    )
                )
                """

            if emplacement:
                query += """
                AND (
                    (M_Warehouse_ID != 1000000 AND 
                     m_storage.M_Locator_ID IN (
                         SELECT M_Locator_ID 
                         FROM M_Locator 
                         WHERE value LIKE :emplacement || '%'
                     ))
                    OR 
                    (M_Warehouse_ID = 1000000 AND 
                     m.M_Locator_ID IN (
                         SELECT M_Locator_ID 
                         FROM M_Locator 
                         WHERE value LIKE :emplacement || '%'
                     ))
                )
                """

            query += """
            ORDER BY sort_order, fournisseur, name
            """

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]

            return data

    except Exception as e:
        logger.error(f"Database error: {e}")
        return {"error": "An error occurred while fetching stock data."}

def fetch_desactivated_lot_data_from_db(fournisseur=None, magasin=None, emplacement=None, name=None):
    """
    Fetch stock data for lots where isactive = 'N' (deactivated lots)
    """
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
            SELECT 
                mati.value AS fournisseur, 
                m.name,  
                SUM(m_storage.qtyonhand) AS qty,
                SUM(M_ATTRIBUTEINSTANCE.valuenumber * m_storage.qtyonhand) AS prix,
                SUM(m_storage.qtyonhand - m_storage.QTYRESERVED) AS qty_dispo, 
                SUM(M_ATTRIBUTEINSTANCE.valuenumber * (m_storage.qtyonhand - m_storage.QTYRESERVED)) AS prix_dispo,
                ml.M_Locator_ID AS locatorid,
                m.m_product_id AS productid,
                1 AS sort_order,
                CASE 
                    WHEN ml.M_Locator_ID = 1000614 THEN 'Préparation'
                    WHEN ml.M_Locator_ID = 1001135 THEN 'HANGAR'
                    WHEN ml.M_Locator_ID = 1001128 THEN 'Dépot_réserve'
                    WHEN ml.M_Locator_ID = 1001136 THEN 'HANGAR_'
                    WHEN ml.M_Locator_ID = 1001020 THEN 'Depot_Vente'
                    ELSE 'Unknown' 
                END AS place
            FROM 
                M_ATTRIBUTEINSTANCE
            JOIN 
                m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
            JOIN 
                M_PRODUCT m ON m.M_PRODUCT_id = m_storage.M_PRODUCT_ID
            JOIN 
                M_Locator ml ON ml.M_Locator_ID = m_storage.M_Locator_ID
            INNER JOIN 
                m_attributeinstance mati ON m_storage.m_attributesetinstance_id = mati.m_attributesetinstance_id
            INNER JOIN 
                m_attributesetinstance mats ON m_storage.m_attributesetinstance_id = mats.m_attributesetinstance_id
            WHERE 
                M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                AND m_storage.qtyonhand > 0
                AND mati.m_attribute_id = 1000508
                AND m_storage.AD_Client_ID = 1000000
                AND NVL(mats.isactive, 'Y') = 'N'
            """

            params = {}

            if fournisseur:
                query += " AND mati.value LIKE :fournisseur || '%'"
                params["fournisseur"] = fournisseur

            if name:
                query += " AND m.name LIKE :name || '%'"
                params["name"] = name

            if magasin:
                query += """
                AND m_storage.M_Locator_ID IN (
                    SELECT M_Locator_ID 
                    FROM M_Locator 
                    WHERE M_Warehouse_ID IN (
                        SELECT M_Warehouse_ID 
                        FROM M_Warehouse 
                        WHERE VALUE LIKE :magasin || '%'
                    )
                )
                """
                params["magasin"] = magasin
            else:
                query += """
                AND m_storage.M_Locator_ID IN (
                    SELECT M_Locator_ID 
                    FROM M_Locator 
                    WHERE M_Warehouse_ID IN (
                        SELECT M_Warehouse_ID 
                        FROM M_Warehouse 
                        WHERE VALUE IN ('HANGAR', '1-Dépôt Principal', '8-Dépot réserve', '88-Dépot Hangar réserve')
                    )
                )
                """

            if emplacement:
                query += """
                AND (
                    (M_Warehouse_ID != 1000000 AND 
                     m_storage.M_Locator_ID IN (
                         SELECT M_Locator_ID 
                         FROM M_Locator 
                         WHERE value LIKE :emplacement || '%'
                     ))
                    OR 
                    (M_Warehouse_ID = 1000000 AND 
                     m.M_Locator_ID IN (
                         SELECT M_Locator_ID 
                         FROM M_Locator 
                         WHERE value LIKE :emplacement || '%'
                     ))
                )
                """
                params["emplacement"] = emplacement

            query += """
            GROUP BY m.name, mati.value, m.m_product_id, ml.M_Locator_ID

            UNION ALL

            SELECT 
                CAST('Total' AS NVARCHAR2(300)) AS fournisseur, 
                CAST('' AS NVARCHAR2(300)) AS name, 
                SUM(m_storage.qtyonhand) AS qty,
                SUM(M_ATTRIBUTEINSTANCE.valuenumber * m_storage.qtyonhand) AS prix,
                SUM(m_storage.qtyonhand - m_storage.QTYRESERVED) AS qty_dispo, 
                SUM(M_ATTRIBUTEINSTANCE.valuenumber * (m_storage.qtyonhand - m_storage.QTYRESERVED)) AS prix_dispo,
                NULL AS locatorid,
                NULL AS productid,
                0 AS sort_order,
                NULL AS place
            FROM 
                M_ATTRIBUTEINSTANCE
            JOIN 
                m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
            JOIN 
                M_PRODUCT m ON m.M_PRODUCT_ID = m_storage.M_PRODUCT_ID
            JOIN 
                M_Locator ml ON ml.M_Locator_ID = m_storage.M_Locator_ID
            INNER JOIN 
                m_attributeinstance mati ON m_storage.m_attributesetinstance_id = mati.m_attributesetinstance_id
            INNER JOIN 
                m_attributesetinstance mats ON m_storage.m_attributesetinstance_id = mats.m_attributesetinstance_id
            WHERE 
                M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                AND m_storage.qtyonhand > 0
                AND mati.m_attribute_id = 1000508
                AND m_storage.AD_Client_ID = 1000000
                AND NVL(mats.isactive, 'Y') = 'N'
            """

            if fournisseur:
                query += " AND mati.value LIKE :fournisseur || '%'"

            if name:
                query += " AND m.name LIKE :name || '%'"

            if magasin:
                query += """
                AND m_storage.M_Locator_ID IN (
                    SELECT M_Locator_ID 
                    FROM M_Locator 
                    WHERE M_Warehouse_ID IN (
                        SELECT M_Warehouse_ID 
                        FROM M_Warehouse 
                        WHERE VALUE LIKE :magasin || '%'
                    )
                )
                """
            else:
                query += """
                AND m_storage.M_Locator_ID IN (
                    SELECT M_Locator_ID 
                    FROM M_Locator 
                    WHERE M_Warehouse_ID IN (
                        SELECT M_Warehouse_ID 
                        FROM M_Warehouse 
                        WHERE VALUE IN ('HANGAR', '1-Dépôt Principal', '8-Dépot réserve', '88-Dépot Hangar réserve')
                    )
                )
                """

            if emplacement:
                query += """
                AND (
                    (M_Warehouse_ID != 1000000 AND 
                     m_storage.M_Locator_ID IN (
                         SELECT M_Locator_ID 
                         FROM M_Locator 
                         WHERE value LIKE :emplacement || '%'
                     ))
                    OR 
                    (M_Warehouse_ID = 1000000 AND 
                     m.M_Locator_ID IN (
                         SELECT M_Locator_ID 
                         FROM M_Locator 
                         WHERE value LIKE :emplacement || '%'
                     ))
                )
                """

            query += """
            ORDER BY sort_order, fournisseur, name
            """

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]

            return data

    except Exception as e:
        logger.error(f"Database error: {e}")
        return {"error": "An error occurred while fetching deactivated lot data."}

# ...existing code...

# New endpoint for deactivated lots
@app.route('/fetch_desactivated_lot_data', methods=['GET'])
def fetch_desactivated_lot_data():
    try:
        fournisseur = request.args.get("fournisseur", None)
        magasin = request.args.get("magasin", None)
        emplacement = request.args.get("emplacement", None)
        name = request.args.get("name", None)

        data = fetch_desactivated_lot_data_from_db(fournisseur, magasin, emplacement, name)
        return jsonify(data)

    except Exception as e:
        logger.error(f"Error fetching deactivated lot data: {e}")
        return jsonify({"error": "Failed to fetch deactivated lot data"}), 500

        
@app.route('/download-stock-excel', methods=['GET'])
def download_stock_excel():
    try:
        # Get query parameters
        fournisseur = request.args.get("fournisseur", None)
        magasin = request.args.get("magasin", None)
        emplacement = request.args.get("emplacement", None)

        # Fetch data from the database
        data = fetch_stock_data_from_db(fournisseur, magasin, emplacement)

        # Check if data contains an error
        if "error" in data:
            return jsonify({"error": data["error"]}), 500

        # Check if data is empty
        if not data:
            return jsonify({"error": "No data available to generate Excel"}), 400

        # Generate Excel file in memory
        excel_output = generate_excel_stock(data)

        # Generate filename based on parameters
        today_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"stock_data_{today_date}.xlsx"

        if fournisseur:
            filename = f"stock_data_fournisseur_{fournisseur}_{today_date}.xlsx"
        if magasin:
            filename = f"stock_data_magasin_{magasin}_{today_date}.xlsx"
        if emplacement:
            filename = f"stock_data_emplacement_{emplacement}_{today_date}.xlsx"
        if fournisseur and magasin:
            filename = f"stock_data_fournisseur_{fournisseur}_magasin_{magasin}_{today_date}.xlsx"
        if fournisseur and emplacement:
            filename = f"stock_data_fournisseur_{fournisseur}_emplacement_{emplacement}_{today_date}.xlsx"
        if magasin and emplacement:
            filename = f"stock_data_magasin_{magasin}_emplacement_{emplacement}_{today_date}.xlsx"
        if fournisseur and magasin and emplacement:
            filename = f"stock_data_fournisseur_{fournisseur}_magasin_{magasin}_emplacement_{emplacement}_{today_date}.xlsx"

        # Send file as a download response
        return send_file(
            excel_output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"Error generating Excel: {e}")
        return jsonify({"error": "Failed to generate Excel file"}), 500


def generate_excel_stock(data):
    # Create a DataFrame from the data
    df = pd.DataFrame(data)

    # Create workbook and worksheet using openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Data"

    # Define header styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Write headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Apply auto filter (using the dimensions of the sheet)
    ws.auto_filter.ref = ws.dimensions

    # Write data rows with alternating row styling for readability
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Optionally, create an Excel table
    table = Table(displayName="DataTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save workbook to a BytesIO stream
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


 

 
def fetch_product_details_data(product_name):
    """
    Fetch detailed product information similar to the marge data structure
    """
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            
            query = """
            SELECT
                *
            FROM
                (
                    SELECT
                        "source"."FOURNISSEUR" "FOURNISSEUR",
                        "source"."PRODUCT" "PRODUCT",
                        "source"."P_ACHAT" "P_ACHAT",
                        "source"."P_VENTE" "P_VENTE",
                        "source"."REM_ACHAT" "REM_ACHAT",
                        "source"."REM_VENTE" "REM_VENTE",
                        "source"."BON_ACHAT" "BON_ACHAT",
                        "source"."BON_VENTE" "BON_VENTE",
                        "source"."REMISE_AUTO" "REMISE_AUTO",
                        "source"."BONUS_AUTO" "BONUS_AUTO",
                        "source"."P_REVIENT" "P_REVIENT",
                        "source"."MARGE" "MARGE",
                        "source"."LABO" "LABO",
                        "source"."LOT" "LOT",
                        "source"."LOT_ACTIVE" "LOT_ACTIVE",

                        "source"."QTY" "QTY",
                        "source"."QTY_DISPO" "QTY_DISPO",
                        "source"."GUARANTEEDATE" "GUARANTEEDATE",
                        "source"."PPA" "PPA",
                        "source"."LOCATION" "LOCATION"

                    FROM
                        (
                            SELECT
                                DISTINCT fournisseur,
                                product,
                                p_achat,
                                p_vente,
                                round(rem_achat, 2) AS rem_achat,
                                rem_vente,
                                round(bon_achat, 2) AS bon_achat,
                                bon_vente,
                                remise_auto,
                                bonus_auto,
                                round(p_revient, 2) AS p_revient,
                                LEAST(round((marge), 2), 100) AS marge,
                                labo,
                                lot,
                                lot_active,
                                qty,
                                qty_dispo,
                                guaranteedate,
                                ppa,
                                CASE 
                                    WHEN m_locator_id = 1000614 THEN 'Préparation'
                                    WHEN m_locator_id = 1001135 THEN 'HANGAR'
                                    WHEN m_locator_id = 1001128 THEN 'Dépot_réserve'
                                    WHEN m_locator_id = 1001136 THEN 'HANGAR_'
                                    WHEN m_locator_id = 1001020 THEN 'Depot_Vente'
                                END AS location
                            FROM
                                (
                                    SELECT
                                        d.*,
                                        LEAST(
                                            round(
                                                (((ventef - ((ventef * nvl(rma, 0)) / 100))) - p_revient) / p_revient * 100,
                                                2
                                            ), 
                                            100
                                        ) AS marge
                                    FROM
                                        (
                                            SELECT
                                                det.*,
                                                (det.p_achat - ((det.p_achat * det.rem_achat) / 100)) / (1 + (det.bon_achat / 100)) p_revient,
                                                (
                                                    det.p_vente - ((det.p_vente * nvl(det.rem_vente, 0)) / 100)
                                                ) / (
                                                    1 + (
                                                        CASE
                                                            WHEN det.bna > 0 THEN det.bna
                                                            ELSE det.bon_vente
                                                        END / 100
                                                    )
                                                ) ventef
                                            FROM
                                                (
                                                    SELECT
                                                        p.name product,
                                                        (
                                                            SELECT
                                                                NAME
                                                            FROM
                                                                XX_Laboratory
                                                            WHERE
                                                                XX_Laboratory_id = p.XX_Laboratory_id
                                                        ) labo,
                                                        mst.qtyonhand qty,
                                                        (mst.qtyonhand - mst.QTYRESERVED) qty_dispo,
                                                        mst.m_locator_id,
                                                        mati.value fournisseur,
                                                        mats.guaranteedate,
                                                        md.name remise_auto,
                                                        sal.description bonus_auto,
                                                        md.flatdiscount rma,
                                                        TO_NUMBER(
                                                            CASE
                                                                WHEN REGEXP_LIKE(sal.name, '^[0-9]+$') THEN sal.name
                                                                ELSE NULL
                                                            END
                                                        ) AS bna,
                                                        (
                                                            SELECT
                                                                valuenumber
                                                            FROM
                                                                m_attributeinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                                AND m_attribute_id = 1000501
                                                        ) p_achat,
                                                        (
                                                            SELECT
                                                                valuenumber
                                                            FROM
                                                                m_attributeinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                                AND m_attribute_id = 1001009
                                                        ) rem_achat,
                                                        (
                                                            SELECT
                                                                valuenumber
                                                            FROM
                                                                m_attributeinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                                AND m_attribute_id = 1000808
                                                        ) bon_achat,
                                                        (
                                                            SELECT
                                                                valuenumber
                                                            FROM
                                                                m_attributeinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                                AND m_attribute_id = 1000502
                                                        ) p_vente,
                                                        (
                                                            SELECT
                                                                valuenumber
                                                            FROM
                                                                m_attributeinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                                AND m_attribute_id = 1001408
                                                        ) rem_vente,
                                                        (
                                                            SELECT
                                                                valuenumber
                                                            FROM
                                                                m_attributeinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                                AND m_attribute_id = 1000908
                                                        ) bon_vente,
                                                        (
                                                            SELECT
                                                                lot
                                                            FROM
                                                                m_attributesetinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                        ) lot,
                                                        (
                                                            SELECT
                                                                isactive
                                                            FROM
                                                                m_attributesetinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                        ) lot_active,
                                                        (
                                                            SELECT
                                                                valuenumber
                                                            FROM
                                                                m_attributeinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                                AND m_attribute_id = 1000503
                                                        ) ppa
                                                    FROM
                                                        m_product p
                                                        INNER JOIN m_storage mst ON p.m_product_id = mst.m_product_id
                                                        INNER JOIN m_attributeinstance mati ON mst.m_attributesetinstance_id = mati.m_attributesetinstance_id
                                                        INNER JOIN m_attributesetinstance mats ON mst.m_attributesetinstance_id = mats.m_attributesetinstance_id
                                                        LEFT JOIN C_BPartner_Product cp ON cp.m_product_id = p.m_product_id
                                                            OR cp.C_BPartner_Product_id IS NULL
                                                        LEFT JOIN M_DiscountSchema md ON cp.M_DiscountSchema_id = md.M_DiscountSchema_id
                                                        LEFT JOIN XX_SalesContext sal ON p.XX_SalesContext_ID = sal.XX_SalesContext_ID
                                                    WHERE
                                                        mati.m_attribute_id = 1000508
                                                        AND mst.m_locator_id IN (1001135, 1000614, 1001128, 1001136, 1001020)
                                                        AND mst.qtyonhand != 0
                                                        AND p.name = :product_name
                                                    ORDER BY
                                                        p.name
                                                ) det
                                            WHERE
                                                det.rem_achat < 200
                                        ) d
                                )
                            GROUP BY
                                fournisseur,
                                product,
                                p_achat,
                                p_vente,
                                rem_achat,
                                rem_vente,
                                bon_achat,
                                bon_vente,
                                remise_auto,
                                bonus_auto,
                                p_revient,
                                marge,
                                labo,
                                lot,
                                lot_active,

                                qty,
                                qty_dispo,
                                guaranteedate,
                                ppa,
                                m_locator_id
                            ORDER BY
                                fournisseur
                        ) "source"
                )
            WHERE
                rownum <= 1048575
            """

            cursor.execute(query, {"product_name": product_name})
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]

            return data

    except Exception as e:
        logger.error(f"Error fetching product details: {e}")
        return {"error": "An error occurred while fetching product details."}

@app.route('/fetch-product-details', methods=['GET'])
def fetch_product_details():
    try:
        product_name = request.args.get("product_name", None)
        
        if not product_name:
            return jsonify({"error": "Product name is required"}), 400

        data = fetch_product_details_data(product_name)
        return jsonify(data)

    except Exception as e:
        logger.error(f"Error fetching product details: {e}")
        return jsonify({"error": "Failed to fetch product details"}), 500


@app.route('/download-product-details-excel', methods=['GET'])
def download_product_details_excel():
    try:
        product_name = request.args.get("product_name", None)
        
        if not product_name:
            return jsonify({"error": "Product name is required"}), 400

        # Fetch data from the database
        data = fetch_product_details_data(product_name)

        # Check if data contains an error
        if isinstance(data, dict) and "error" in data:
            return jsonify(data), 500

        # Check if data is empty
        if not data:
            return jsonify({"error": "No data found for the specified product"}), 404

        # Generate Excel file in memory
        excel_output = generate_excel_product_details(data, product_name)

        # Generate filename
        today_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"product_details_{product_name}_{today_date}.xlsx"

        # Send file as a download response
        return send_file(
            excel_output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Error generating product details Excel: {e}")
        return jsonify({"error": "Failed to generate Excel file"}), 500

def generate_excel_product_details(data, product_name):
    """Generate Excel file for product details data"""
    if not data:
        # Create empty workbook if no data
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Details"
        ws.append(["No data available"])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    df = pd.DataFrame(data)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Details - {product_name[:20]}"

    # Define header styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Write headers
    if not df.empty:
        ws.append(df.columns.tolist())
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Apply auto filter
        ws.auto_filter.ref = ws.dimensions

        # Write data rows
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            ws.append(row)

        # Create table
        table = Table(displayName="ProductDetailsTable", ref=ws.dimensions)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


    
def fetch_magasins_from_db(magasin=None, emplacement=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT DISTINCT m.value AS MAGASIN
                FROM M_Locator ml
                JOIN M_Warehouse m ON m.M_WAREHOUSE_ID = ml.M_WAREHOUSE_ID
                WHERE m.ISACTIVE = 'Y'
                  AND m.AD_Client_ID = 1000000
                  AND ml.ISACTIVE = 'Y'
                  AND ml.AD_Client_ID = 1000000
            """

            params = {}

            # Dynamically add filters
            if magasin:
                query += " AND m.value LIKE :magasin || '%'"
                params["magasin"] = magasin

            if emplacement:
                query += " AND ml.value LIKE :emplacement || '%'"
                params["emplacement"] = emplacement

            query += " ORDER BY m.value"

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]

            return data

    except Exception as e:
        logger.error(f"Error fetching magasins: {e}")
        return {"error": "An error occurred while fetching magasins."}
    

@app.route('/fetch-magasins', methods=['GET'])
def fetch_magasins():
    try:
        magasin = request.args.get("magasin", None)
        emplacement = request.args.get("emplacement", None)

        data = fetch_magasins_from_db(magasin, emplacement)
        return jsonify(data)

    except Exception as e:
        logger.error(f"Error fetching magasins: {e}")
        return jsonify({"error": "Failed to fetch magasins"}), 500
    




@app.route('/fetch-emplacements', methods=['GET'])
def fetch_emplacements():
    try:
        magasin = request.args.get("magasin", None)
        emplacement = request.args.get("emplacement", None)

        data = fetch_emplacements_from_db(magasin, emplacement)
        return jsonify(data)

    except Exception as e:
        logger.error(f"Error fetching emplacements: {e}")
        return jsonify({"error": "Failed to fetch emplacements"}), 500
    



def fetch_emplacements_from_db(magasin=None, emplacement=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT ml.value AS EMPLACEMENT
                FROM M_Locator ml
                JOIN M_Warehouse m ON m.M_WAREHOUSE_ID = ml.M_WAREHOUSE_ID
                WHERE m.ISACTIVE = 'Y'
                  AND m.AD_Client_ID = 1000000
                  AND ml.ISACTIVE = 'Y'
                  AND ml.AD_Client_ID = 1000000
            """

            params = {}

            # Dynamically add filters
            if magasin:
                query += " AND m.value LIKE :magasin || '%'"
                params["magasin"] = magasin

            if emplacement:
                query += " AND ml.value LIKE :emplacement || '%'"
                params["emplacement"] = emplacement

            query += " ORDER BY m.value"

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]

            return data

    except Exception as e:
        logger.error(f"Error fetching emplacements: {e}")
        return {"error": "An error occurred while fetching emplacements."}

# ---------- Expiring products endpoint ----------
def parse_ref_date(date_str: str | None) -> datetime:
    """Parse a reference date from query params; defaults to today if missing/invalid."""
    if not date_str:
        return datetime.today()
    # Try a few common formats
    fmts = ["%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"]
    for fmt in fmts:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return datetime.today()


def fetch_expiring_from_db(magasin=None, emplacement=None, fournisseur=None, ref_date=None, within_months=6):
    """
    Fetch products with guaranteed date near expiry.
    Categories: expired, 1mths, 3mths, 6mths (relative to ref_date; default today).
    """
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            # Base query: join storage, product, locator, warehouse, attribute set instance
            query = """
                SELECT 
                    m.m_product_id              AS productid,
                    m.name                      AS product_name,
                    SUM(ms.qtyonhand)           AS qty,
                    set_att.GUARANTEEDATE       AS expire_date,
                    set_att.lot                 AS lot,
                    ml.value                    AS emplacement,
                    mw.value                    AS magasin,
                    sup.value                   AS fournisseur,
                    MAX(pri.valuenumber)        AS price,
                    CASE
                        WHEN set_att.GUARANTEEDATE < :ref_date THEN 'expired'
                        WHEN set_att.GUARANTEEDATE < ADD_MONTHS(:ref_date, 1) THEN '1mths'
                        WHEN set_att.GUARANTEEDATE < ADD_MONTHS(:ref_date, 3) THEN '3mths'
                        WHEN set_att.GUARANTEEDATE < ADD_MONTHS(:ref_date, :within_months) THEN '6mths'
                        ELSE 'later'
                    END AS category
                FROM m_storage ms
                JOIN m_product m ON m.m_product_id = ms.m_product_id
                JOIN m_locator ml ON ml.m_locator_id = ms.m_locator_id
                JOIN m_warehouse mw ON mw.m_warehouse_id = ml.m_warehouse_id
                JOIN m_attributesetinstance set_att ON set_att.m_attributesetinstance_id = ms.m_attributesetinstance_id
                LEFT JOIN m_attributeinstance sup 
                       ON sup.m_attributesetinstance_id = ms.m_attributesetinstance_id
                      AND sup.m_attribute_id = 1000508 -- fournisseur attribute
                LEFT JOIN m_attributeinstance pri
                       ON pri.m_attributesetinstance_id = ms.m_attributesetinstance_id
                      AND pri.m_attribute_id = 1000504 -- prix (unit) attribute
                WHERE 
                    ms.AD_Client_ID = 1000000
                    AND mw.ISACTIVE = 'Y'
                    AND ml.ISACTIVE = 'Y'
                    AND ms.qtyonhand > 0
                    AND set_att.GUARANTEEDATE IS NOT NULL
                    AND set_att.GUARANTEEDATE < ADD_MONTHS(:ref_date, :within_months)
            """

            params = {
                "ref_date": ref_date or datetime.today(),
                "within_months": within_months,
            }

            if magasin:
                query += " AND UPPER(mw.value) LIKE UPPER(:magasin) || '%'"
                params["magasin"] = magasin

            if emplacement:
                query += " AND UPPER(ml.value) LIKE UPPER(:emplacement) || '%'"
                params["emplacement"] = emplacement

            if fournisseur:
                query += " AND UPPER(sup.value) LIKE UPPER(:fournisseur) || '%'"
                params["fournisseur"] = fournisseur

            # Grouping and filtering out 'later' category via WHERE already; group by selected cols
            query += """
                GROUP BY 
                    m.m_product_id, m.name, set_att.GUARANTEEDATE, set_att.lot, ml.value, mw.value, sup.value,
                    CASE
                        WHEN set_att.GUARANTEEDATE < :ref_date THEN 'expired'
                        WHEN set_att.GUARANTEEDATE < ADD_MONTHS(:ref_date, 1) THEN '1mths'
                        WHEN set_att.GUARANTEEDATE < ADD_MONTHS(:ref_date, 3) THEN '3mths'
                        WHEN set_att.GUARANTEEDATE < ADD_MONTHS(:ref_date, :within_months) THEN '6mths'
                        ELSE 'later'
                    END
                HAVING 
                    CASE
                        WHEN set_att.GUARANTEEDATE < :ref_date THEN 'expired'
                        WHEN set_att.GUARANTEEDATE < ADD_MONTHS(:ref_date, 1) THEN '1mths'
                        WHEN set_att.GUARANTEEDATE < ADD_MONTHS(:ref_date, 3) THEN '3mths'
                        WHEN set_att.GUARANTEEDATE < ADD_MONTHS(:ref_date, :within_months) THEN '6mths'
                        ELSE 'later'
                    END <> 'later'
                ORDER BY 
                    CASE 
                        WHEN set_att.GUARANTEEDATE < :ref_date THEN 0
                        WHEN set_att.GUARANTEEDATE < ADD_MONTHS(:ref_date, 1) THEN 1
                        WHEN set_att.GUARANTEEDATE < ADD_MONTHS(:ref_date, 3) THEN 2
                        WHEN set_att.GUARANTEEDATE < ADD_MONTHS(:ref_date, :within_months) THEN 3
                        ELSE 4
                    END,
                    mw.value, ml.value, m.name, set_att.GUARANTEEDATE
            """

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0].lower() for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]

            # Convert datetime to ISO strings for JSON safety
            for item in data:
                if item.get("expire_date") and isinstance(item["expire_date"], (datetime,)):
                    item["expire_date"] = item["expire_date"].strftime("%Y-%m-%d")
            return data

    except Exception as e:
        logger.error(f"Error fetching expiring products: {e}")
        return {"error": "An error occurred while fetching expiring products."}


@app.route('/expiring', methods=['GET'])
def expiring_products():
    """Return products that are expired or will expire within 6 months, categorized."""
    try:
        magasin = request.args.get("magasin")
        emplacement = request.args.get("emplacement")
        fournisseur = request.args.get("fournisseur")
        date_str = request.args.get("date")
        within_months = int(request.args.get("within_months", "6"))

        ref_date = parse_ref_date(date_str)

        data = fetch_expiring_from_db(
            magasin=magasin,
            emplacement=emplacement,
            fournisseur=fournisseur,
            ref_date=ref_date,
            within_months=within_months,
        )

        return jsonify(data)

    except Exception as e:
        logger.error(f"Error in /expiring route: {e}")
        return jsonify({"error": "Failed to fetch expiring products"}), 500






# Helper function to generate Excel file


# Fetch total recap data

# Fetch fournisseur data


def generate_excel_totalrecap_facturation(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Total Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="TotalRecapTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



 


def fetch_bccb_recap_fact(start_date, end_date, fournisseur, product, client, operateur, bccb, zone):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT DOCUMENTNO, DATEORDERED, GRANDTOTAL, 
      round(avg( CASE WHEN marge < 0 THEN 0 ELSE marge END),2) AS marge
FROM (  
    SELECT det.*, 
           ROUND((det.ventef - det.p_revient) / det.p_revient * 100, 2) AS marge
    FROM (  
        SELECT lot.*,  
               (lot.priceentered - ((lot.priceentered * NVL(lot.remise_vente, 0)) / 100)) / (1 + (lot.bonus_vente / 100)) AS ventef
        FROM (  
            SELECT ol.priceentered, 
                   (SELECT valuenumber 
                    FROM m_attributeinstance 
                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                          AND m_attribute_id = 1001519) AS p_revient,
                   
                   (SELECT valuenumber 
                    FROM m_attributeinstance 
                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                          AND m_attribute_id = 1001523) AS bonus_vente,
                   
                   (SELECT valuenumber 
                    FROM m_attributeinstance 
                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                          AND m_attribute_id = 1001532) AS remise_vente,
                   
                   c.DOCUMENTNO, 
                   c.DATEORDERED, 
                   c.GRANDTOTAL
            FROM C_Order c
            JOIN C_OrderLine ol ON c.C_Order_ID = ol.C_Order_ID 
            JOIN M_InOut mi ON mi.C_Order_ID = c.C_Order_ID
            JOIN xx_ca_fournisseur xf ON xf.DOCUMENTNO = mi.DOCUMENTNO
            JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
            JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
            JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
            JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
            WHERE c.AD_Org_ID = 1000012 
                  AND ol.qtyentered > 0
                  AND xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                      AND TO_DATE(:end_date, 'YYYY-MM-DD')
                  AND (c.DOCSTATUS = 'CL' OR c.DOCSTATUS = 'CO')
                  AND c.C_DocType_ID IN (1002845, 1002846)
                  AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                  AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                  AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                  AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                  AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                  AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
            GROUP BY c.DOCUMENTNO, c.DATEORDERED, c.GRANDTOTAL, ol.priceentered, ol.m_attributesetinstance_id
            ORDER BY c.DOCUMENTNO
        )  lot  
    )  det  
)  
group by DOCUMENTNO, DATEORDERED, GRANDTOTAL
order by DOCUMENTNO

            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None
            }
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching BCCB recap: {e}")
        return {"error": "An error occurred while fetching BCCB recap."}

@app.route('/fetchBCCBRecapfact', methods=['GET'])
def fetch_bccb_fact():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    data = fetch_bccb_recap_fact(start_date, end_date, fournisseur, product, client, operateur, bccb, zone)
    return jsonify(data)





















@app.route('/download-BCCB-excel-fac', methods=['GET'])
def download_bccb_excelf():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data
    data = fetch_bccb_recap_fact(start_date, end_date, fournisseur, product, client, operateur, bccb, zone)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"BCCBRecap_Facturation_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_bccbf(data, filename)

def generate_excel_bccbf(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "BCCB Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="BCCBRecapTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")




def fetch_total_recap_achat(start_date, end_date, fournisseur, product):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT
                    SUM(CASE 
                        WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                        ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                    END) AS chiffre
                FROM M_InOut xf
                JOIN M_InOutline mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
                JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
                LEFT JOIN C_InvoiceLine ci ON ci.M_INOUTLINE_ID = mi.M_INOUTLINE_ID
                JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID = mi.M_ATTRIBUTESETINSTANCE_ID
                JOIN M_PRODUCT m ON m.M_PRODUCT_id = mi.M_PRODUCT_id
                WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND ma.M_Attribute_ID = 1000504
                    AND xf.AD_Org_ID = 1000000
                    AND xf.C_DocType_ID IN (1000013, 1000646)
                    AND xf.M_Warehouse_ID IN (1000724, 1000000, 1000720, 1000725)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None
            }

            cursor.execute(query, params)
            row = cursor.fetchone()
            
            return {"chiffre": row[0] if row and row[0] is not None else 0}
    
    except Exception as e:
        logger.error(f"Error fetching total recap achat: {e}")
        return {"error": "An error occurred while fetching total recap achat."}
@app.route('/fetchTotalRecapAchat', methods=['GET'])
def fetch_total_achat():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    data = fetch_total_recap_achat(start_date, end_date, fournisseur, product)
    return jsonify(data)

 
# PRIX DE VENT 1001519   1000084 DOCTYPE  1002854 RETOUR
def fetch_total_recap_achat_fact(start_date, end_date, fournisseur, product):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT
                    SUM(CASE 
                        WHEN xf.C_DocType_ID = 1002854 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                        ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                    END) AS chiffre
                FROM M_InOut xf
                JOIN M_InOutline mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
                JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
                LEFT JOIN C_InvoiceLine ci ON ci.M_INOUTLINE_ID = mi.M_INOUTLINE_ID
                JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID = mi.M_ATTRIBUTESETINSTANCE_ID
                JOIN M_PRODUCT m ON m.M_PRODUCT_id = mi.M_PRODUCT_id
                WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND ma.M_Attribute_ID = 1001519 
                    AND xf.AD_Org_ID = 1000012
                    AND XF.DOCSTATUS!='RE'
                    AND xf.C_DocType_ID IN (1000084, 1002854)
                    AND xf.M_Warehouse_ID IN (1000014, 1000721)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None
            }

            cursor.execute(query, params)
            row = cursor.fetchone()
            
            return {"chiffre": row[0] if row and row[0] is not None else 0}
    
    except Exception as e:
        logger.error(f"Error fetching total recap achat: {e}")
        return {"error": "An error occurred while fetching total recap achat."}
@app.route('/fetchTotalRecapAchat_fact', methods=['GET'])
def fetch_total_achat_fact():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    data = fetch_total_recap_achat_fact(start_date, end_date, fournisseur, product)
    return jsonify(data)

#facturation
def fetch_fournisseur_recap_achat_fact(start_date, end_date, fournisseur, product):
    try:
        # Acquire a connection from the pool
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT
                    CAST(cb.name AS VARCHAR2(300)) AS FOURNISSEUR,   
                    SUM(CASE 
                        WHEN xf.C_DocType_ID = 1002854 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                        ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                    END) AS chiffre,
                    0 AS sort_order
                FROM 
                    M_InOut  xf
                JOIN M_INOUTLINE mi ON mi.M_INOUT_ID=xf.M_INOUT_ID
                JOIN C_BPartner cb ON cb.C_BPARTNER_ID=xf.C_BPARTNER_ID
                LEFT JOIN C_InvoiceLine ci ON ci.M_INOUTLINE_ID=mi.M_INOUTLINE_ID
                JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID=mi.M_ATTRIBUTESETINSTANCE_ID
                JOIN M_PRODUCT m ON m.M_PRODUCT_id=mi.M_PRODUCT_id
                WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND ma.M_Attribute_ID=1001519
                    AND xf.AD_Org_ID = 1000012
                    AND XF.DOCSTATUS!='RE'
                    AND xf.C_DocType_ID IN (1000084, 1002854)
                    AND xf.M_Warehouse_ID IN (1000014, 1000721)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
                GROUP BY 
                    cb.name

                UNION ALL

                SELECT
                    CAST('Total' AS VARCHAR2(300)) AS FOURNISSEUR, 
                    SUM(CASE 
                        WHEN xf.C_DocType_ID = 1002854 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                        ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                    END) AS chiffre,
                    1 AS sort_order
                FROM 
                    M_InOut  xf
                JOIN M_INOUTLINE mi ON mi.M_INOUT_ID=xf.M_INOUT_ID
                JOIN C_BPartner cb ON cb.C_BPARTNER_ID=xf.C_BPARTNER_ID
                LEFT JOIN C_InvoiceLine ci ON ci.M_INOUTLINE_ID=mi.M_INOUTLINE_ID
                JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID=mi.M_ATTRIBUTESETINSTANCE_ID
                JOIN M_PRODUCT m ON m.M_PRODUCT_id=mi.M_PRODUCT_id
                WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND ma.M_Attribute_ID=1001519
                    AND xf.AD_Org_ID = 1000012
                    AND XF.DOCSTATUS!='RE'
                    AND xf.C_DocType_ID IN (1000084, 1002854)
                    AND xf.M_Warehouse_ID IN (1000014, 1000721)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None
            }

            # Execute the query with the provided parameters
            cursor.execute(query, params)

            # Fetch the results
            rows = cursor.fetchall()

            # Format the results into a list of dictionaries
            data = [{"FOURNISSEUR": row[0], "CHIFFRE": row[1], "SORT_ORDER": row[2]} for row in rows]

            return data
    
    except Exception as e:
        logger.error(f"Error fetching fournisseur recap achat: {e}")
        return {"error": "An error occurred while fetching fournisseur recap achat."}

# Flask route to handle the request
@app.route('/fetchfourisseurRecapAchat_fact', methods=['GET'])
def fetch_fournisseur_achat_fact():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    # Ensure both start_date and end_date are provided
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data from the database
    data = fetch_fournisseur_recap_achat_fact(start_date, end_date, fournisseur, product)

    # Return the result as a JSON response
    return jsonify(data)


def fetch_product_achat_recap_fact(start_date, end_date, fournisseur, product):
    try:
        # Acquire a connection from the pool
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT 
                    CAST(m.name AS VARCHAR2(300)) AS produit,   
                    SUM(CASE 
                            WHEN xf.C_DocType_ID = 1002854 THEN -1 * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(mi.QTYENTERED)
                        END) AS qty,
                    
                    SUM(CASE 
                            WHEN xf.C_DocType_ID = 1002854 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                        END) AS chiffre,
                    0 AS sort_order
                FROM 
                    M_InOut xf
                    JOIN M_INOUTLINE mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
                    JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID = mi.M_ATTRIBUTESETINSTANCE_ID
                    JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
                    JOIN M_PRODUCT m ON m.M_PRODUCT_id = mi.M_PRODUCT_id
                WHERE 
                    xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND xf.AD_Org_ID = 1000012
                    AND XF.DOCSTATUS!='RE'
                    AND xf.C_DocType_ID IN (1000084, 1002854)
                    AND ma.M_Attribute_ID = 1001519
                    AND xf.DOCSTATUS = 'CO'
                    AND xf.M_Warehouse_ID IN (1000014, 1000721)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
                GROUP BY 
                    m.name

                UNION ALL

                SELECT
                    CAST('Total' AS VARCHAR2(300)) AS produit, 
                    SUM(CASE 
                            WHEN xf.C_DocType_ID = 1002854 THEN -1 * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(mi.QTYENTERED)
                        END) AS qty,
                    SUM(CASE 
                            WHEN xf.C_DocType_ID = 1002854 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                        END) AS chiffre,
                    1 AS sort_order
                FROM 
                    M_InOut xf
                    JOIN M_INOUTLINE mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
                    JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID = mi.M_ATTRIBUTESETINSTANCE_ID
                    JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
                    JOIN M_PRODUCT m ON m.M_PRODUCT_id = mi.M_PRODUCT_id
                WHERE 
                    xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND xf.AD_Org_ID = 1000012
                    AND XF.DOCSTATUS!='RE'
                    AND xf.C_DocType_ID IN (1000084, 1002854)
                    AND ma.M_Attribute_ID = 1001519
                    AND xf.DOCSTATUS = 'CO'
                    AND xf.M_Warehouse_ID IN (1000014, 1000721)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None
            }

            # Execute the query with the provided parameters
            cursor.execute(query, params)

            # Fetch the results
            rows = cursor.fetchall()

            # Format the results into a list of dictionaries
            data = [{"PRODUIT": row[0], "QTY": row[1], "CHIFFRE": row[2], "SORT_ORDER": row[3]} for row in rows]

            return data
    
    except Exception as e:
        logger.error(f"Error fetching product recap achat: {e}")
        return {"error": "An error occurred while fetching product recap achat."}

@app.route('/fetchProductRecapAchat_fact', methods=['GET'])
def fetch_product_achat_fact():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    # Ensure both start_date and end_date are provided
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data from the database
    data = fetch_product_achat_recap_fact(start_date, end_date, fournisseur, product)

    # Return the result as a JSON response
    return jsonify(data)



def fetch_fournisseur_recap_achat(start_date, end_date, fournisseur, product):
    try:
        # Acquire a connection from the pool
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT
                    CAST(cb.name AS VARCHAR2(300)) AS FOURNISSEUR,   
                    SUM(CASE 
                        WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                        ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                    END) AS chiffre,
                    0 AS sort_order
                FROM 
                    M_InOut  xf
                JOIN M_INOUTLINE mi ON mi.M_INOUT_ID=xf.M_INOUT_ID
                JOIN C_BPartner cb ON cb.C_BPARTNER_ID=xf.C_BPARTNER_ID
                LEFT JOIN C_InvoiceLine ci ON ci.M_INOUTLINE_ID=mi.M_INOUTLINE_ID
                JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID=mi.M_ATTRIBUTESETINSTANCE_ID
                JOIN M_PRODUCT m ON m.M_PRODUCT_id=mi.M_PRODUCT_id
                WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND ma.M_Attribute_ID=1000504
                    AND xf.AD_Org_ID = 1000000
                    AND xf.C_DocType_ID IN (1000013, 1000646)
                    AND xf.M_Warehouse_ID IN (1000724, 1000000, 1000720, 1000725)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
                GROUP BY 
                    cb.name

                UNION ALL

                SELECT
                    CAST('Total' AS VARCHAR2(300)) AS FOURNISSEUR, 
                    SUM(CASE 
                        WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                        ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                    END) AS chiffre,
                    1 AS sort_order
                FROM 
                    M_InOut  xf
                JOIN M_INOUTLINE mi ON mi.M_INOUT_ID=xf.M_INOUT_ID
                JOIN C_BPartner cb ON cb.C_BPARTNER_ID=xf.C_BPARTNER_ID
                LEFT JOIN C_InvoiceLine ci ON ci.M_INOUTLINE_ID=mi.M_INOUTLINE_ID
                JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID=mi.M_ATTRIBUTESETINSTANCE_ID
                JOIN M_PRODUCT m ON m.M_PRODUCT_id=mi.M_PRODUCT_id
                WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND ma.M_Attribute_ID=1000504
                    AND xf.AD_Org_ID = 1000000
                    AND xf.C_DocType_ID IN (1000013, 1000646)
                    AND xf.M_Warehouse_ID IN (1000724, 1000000, 1000720, 1000725)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None
            }

            # Execute the query with the provided parameters
            cursor.execute(query, params)

            # Fetch the results
            rows = cursor.fetchall()

            # Format the results into a list of dictionaries
            data = [{"FOURNISSEUR": row[0], "CHIFFRE": row[1], "SORT_ORDER": row[2]} for row in rows]

            return data
    
    except Exception as e:
        logger.error(f"Error fetching fournisseur recap achat: {e}")
        return {"error": "An error occurred while fetching fournisseur recap achat."}

# Flask route to handle the request
@app.route('/fetchfourisseurRecapAchat', methods=['GET'])
def fetch_fournisseur_achat():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    # Ensure both start_date and end_date are provided
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data from the database
    data = fetch_fournisseur_recap_achat(start_date, end_date, fournisseur, product)

    # Return the result as a JSON response
    return jsonify(data)


def fetch_product_achat_recap(start_date, end_date, fournisseur, product):
    try:
        # Acquire a connection from the pool
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT 
                    CAST(m.name AS VARCHAR2(300)) AS produit,   
                    SUM(CASE 
                            WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(mi.QTYENTERED)
                        END) AS qty,
                    
                    SUM(CASE 
                            WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                        END) AS chiffre,
                    0 AS sort_order
                FROM 
                    M_InOut xf
                    JOIN M_INOUTLINE mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
                    JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID = mi.M_ATTRIBUTESETINSTANCE_ID
                    JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
                    JOIN M_PRODUCT m ON m.M_PRODUCT_id = mi.M_PRODUCT_id
                WHERE 
                    xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND xf.AD_Org_ID = 1000000
                    AND xf.C_DocType_ID IN (1000013, 1000646)
                    AND ma.M_Attribute_ID = 1000504
                    AND xf.DOCSTATUS = 'CO'
                    AND xf.M_Warehouse_ID IN (1000724, 1000000, 1000720, 1000725)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
                GROUP BY 
                    m.name

                UNION ALL

                SELECT
                    CAST('Total' AS VARCHAR2(300)) AS produit, 
                    SUM(CASE 
                            WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(mi.QTYENTERED)
                        END) AS qty,
                    SUM(CASE 
                            WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                        END) AS chiffre,
                    1 AS sort_order
                FROM 
                    M_InOut xf
                    JOIN M_INOUTLINE mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
                    JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID = mi.M_ATTRIBUTESETINSTANCE_ID
                    JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
                    JOIN M_PRODUCT m ON m.M_PRODUCT_id = mi.M_PRODUCT_id
                WHERE 
                    xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND xf.AD_Org_ID = 1000000
                    AND xf.C_DocType_ID IN (1000013, 1000646)
                    AND ma.M_Attribute_ID = 1000504
                    AND xf.DOCSTATUS = 'CO'
                    AND xf.M_Warehouse_ID IN (1000724, 1000000, 1000720, 1000725)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None
            }

            # Execute the query with the provided parameters
            cursor.execute(query, params)

            # Fetch the results
            rows = cursor.fetchall()

            # Format the results into a list of dictionaries
            data = [{"PRODUIT": row[0], "QTY": row[1], "CHIFFRE": row[2], "SORT_ORDER": row[3]} for row in rows]

            return data
    
    except Exception as e:
        logger.error(f"Error fetching product recap achat: {e}")
        return {"error": "An error occurred while fetching product recap achat."}

@app.route('/fetchProductRecapAchat', methods=['GET'])
def fetch_product_achat():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    # Ensure both start_date and end_date are provided
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data from the database
    data = fetch_product_achat_recap(start_date, end_date, fournisseur, product)

    # Return the result as a JSON response
    return jsonify(data)


@app.route('/download-recap-product-achat-excel', methods=['GET'])
def download_recap_product_achat_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
 

    # Ensure both start and end dates are provided
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch the data using the fetch_product_achat_recap function
    data = fetch_product_achat_recap(start_date, end_date, fournisseur, product)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"ProductRecapAchat_{start_date}_to_{end_date}_{today_date}.xlsx"

    # Generate and return the Excel file
    return generate_excel_product_achat(data, filename)

def generate_excel_product_achat(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Convert the data into a DataFrame
    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    # Create an Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers to the sheet
    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table for better presentation
    table = Table(displayName="ProductRecapTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send the file to the client for download
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

  
@app.route('/download-recap-fournisseur-achat-excel', methods=['GET'])
def download_recap_fournisseur_achat_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    # Ensure both start and end dates are provided
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data using the fetch_fournisseur_recap_achat function
    data = fetch_fournisseur_recap_achat(start_date, end_date, fournisseur, product)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"FournisseurRecapAchat_{start_date}_to_{end_date}_{today_date}.xlsx"

    # Generate and return the Excel file
    return generate_excel_fournisseur_achat(data, filename)

def generate_excel_fournisseur_achat(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Convert the data into a DataFrame
    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    # Create an Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Fournisseur Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers to the sheet
    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table for better presentation
    table = Table(displayName="FournisseurRecapAchatTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send the file to the client for download
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


#download achat fact

@app.route('/download-recap-fournisseur-achat_facturation-excel', methods=['GET'])
def download_recap_fournisseur_achatf_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    # Ensure both start_date and end_date are provided
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data using the updated fetch function
    data = fetch_fournisseur_recap_achat_fact(start_date, end_date, fournisseur, product)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"FournisseurRecapAchat_Facturation_{start_date}_to_{end_date}_{today_date}.xlsx"

    # Generate and return the Excel file
    return generate_excel_fournisseur_achatf(data, filename)

def generate_excel_fournisseur_achatf(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Convert the data into a DataFrame
    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    # Create an Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Fournisseur Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers to the sheet
    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table for better presentation
    table = Table(displayName="FournisseurRecapAchatTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send the file to the client for download
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



@app.route('/download-recap-product-achat_facturation-excel', methods=['GET'])
def download_recap_product_achatf_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    # Ensure both start_date and end_date are provided
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data using the updated fetch function
    data = fetch_product_achat_recap_fact(start_date, end_date, fournisseur, product)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"Product_RecapAchat_Facturation_{start_date}_to_{end_date}_{today_date}.xlsx"

    # Generate and return the Excel file
    return generate_excel_prdct_achatf(data, filename)

def generate_excel_prdct_achatf(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Convert the data into a DataFrame
    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    # Create an Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Fournisseur Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers to the sheet
    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table for better presentation
    table = Table(displayName="FournisseurRecapAchatTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send the file to the client for download
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")





def fetch_bccb_productfact(bccb, ad_org_id):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT product, qty, remise, 
                       CASE WHEN marge < 0 THEN 0 ELSE marge END AS marge 
                FROM (
                    SELECT det.*, 
                           ROUND((det.ventef - det.p_revient) / det.p_revient * 100, 2) AS marge 
                    FROM (
                        SELECT lot.*, 
                               (lot.priceentered - ((lot.priceentered * NVL(lot.remise_vente, 0)) / 100)) / 
                               (1 + (lot.bonus_vente / 100)) AS ventef 
                        FROM (
                            SELECT ol.priceentered AS priceentered, 
                                   ol.qtyentered AS qty, 
                                   mp.name AS product, 
                                   ol.discount / 100 AS remise, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1001519) AS p_revient, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1001523) AS bonus_vente, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1001532) AS remise_vente 
                            FROM c_orderline ol 
                            INNER JOIN c_order o ON o.c_order_id = ol.c_order_id 
                            INNER JOIN m_product mp ON ol.m_product_id = mp.m_product_id 
                            WHERE ol.qtyentered > 0 
                                  AND (:bccb IS NULL OR UPPER(o.documentno) LIKE UPPER(:bccb) || '%')
                                  AND o.AD_Org_ID = :ad_org_id
                        ) lot
                    ) det
                )
            """
            
            params = {
                'bccb': bccb or None,
                'ad_org_id': ad_org_id
            }
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching BCCB product data: {e}")
        return {"error": "An error occurred while fetching BCCB product data."}

@app.route('/fetchBCCBProductfact', methods=['GET'])
def fetch_bccb_pf():
    bccb = request.args.get('bccb')
    ad_org_id = request.args.get('ad_org_id')

    if not ad_org_id:
        return jsonify({"error": "Missing ad_org_id parameter"}), 400

    data = fetch_bccb_productfact(bccb, ad_org_id)
    return jsonify(data)



@app.route('/download-bccb-product-excel-f', methods=['GET'])
def download_bccb_product_excelf():
    bccb = request.args.get('bccb')
    ad_org_id = request.args.get('ad_org_id', '1000012')

    if not bccb:
        return jsonify({"error": "Missing BCCB parameter"}), 400

    try:
        ad_org_id = int(ad_org_id)
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_bccb_productfact(bccb, ad_org_id)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"BCCB_Product_Facturation_{bccb}_{today_date}.xlsx"

    return generate_excel_bccb_productf(data, filename)


def generate_excel_bccb_productf(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "BCCB Product Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="BCCBProductTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def fetch_rotation_product_data():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
            select M_product_id as M_PRODUCT_ID, name as NAME from M_PRODUCT
                WHERE AD_Client_ID = 1000000
                AND AD_Org_ID = 1000000
                AND ISACTIVE = 'Y'
                ORDER BY name
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching product data: {e}")
        return {"error": "An error occurred while fetching product data."}
    
@app.route('/fetch-rotation-product-data', methods=['GET'])
def fetch_rotation_product_data_endpoint():
    data = fetch_rotation_product_data()
    return jsonify(data)



def histogram(start_date, end_date, product_id):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
WITH 
date_range AS (
  SELECT 
    CASE 
      WHEN TO_DATE(:end_date, 'YYYY-MM-DD') - TO_DATE(:start_date, 'YYYY-MM-DD') < 30 THEN 
        TO_CHAR(TO_DATE(:start_date, 'YYYY-MM-DD') + LEVEL - 1, 'YYYY-MM-DD')
      ELSE 
        TO_CHAR(ADD_MONTHS(TO_DATE(:start_date, 'YYYY-MM-DD'), LEVEL - 1), 'YYYY-MM')
    END AS invoice_period
  FROM DUAL
  CONNECT BY 
    CASE 
      WHEN TO_DATE(:end_date, 'YYYY-MM-DD') - TO_DATE(:start_date, 'YYYY-MM-DD') < 30 THEN 
        TO_DATE(:start_date, 'YYYY-MM-DD') + LEVEL - 1
      ELSE 
        ADD_MONTHS(TO_DATE(:start_date, 'YYYY-MM-DD'), LEVEL - 1)
    END <= TO_DATE(:end_date, 'YYYY-MM-DD')
),

invoice_data AS (
  SELECT
    dr.invoice_period,
    NVL(SUM(ff.QTYENTERED), 0) AS qty_vendu
  FROM
    date_range dr
  LEFT JOIN
    xx_ca_fournisseur_facture ff
  ON
    CASE 
      WHEN TO_DATE(:end_date, 'YYYY-MM-DD') - TO_DATE(:start_date, 'YYYY-MM-DD') < 30 THEN 
        TO_CHAR(ff.DATEINVOICED, 'YYYY-MM-DD')
      ELSE 
        TO_CHAR(ff.DATEINVOICED, 'YYYY-MM')
    END = dr.invoice_period
    AND ff.DATEINVOICED BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
    AND ff.AD_ORG_ID = 1000000
    AND (:product_id IS NULL OR ff.M_Product_ID = :product_id)
  GROUP BY
    dr.invoice_period
),

purchase_data AS (
  SELECT
    dr.invoice_period,
    SUM(
      CASE 
        WHEN xf.C_DocType_ID = 1000646 THEN -1 * mi.QTYENTERED
        ELSE mi.QTYENTERED
      END
    ) AS qty_acheté
  FROM
    date_range dr
  LEFT JOIN
    M_InOut xf ON
      CASE 
        WHEN TO_DATE(:end_date, 'YYYY-MM-DD') - TO_DATE(:start_date, 'YYYY-MM-DD') < 30 THEN 
          TO_CHAR(xf.MOVEMENTDATE, 'YYYY-MM-DD')
        ELSE 
          TO_CHAR(xf.MOVEMENTDATE, 'YYYY-MM')
      END = dr.invoice_period
    AND xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
    AND xf.AD_Org_ID = 1000000
    AND xf.C_DocType_ID IN (1000013, 1000646)
    AND xf.M_Warehouse_ID != 1000521
  JOIN M_INOUTLINE mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
    AND mi.AD_Org_ID = 1000000
  JOIN C_InvoiceLine ci ON ci.M_INOUTLINE_ID = mi.M_INOUTLINE_ID
  JOIN M_PRODUCT m ON m.M_PRODUCT_ID = mi.M_PRODUCT_ID
    AND m.AD_Org_ID = 1000000
    AND (:product_id IS NULL OR m.M_PRODUCT_ID = :product_id)
  LEFT JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
  GROUP BY
    dr.invoice_period
),

combined_data AS (
  SELECT
    COALESCE(id.invoice_period, pd.invoice_period) AS invoice_period,
    id.qty_vendu,
    pd.qty_acheté,
    1 AS sort_order
  FROM
    invoice_data id
  FULL OUTER JOIN
    purchase_data pd
  ON
    id.invoice_period = pd.invoice_period
)

SELECT
  invoice_period AS period,
  COALESCE(SUM(qty_acheté), 0) AS qty_acheté,
  COALESCE(SUM(qty_vendu), 0) AS qty_vendu
FROM
  combined_data
GROUP BY
  invoice_period,
  sort_order
ORDER BY
  invoice_period
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
                'product_id': product_id or None
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()

            data = [
                {
                    "PERIOD": row[0],
                    "QTY_ACHETÉ": row[1],
                    "QTY_VENDU": row[2]
                }
                for row in rows
            ]

            return data

    except Exception as e:
        logger.error(f"Error fetching histogram data: {e}")
        return {"error": "An error occurred while fetching histogram data."}

@app.route('/histogram', methods=['GET'])
def histogram_route():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    product_id = request.args.get('product_id', '')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    data = histogram(start_date, end_date, product_id)
    return jsonify(data)


def fetch_historique_rotation(product_id):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                WITH Latest_Purchase AS (
                    SELECT 
                        cl.M_Product_ID,
                        SUM(cl.qtyentered) AS last_purchase_qty,
                        c.dateinvoiced,
                        ROW_NUMBER() OVER (PARTITION BY cl.M_Product_ID ORDER BY c.dateinvoiced DESC) AS rn
                    FROM 
                        C_InvoiceLine cl
                    JOIN C_Invoice c ON c.C_Invoice_id = cl.C_Invoice_id
                    JOIN M_INOUTLINE ml ON ml.M_INOUTLINE_id = cl.M_INOUTLINE_ID
                    WHERE 
                        c.dateinvoiced BETWEEN TO_DATE('01/01/2020', 'DD/MM/YYYY') AND SYSDATE
                        AND c.AD_Client_ID = 1000000
                        AND c.AD_Org_ID = 1000000
                        AND c.ISSOTRX = 'N'
                        AND c.DOCSTATUS in ('CO','CL')
                        AND ml.M_Locator_ID != 1001020
                    GROUP BY 
                        cl.M_Product_ID, c.dateinvoiced
                ),
                Filtered_Latest_Purchase AS (
                    SELECT 
                        M_Product_ID,
                        last_purchase_qty,
                        dateinvoiced
                    FROM 
                        Latest_Purchase
                    WHERE 
                        rn = 1
                ),
                On_Hand_Quantity AS (
                    SELECT  
                        m.M_Product_ID as midp,
                        m.name AS product_name,
                        SUM(s.QTYONHAND) - SUM(s.QTYRESERVED) AS QTYONHAND
                    FROM 
                        m_product m
                    JOIN m_storage s ON s.M_PRODUCT_ID = m.M_PRODUCT_ID
                    JOIN M_Locator ml ON ml.M_Locator_ID = s.M_Locator_ID
                    WHERE 
                        s.AD_Client_ID = 1000000
                        AND m.AD_Client_ID = 1000000
                        AND s.M_Locator_ID IN (1001135, 1000614, 1001128, 1001136)
                        AND (:product_id IS NULL OR m.M_PRODUCT_ID = :product_id)
                    GROUP BY 
                        m.M_Product_ID, m.name
                ),
                Stock_Principale AS (
                    SELECT 
                        ROUND(SUM(M_ATTRIBUTEINSTANCE.valuenumber * (m_storage.qtyonhand - m_storage.QTYRESERVED)), 2) AS stock_principale
                    FROM 
                        M_ATTRIBUTEINSTANCE
                    JOIN m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
                    WHERE 
                        M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                        AND m_storage.qtyonhand > 0
                        AND m_storage.M_Locator_ID IN (1001135, 1000614, 1001128, 1001136)
                        AND (:product_id IS NULL OR m_storage.M_Product_ID = :product_id)
                )
                SELECT 
                    oq.midp,
                    oq.product_name,
                    oq.QTYONHAND AS "QTY DISPO",
                    COALESCE(fp.last_purchase_qty, 0) AS "DERNIER ACHAT",
                    fp.dateinvoiced AS "DATE",
                    sp.stock_principale AS "VALEUR"
                FROM 
                    On_Hand_Quantity oq
                LEFT JOIN 
                    Filtered_Latest_Purchase fp ON oq.midp = fp.M_Product_ID
                CROSS JOIN 
                    Stock_Principale sp
                ORDER BY 
                    oq.product_name
            """

            params = {
                'product_id': product_id or None
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()

            data = [
                {
                    "PRODUCT_ID": row[0],
                    "PRODUCT_NAME": row[1],
                    "QTY_DISPO": row[2],
                    "DERNIER_ACHAT": row[3],
                    "DATE": row[4],
                    "VALEUR": row[5]
                }
                for row in rows
            ]

            return data
    
    except Exception as e:
        logger.error(f"Error fetching historique rotation: {e}")
        return {"error": "An error occurred while fetching historique rotation."}

@app.route('/fetchHistoriqueRotation', methods=['GET'])
def fetch_historique():
    product_id = request.args.get('product_id')
    
    # If product_id is "undefined", set it to None
    if product_id == "undefined":
        product_id = None

    data = fetch_historique_rotation(product_id)
    return jsonify(data)




@app.route('/download-rotation-par-mois-excel', methods=['GET'])
def download_rotation_par_mois_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    product_id = request.args.get('product_id', 'All_Products')  # Default if no product is provided
    
    # If product_id is "undefined", set it to None
    if product_id == "undefined":
        product_id = None

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data
    data = rotation_par_mois(start_date, end_date, product_id)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename with product ID, date range, and download date & time
    download_datetime = datetime.now().strftime("%d-%m-%Y_%H-%M")  # Day-Month-Year_Hour-Minute
    sanitized_product = str(product_id).replace(" ", "_").replace("/", "-")  # Replace spaces & slashes
    filename = f"Rotation_{sanitized_product}_{start_date}_to_{end_date}_{download_datetime}.xlsx"

    return generate_excel_rotation(data, filename)


def generate_excel_rotation(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Convert data to DataFrame
    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Rotation Par Mois"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="RotationTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send the file to the client
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def rotation_par_mois(start_date, end_date, product_id):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
WITH date_range AS (
    SELECT 
        CASE 
            WHEN TO_DATE(:end_date, 'YYYY-MM-DD') - TO_DATE(:start_date, 'YYYY-MM-DD') < 30 THEN 
                TO_CHAR(TO_DATE(:start_date, 'YYYY-MM-DD') + LEVEL - 1, 'YYYY-MM-DD')
            ELSE 
                TO_CHAR(ADD_MONTHS(TO_DATE(:start_date, 'YYYY-MM-DD'), LEVEL - 1), 'YYYY-MM')
        END AS invoice_period
    FROM DUAL
    CONNECT BY 
        CASE 
            WHEN TO_DATE(:end_date, 'YYYY-MM-DD') - TO_DATE(:start_date, 'YYYY-MM-DD') < 30 THEN 
                TO_DATE(:start_date, 'YYYY-MM-DD') + LEVEL - 1
            ELSE 
                ADD_MONTHS(TO_DATE(:start_date, 'YYYY-MM-DD'), LEVEL - 1)
        END <= TO_DATE(:end_date, 'YYYY-MM-DD')
),

invoice_data AS (
    SELECT
        dr.invoice_period,
        NVL(SUM(ff.QTYENTERED), 0) AS qty_vendu
    FROM
        date_range dr
    LEFT JOIN
        xx_ca_fournisseur_facture ff
    ON
        CASE 
            WHEN TO_DATE(:end_date, 'YYYY-MM-DD') - TO_DATE(:start_date, 'YYYY-MM-DD') < 30 THEN 
                TO_CHAR(ff.DATEINVOICED, 'YYYY-MM-DD')
            ELSE 
                TO_CHAR(ff.DATEINVOICED, 'YYYY-MM')
        END = dr.invoice_period
        AND ff.DATEINVOICED BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
        AND ff.AD_ORG_ID = 1000000
        AND (:product_id IS NULL OR ff.M_Product_ID = :product_id)
        AND NOT (ff.docstatus = 'RE' or ff.docstatus = 'VO')
    GROUP BY
        dr.invoice_period
),

purchase_data AS (
    SELECT
        dr.invoice_period,
        SUM(
            CASE 
                WHEN xf.C_DocType_ID = 1000646 THEN -1 * mi.QTYENTERED
                ELSE mi.QTYENTERED
            END
        ) AS qty_acheté
    FROM
        date_range dr
    LEFT JOIN
        M_InOut xf ON
            CASE 
                WHEN TO_DATE(:end_date, 'YYYY-MM-DD') - TO_DATE(:start_date, 'YYYY-MM-DD') < 30 THEN 
                    TO_CHAR(xf.MOVEMENTDATE, 'YYYY-MM-DD')
                ELSE 
                    TO_CHAR(xf.MOVEMENTDATE, 'YYYY-MM')
            END = dr.invoice_period
        AND xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
        AND xf.AD_Org_ID = 1000000
        AND xf.C_DocType_ID IN (1000013, 1000646)
        AND xf.M_Warehouse_ID != 1000521
        AND NOT (xf.docstatus = 'RE' OR xf.docstatus = 'VO')
    JOIN M_INOUTLINE mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
        AND mi.AD_Org_ID = 1000000
    left outer JOIN C_InvoiceLine ci ON ci.M_INOUTLINE_ID = mi.M_INOUTLINE_ID
    JOIN M_PRODUCT m ON m.M_PRODUCT_ID = mi.M_PRODUCT_ID
        AND m.AD_Org_ID = 1000000
        AND (:product_id IS NULL OR m.M_PRODUCT_ID = :product_id)
    LEFT JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
    GROUP BY
        dr.invoice_period
),

sold_totals AS (
    SELECT
        'TOTAL' AS invoice_period,
        SUM(qty_vendu) AS qty_vendu,
        NULL AS qty_acheté,
        2 AS sort_order
    FROM
        invoice_data
),
sold_averages AS (
    SELECT
        'MOYENNE' AS invoice_period,
        TRUNC(AVG(qty_vendu)) AS qty_vendu,
        NULL AS qty_acheté,
        3 AS sort_order
    FROM
        invoice_data
),

purchased_totals AS (
    SELECT
        'TOTAL' AS invoice_period,
        NULL AS qty_vendu,
        SUM(qty_acheté) AS qty_acheté,
        2 AS sort_order
    FROM
        purchase_data
),
purchased_averages AS (
    SELECT
        'MOYENNE' AS invoice_period,
        NULL AS qty_vendu,
        TRUNC(AVG(qty_acheté)) AS qty_acheté,
        3 AS sort_order
    FROM
        purchase_data
),

combined_data AS (
    SELECT
        COALESCE(id.invoice_period, pd.invoice_period) AS invoice_period,
        id.qty_vendu,
        pd.qty_acheté,
        1 AS sort_order
    FROM
        invoice_data id
    FULL OUTER JOIN
        purchase_data pd
    ON
        id.invoice_period = pd.invoice_period
),

final_results AS (
    SELECT
        invoice_period,
        COALESCE(qty_vendu, 0) AS qty_vendu,
        COALESCE(qty_acheté, 0) AS qty_acheté,
        sort_order
    FROM
        combined_data
    UNION ALL
    SELECT
        invoice_period,
        qty_vendu,
        qty_acheté,
        sort_order
    FROM
        sold_totals
    UNION ALL
    SELECT
        invoice_period,
        qty_vendu,
        qty_acheté,
        sort_order
    FROM
        sold_averages
    UNION ALL
    SELECT
        invoice_period,
        qty_vendu,
        qty_acheté,
        sort_order
    FROM
        purchased_totals
    UNION ALL
    SELECT
        invoice_period,
        qty_vendu,
        qty_acheté,
        sort_order
    FROM
        purchased_averages
)

SELECT
    invoice_period AS period,
    SUM(qty_vendu) AS qty_vendu,
    SUM(qty_acheté) AS qty_acheté
FROM
    final_results
GROUP BY
    invoice_period,
    sort_order
ORDER BY
    sort_order, invoice_period
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
                'product_id': product_id or None
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()

            data = []
            for row in rows:
                period = row[0]
                qty_vendu = row[1]
                qty_achete = row[2]
                
                # Calculate QTY_INITIAL for each period (except TOTAL and MOYENNE)
                qty_initial = 0
                if period not in ['TOTAL', 'MOYENNE'] and product_id:
                    try:
                        # Convert period to date format for initial stock calculation
                        if '-' in period:
                            if len(period.split('-')) == 2:
                                # Format: "2024-01" -> "2024-01-01"
                                target_date = period + '-01'
                            else:
                                # Format: "2024-01-15" -> use as is
                                target_date = period
                        else:
                            # Handle other formats
                            target_date = period + '-01-01'
                        
                        # Call the existing get_initial_stock_for_date function
                        initial_stock_result = get_initial_stock_for_date(target_date, product_id)
                        
                        if 'error' not in initial_stock_result:
                            qty_initial = initial_stock_result.get('initial_stock', 0)
                        else:
                            logger.warning(f"Error getting initial stock for period {period}: {initial_stock_result['error']}")
                            qty_initial = 0
                    except Exception as e:
                        logger.error(f"Error calculating initial stock for period {period}: {e}")
                        qty_initial = 0
                elif period == 'TOTAL':
                    # For TOTAL, sum up all the initial stocks from regular periods
                    try:
                        total_initial = 0
                        for data_row in data:
                            if data_row['PERIOD'] not in ['TOTAL', 'MOYENNE']:
                                total_initial += data_row.get('QTY_INITIAL', 0)
                        qty_initial = total_initial
                    except Exception as e:
                        logger.error(f"Error calculating total initial stock: {e}")
                        qty_initial = 0
                elif period == 'MOYENNE':
                    # For MOYENNE, calculate average of all initial stocks from regular periods
                    try:
                        regular_periods = [d for d in data if d['PERIOD'] not in ['TOTAL', 'MOYENNE']]
                        if regular_periods:
                            avg_initial = sum(d.get('QTY_INITIAL', 0) for d in regular_periods) / len(regular_periods)
                            qty_initial = round(avg_initial)
                        else:
                            qty_initial = 0
                    except Exception as e:
                        logger.error(f"Error calculating average initial stock: {e}")
                        qty_initial = 0
                
                data.append({
                    "PERIOD": period,
                    "QTY_VENDU": qty_vendu,
                    "QTY_ACHETÉ": qty_achete,
                    "QTY_INITIAL": qty_initial
                })

            return data

    except Exception as e:
        logger.error(f"Error fetching rotation par mois: {e}")
        return {"error": "An error occurred while fetching rotation data."}

@app.route('/rotationParMois', methods=['GET'])
def fetch_rotation():
    product_id = request.args.get('product_id')  # Ensure product_id is received first
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # Debugging
    logger.info(f"Received parameters: product_id={product_id}, start_date={start_date}, end_date={end_date}")

    data = rotation_par_mois(start_date, end_date, product_id)
    return jsonify(data)



@app.route('/fetchZonerotation', methods=['GET'])
def fetch_zone_rotation_endpoint():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    product_id = request.args.get('product_id')
    
    # If product_id is "undefined", set it to None
    if product_id == "undefined":
        product_id = None

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    data = fetch_zone_rotation(start_date, end_date, product_id)
    return jsonify(data)


def fetch_zone_rotation(start_date, end_date, product_id):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT 
                    sr.name AS "ZONE",
                    SUM(xf.qtyentered) AS "QTY"
                FROM C_SalesRegion sr
                JOIN C_BPartner_Location bpl ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                JOIN xx_ca_fournisseur xf ON bpl.C_BPartner_ID = xf.CLIENTID
                WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                      AND TO_DATE(:end_date, 'YYYY-MM-DD')
                      AND xf.AD_Org_ID = 1000000
                      AND xf.DOCSTATUS != 'RE'
                      AND (:product_id IS NULL OR xf.M_Product_ID = :product_id)
                GROUP BY sr.name
                ORDER BY SUM(xf.qtyentered) DESC
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'product_id': product_id or None
            }
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching zone recap: {e}")
        return {"error": "An error occurred while fetching zone recap."}








def journal_vente(start_date, end_date, client):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
            SELECT ci.DocumentNo,
                   ci.DateInvoiced,
                   cb.name as client,
                   CASE WHEN ci.isreturntrx = 'Y'
                        THEN -(ci.GRANDTOTAL - getinvoicetaxamt(ci.c_invoice_id)) 
                        ELSE ci.GRANDTOTAL - getinvoicetaxamt(ci.c_invoice_id) 
                   END as TotalHT,
                   CASE WHEN ci.isreturntrx = 'Y'           
                        THEN -getinvoicetaxamt(ci.c_invoice_id) 
                        ELSE getinvoicetaxamt(ci.c_invoice_id) 
                   END as TotalTVA,
                   CASE WHEN ci.isreturntrx = 'Y'
                        THEN -ci.chargeamt 
                        ELSE ci.chargeamt 
                   END as TotalDT,
                   CASE WHEN ci.isreturntrx = 'Y'
                        THEN -ci.GRANDTOTAL 
                        ELSE ci.GRANDTOTAL 
                   END as TotalTTC,
                   CASE WHEN ci.isreturntrx = 'Y'
                        THEN -ci.GRANDTOTAL - ci.chargeamt 
                        ELSE ci.GRANDTOTAL + ci.chargeamt 
                   END as NETAPAYER,   
                   sr.name as region,
                   adc.name as Entreprise,
                   COALESCE((
                        SELECT sum(tax.taxbaseamt)
                        FROM C_InvoiceTax tax 
                        WHERE tax.C_Invoice_ID = ci.C_Invoice_ID 
                        AND tax.taxamt = 0
                    ), 0) as Montant_Exonere
            FROM C_INVOICE ci
            INNER JOIN AD_CLIENT adc ON (adc.ad_client_id = ci.ad_client_id)
            INNER JOIN AD_ORG ado ON (ado.ad_org_id = ci.ad_org_id)
            INNER JOIN C_BPARTNER cb ON (cb.c_bpartner_id = ci.c_bpartner_id) 
            INNER JOIN C_BPARTNER_Location bpl ON (bpl.C_BPARTNER_Location_id = ci.C_BPARTNER_Location_id) 
            LEFT OUTER JOIN c_salesregion sr ON (bpl.C_SalesRegion_ID = sr.C_SalesRegion_ID)
            WHERE ci.dateInvoiced BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                     AND TO_DATE(:end_date, 'YYYY-MM-DD')
            AND ci.ad_Org_id = 1000012
            AND ci.ISSOTRX = 'Y' 
            AND ci.docstatus IN ('CO', 'CL')
            AND ci.c_doctype_id IN (SELECT c_doctype_id FROM c_doctype WHERE Xx_Excluejournalvente = 'N')
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
            }

            if client:
                query += " AND cb.name LIKE :client"
                params['client'] = f"%{client}%"

            query += " ORDER BY ci.DateInvoiced"

            cursor.execute(query, params)
            rows = cursor.fetchall()

            data = [
                {
                    "DocumentNo": row[0],
                    "DateInvoiced": row[1],
                    "Client": row[2],
                    "TotalHT": row[3],
                    "TotalTVA": row[4],
                    "TotalDT": row[5],
                    "TotalTTC": row[6],
                    "NETAPAYER": row[7],
                    "Region": row[8],
                    "Entreprise": row[9],
                    "Montant_Exonere": row[10]
                }
                for row in rows
            ]

            return data

    except Exception as e:
        logger.error(f"Error fetching journal de vente: {e}")
        return {"error": "An error occurred while fetching sales journal data."}



@app.route('/journalVente', methods=['GET'])
def fetch_journal():
    client = request.args.get('client')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    data = journal_vente(start_date, end_date, client)
    return jsonify(data)

def total_journal(start_date, end_date, client):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
            SELECT 
                SUM(CASE 
                    WHEN ci.isreturntrx = 'Y' THEN -(ci.GRANDTOTAL - getinvoicetaxamt(ci.c_invoice_id)) 
                    ELSE ci.GRANDTOTAL - getinvoicetaxamt(ci.c_invoice_id) 
                END) AS TotalHT,
                SUM(CASE 
                    WHEN ci.isreturntrx = 'Y' THEN -getinvoicetaxamt(ci.c_invoice_id) 
                    ELSE getinvoicetaxamt(ci.c_invoice_id) 
                END) AS TotalTVA,
                SUM(CASE 
                    WHEN ci.isreturntrx = 'Y' THEN -ci.chargeamt 
                    ELSE ci.chargeamt 
                END) AS TotalDT,
                SUM(CASE 
                    WHEN ci.isreturntrx = 'Y' THEN -ci.GRANDTOTAL  
                    ELSE ci.GRANDTOTAL 
                END) AS TotalTTC,
                SUM(CASE 
                    WHEN ci.isreturntrx = 'Y' THEN -ci.GRANDTOTAL - ci.chargeamt
                    ELSE ci.GRANDTOTAL + ci.chargeamt
                END) AS NETAPAYER
            FROM C_INVOICE ci
            INNER JOIN AD_CLIENT adc ON adc.ad_client_id = ci.ad_client_id
            INNER JOIN AD_ORG ado ON ado.ad_org_id = ci.ad_org_id
            INNER JOIN C_BPARTNER cb ON cb.c_bpartner_id = ci.c_bpartner_id 
            INNER JOIN C_BPARTNER_Location bpl ON bpl.C_BPARTNER_Location_id = ci.C_BPARTNER_Location_id 
            LEFT OUTER JOIN C_SALESREGION sr ON bpl.C_SalesRegion_ID = sr.C_SalesRegion_ID
            WHERE ci.dateInvoiced BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                     AND TO_DATE(:end_date, 'YYYY-MM-DD')
            AND ci.ad_Org_id = 1000012
            AND ci.ISSOTRX = 'Y' 
            AND ci.docstatus IN ('CO', 'CL')
            AND ci.c_doctype_id IN (SELECT c_doctype_id FROM c_doctype WHERE Xx_Excluejournalvente = 'N')
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
            }

            if client:
                query += " AND cb.name LIKE :client"
                params['client'] = f"%{client}%"

            cursor.execute(query, params)
            row = cursor.fetchone()

            # Assuming there's only one row of results, we return the aggregated totals
            total_data = {
                "TotalHT": row[0],
                "TotalTVA": row[1],
                "TotalDT": row[2],
                "TotalTTC": row[3],
                "NETAPAYER": row[4]
            }

            return total_data

    except Exception as e:
        logger.error(f"Error fetching total journal data: {e}")
        return {"error": "An error occurred while fetching total journal data."}

@app.route('/totalJournal', methods=['GET'])
def fetch_total_journal():
    client = request.args.get('client')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    total_data = total_journal(start_date, end_date, client)
    return jsonify(total_data)



@app.route('/download-journal-vente-excel', methods=['GET'])
def download_journal_vente_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    client = request.args.get('client', 'All_Clients')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data
    data = journal_vente(start_date, end_date, client)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    download_datetime = datetime.now().strftime("%d-%m-%Y_%H-%M")
    sanitized_client = client.replace(" ", "_").replace("/", "-")
    filename = f"JournalVente_{sanitized_client}_{start_date}_to_{end_date}_{download_datetime}.xlsx"

    return generate_excel_journal(data, filename)


def generate_excel_journal(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Journal Vente"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="JournalVenteTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send the file to the client
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")







# Assuming DB_POOL is a valid database connection pool


def fetch_etat_fournisseur():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            # Static SQL query without dynamic filtering
            query = """
                SELECT 
                    ROUND(SUM(total_echu), 2) AS "TOTAL ECHU",
                    ROUND(SUM(total_dette), 2) AS "TOTAL DETTE",
                    ROUND(SUM(STOCK), 2) AS "TOTAL STOCK"
                FROM (
                    -- TOTAL ECHU Calculation
                    SELECT  
                        SUM(COALESCE(invoiceOpen(inv.C_Invoice_ID, 0), 0)) AS total_echu,
                        0 AS total_dette,
                        0 AS STOCK
                    FROM C_Invoice inv
                    INNER JOIN c_bpartner bp ON bp.c_bpartner_id = inv.c_bpartner_id
                    LEFT OUTER JOIN C_BPARTNER_LOCATION bpl ON bp.c_bpartner_id = bpl.c_bpartner_id
                    LEFT OUTER JOIN C_SalesRegion SR ON SR.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    INNER JOIN C_PAYMENTTERM pt ON inv.C_PaymentTerm_ID = pt.C_PaymentTerm_ID
                    WHERE inv.docstatus IN ('CO', 'CL')
                      AND inv.ad_client_id = 1000000
                      AND inv.ISSOTRX = 'N'
                      AND bp.isactive = 'Y'
                      AND bp.isactive = 'Y'

                      AND bp.isvendor = 'Y'
                      AND COALESCE(invoiceOpen(inv.C_Invoice_ID, 0), 0) <> 0
                      AND inv.AD_Org_ID = 1000000
                      AND inv.AD_Client_ID = 1000000
                      AND (inv.dateinvoiced + pt.netdays) BETWEEN TO_DATE('01/01/2000', 'DD/MM/YYYY') AND SYSDATE
                      AND bp.name NOT LIKE 'solde initial%'

                    UNION ALL

                    -- TOTAL DETTE Calculation
                    SELECT  
                        0 AS total_echu,
                        SUM(cs.grandtotal - cs.verse_fact) AS total_dette,
                        0 AS STOCK
                    FROM (
                        SELECT cs.grandtotal, cs.verse_fact, cs.C_Invoice_ID,
                               ROW_NUMBER() OVER (PARTITION BY cs.name, cs.C_Invoice_ID ORDER BY cs.dateinvoiced DESC) AS rn
                        FROM xx_vendor_status cs
                        WHERE cs.AD_Client_ID = 1000000
                          AND cs.AD_Org_ID = 1000000
                          AND cs.dateinvoiced BETWEEN TO_DATE('01/01/2000', 'DD/MM/YYYY') AND TO_DATE('30/12/3000', 'DD/MM/YYYY')
                          AND cs.name NOT LIKE 'solde initial%'
                    ) cs
                    WHERE cs.rn = 1

                    UNION ALL

                    -- TOTAL STOCK Calculation
                    SELECT 
                        0 AS total_echu,
                        0 AS total_dette,
                        ROUND(SUM(asi.valuenumber * (ms.qtyonhand - ms.QTYRESERVED)), 2) AS STOCK
                    FROM 
                        M_ATTRIBUTEINSTANCE asi
                    JOIN 
                        m_storage ms ON ms.M_ATTRIBUTEsetINSTANCE_id = asi.M_ATTRIBUTEsetINSTANCE_id
                    LEFT JOIN 
                        C_bpartner bp ON bp.c_bpartner_id = ValueNUMBER_of_ASI('Fournisseur', asi.m_attributesetinstance_id)
                    WHERE 
                        asi.M_Attribute_ID = 1000504
                        AND ms.qtyonhand > 0
                        AND ms.m_locator_id IN (1000614, 1001128, 1001135, 1001136)
                        AND bp.name NOT LIKE 'solde initial%'
                ) temp
                HAVING ROUND(SUM(total_echu), 2) <> 0 OR ROUND(SUM(total_dette), 2) <> 0
                ORDER BY "TOTAL DETTE" DESC, "TOTAL ECHU" DESC, "TOTAL STOCK" DESC
            """

            # Execute the query
            cursor.execute(query)
            row = cursor.fetchone()

            # Return the results
            return {
                "TOTAL ECHU": row[0] if row else 0,
                "TOTAL DETTE": row[1] if row else 0,
                "TOTAL STOCK": row[2] if row else 0
            }

    except Exception as e:
        logging.error(f"Error fetching etat fournisseur: {e}")
        return {"error": "An error occurred while fetching etat fournisseur."}



@app.route('/etat_fournisseur', methods=['GET'])
def get_etat_fournisseur():
    result = fetch_etat_fournisseur()
    return jsonify(result)

 

def fetch_fournisseur_dette(fournisseur=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            fournisseur_filter = "AND bp.name LIKE :fournisseur || '%'" if fournisseur else ""
            fournisseur_filter_cs = "AND cs.name LIKE :fournisseur || '%'" if fournisseur else ""

            query = f"""
                SELECT 
                    fournisseur,
                    ROUND(SUM(total_echu), 2) AS "TOTAL ECHU",
                    ROUND(SUM(total_dette), 2) AS "TOTAL DETTE",
                    ROUND(SUM(STOCK), 2) AS "TOTAL STOCK"
                FROM (
                    SELECT  
                        bp.name AS fournisseur,
                        SUM(COALESCE(invoiceOpen(inv.C_Invoice_ID, 0), 0)) AS total_echu,
                        0 AS total_dette,
                        0 AS STOCK
                    FROM C_Invoice inv
                    INNER JOIN c_bpartner bp ON bp.c_bpartner_id = inv.c_bpartner_id
                    LEFT OUTER JOIN C_BPARTNER_LOCATION bpl ON bp.c_bpartner_id = bpl.c_bpartner_id
                    LEFT OUTER JOIN C_SalesRegion SR ON SR.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    INNER JOIN C_PAYMENTTERM pt ON inv.C_PaymentTerm_ID = pt.C_PaymentTerm_ID
                    WHERE inv.docstatus IN ('CO', 'CL')
                      AND inv.ad_client_id = 1000000
                      AND inv.ISSOTRX = 'N'
                      AND bp.isactive = 'Y'
                      AND bp.isvendor = 'Y'
                      AND COALESCE(invoiceOpen(inv.C_Invoice_ID, 0), 0) <> 0
                      AND inv.AD_Org_ID = 1000000
                      AND (inv.dateinvoiced + pt.netdays) BETWEEN TO_DATE('01/01/2020', 'DD/MM/YYYY') AND SYSDATE
                      AND bp.name NOT LIKE 'solde initial%'
                      {fournisseur_filter}
                    GROUP BY bp.name

                    UNION ALL

                    SELECT  
                        cs.name AS fournisseur,
                        0 AS total_echu,
                        SUM(cs.grandtotal - cs.verse_fact) AS total_dette,
                        0 AS STOCK
                    FROM (
                        SELECT cs.name, cs.grandtotal, cs.verse_fact, cs.C_Invoice_ID,
                               ROW_NUMBER() OVER (PARTITION BY cs.name, cs.C_Invoice_ID ORDER BY cs.dateinvoiced DESC) AS rn
                        FROM xx_vendor_status cs
                        WHERE cs.AD_Client_ID = 1000000
                          AND cs.AD_Org_ID = 1000000
                          AND cs.dateinvoiced BETWEEN TO_DATE('01/01/2015', 'DD/MM/YYYY') AND TO_DATE('30/12/3000', 'DD/MM/YYYY')
                          AND cs.name NOT LIKE 'solde initial%'
                          {fournisseur_filter_cs}
                    ) cs
                    WHERE cs.rn = 1
                    GROUP BY cs.name

                    UNION ALL

                    SELECT 
                        bp.name AS fournisseur,
                        0 AS total_echu,
                        0 AS total_dette,
                        ROUND(SUM(asi.valuenumber * (ms.qtyonhand - ms.QTYRESERVED)), 2) AS STOCK
                    FROM 
                        M_ATTRIBUTEINSTANCE asi
                    JOIN 
                        m_storage ms ON ms.M_ATTRIBUTEsetINSTANCE_id = asi.M_ATTRIBUTEsetINSTANCE_id
                    LEFT JOIN 
                        C_bpartner bp ON bp.c_bpartner_id = ValueNUMBER_of_ASI('Fournisseur', asi.m_attributesetinstance_id)
                    WHERE 
                        asi.M_Attribute_ID = 1000504
                        AND ms.qtyonhand > 0
                        AND ms.m_locator_id IN (1000614, 1001128, 1001135, 1001136)
                        AND bp.name NOT LIKE 'solde initial%'
                        {fournisseur_filter}
                    GROUP BY bp.name
                ) temp
                GROUP BY fournisseur
                HAVING ROUND(SUM(total_echu), 2) <> 0 OR ROUND(SUM(total_dette), 2) <> 0
                ORDER BY "TOTAL DETTE" DESC, "TOTAL ECHU" DESC, "TOTAL STOCK" DESC
            """

            params = {'fournisseur': fournisseur} if fournisseur else {}
            cursor.execute(query, params)
            rows = cursor.fetchall()

            result = []
            for row in rows:
                result.append({
                    "FOURNISSEUR": row[0],
                    "TOTAL ECHU": row[1],
                    "TOTAL DETTE": row[2],
                    "TOTAL STOCK": row[3]
                })

            return result

    except Exception as e:
        logging.error(f"Error fetching fournisseur dette: {e}")
        return {"error": "An error occurred while fetching fournisseur dette."}


@app.route('/fetchFournisseurDette', methods=['GET'])
def fetch_fournisseur_dette_api():
    fournisseur = request.args.get('fournisseur')  # Get the fournisseur parameter from the request

    data = fetch_fournisseur_dette(fournisseur)  # Call the function to get the data
    return jsonify(data)  # Return the result as JSON



def fetch_order_confirmed():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            
            query = """
                SELECT * FROM (
                    SELECT 
                        CAST(org.name AS VARCHAR2(300)) AS organisation,
                        CAST(co.documentno AS VARCHAR2(50)) AS ndocument,
                        CAST(cb.name AS VARCHAR2(300)) AS tier,
                        co.dateordered AS datecommande,
                        CAST(us.name AS VARCHAR2(100)) AS vendeur,
                        ROUND(((co.totallines / (SELECT SUM(mat.valuenumber * li.qtyentered) 
                             FROM c_orderline li 
                             INNER JOIN m_attributeinstance mat ON mat.m_attributesetinstance_id = li.m_attributesetinstance_id
                             WHERE mat.m_attribute_id = 1000504 
                               AND li.c_order_id = co.c_order_id 
                               AND li.qtyentered > 0 
                             GROUP BY li.c_order_id)) - 1) * 100, 2) AS marge,
                        ROUND(co.totallines, 2) AS montant,
                        1 AS sort_order
                    FROM 
                        c_order co
                    INNER JOIN ad_org org ON co.ad_org_id = org.ad_org_id
                    INNER JOIN c_bpartner cb ON co.c_bpartner_id = cb.c_bpartner_id
                    INNER JOIN ad_user us ON co.salesrep_id = us.ad_user_id
                    INNER JOIN c_orderconfirm conf ON co.c_order_id = conf.c_order_id
                    WHERE 
                        conf.c_orderconfirm_id IS NOT NULL 
                        AND co.docstatus = 'IP' 
                        AND co.docaction IN ('CO', 'PR') 
                        AND co.c_doctypetarget_id = 1000539 
                        AND co.ad_org_id = 1000000
                    
                    UNION ALL
                    
                    SELECT 
                        CAST('Total' AS VARCHAR2(300)) AS organisation,
                        CAST(NULL AS VARCHAR2(50)) AS ndocument,
                        CAST(NULL AS VARCHAR2(300)) AS tier,
                        NULL AS datecommande,
                        CAST(NULL AS VARCHAR2(100)) AS vendeur,
                        ROUND(AVG(ROUND(((co.totallines / (SELECT SUM(mat.valuenumber * li.qtyentered) 
                             FROM c_orderline li 
                             INNER JOIN m_attributeinstance mat ON mat.m_attributesetinstance_id = li.m_attributesetinstance_id
                             WHERE mat.m_attribute_id = 1000504 
                               AND li.c_order_id = co.c_order_id 
                               AND li.qtyentered > 0 
                             GROUP BY li.c_order_id)) - 1) * 100, 2)), 2) AS marge,
                        ROUND(SUM(co.totallines), 2) AS montant,
                        0 AS sort_order
                    FROM 
                        c_order co
                    INNER JOIN ad_org org ON co.ad_org_id = org.ad_org_id
                    INNER JOIN c_bpartner cb ON co.c_bpartner_id = cb.c_bpartner_id
                    INNER JOIN ad_user us ON co.salesrep_id = us.ad_user_id
                    INNER JOIN c_orderconfirm conf ON co.c_order_id = conf.c_order_id
                    WHERE 
                        conf.c_orderconfirm_id IS NOT NULL 
                        AND co.docstatus = 'IP' 
                        AND co.docaction IN ('CO', 'PR') 
                        AND co.c_doctypetarget_id = 1000539 
                        AND co.ad_org_id = 1000000
                )
                ORDER BY sort_order, montant DESC
            """
            
            cursor.execute(query)
            rows = cursor.fetchall()
            
            columns = [col[0] for col in cursor.description]
            result = [dict(zip(columns, row)) for row in rows]
            
            return result

    except Exception as e:
        logging.error(f"Error fetching order confirmed: {e}")
        return {"error": "An error occurred while fetching order confirmed."}

@app.route('/order_confirmed', methods=['GET'])
def get_order_confirmed():
    result = fetch_order_confirmed()
    return jsonify(result)




# Updated function that returns only the total price
def fetch_total_stock_price():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
            SELECT 
                SUM(M_ATTRIBUTEINSTANCE.valuenumber * m_storage.qtyonhand) AS prix_total
            FROM 
                M_ATTRIBUTEINSTANCE
            JOIN 
                m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
            JOIN 
                M_PRODUCT m ON m.M_PRODUCT_ID = m_storage.M_PRODUCT_ID
            JOIN 
                M_Locator ml ON ml.M_Locator_ID = m_storage.M_Locator_ID
            INNER JOIN 
                m_attributeinstance mati ON m_storage.M_ATTRIBUTEsetINSTANCE_id = mati.M_ATTRIBUTEsetINSTANCE_id
            WHERE 
                M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                AND m_storage.qtyonhand > 0
                AND mati.m_attribute_id = 1000508
                AND m_storage.AD_Client_ID = 1000000
                AND m_storage.M_Locator_ID IN (
                    SELECT M_Locator_ID 
                    FROM M_Locator 
                    WHERE M_Warehouse_ID IN (
                        SELECT M_Warehouse_ID 
                        FROM M_Warehouse 
                        WHERE VALUE IN (
                            'HANGAR', '1-Dépôt Principal', '8-Dépot réserve', '88-Dépot Hangar réserve'
                        )
                    )
                )
            """

            cursor.execute(query)
            result = cursor.fetchone()

            return {"prix_total": result[0] if result else 0}

    except Exception as e:
        logger.error(f"Database error: {e}")
        return {"error": "An error occurred while fetching total price."}



@app.route('/total-stock', methods=['GET'])
def get_total_stock_price():
    try:
        total_data = fetch_total_stock_price()
        return jsonify(total_data)
    except Exception as e:
        logger.error(f"Error fetching total stock price: {e}")
        return jsonify({"error": "Failed to fetch total price"}), 500





@app.route('/download-quota-excel', methods=['GET'])
def download_quota_excel():
    data = fetch_quota_product()  # Fetch all data
    produit_filter = request.args.get("produit", "").strip()

    # Apply filter if provided (already applied inside fetch_quota_product, but double check here)
    if produit_filter:
        data = [row for row in data if produit_filter.lower() in row["NAME"].lower()]

    return generate_excel_quota(data, produit_filter)

#----------------------------
def generate_excel_quota(data, produit_filter):
    if not data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered Quota"

    # Apply header styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(df.columns.tolist())
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Append data with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Generate filename dynamically based on filter
    today_date = datetime.now().strftime("%d-%m-%Y")
    filter_text = f"produit-{produit_filter}" if produit_filter else ""
    filename = f"Quota_{filter_text}_{today_date}.xlsx" if filter_text else f"Quota_{today_date}.xlsx"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Fetch product quota data from Oracle DB
def fetch_quota_product():
    try:
        produit = request.args.get('produit', '')  # Get 'produit' param from URL, default to empty string

        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            # Base SQL with optional filtering
            query = """
                SELECT m.name, SUM(s.QTYONHAND) AS QTY , sum(s.QTYONHAND * at.valuenumber) as Prix
                
                FROM m_product m

                JOIN m_storage s ON s.M_PRODUCT_ID = m.M_PRODUCT_ID
                INNER JOIN m_attributeinstance at
                ON at.m_attributesetinstance_id = s.m_attributesetinstance_id



                WHERE m.XX_SalesContext_ID = 1000100
                    AND m.AD_Client_ID = 1000000
                    AND s.QTYONHAND > 0
                    AND s.M_Locator_ID = 1000614
                    and at.M_Attribute_ID=1000502
            """

            # If produit filter is provided, add it to the query
            if produit:
                query += " AND LOWER(m.name) LIKE :produit"

            query += """
                GROUP BY m.name
                ORDER BY m.name
            """

            # Bind parameter if filtering
            if produit:
                cursor.execute(query, {"produit": produit.lower() + '%'})
            else:
                cursor.execute(query)

            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            return data

    except Exception as e:
        logger.error(f"Error fetching product quota data: {e}")
        return {"error": "An error occurred while fetching product quota data."}





@app.route('/quota-product', methods=['GET'])
def quota_product():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_quota_product()
    return jsonify(data)



def fetch_quota_operator():
    try:
        produit = request.args.get('produit', '')

        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            # SQL query with optional filtering
            query = """
                SELECT a.name, SUM(q.quantity - q.QTYORDERED) AS qty
                FROM AD_USER a
                JOIN XX_Quota q ON q.AD_USER_id = a.AD_USER_id
                JOIN M_PRODUCT m ON m.M_PRODUCT_id = q.M_PRODUCT_ID
                WHERE q.M_Warehouse_ID = 1000000
            """

            # Apply filter if produit is provided
            if produit:
                query += " AND LOWER(m.name) LIKE :produit"

            query += """
                GROUP BY a.name
                ORDER BY qty DESC
            """

            # Execute with binding
            if produit:
                cursor.execute(query, {"produit": f"%{produit.lower()}%"})
            else:
                cursor.execute(query)

            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            return data

    except Exception as e:
        logger.error(f"Error fetching operator quota data: {e}")
        return {"error": "An error occurred while fetching operator quota data."}

@app.route('/quota-operator', methods=['GET'])
def quota_operator():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_quota_operator()
    return jsonify(data)










from flask import request, jsonify

@app.route('/fetchFournisseurDataByYear', methods=['GET'])
def fetch_fournisseur_by_year():

    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    years = request.args.getlist('years')  # Get all years parameters
    month = request.args.get('month')

    try:
        years = [int(y) for y in years] if years else None
        month = int(month) if month else None
    except ValueError:
        return jsonify({"error": "Invalid year or month format"}), 400

    if not years:
        # If no years specified, get current year
        current_year = datetime.now().year
        years = [current_year]

    full_data = {}
    for year in years:
        yearly_data = fetch_fournisseur_data_by_year(
            fournisseur=fournisseur,
            product=product,
            client=client,
            operateur=operateur,
            bccb=bccb,
            zone=zone,
            year=year,
            month=month
        )
        full_data[str(year)] = yearly_data

    return jsonify(full_data)



def fetch_fournisseur_data_by_year(fournisseur=None, product=None, client=None, operateur=None,
                                   bccb=None, zone=None, year=None, month=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            ad_org_id = 1000000

            if year and month:
                start_date = f"{year}-{month:02d}-01"
                last_day = 31 if month in [1, 3, 5, 7, 8, 10, 12] else 30
                if month == 2:
                    last_day = 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28
                end_date = f"{year}-{month:02d}-{last_day}"
            elif year:
                start_date = f"{year}-01-01"
                end_date = f"{year}-12-31"
            else:
                current_year = datetime.now().year
                start_date = f"{current_year}-01-01"
                end_date = f"{current_year}-12-31"
                year = current_year

            query = """
                SELECT 
                    TO_CHAR(xf.MOVEMENTDATE, 'YYYY') AS year,
                    TO_CHAR(xf.MOVEMENTDATE, 'MM') AS month,
                    SUM(xf.TOTALLINE) AS total,
                    SUM(xf.qtyentered) AS QTY,
                    ROUND(
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            WHEN SUM(xf.CONSOMATION) < 0 THEN ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION) * -1) * 100
                            ELSE ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION)) * 100
                        END, 4) AS marge
                FROM xx_ca_fournisseur xf
                JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                JOIN C_ORDER C ON mi.C_ORDER_ID = c.C_ORDER_ID
                WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND xf.AD_Org_ID = :ad_org_id
                    AND xf.DOCSTATUS != 'RE'
                    AND (:fournisseur IS NULL OR xf.name LIKE :fournisseur || '%')
                    AND (:product IS NULL OR xf.product LIKE :product || '%')
                    AND (:client IS NULL OR cb.name LIKE :client || '%')
                    AND (:operateur IS NULL OR au.name LIKE :operateur || '%')
                    AND (:bccb IS NULL OR C.DOCUMENTNO LIKE :bccb || '%')
                    AND (:zone IS NULL OR sr.name LIKE :zone || '%')
                GROUP BY ROLLUP(TO_CHAR(xf.MOVEMENTDATE, 'YYYY'), TO_CHAR(xf.MOVEMENTDATE, 'MM'))
                HAVING TO_CHAR(xf.MOVEMENTDATE, 'YYYY') = :year_str OR TO_CHAR(xf.MOVEMENTDATE, 'YYYY') IS NULL
                ORDER BY year NULLS LAST, month NULLS LAST
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'ad_org_id': ad_org_id,
                'year_str': str(year)
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in rows]

    except Exception as e:
        logger.error(f"Error fetching total fournisseur data: {e}")
        return {"error": "An error occurred while fetching total data."}

@app.route('/listproduct')
def listproduct():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
select name from M_PRODUCT
WHERE AD_Client_ID = 1000000
AND AD_Org_ID = 1000000
  AND ISACTIVE = 'Y'
ORDER BY name


            """
            cursor.execute(query)
            result = [row[0] for row in cursor.fetchall()]
            return jsonify(result)
    except Exception as e:
        logger.error(f"Error fetching products: {e}")
        return jsonify({"error": "Could not fetch products list"}), 500
    






@app.route('/listfournisseur')
def list_fournisseur():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT cb.name
                FROM C_BPartner cb
                WHERE cb.AD_Client_ID = 1000000
                  AND cb.ISVENDOR = 'Y'
                  AND cb.ISACTIVE = 'Y'
                ORDER BY cb.name
            """
            cursor.execute(query)
            result = [row[0] for row in cursor.fetchall()]
            return jsonify(result)
    except Exception as e:
        logger.error(f"Error fetching fournisseurs: {e}")
        return jsonify({"error": "Could not fetch fournisseur list"}), 500


@app.route('/listregion')
def list_region():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT name
                FROM C_SalesRegion
                WHERE ISACTIVE = 'Y'
                  AND AD_Client_ID = 1000000
            """
            cursor.execute(query)
            result = [row[0] for row in cursor.fetchall()]
            return jsonify(result)
    except Exception as e:
        logger.error(f"Error fetching regions: {e}")
        return jsonify({"error": "Could not fetch region list"}), 500


@app.route('/listclient')
def list_client():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT name
                FROM C_BPartner
                WHERE iscustomer = 'Y'
                  AND AD_Client_ID = 1000000
                  AND AD_Org_ID = 1000000
                ORDER BY name
            """
            cursor.execute(query)
            result = [row[0] for row in cursor.fetchall()]
            return jsonify(result)
    except Exception as e:
        logger.error(f"Error fetching clients: {e}")
        return jsonify({"error": "Could not fetch client list"}), 500



@app.route('/fetchFournisseurRecapAchatByYear', methods=['GET'])
def fetch_fournisseur_recap_achat_by_year():
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    years = request.args.getlist('years')  # Get all years parameters
    month = request.args.get('month')

    try:
        years = [int(y) for y in years] if years else None
        month = int(month) if month else None
    except ValueError:
        return jsonify({"error": "Invalid year or month format"}), 400

    if not years:
        # If no years specified, get current year
        current_year = datetime.now().year
        years = [current_year]

    full_data = {}
    for year in years:
        yearly_data = fetch_fournisseur_recap_achat_data_by_year(
            fournisseur=fournisseur,
            product=product,
            year=year,
            month=month
        )
        full_data[str(year)] = yearly_data

    return jsonify(full_data)

def fetch_fournisseur_recap_achat_data_by_year(fournisseur=None, product=None, year=None, month=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            if year and month:
                start_date = f"{year}-{month:02d}-01"
                last_day = 31 if month in [1, 3, 5, 7, 8, 10, 12] else 30
                if month == 2:
                    last_day = 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28
                end_date = f"{year}-{month:02d}-{last_day}"
            elif year:
                start_date = f"{year}-01-01"
                end_date = f"{year}-12-31"
            else:
                current_year = datetime.now().year
                start_date = f"{current_year}-01-01"
                end_date = f"{current_year}-12-31"
                year = current_year

            query = """
                SELECT 
                    TO_CHAR(xf.MOVEMENTDATE, 'YYYY') AS year,
                    TO_CHAR(xf.MOVEMENTDATE, 'MM') AS month,
                    SUM(CASE 
                        WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                        ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                    END) AS chiffre,
                    SUM(CASE 
                        WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(mi.QTYENTERED)
                        ELSE TO_NUMBER(mi.QTYENTERED)
                    END) AS qty
                FROM 
                    M_InOut  xf
                JOIN M_INOUTLINE mi ON mi.M_INOUT_ID=xf.M_INOUT_ID
                JOIN C_BPartner cb ON cb.C_BPARTNER_ID=xf.C_BPARTNER_ID
                LEFT JOIN C_InvoiceLine ci ON ci.M_INOUTLINE_ID=mi.M_INOUTLINE_ID
                JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID=mi.M_ATTRIBUTESETINSTANCE_ID
                JOIN M_PRODUCT m ON m.M_PRODUCT_id=mi.M_PRODUCT_id
                WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND ma.M_Attribute_ID=1000504
                    AND xf.AD_Org_ID = 1000000
                    AND xf.C_DocType_ID IN (1000013, 1000646)
                    AND xf.M_Warehouse_ID IN (1000724, 1000000, 1000720, 1000725)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
                GROUP BY ROLLUP(TO_CHAR(xf.MOVEMENTDATE, 'YYYY'), TO_CHAR(xf.MOVEMENTDATE, 'MM'))
                HAVING TO_CHAR(xf.MOVEMENTDATE, 'YYYY') = :year_str OR TO_CHAR(xf.MOVEMENTDATE, 'YYYY') IS NULL
                ORDER BY year NULLS LAST, month NULLS LAST
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'year_str': str(year)
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in rows]

    except Exception as e:
        logger.error(f"Error fetching fournisseur recap achat data by year: {e}")
        return {"error": "An error occurred while fetching recap achat data by year."}

@app.route('/download-fournisseur-recap-excel', methods=['GET'])
def download_fournisseur_recap_excel():
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    years = request.args.getlist('years')
    month = request.args.get('month')

    try:
        years = [int(y) for y in years] if years else []
        month = int(month) if month else None
    except ValueError:
        return {"error": "Invalid year or month format"}, 400

    if not years:
        years = [datetime.now().year]

    all_data = []
    for year in years:
        data = fetch_fournisseur_recap_achat_data_by_year(
            fournisseur=fournisseur,
            product=product,
            year=year,
            month=month
        )
        for row in data:
            row['YEAR'] = year  # Add year for clarity if not in row
        all_data.extend(data)

    return generate_excel_fournisseur_recap(all_data, fournisseur, product, years, month)


def generate_excel_fournisseur_recap(data, fournisseur, product, years, month):
    if not data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    wb = Workbook()
    ws = wb.active
    ws.title = "Fournisseur Recap"

    # Style header
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(df.columns.tolist())
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Append data rows with alternating colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Filename generation
    today = datetime.now().strftime("%d-%m-%Y")
    filters = []
    if fournisseur:
        filters.append(f"fournisseur-{fournisseur}")
    if product:
        filters.append(f"product-{product}")
    if years:
        filters.append(f"years-{'-'.join(map(str, years))}")
    if month:
        filters.append(f"month-{month}")
    
    filter_text = "_".join(filters)
    filename = f"FournisseurRecap_{filter_text}_{today}.xlsx" if filters else f"FournisseurRecap_{today}.xlsx"

    # Return file
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

FEEDBACK_FILE = 'json_files/feedback.json'

@app.route('/submit_feedback', methods=['POST'])
def submit_feedback():
    data = request.get_json()
    data['timestamp'] = datetime.now().isoformat()

    # Load existing feedback
    if os.path.exists(FEEDBACK_FILE):
        with open(FEEDBACK_FILE, 'r') as f:
            feedback_list = json.load(f)
    else:
        feedback_list = []

    feedback_list.append(data)

    # Save back to JSON
    with open(FEEDBACK_FILE, 'w') as f:
        json.dump(feedback_list, f, indent=2)

    return jsonify({'status': 'success'}), 200


def get_last_recouvrement_value():
    """Get the last recouvrement value from the MySQL database"""
    try:
        conn = get_localdb_connection()
        if not conn:
            logger.error("Failed to connect to MySQL database for last recouvrement")
            return 481114494.36  # Default fallback value
        
        cursor = conn.cursor()
        
        try:
            # Get the last recouvrement value from the table
            cursor.execute("""
                SELECT value 
                FROM recouvrement 
                ORDER BY date DESC 
                LIMIT 1
            """)
            
            row = cursor.fetchone()
            if row:
                return float(row[0])
            else:
                return 481114494.36  # Default fallback value
                
        except Exception as e:
            logger.error(f"Error fetching last recouvrement data: {e}")
            return 481114494.36  # Default fallback value
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        logger.error(f"Error in get_last_recouvrement_value: {e}")
        return 481114494.36  # Default fallback value

@app.route('/api/recouvrement/last-value', methods=['GET'])
def last_recouvrement_value():
    """
    API endpoint to get the last recouvrement value from database
    Returns JSON response with the value or default fallback
    """
    try:
        value = get_last_recouvrement_value()
        return jsonify({
            'status': 'success',
            'value': value,
            'message': 'Last recouvrement value retrieved successfully'
        }), 200
    except Exception as e:
        logger.error(f"Error in last_recouvrement_value endpoint: {e}")
        return jsonify({
            'status': 'error',
            'value': 481114494.36,
            'message': 'Failed to retrieve recouvrement value, using fallback'
        }), 500

@app.route('/recouvrement', methods=['GET'])
def recouvrement():
    # Get the current goal from the database instead of using a static value
    current_goal = get_last_recouvrement_value()
    
    try:
        # Rest of your existing query code remains the same
        today = datetime.now()
        
        # If it's before 12:00 PM on the first day of the month, use the previous month
        if today.hour < 12 and today.day == 1:
            # Use previous month
            if today.month == 1:
                year = today.year - 1
                month = 12
            else:
                year = today.year
                month = today.month - 1
        else:
            # Use current month (after 12:00 PM or not the first day)
            year = today.year
            month = today.month
            
        start_date = f"{year}-{month:02d}-01"
        last_day = calendar.monthrange(year, month)[1]
        end_date = f"{year}-{month:02d}-{last_day}"

        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT 
                    ROUND(:objectif, 2) AS "OBJECTIF_MENSUEL", 
                    ROUND(SUM(p.payamt), 2) AS "TOTAL_RECOUVREMENT",
                    ROUND(SUM(p.payamt)/:objectif, 4) AS "POURCENTAGE"
                FROM 
                    C_Payment p
                    JOIN C_BPartner b ON b.C_BPartner_id = p.C_BPartner_id
                WHERE 
                    b.iscustomer = 'Y'
                    AND b.C_PaymentTerm_ID != 1000000
                    AND p.AD_Client_ID = 1000000
                    AND p.docstatus in ('CO','CL')
                    AND p.ZSubPaymentRule_ID in (1000007,1000016)
                    AND p.datetrx >= TO_DATE(:start_date, 'YYYY-MM-DD')
                    AND p.datetrx <= TO_DATE(:end_date, 'YYYY-MM-DD')
            """

            cursor.execute(query, {
                'objectif': current_goal,
                'start_date': start_date,
                'end_date': end_date
            })

            row = cursor.fetchone()
            if not row:
                return jsonify({
                    "OBJECTIF_MENSUEL": current_goal,
                    "TOTAL_RECOUVREMENT": 0,
                    "POURCENTAGE": 0
                })
                
            columns = [col[0] for col in cursor.description]
            result = dict(zip(columns, row))
            
            # Ensure we handle null values
            if result["TOTAL_RECOUVREMENT"] is None:
                result["TOTAL_RECOUVREMENT"] = 0
            if result["POURCENTAGE"] is None:
                result["POURCENTAGE"] = 0
                
            return jsonify(result)

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({
            "error": "Erreur serveur",
            "OBJECTIF_MENSUEL": current_goal,
            "TOTAL_RECOUVREMENT": 0,
            "POURCENTAGE": 0
        }), 500





@app.route('/pending_documents', methods=['GET'])
def pending_documents():
    try:
        # Get sales_region parameter from request, default to None (fetch all)
        sales_region = request.args.get('sales_region')
        
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT 
                    IO.DOCUMENTNO, 
                    IO.CREATED, 
                    IO.DESCRIPTION, 
                    IO.M_InOut_ID, 
                    BP.NAME,
                    SR.NAME AS SALES_REGION
                FROM 
                    M_InOut IO
                    JOIN C_BPartner BP ON IO.C_BPartner_ID = BP.C_BPartner_ID
                    JOIN C_BPartner_Location BPL ON IO.C_BPartner_ID = BPL.C_BPartner_ID
                    JOIN C_SalesRegion SR ON BPL.C_SalesRegion_ID = SR.C_SalesRegion_ID
                WHERE 
                    IO.ISACTIVE = 'Y'
                    AND IO.DOCACTION = 'CO'
                    AND IO.DOCSTATUS = 'IP'
                    AND IO.PROCESSED = 'N'
                    AND IO.ISAPPROVED = 'N'
                    AND IO.XX_CONTROLEUR_ID IS NULL
                    AND IO.XX_PREPARATEUR_ID IS NULL
                    AND IO.XX_CONTROLEUR_CH_ID IS NULL
                    AND IO.XX_PREPARATEUR_CH_ID IS NULL
                    AND IO.XX_EMBALEUR_CH_ID IS NULL
                    AND IO.XX_EMBALEUR_ID IS NULL
                    AND IO.C_DocType_ID='1002733'
            """
            
            # Add sales region filter if provided
            params = []
            if sales_region:
                query += " AND SR.NAME = :1"
                params.append(sales_region)
            
            cursor.execute(query, params)
            results = cursor.fetchall()
            
            documents = []
            for row in results:
                doc_no, created, description, inoutid, bp_name, sales_region = row
                # Format date as "26 May 2025 09:58"
                formatted_date = created.strftime('%d %b %Y %H:%M')
                documents.append({
                    'documentNo': doc_no,
                    'createdDate': formatted_date,
                    'description': description if description else '-',
                    'inoutid': inoutid,
                    'businessPartner': bp_name,
                    'salesRegion': sales_region,
                    'status': 'Created'
                })
            
            return jsonify({
                "count": len(documents),
                "documents": documents,
                "lastUpdated": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "salesRegionFilter": sales_region if sales_region else "All"
            })

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({
            "error": "Server error",
            "message": str(e),
            "count": 0,
            "documents": []
        }), 500



@app.route('/regions', methods=['GET'])
def select_regions():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """

SELECT C_SALESREGION_ID, NAME 
                FROM C_SalesRegion
                     where ISACTIVE= 'Y'
                     and AD_Client_ID = 1000000
                ORDER BY NAME
           
            """
            cursor.execute(query)
            results = cursor.fetchall()
            
            regions = []
            for row in results:
                region_id, name = row
                regions.append({
                    'id': region_id,
                    'name': name
                })
            
            return jsonify({
                "count": len(regions),
                "regions": regions
            })

    except Exception as e:
        print(f"Error fetching regions: {e}")
        return jsonify({
            "error": "Server error",
            "message": str(e),
            "count": 0,
            "regions": []
        }), 500


@app.route('/inout_lines', methods=['GET'])
def get_inout_lines():
    try:
        m_inout_id = request.args.get('m_inout_id')
        
        if not m_inout_id:
            return jsonify({
                "error": "Missing parameter",
                "message": "m_inout_id parameter is required"
            }), 400

        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT P.NAME, ML.MOVEMENTQTY 
                FROM M_InOutline ML
                JOIN M_Product P ON ML.M_Product_ID = P.M_Product_ID
                WHERE ML.M_InOut_ID = :m_inout_id
            """
            cursor.execute(query, {'m_inout_id': m_inout_id})
            results = cursor.fetchall()
            
            lines = []
            for name, movement_qty in results:
                lines.append({
                    'product_name': name,
                    'quantity': float(movement_qty) if movement_qty else 0.0
                })
            
            return jsonify({
                "m_inout_id": m_inout_id,
                "count": len(lines),
                "lines": lines
            })

    except Exception as e:
        print(f"Error fetching inout lines: {e}")
        return jsonify({
            "error": "Server error",
            "message": str(e),
            "count": 0,
            "lines": []
        }), 500

        












def fetch_total_stock_by_location():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                WITH stock_principale_data AS (
                    SELECT 
                        ROUND(SUM(M_ATTRIBUTEINSTANCE.valuenumber * m_storage.qtyonhand), 2) AS stock_principale
                    FROM 
                        M_ATTRIBUTEINSTANCE
                    JOIN 
                        m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
                    WHERE 
                        M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                        AND m_storage.qtyonhand > 0
                        AND m_storage.m_locator_id = 1000614
                ),
                hangar_data AS (
                    SELECT 
                        ROUND(SUM(M_ATTRIBUTEINSTANCE.valuenumber * m_storage.qtyonhand), 2) AS hangar
                    FROM 
                        M_ATTRIBUTEINSTANCE
                    JOIN 
                        m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
                    WHERE 
                        M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                        AND m_storage.qtyonhand > 0
                        AND m_storage.m_locator_id = 1001135
                ),
                hangarresrev_data AS (
                    SELECT 
                        ROUND(SUM(M_ATTRIBUTEINSTANCE.valuenumber * m_storage.qtyonhand), 2) AS hangarresrev
                    FROM 
                        M_ATTRIBUTEINSTANCE
                    JOIN 
                        m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
                    WHERE 
                        M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                        AND m_storage.qtyonhand > 0
                        AND m_storage.m_locator_id = 1001136
                ),
                depot_reserver_data AS (
                    SELECT 
                        ROUND(SUM(M_ATTRIBUTEINSTANCE.valuenumber * m_storage.qtyonhand), 2) AS depot_reserver
                    FROM 
                        M_ATTRIBUTEINSTANCE
                    JOIN 
                        m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
                    WHERE 
                        M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                        AND m_storage.qtyonhand > 0
                        AND m_storage.m_locator_id = 1001128
                )

                SELECT 
                    ROUND(SUM(sp.stock_principale), 2) AS STOCK_principale,
                    ROUND(SUM(h.hangar), 2) AS hangar,
                    ROUND(SUM(hr.hangarresrev), 2) AS hangarréserve,
                    ROUND(SUM(dr.depot_reserver), 2) AS depot_reserver,
                    ROUND(
                        NVL(SUM(sp.stock_principale), 0) + 
                        NVL(SUM(h.hangar), 0) + 
                        NVL(SUM(hr.hangarresrev), 0) + 
                        NVL(SUM(dr.depot_reserver), 0), 
                        2
                    ) AS total_stock
                FROM 
                    stock_principale_data sp,
                    hangar_data h,
                    hangarresrev_data hr,
                    depot_reserver_data dr
            """

            cursor.execute(query)
            result = cursor.fetchone()

            return {
                "STOCK_principale": result[0] or 0,
                "hangar": result[1] or 0,
                "hangarréserve": result[2] or 0,
                "depot_reserver": result[3] or 0,
                "total_stock": result[4] or 0
            }

    except Exception as e:
        logger.error(f"Database error: {e}")
        return {"error": "An error occurred while fetching stock data."}



@app.route('/stock-summary', methods=['GET'])
def get_stock_summary():
    try:
        stock_data = fetch_total_stock_by_location()
        return jsonify(stock_data)
    except Exception as e:
        logger.error(f"Error fetching stock summary: {e}")
        return jsonify({"error": "Failed to fetch stock summary"}), 500






def fetch_credit_client():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT 
                    ROUND(SUM(SoldeFact + SoldeBL), 2) AS credit_client
                FROM (
                    SELECT 
                        bp.c_bpartner_id, 
                        (
                            SELECT COALESCE(SUM(invoiceOpen(inv.C_Invoice_ID, 0)), 0) 
                            FROM C_Invoice inv 
                            WHERE bp.c_bpartner_id = inv.c_bpartner_id 
                            AND paymentTermDueDays(inv.C_PAYMENTTERM_ID, inv.DATEINVOICED, TO_DATE('01/01/3000', 'DD/MM/YYYY')) >= 0
                            AND inv.docstatus IN ('CO', 'CL') 
                            AND inv.AD_ORGTRX_ID = inv.ad_org_id 
                            AND inv.ad_client_id = 1000000
                            AND inv.C_PaymentTerm_ID != 1000000
                        ) AS SoldeFact,
                        (
                            SELECT COALESCE(SUM(invoiceOpen(inv.C_Invoice_ID, 0)), 0) 
                            FROM C_Invoice inv 
                            WHERE bp.c_bpartner_id = inv.c_bpartner_id 
                            AND paymentTermDueDays(inv.C_PAYMENTTERM_ID, inv.DATEINVOICED, TO_DATE('01/01/3000', 'DD/MM/YYYY')) >= 0
                            AND inv.docstatus IN ('CO', 'CL') 
                            AND inv.AD_ORGTRX_ID <> inv.ad_org_id 
                            AND inv.ad_client_id = 1000000
                            AND inv.C_PaymentTerm_ID != 1000000
                        ) AS SoldeBL
                    FROM 
                        c_bpartner bp 
                        INNER JOIN C_BPartner_Location bpl ON bp.C_BPartner_id = bpl.C_BPartner_id
                        INNER JOIN ad_user u ON bp.salesrep_id = u.ad_user_id
                        LEFT OUTER JOIN AD_User u2 ON u2.AD_User_ID = bp.XX_TempSalesRep_ID
                        LEFT OUTER JOIN C_SalesRegion sr ON sr.C_SalesRegion_id = bpl.C_SalesRegion_id
                        LEFT OUTER JOIN C_City sr2 ON sr2.C_City_id = bpl.C_City_id
                    WHERE 
                        bp.iscustomer = 'Y' 
                        AND bp.C_BP_Group_ID IN (1000003, 1000926, 1001330)
                        AND bp.isactive = 'Y' 
                        AND sr.isactive = 'Y'
                        AND bp.C_PaymentTerm_ID != 1000000
                        AND bpl.c_salesregion_id IN (
                            101, 102, 1000032, 1001776, 1001777, 1001778, 1001779, 1001780, 1001781,
                            1001782, 1001783, 1001784, 1001785, 1001786, 1001787, 1001788, 1001789,
                            1001790, 1001791, 1001792, 1001793, 1001794, 1002076, 1002077, 1002078,
                            1002079, 1002080, 1002176, 1002177, 1002178, 1002179, 1002180, 1002181,
                            1002283, 1002285, 1002286, 1002287, 1002288
                        )
                ) subquery
            """

            cursor.execute(query)
            row = cursor.fetchone()
            credit_client = row[0] if row else 0.0
            return {"credit_client": credit_client}

    except Exception as e:
        logger.error(f"Error fetching credit client data: {e}")
        return {"error": "An error occurred while fetching credit client data."}




@app.route('/credit-client', methods=['GET'])
def credit_client():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_credit_client()
    return jsonify(data)

def fetch_caisse():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT 
                    ROUND(endingbalance, 2) AS caisse
                FROM 
                    C_BankStatement
                WHERE 
                    C_BankAccount_ID = 1000205
                    AND docstatus = 'CO'
                    AND AD_Client_ID = 1000000
                ORDER BY 
                    statementdate DESC
                FETCH FIRST 1 ROW ONLY
            """

            cursor.execute(query)
            row = cursor.fetchone()
            caisse = row[0] if row else 0.0
            return {"caisse": caisse}

    except Exception as e:
        logger.error(f"Error fetching caisse data: {e}")
        return {"error": "An error occurred while fetching caisse data."}


@app.route('/caisse', methods=['GET'])
def caisse():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_caisse()
    return jsonify(data)



def fourniseurdettfond():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT 
                    ROUND(SUM(invopenamt), 2) AS credit_fournisseur
                FROM (
                    SELECT 
                        COALESCE(invoiceOpen(inv.C_Invoice_ID, 0), 0) AS invopenamt
                    FROM
                        C_Invoice inv
                        INNER JOIN c_bpartner bp ON bp.c_bpartner_id = inv.c_bpartner_id
                        LEFT OUTER JOIN C_BPARTNER_LOCATION bpl ON bp.c_bpartner_id = bpl.c_bpartner_id
                        LEFT OUTER JOIN C_SalesRegion SR ON SR.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                        INNER JOIN C_PAYMENTTERM pt ON inv.C_PaymentTerm_ID = pt.C_PaymentTerm_ID
                        LEFT OUTER JOIN ad_user usr ON usr.ad_user_id = inv.salesrep_id
                        LEFT OUTER JOIN ad_user usr2 ON usr2.ad_user_id = bp.salesrep_id
                        LEFT OUTER JOIN C_City ct ON ct.C_City_ID = bpl.C_City_ID
                    WHERE
                        inv.docstatus IN ('CO', 'CL')
                        AND inv.ad_client_id = 1000000
                        --AND inv.ISSOTRX = 'N'
                        AND bp.isactive = 'Y'
                        AND bp.isvendor = 'Y'
                )
            """

            cursor.execute(query)
            result = cursor.fetchone()
            return result[0] if result else 0.0

    except Exception as e:
        print(f"An error occurred: {e}")
        return None


@app.route('/fourniseurdettfond', methods=['GET'])
def get_fourniseurdettfond():
    result = fourniseurdettfond()
    return jsonify({"value": result})  # Now the frontend gets a key



# Path to store the previous values
DATA_FILE = os.path.join(os.path.dirname(__file__), 'kpi_data.json')

# Initialize data file if it doesn't exist
if not os.path.exists(DATA_FILE):
    with open(DATA_FILE, 'w') as f:
        json.dump({"fonds_propre": []}, f)

@app.route('/previous-fonds-propre', methods=['GET'])
def get_previous_fonds_propre():
    try:
        with open(DATA_FILE, 'r') as f:
            data = json.load(f)
        
        last_value = data['fonds_propre'][-1] if data['fonds_propre'] else None
        return jsonify(last_value)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/store-fonds-propre', methods=['POST'])
def store_fonds_propre():
    try:
        with open(DATA_FILE, 'r') as f:
            data = json.load(f)
        
        new_data = request.get_json()
        value = new_data.get('value')
        date = new_data.get('date', datetime.now().isoformat())
        
        # Keep only the last 12 months of data
        if len(data['fonds_propre']) >= 12:
            data['fonds_propre'].pop(0)
        
        # Ensure that the data is stored correctly
        data['fonds_propre'].append({"value": value, "date": date})
        
        with open(DATA_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/fonds-propre-chart-data')
def get_fonds_propre_chart_data():
    try:
        with open(DATA_FILE, 'r') as f:
            data = json.load(f)
        
        sorted_data = sorted(data['fonds_propre'], key=lambda x: x['date'])
        labels = [entry['date'] for entry in sorted_data]
        values = [entry['value'] for entry in sorted_data]
        
        return jsonify({
            "labels": labels,
            "values": values,
            "min_date": labels[0] if labels else None,
            "max_date": labels[-1] if labels else None
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/dette-fournisseur-chart-data')
def get_dette_fournisseur_chart_data():
    try:
        with open(DETTE_FILE, 'r') as f:
            data = json.load(f)
        
        sorted_data = sorted(data['dette_fournisseur'], key=lambda x: x['date'])
        labels = [entry['date'] for entry in sorted_data]
        values = [entry['value'] for entry in sorted_data]
        
        return jsonify({
            "labels": labels,
            "values": values,
            "min_date": labels[0] if labels else None,
            "max_date": labels[-1] if labels else None
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Add similar endpoints for other metrics (credit-client, caisse, tresorerie)


# Update your existing DATA_FILE path or create a new one
DETTE_FILE = 'kpi_dette.json'

if not os.path.exists(DETTE_FILE):
    with open(DETTE_FILE, 'w') as f:
        json.dump({"dette_fournisseur": []}, f)

@app.route('/previous-dette-fournisseur', methods=['GET'])
def get_previous_dette_fournisseur():
    try:
        with open(DETTE_FILE, 'r') as f:
            data = json.load(f)
        
        last_value = data['dette_fournisseur'][-1] if data['dette_fournisseur'] else None
        return jsonify(last_value)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/store-dette-fournisseur', methods=['POST'])
def store_dette_fournisseur():
    try:
        with open(DETTE_FILE, 'r') as f:
            data = json.load(f)
        
        new_data = request.get_json()
        value = new_data.get('value')
        date = new_data.get('date', datetime.now().isoformat())
        
        # Keep only the last 12 months of data
        if len(data['dette_fournisseur']) >= 12:
            data['dette_fournisseur'].pop(0)
        
        data['dette_fournisseur'].append({"value": value, "date": date})
        
        with open(DETTE_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


STOCK_FILE = 'kpi_stock.json'

if not os.path.exists(STOCK_FILE):
    with open(STOCK_FILE, 'w') as f:
        json.dump({
            "total_stock": [],
            "stock_principale": [],
            "hangar": [],
            "hangar_reserve": [],
            "depot_reserver": []
        }, f)

@app.route('/previous-stock/<stock_type>', methods=['GET'])
def get_previous_stock(stock_type):
    try:
        with open(STOCK_FILE, 'r') as f:
            data = json.load(f)
        
        key = {
            'total': 'total_stock',
            'principale': 'stock_principale',
            'hangar': 'hangar',
            'hangar_reserve': 'hangar_reserve',
            'depot': 'depot_reserver'
        }.get(stock_type)
        
        if not key:
            return jsonify({"error": "Invalid stock type"}), 400
            
        last_value = data[key][-1] if data[key] else None
        return jsonify(last_value)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/store-stock', methods=['POST'])
def store_stock():
    try:
        with open(STOCK_FILE, 'r') as f:
            data = json.load(f)
        
        new_data = request.get_json()
        
        # Store all stock values
        stock_types = ['total_stock', 'stock_principale', 'hangar', 'hangar_reserve', 'depot_reserver']
        current_date = datetime.now().isoformat()
        
        for stype in stock_types:
            value = new_data.get(stype)
            if value is not None:
                # Keep only the last 12 months of data
                if len(data[stype]) >= 12:
                    data[stype].pop(0)
                
                data[stype].append({
                    "value": value,
                    "date": current_date
                })
        
        with open(STOCK_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
        
# Add this to your Flask app
CREDIT_FILE = 'kpi_credit.json'

if not os.path.exists(CREDIT_FILE):
    with open(CREDIT_FILE, 'w') as f:
        json.dump({"credit_client": []}, f)

@app.route('/previous-credit-client', methods=['GET'])
def get_previous_credit_client():
    try:
        with open(CREDIT_FILE, 'r') as f:
            data = json.load(f)
        
        last_value = data['credit_client'][-1] if data['credit_client'] else None
        return jsonify(last_value)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/store-credit-client', methods=['POST'])
def store_credit_client():
    try:
        with open(CREDIT_FILE, 'r') as f:
            data = json.load(f)
        
        new_data = request.get_json()
        value = new_data.get('value')
        date = new_data.get('date', datetime.now().isoformat())
        
        # Keep only the last 12 months of data
        if len(data['credit_client']) >= 12:
            data['credit_client'].pop(0)
        
        data['credit_client'].append({"value": value, "date": date})
        
        with open(CREDIT_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    

CAISSE_FILE = 'kpi_caisse.json'

if not os.path.exists(CAISSE_FILE):
    with open(CAISSE_FILE, 'w') as f:
        json.dump({"caisse": []}, f)

@app.route('/previous-caisse', methods=['GET'])
def get_previous_caisse():
    try:
        with open(CAISSE_FILE, 'r') as f:
            data = json.load(f)
        
        last_value = data['caisse'][-1] if data['caisse'] else None
        return jsonify(last_value)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/store-caisse', methods=['POST'])
def store_caisse():
    try:
        with open(CAISSE_FILE, 'r') as f:
            data = json.load(f)
        
        new_data = request.get_json()
        value = new_data.get('value')
        date = new_data.get('date', datetime.now().isoformat())
        
        # Keep only the last 12 months of data
        if len(data['caisse']) >= 12:
            data['caisse'].pop(0)
        
        data['caisse'].append({"value": value, "date": date})
        
        with open(CAISSE_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/test-caisse', methods=['GET'])
def test_caisse():
    try:
        # Return a fixed value of 10000 as JSON
        return jsonify({"caisse": 10000})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/test-fournisseur-dette', methods=['GET'])
def test_fournisseur_dette():
    try:
        # Return a fixed test value of 20000 as JSON
        return jsonify({"value": 20000})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


TRESORERIE_FILE = 'kpi_tresorerie.json'

if not os.path.exists(TRESORERIE_FILE):
    with open(TRESORERIE_FILE, 'w') as f:
        json.dump({"tresorerie": []}, f)

@app.route('/previous-tresorerie', methods=['GET'])
def get_previous_tresorerie():
    try:
        with open(TRESORERIE_FILE, 'r') as f:
            data = json.load(f)
        
        last_value = data['tresorerie'][-1] if data['tresorerie'] else None
        return jsonify(last_value)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/store-tresorerie', methods=['POST'])
def store_tresorerie():
    try:
        with open(TRESORERIE_FILE, 'r') as f:
            data = json.load(f)
        
        new_data = request.get_json()
        value = new_data.get('value')
        date = new_data.get('date', datetime.now().isoformat())
        
        # Keep only the last 12 months of data
        if len(data['tresorerie']) >= 12:
            data['tresorerie'].pop(0)
        
        data['tresorerie'].append({"value": value, "date": date})
        
        with open(TRESORERIE_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/credit-client-chart-data')
def get_credit_client_chart_data():
    try:
        with open(CREDIT_FILE, 'r') as f:
            data = json.load(f)

        sorted_data = sorted(data['credit_client'], key=lambda x: x['date'])
        labels = [entry['date'] for entry in sorted_data]
        values = [entry['value'] for entry in sorted_data]

        return jsonify({
            "labels": labels,
            "values": values,
            "min_date": labels[0] if labels else None,
            "max_date": labels[-1] if labels else None
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
@app.route('/caisse-chart-data')
def get_caisse_chart_data():
    try:
        with open('kpi_caisse.json', 'r') as f:
            data = json.load(f)

        sorted_data = sorted(data['caisse'], key=lambda x: x['date'])
        labels = [entry['date'] for entry in sorted_data]
        values = [entry['value'] for entry in sorted_data]

        return jsonify({
            "labels": labels,
            "values": values,
            "min_date": labels[0] if labels else None,
            "max_date": labels[-1] if labels else None
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
@app.route('/tresorerie-chart-data')
def get_tresorerie_chart_data():
    try:
        with open('kpi_tresorerie.json', 'r') as f:
            data = json.load(f)

        sorted_data = sorted(data['tresorerie'], key=lambda x: x['date'])
        labels = [entry['date'] for entry in sorted_data]
        values = [entry['value'] for entry in sorted_data]

        return jsonify({
            "labels": labels,
            "values": values,
            "min_date": labels[0] if labels else None,
            "max_date": labels[-1] if labels else None
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/combined-stock-chart-data')
def get_combined_stock_chart_data():
    try:
        with open(STOCK_FILE, 'r') as f:
            data = json.load(f)
        
        # Get all dates from all stock types
        all_dates = set()
        for stock_type in ['total_stock', 'stock_principale', 'hangar', 'hangar_reserve', 'depot_reserver']:
            for entry in data[stock_type]:
                all_dates.add(entry['date'])
        
        sorted_dates = sorted(all_dates)
        
        # Prepare datasets for each stock type
        datasets = []
        colors = ['#4e73df', '#e74a3b', '#f6c23e', '#1cc88a', '#36b9cc']
        stock_types = [
            ('total_stock', 'Total Stock'),
            ('stock_principale', 'Stock Principale'),
            ('hangar', 'Hangar'),
            ('hangar_reserve', 'Hangar Reserve'),
            ('depot_reserver', 'Dépôt Reserve')
        ]
        
        for i, (stock_type, label) in enumerate(stock_types):
            values = []
            stock_data = {entry['date']: entry['value'] for entry in data[stock_type]}
            
            for date in sorted_dates:
                values.append(stock_data.get(date, None))
            
            datasets.append({
                "label": label,
                "data": values,
                "borderColor": colors[i],
                "backgroundColor": f"{colors[i]}20",
                "pointBackgroundColor": colors[i],
                "pointBorderColor": '#fff',
                "pointHoverRadius": 5,
                "pointHoverBackgroundColor": colors[i],
                "pointHoverBorderColor": '#fff',
                "pointHitRadius": 10,
                "pointBorderWidth": 2,
                "borderWidth": 2,
                "tension": 0.3,
                "fill": i == 0  # Only fill the first dataset for better visibility
            })
        
        return jsonify({
            "labels": sorted_dates,
            "datasets": datasets,
            "min_date": sorted_dates[0] if sorted_dates else None,
            "max_date": sorted_dates[-1] if sorted_dates else None
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def fetch_paiment():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
SELECT
    NVL(SUM(
        CASE 
            WHEN z.name = 'Encaiss: Espèces' THEN ROUND(ci.PayAmt, 2)
            WHEN z.name = 'Décaiss: Espèces' THEN -ROUND(ci.PayAmt, 2)
            ELSE 0 
        END
    ), 0) AS total_difference
FROM 
    C_Payment ci
    INNER JOIN ZSubPaymentRule z ON ci.ZSubPaymentRule_ID = z.ZSubPaymentRule_ID
WHERE 
    TRUNC(ci.DATETRX) = TRUNC(SYSDATE)
    AND ci.DOCACTION IN ('CO', 'CL', 'co', 'cl')
    AND ci.AD_Client_ID = 1000000
    AND z.name IN ('Encaiss: Espèces', 'Décaiss: Espèces')


            """

            cursor.execute(query)
            row = cursor.fetchone()
            paiment = row[0] if row else 0.0
            return {"Total_Paiment": paiment}

    except Exception as e:
        logger.error(f"Error fetching paiment data: {e}")
        return {"error": "An error occurred while fetching paiment data."}


@app.route('/paiement-net', methods=['GET'])
def paiment():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_paiment()
    return jsonify(data)

def get_total_checks():
    try:
        # Read the bank.json file
        with open('bank.json', 'r') as f:
            bank_data = json.load(f)
            
        # Get the latest entry
        latest_entry = bank_data[-1]
        
        # Get individual check amounts
        bna_checks = latest_entry['bna_check']
        baraka_checks = latest_entry['baraka_check']
        
        # Calculate total checks
        total_checks = bna_checks + baraka_checks
        
        return {
            "total_checks": total_checks,
            "details": {
                "BNA": {
                    "checks": bna_checks
                },
                "Baraka": {
                    "checks": baraka_checks
                }
            }
        }
    except Exception as e:
        logger.error(f"Error calculating total checks: {e}")
        return {"error": "An error occurred while calculating total checks"}

@app.route('/total-checks', methods=['GET'])
def total_checks():
    try:
        result = get_total_checks()
        return jsonify(result)
    except Exception as e:
        logger.error(f"Error in total checks route: {e}")
        return jsonify({"error": "Failed to get total checks"}), 500


def get_total_bank():
    try:
        # Read the bank.json file
        with open('bank.json', 'r') as f:
            bank_data = json.load(f)
            
        # Get the latest entry
        latest_entry = bank_data[-1]
        
        # Calculate individual bank amounts
        bna_total = latest_entry['bna_sold'] + latest_entry['bna_remise']
        baraka_total = latest_entry['baraka_sold'] + latest_entry['baraka_remise']
        
        # Calculate total bank amount
        total_bank = bna_total + baraka_total
        
        return {
            "total_bank": total_bank,
            "details": {
                "BNA": {
                    "sold": latest_entry['bna_sold'],
                    "remise": latest_entry['bna_remise'],
                    "total": bna_total
                },
                "Baraka": {
                    "sold": latest_entry['baraka_sold'],
                    "remise": latest_entry['baraka_remise'],
                    "total": baraka_total
                }
            }
        }
    except Exception as e:
        logger.error(f"Error calculating total bank: {str(e)}")
        return {"error": "An error occurred while calculating total bank"}

@app.route('/total-bank', methods=['GET'])
def total_bank():
    try:
        result = get_total_bank()
        return jsonify(result)
    except Exception as e:
        logger.error(f"Error in total bank route: {e}")
        return jsonify({"error": "Failed to get total bank"}), 500



JSON_FILE = 'json_files/all_performance_data.json'

def get_filtered_data(start_date=None, end_date=None):
    try:
        with open(JSON_FILE, 'r') as f:
            data = json.load(f)

        # Convert dates if provided
        start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00')) if start_date else None
        end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00')) if end_date else None

        # Filter entries if dates provided
        entries = []
        for timestamp, entry in data.items():
            entry_dt = datetime.fromisoformat(entry['timestamp'].replace('Z', '+00:00'))
            if ((not start_dt or entry_dt >= start_dt) and 
                (not end_dt or entry_dt <= end_dt)):
                entries.append(entry)

        # Sort entries by timestamp
        entries.sort(key=lambda x: x['timestamp'])

        return entries

    except Exception as e:
        print(f"Error reading data: {e}")
        return []

@app.route('/kpi-trends-data', methods=['GET'])
def get_kpi_trends_data():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        entries = get_filtered_data(start_date, end_date)
        
        if not entries:
            return jsonify({
                "error": "No data found for the specified period"
            }), 404

        # Extract data series
        labels = [entry['timestamp'] for entry in entries]
        dette = [float(entry['dette']['total']) if 'dette' in entry else None for entry in entries]
        stock = [float(entry['stock']['total']) if 'stock' in entry else None for entry in entries]
        tresorerie = [float(entry['tresorerie']['total']) if 'tresorerie' in entry else None for entry in entries]
        credit = [float(entry['credit_client']) if 'credit_client' in entry else None for entry in entries]

        # Build consolidated response
        datasets = [
            {
                "label": "Dette Fournisseur",
                "data": dette,
                "borderColor": "#ff0033",
                "backgroundColor": "rgba(255, 0, 51, 0.1)",
                "tension": 0.3,
                "borderWidth": 2
            },
            {
                "label": "Stock Total",
                "data": stock,
                "borderColor": "#1cc88a",
                "backgroundColor": "rgba(28, 200, 138, 0.1)",
                "tension": 0.3,
                "borderWidth": 2
            },
            {
                "label": "Trésorerie",
                "data": tresorerie,
                "borderColor": "#36b9cc",
                "backgroundColor": "rgba(54, 185, 204, 0.1)",
                "tension": 0.3,
                "borderWidth": 2
            },
            {
                "label": "Crédit Client",
                "data": credit,
                "borderColor": "#f6c23e",
                "backgroundColor": "rgba(246, 194, 62, 0.1)",
                "tension": 0.3,
                "borderWidth": 2
            }
        ]

        return jsonify({
            "labels": labels,
            "datasets": datasets,
            "min_date": labels[0] if labels else None,
            "max_date": labels[-1] if labels else None
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# Stock Movement functionality
def calculate_stock_initial_for_multiple_emplacements(start_date, product, fournisseur, v2_mode=False):
    """
    Calculate stock initial for multiple emplacements by summing each emplacement individually
    This solves the issue when emplacement is empty and we need to aggregate stock from all main locations
    Returns both total and breakdown by emplacement
    """
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            
            # Define the main emplacements
            emplacements = ['Préparation', 'HANGAR', 'Dépot Hangar réserve', 'Dépot réserve']
            total_stock_initial = 0
            emplacement_breakdown = {}
            
            # Determine the exclusion clause based on V2 mode
            if v2_mode:
                exclusion_clause = "WHEN COALESCE(inv.docstatus, m.docstatus, io.docstatus) IN ('RE','VO')"
            else:
                exclusion_clause = "WHEN COALESCE(inv.docstatus, m.docstatus, io.docstatus) IN ('RE')"
            
            for emp in emplacements:
                query = f"""
                SELECT SUM(
                    CASE
                        {exclusion_clause}
                        THEN 0
                        ELSE s.movementqty
                    END
                ) AS initial_stock
                FROM m_transaction s
                INNER JOIN m_product p ON (s.m_product_id = p.m_product_id)
                INNER JOIN m_locator l ON (l.m_locator_id = s.m_locator_id)
                INNER JOIN M_attributeinstance att ON (att.m_attributesetinstance_id = s.m_attributesetinstance_id)
                LEFT OUTER JOIN M_InventoryLine il ON (s.M_InventoryLine_ID=il.M_InventoryLine_ID)
                LEFT OUTER JOIN M_Inventory inv ON (inv.m_inventory_id = il.m_inventory_id)
                LEFT OUTER JOIN M_MovementLine ml ON (s.M_MovementLine_ID=ml.M_MovementLine_ID)
                LEFT OUTER JOIN M_Movement m ON (m.M_Movement_ID=ml.M_Movement_ID)
                LEFT OUTER JOIN M_InOutLine iol ON (s.M_InOutLine_ID=iol.M_InOutLine_ID)
                LEFT OUTER JOIN M_Inout io ON (iol.M_InOut_ID=io.M_InOut_ID)
                WHERE s.movementdate < TO_DATE(:start_date, 'YYYY-MM-DD')
                AND (:product IS NULL OR p.M_Product_ID = :product)
                AND (:fournisseur IS NULL OR att.value LIKE :fournisseur || '%')
                AND l.value LIKE :emplacement || '%'
                AND att.m_attribute_id = 1000508
                AND s.AD_Client_ID = 1000000
                """
                
                params = {
                    'start_date': start_date,
                    'product': product,
                    'fournisseur': fournisseur,
                    'emplacement': emp
                }
                
                cursor.execute(query, params)
                result = cursor.fetchone()
                emplacement_stock = result[0] if result and result[0] else 0
                
                total_stock_initial += emplacement_stock
                emplacement_breakdown[emp] = emplacement_stock
               
            
            return {
                'total': total_stock_initial,
                'breakdown': emplacement_breakdown
            }
            
    except Exception as e:
        logger.error(f"Error calculating stock initial for multiple emplacements: {e}")
        return {
            'total': 0,
            'breakdown': {}
        }

def fetch_stock_movement_data(start_date=None, end_date=None, product=None, fournisseur=None, emplacement=None, v2_mode=False):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            
            # For empty or None emplacement, we want to fetch all main locations
            if emplacement == '' or emplacement is None:
                # Calculate stock initial separately for multiple emplacements
                stock_initial_result = calculate_stock_initial_for_multiple_emplacements(start_date, product, fournisseur, v2_mode)
                calculated_stock_initial = stock_initial_result['total']
                emplacement_breakdown = stock_initial_result['breakdown']
                
                emplacement_condition = """
                AND (l.value LIKE 'Préparation%' OR l.value LIKE 'HANGAR%' OR l.value LIKE 'Dépot Hangar réserve%' OR l.value LIKE 'Dépot réserve%')
                """
                # Use the calculated stock initial instead of subquery
                stock_initial_value = str(calculated_stock_initial)
                  # Exclude internal transfers when no specific emplacement is selected
                movement_line_condition = ""
                internal_transfer_condition = """
                AND NOT (
                    t.M_MovementLine_ID IS NOT NULL AND (
                        (ml.M_Locator_ID = 1001135 AND ml.M_LocatorTo_ID = 1000614) OR
                        (ml.M_Locator_ID = 1000614 AND ml.M_LocatorTo_ID = 1001135) OR
                        (ml.M_Locator_ID = 1001136 AND ml.M_LocatorTo_ID = 1001135) OR
                        (ml.M_Locator_ID = 1001135 AND ml.M_LocatorTo_ID = 1001136) OR
                        (ml.M_Locator_ID = 1001136 AND ml.M_LocatorTo_ID = 1001128) OR
                        (ml.M_Locator_ID = 1001128 AND ml.M_LocatorTo_ID = 1001136) OR
                        (ml.M_Locator_ID = 1001136 AND ml.M_LocatorTo_ID = 1000614) OR
                        (ml.M_Locator_ID = 1000614 AND ml.M_LocatorTo_ID = 1001136) OR
                        (ml.M_Locator_ID = 1001128 AND ml.M_LocatorTo_ID = 1001135) OR
                        (ml.M_Locator_ID = 1001135 AND ml.M_LocatorTo_ID = 1001128) OR
                        (ml.M_Locator_ID = 1001128 AND ml.M_LocatorTo_ID = 1000614) OR
                        (ml.M_Locator_ID = 1000614 AND ml.M_LocatorTo_ID = 1001128)
                    )
                )
                """
                # Don't include emplacement in params for this case
                params = {
                'start_date': start_date,
                'end_date': end_date,
                'product': product,
                'fournisseur': fournisseur
                }
            else:
                # For specific emplacement, use the original subquery method
                emplacement_breakdown = {}  # No breakdown for specific emplacement
                emplacement_condition = """
                AND l.value LIKE :emplacement || '%'
                """
                
                # Determine the exclusion clause based on V2 mode
                if v2_mode:
                    exclusion_clause = "WHEN COALESCE(inv.docstatus, m.docstatus, io.docstatus) IN ('RE','VO')"
                else:
                    exclusion_clause = "WHEN COALESCE(inv.docstatus, m.docstatus, io.docstatus) IN ('RE')"
                
                stock_initial_value = f"""coalesce((SELECT SUM(
                    CASE
                        {exclusion_clause}
                        THEN 0
                        ELSE s.movementqty
                    END
                )
                FROM m_transaction s
                inner join m_product p on (s.m_product_id = p.m_product_id)
                inner join m_locator l on (l.m_locator_id = s.m_locator_id)
                inner join M_attributeinstance att on (att.m_attributesetinstance_id = s.m_attributesetinstance_id)
                LEFT OUTER JOIN M_InventoryLine il ON (s.M_InventoryLine_ID=il.M_InventoryLine_ID)
                LEFT OUTER JOIN M_Inventory inv ON (inv.m_inventory_id = il.m_inventory_id)
                LEFT OUTER JOIN M_MovementLine ml ON (s.M_MovementLine_ID=ml.M_MovementLine_ID)
                LEFT OUTER JOIN M_Movement m ON (m.M_Movement_ID=ml.M_Movement_ID)
                LEFT OUTER JOIN M_InOutLine iol ON (s.M_InOutLine_ID=iol.M_InOutLine_ID)
                LEFT OUTER JOIN M_Inout io ON (iol.M_InOut_ID=io.M_InOut_ID)
                WHERE s.movementdate <  TO_DATE(:start_date, 'YYYY-MM-DD')
                AND (:product IS NULL OR p.M_Product_ID = :product)
                AND (:fournisseur IS NULL OR att.value like :fournisseur || '%')
                AND l.value LIKE :emplacement || '%'
                AND att.m_attribute_id = 1000508
                AND s.AD_Client_ID = 1000000
                ), 0)"""
                  # Don't exclude internal transfers when specific emplacement is selected
                movement_line_condition = ""
                internal_transfer_condition = ""
                params = {
                    'start_date': start_date,
                    'end_date': end_date,
                    'product': product,
                    'fournisseur': fournisseur,
                    'emplacement': emplacement                }
            
            query = """
            SELECT
                t.MovementDate AS MovementDate,
                nvl(nvl(io.documentno,inv.documentno),m.documentno) as documentno,
                nvl(bp.name, nvl(inv.description,m.description)) as name,
                p.name AS productname,
                CASE WHEN t.movementqty > 0 then t.movementqty else 0 end as ENTREE,
                CASE WHEN t.movementqty < 0 then ABS(t.movementqty) else 0 end as SORTIE,
                """ + stock_initial_value + """ AS StockInitial,
                asi.lot,
                l_from.value AS locator_from,
                l_to.value AS locator_to,
                COALESCE(io.docstatus, m.docstatus, inv.docstatus, 'N/A') AS docstatus
            FROM M_Transaction t
            INNER JOIN ad_org org
            ON org.ad_org_id = t.ad_org_id
            LEFT JOIN ad_orginfo oi
            ON oi.ad_org_id = org.ad_org_id
            LEFT JOIN c_location orgloc
            ON orgloc.c_location_id = oi.c_location_id
            INNER JOIN M_Locator l
            ON (t.M_Locator_ID=l.M_Locator_ID)
            INNER JOIN M_Product p
            ON (t.M_Product_ID=p.M_Product_ID)
            LEFT OUTER JOIN M_InventoryLine il
            ON (t.M_InventoryLine_ID=il.M_InventoryLine_ID)
            LEFT OUTER JOIN M_Inventory inv
            ON (inv.m_inventory_id = il.m_inventory_id )
            LEFT OUTER JOIN M_MovementLine ml
            ON (t.M_MovementLine_ID=ml.M_MovementLine_ID""" + movement_line_condition + """
            )
            LEFT OUTER JOIN M_Movement m
            ON (m.M_Movement_ID=ml.M_Movement_ID )
            LEFT OUTER JOIN M_InOutLine iol
            ON (t.M_InOutLine_ID=iol.M_InOutLine_ID)
            LEFT OUTER JOIN M_Inout io
            ON (iol.M_InOut_ID=io.M_InOut_ID )
            LEFT OUTER JOIN C_BPartner bp
            ON (bp.C_BPartner_ID = io.C_BPartner_ID)
            INNER JOIN M_attributesetinstance asi on t.m_attributesetinstance_id = asi.m_attributesetinstance_id
            INNER JOIN M_attributeinstance att on (att.m_attributesetinstance_id = asi.m_attributesetinstance_id)
            -- Add joins for from and to locators
            LEFT JOIN M_Locator l_from ON (
                CASE 
                    WHEN t.M_MovementLine_ID IS NOT NULL THEN ml.M_Locator_ID 
                    WHEN t.M_InOutLine_ID IS NOT NULL THEN 
                        CASE WHEN t.MovementQty > 0 THEN iol.M_Locator_ID ELSE NULL END
                    ELSE NULL 
                END = l_from.M_Locator_ID
            )
            LEFT JOIN M_Locator l_to ON (
                CASE 
                    WHEN t.M_MovementLine_ID IS NOT NULL THEN ml.M_LocatorTo_ID 
                    WHEN t.M_InOutLine_ID IS NOT NULL THEN 
                        CASE WHEN t.MovementQty < 0 THEN iol.M_Locator_ID ELSE NULL END
                    ELSE NULL 
                END = l_to.M_Locator_ID
            )            WHERE
            att.m_attribute_id = 1000508
            AND (:end_date IS NULL OR t.movementdate <= TO_DATE(:end_date, 'YYYY-MM-DD'))
            AND (:start_date IS NULL OR t.movementdate >= TO_DATE(:start_date, 'YYYY-MM-DD'))
            AND (:product IS NULL OR p.M_Product_ID = :product)
            AND (:fournisseur IS NULL OR att.value like :fournisseur || '%')
            AND NOT (t.movementqty = 0)
            """ + emplacement_condition + internal_transfer_condition + """
            AND t.AD_Client_ID = 1000000
            ORDER BY t.MovementDate DESC
            """

           
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            
            # Add emplacement breakdown to the response
            return {
                'data': data,
                'emplacement_breakdown': emplacement_breakdown
            }

    except Exception as e:
        logger.error(f"Database error in fetch_stock_movement_data: {e}")
        return {"error": "An error occurred while fetching stock movement data."}

@app.route('/fetch-stock-movement-data', methods=['GET'])
def fetch_stock_movement():
    try:
        start_date = request.args.get("start_date", None)
        end_date = request.args.get("end_date", None)
        product = request.args.get("product", None)
        fournisseur = request.args.get("fournisseur", None)
        emplacement = request.args.get("emplacement", None)
        v2_mode = request.args.get("v2_mode", "false").lower() == "true"

        # Log the product parameter for debugging
        if product:
            logger.info(f"Fetching stock movement data for product ID: {product}")
        
        data = fetch_stock_movement_data(start_date, end_date, product, fournisseur, emplacement, v2_mode)
        return jsonify(data)

    except Exception as e:
        logger.error(f"Error in fetch_stock_movement endpoint: {e}")
        return jsonify({"error": "Failed to fetch stock movement data"}), 500

@app.route('/download-stock-movement-excel', methods=['GET'])
def download_stock_movement_excel():
    try:
        start_date = request.args.get("start_date", None)
        end_date = request.args.get("end_date", None)
        product = request.args.get("product", None)
        fournisseur = request.args.get("fournisseur", None)
        emplacement = request.args.get("emplacement", None)
        v2_mode = request.args.get("v2_mode", "false").lower() == "true"

        data_result = fetch_stock_movement_data(start_date, end_date, product, fournisseur, emplacement, v2_mode)
        
        if isinstance(data_result, dict) and "error" in data_result:
            return jsonify(data_result), 500

        # Extract the data part for Excel generation
        data = data_result.get('data', []) if isinstance(data_result, dict) else data_result

        # Generate Excel file
        excel_output = generate_excel_stock_movement(data)
        
        # Generate filename
        today_date = datetime.now().strftime("%d-%m-%Y")
        filename = f"Stock_Movement_{today_date}.xlsx"

        return send_file(excel_output, as_attachment=True, download_name=filename, 
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        logger.error(f"Error in download_stock_movement_excel: {e}")
        return jsonify({"error": "Failed to generate Excel file"}), 500

def generate_excel_stock_movement(data):
    """Generate Excel file for stock movement data"""
    df = pd.DataFrame(data)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Movement"

    # Define header styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Write headers
    if not df.empty:
        ws.append(df.columns.tolist())
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Apply auto filter
        ws.auto_filter.ref = ws.dimensions

        # Write data rows
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            ws.append(row)

        # Create table
        table = Table(displayName="StockMovementTable", ref=ws.dimensions)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def get_initial_stock_for_date(target_date, product_id, v2_mode=False):
    """
    Calculate the initial stock for a given product at a specific date.
    Uses the same logic as the existing stock calculation functions.
    """
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            
            # Determine the exclusion clause based on V2 mode
            if v2_mode:
                exclusion_clause = "WHEN COALESCE(inv.docstatus, m.docstatus, io.docstatus) IN ('RE','VO')"
            else:
                exclusion_clause = "WHEN COALESCE(inv.docstatus, m.docstatus, io.docstatus) IN ('RE')"
            
            # Query to get the stock before the target date using same logic as stock movement
            # Note: We do NOT filter by DOCSTATUS here because initial stock calculation 
            # should include ALL movements before the target date, regardless of document status
            # This matches the logic used in the working stock movement queries
            query = f"""
                 SELECT SUM(
                    CASE 
                        {exclusion_clause}
                        THEN 0
                        ELSE s.movementqty 
                    END
                ) AS initial_stock
                FROM m_transaction s
                INNER JOIN m_product p ON (s.m_product_id = p.m_product_id)
                INNER JOIN m_locator l ON (l.m_locator_id = s.m_locator_id)
               
                LEFT OUTER JOIN M_InventoryLine il
                ON (s.M_InventoryLine_ID=il.M_InventoryLine_ID)
                
                LEFT OUTER JOIN M_Inventory inv
                ON (inv.m_inventory_id = il.m_inventory_id)

                LEFT OUTER JOIN M_MovementLine ml
                ON (s.M_MovementLine_ID=ml.M_MovementLine_ID)

                LEFT OUTER JOIN M_Movement m
                ON (m.M_Movement_ID=ml.M_Movement_ID)

                LEFT OUTER JOIN M_InOutLine iol
                ON (s.M_InOutLine_ID=iol.M_InOutLine_ID)

                LEFT OUTER JOIN M_Inout io
                ON (iol.M_InOut_ID=io.M_InOut_ID)
                
                WHERE s.movementdate < TO_DATE(:target_date, 'YYYY-MM-DD')
                AND p.m_product_id = :product_id
                AND (l.value LIKE 'Préparation%' OR l.value LIKE 'HANGAR%')
                AND s.AD_Client_ID = 1000000

            """
            
            cursor.execute(query, {
                'product_id': product_id,
                'target_date': target_date
            })
            
            result = cursor.fetchone()
            initial_stock = float(result[0]) if result and result[0] is not None else 0.0
            
            return {
                'initial_stock': initial_stock,
                'product_id': product_id,
                'target_date': target_date,
                'v2_mode': v2_mode
            }
            
    except Exception as e:
        logger.error(f"Error in get_initial_stock_for_date: {e}")
        return {'error': str(e)}

@app.route('/getInitialStock', methods=['GET'])
def get_initial_stock_endpoint():
    """
    API endpoint to get initial stock for a specific date and product.
    Expected parameters: date (YYYY-MM-DD format), product_id, v2_mode (optional, default false)
    """
    try:
        target_date = request.args.get('date')
        product_id = request.args.get('product_id')
        v2_mode = request.args.get('v2_mode', 'false').lower() == 'true'
        
        if not target_date or not product_id:
            return jsonify({'error': 'Both date and product_id parameters are required'}), 400
        
        # Validate date format
        try:
            datetime.strptime(target_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
        
        # Validate product_id is numeric
        try:
            int(product_id)
        except ValueError:
            return jsonify({'error': 'product_id must be numeric'}), 400
        
        result = get_initial_stock_for_date(target_date, product_id, v2_mode)
        
        if 'error' in result:
            return jsonify(result), 500
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Error in get_initial_stock_endpoint: {e}")
        return jsonify({'error': 'Internal server error'}), 500





@app.route('/fetch-emplacements-stock')
def fetch_emplacements_stock():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT ml.value AS EMPLACEMENT
                FROM M_Locator ml
                JOIN M_Warehouse m ON m.M_WAREHOUSE_ID = ml.M_WAREHOUSE_ID
                WHERE m.ISACTIVE = 'Y'
                  AND m.AD_Client_ID = 1000000
                  AND ml.ISACTIVE = 'Y'
                  AND ml.AD_Client_ID = 1000000
                ORDER BY m.value
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            return jsonify(data)
    except Exception as e:
        logger.error(f"Error fetching emplacements: {e}")
        return jsonify({"error": "Could not fetch emplacement list"}), 500
    



@app.route('/rotation_monthly_achat_pdf', methods=['GET'])
def download_product_achat_pdf():
    from flask import request, jsonify

    years = request.args.get('years')
    fournisseur_param = request.args.get('fournisseur')
    product_id = request.args.get('product_id')
    
    # If product_id is "undefined", set it to None
    if product_id == "undefined":
        product_id = None

    if not years:
        years = str(datetime.now().year)

    try:
        year_list = [int(y.strip()) for y in years.split(',')]
    except ValueError:
        return jsonify({"error": "Invalid years format"}), 400

    # Handle multiple suppliers (comma-separated or single)
    if fournisseur_param:
        fournisseurs = [f.strip() for f in fournisseur_param.split(',') if f.strip()]
    else:
        return jsonify({"error": "Fournisseur parameter is required"}), 400

    data = fetch_rotation_monthly_achat(year_list, fournisseurs, product_id)
    if 'error' in data:
        return jsonify(data), 500

    # Create descriptive filename
    supplier_text = f"{len(fournisseurs)}_suppliers" if len(fournisseurs) > 1 else fournisseurs[0][:20]
    filename = f"achat_recap_{years.replace(',', '-')}_{supplier_text}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
    elements = []

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    normal_style = styles["Normal"]
    table_total_style = ParagraphStyle("row", parent=normal_style, alignment=1, textColor=colors.blue, fontName="Helvetica-Bold", fontSize=6)

    table_header_style = ParagraphStyle("header", parent=normal_style, alignment=1, textColor=colors.white, fontSize=8)
    table_row_style = ParagraphStyle("row", parent=normal_style, alignment=1, fontSize=6)

    elements.append(Paragraph("Récapitulatif des Achats", title_style))
    
    # Create supplier display text
    supplier_display = ", ".join(fournisseurs) if len(fournisseurs) <= 3 else f"{len(fournisseurs)} suppliers"
    
    # Try to get product name if product_id is provided
    product_name = "Tous"
    if product_id:
        try:
            with DB_POOL.acquire() as connection:
                cursor = connection.cursor()
                cursor.execute("SELECT name FROM M_PRODUCT WHERE M_PRODUCT_ID = :product_id", {'product_id': product_id})
                result = cursor.fetchone()
                if result:
                    product_name = result[0]
        except Exception as e:
            logger.error(f"Error fetching product name: {e}")
    
    elements.append(Paragraph(f"Fournisseurs: {supplier_display} | Produit: {product_name} | Années: {years}", styles["Heading2"]))
    elements.append(Spacer(1, 0.25 * inch))

    for year_idx, (year, months_data) in enumerate(data.items()):
        year_color = YEAR_COLORS[year_idx % len(YEAR_COLORS)]
        year_style = ParagraphStyle("year", parent=styles["Heading1"], textColor=colors.white, backColor=year_color)
        elements.append(Paragraph(f"Année: {year}", year_style))

        for part_label, month_range in [("Janvier à Juin", range(1, 7)), ("Juillet à Décembre", range(7, 13))]:
            # For multiple suppliers, create a combined product map with supplier info
            if len(fournisseurs) > 1:
                combined_products = {}
                for m in month_range:
                    m_str = str(m).zfill(2)
                    if m_str in months_data:
                        for item in months_data[m_str]["details"]:
                            product_key = f"{item['PRODUIT']} ({item['FOURNISSEUR']})"
                            if product_key not in combined_products:
                                combined_products[product_key] = {}
                            combined_products[product_key][m] = item
                unique_products = list(combined_products.keys())
            else:
                # Single supplier - use original logic
                unique_products = set()
                for m in month_range:
                    m_str = str(m).zfill(2)
                    if m_str in months_data:
                        for item in months_data[m_str]["details"]:
                            unique_products.add(item["PRODUIT"])
                unique_products = list(unique_products)

            if not unique_products:
                continue

            # Header rows
            header_row1 = [Paragraph("Produit", table_header_style)]
            for m in month_range:
                m_str = str(m).zfill(2)
                month_name = months_data.get(m_str, {}).get("month_name", datetime(2025, m, 1).strftime("%b")).capitalize()
                header_row1.extend([Paragraph(month_name, table_header_style), ""])

            header_row2 = [""]
            for _ in month_range:
                header_row2.extend([
                    Paragraph("Qty", table_header_style),
                    Paragraph("Prix", table_header_style)
                ])

            table_data = [header_row1, header_row2]

            # Product rows
            for prod in sorted(unique_products):
                row = [Paragraph(prod, table_row_style)]
                for m in month_range:
                    m_str = str(m).zfill(2)
                    month_data = months_data.get(m_str, {})
                    
                    if len(fournisseurs) > 1:
                        # For combined view, find the specific product entry
                        detail = combined_products.get(prod, {}).get(m)
                        qty = f"{detail['QTY']:,.0f}" if detail else "-"
                        amt = f"{detail['CHIFFRE']:,.2f}" if detail else "-"
                    else:
                        # Single supplier - use original logic
                        detail = next((d for d in month_data.get("details", []) if d["PRODUIT"] == prod), None)
                        qty = f"{detail['QTY']:,.0f}" if detail else "-"
                        amt = f"{detail['CHIFFRE']:,.2f}" if detail else "-"
                    
                    row.append(Paragraph(qty, table_row_style))
                    row.append(Paragraph(amt, table_row_style))
                table_data.append(row)

            # Total row
            # Total row
            total_row = [Paragraph("TOTAL", table_total_style)]
            for m in month_range:
                m_str = str(m).zfill(2)
                month_data = months_data.get(m_str, {})
                total = month_data.get("total", {})

                qty_val = total.get("QTY")
                amt_val = total.get("CHIFFRE")

                qty = f"{qty_val:,.0f}" if qty_val is not None else "-"
                amt = f"{amt_val:,.2f}" if amt_val is not None else "-"

                total_row.append(Paragraph(qty, table_total_style))
                total_row.append(Paragraph(amt, table_total_style))
            table_data.append(total_row)

            # Table construction
            col_widths = [1.9 * inch] + [0.5 * inch] * (len(month_range) * 2)
            table = ReportLabTable(table_data, colWidths=col_widths)

            # Create span styles for merged headers
            span_style = []
            for i in range(len(month_range)):
                col_start = 1 + i * 2
                span_style.append(('SPAN', (col_start, 0), (col_start + 1, 0)))

            # Final style
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 1), year_color),
                ('TEXTCOLOR', (0, 0), (-1, 1), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 2),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 2), (-1, -2), [colors.whitesmoke, colors.white]),
            ] + span_style))

            elements.append(Paragraph(f"{part_label}", styles["Heading3"]))
            elements.append(table)
            elements.append(Spacer(1, 0.3 * inch))

    doc.build(elements)
    buffer.seek(0)
    response = make_response(buffer.getvalue())
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f"attachment; filedisplayName={filename}"

    return response


@app.route('/rotation_monthly_achat', methods=['GET'])
def fetch_rotation_achat():
    years = request.args.get('years')  # Comma-separated list of years
    fournisseur_param = request.args.get('fournisseur')  # Can be comma-separated
    product_id = request.args.get('product_id')
    
    # If product_id is "undefined", set it to None
    if product_id == "undefined":
        product_id = None

    # If years is not provided, use current year
    if not years:
        current_year = datetime.now().year
        years = str(current_year)
    
    # Handle multiple suppliers (comma-separated or single)
    if fournisseur_param:
        fournisseurs = [f.strip() for f in fournisseur_param.split(',') if f.strip()]
    else:
        return jsonify({"error": "At least one supplier is required"}), 400

    try:
        # Split the years string into a list and convert to integers
        year_list = [int(year.strip()) for year in years.split(',')]
    except ValueError:
        return jsonify({"error": "Invalid years format. Please provide comma-separated years (e.g., 2022,2023,2024)"}), 400

    # Fetch data from the database with the list of suppliers
    data = fetch_rotation_monthly_achat(year_list, fournisseurs, product_id)

    # Return the result as a JSON response
    return jsonify(data)


def fetch_rotation_monthly_achat(year_list, fournisseurs, product_id=None):
    try:
        # Acquire a connection from the pool
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            # Build the supplier condition
            if isinstance(fournisseurs, str):
                fournisseurs = [fournisseurs]

            supplier_conditions = []
            for i, f in enumerate(fournisseurs):
                bind_var = f'fournisseur_{i}'
                supplier_conditions.append(f"UPPER(cb.name) LIKE UPPER(:{bind_var}) || '%'")

            supplier_condition = ' OR '.join(supplier_conditions) if supplier_conditions else '1=0'

            # Create the years parameter list
            year_placeholders = ', '.join([f':year_{i}' for i in range(len(year_list))])
            
            query = """
                WITH monthly_data AS (
                    SELECT 
                        TO_CHAR(xf.MOVEMENTDATE, 'YYYY') AS year,
                        TO_CHAR(xf.MOVEMENTDATE, 'MM') AS month_num,
                        TO_CHAR(xf.MOVEMENTDATE, 'Month') AS month_name,
                        m.name AS produit,
                        cb.name AS fournisseur,
                        SUM(CASE 
                            WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(mi.QTYENTERED)
                        END) AS qty,
                        SUM(CASE 
                            WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                        END) AS chiffre
                    FROM 
                        M_InOut xf
                        JOIN M_INOUTLINE mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
                        JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID = mi.M_ATTRIBUTESETINSTANCE_ID
                        JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
                        JOIN M_PRODUCT m ON m.M_PRODUCT_id = mi.M_PRODUCT_id
                    WHERE 
                        TO_CHAR(xf.MOVEMENTDATE, 'YYYY') IN (""" + year_placeholders + """)
                        AND xf.AD_Org_ID = 1000000
                        AND xf.C_DocType_ID IN (1000013, 1000646)
                        AND ma.M_Attribute_ID = 1000504
                        AND xf.DOCSTATUS = 'CO'
                        AND xf.M_Warehouse_ID IN (1000724, 1000000, 1000720, 1000725)
                        AND (""" + supplier_condition + """)
                        AND (:product_id IS NULL OR m.M_PRODUCT_ID = :product_id)
                    GROUP BY 
                        TO_CHAR(xf.MOVEMENTDATE, 'YYYY'),
                        TO_CHAR(xf.MOVEMENTDATE, 'MM'),
                        TO_CHAR(xf.MOVEMENTDATE, 'Month'),
                        m.name,
                        cb.name
                )
                SELECT * FROM monthly_data
                ORDER BY year, month_num, produit
            """

            # Prepare parameters
            params = {'product_id': product_id}
            
            # Add supplier parameters
            for i, f in enumerate(fournisseurs):
                params[f'fournisseur_{i}'] = f

            # Add year parameters
            for i, year in enumerate(year_list):
                params[f'year_{i}'] = str(year)

            # Execute the query with the parameters
            cursor.execute(query, params)

            # Fetch the results
            rows = cursor.fetchall()

            # Format the results into a nested dictionary structure
            result = {}
            
            # Process each row and organize by year and month
            for row in rows:
                year, month_num, month_name, produit, fournisseur, qty, chiffre = row
                month_num = month_num.zfill(2)  # Ensure two digits for month number
                
                if year not in result:
                    result[year] = {}
                
                if month_num not in result[year]:
                    result[year][month_num] = {
                        'month_name': month_name.strip(),
                        'details': [],
                        'total': {'QTY': 0, 'CHIFFRE': 0}
                    }
                
                # Add product details with supplier info
                result[year][month_num]['details'].append({
                    'PRODUIT': produit,
                    'FOURNISSEUR': fournisseur,
                    'QTY': float(qty) if qty is not None else 0,
                    'CHIFFRE': float(chiffre) if chiffre is not None else 0
                })
                
                # Update monthly totals
                result[year][month_num]['total']['QTY'] += float(qty) if qty is not None else 0
                result[year][month_num]['total']['CHIFFRE'] += float(chiffre) if chiffre is not None else 0

            # Sort years and months numerically
            sorted_result = {
                str(year): dict(sorted(months.items()))
                for year, months in sorted(result.items())
            }
            return sorted_result

    except Exception as e:
        logger.error(f"Error fetching product recap achat: {e}")
        return {"error": "An error occurred while fetching product recap achat."}

def fetch_fournisseur_by_product(product_id):
    try:
        # Check if product_id is valid
        if not product_id or product_id == "undefined":
            return []
            
        # Ensure product_id is a valid number
        try:
            product_id = int(product_id)
        except ValueError:
            logger.error(f"Invalid product_id format: {product_id}")
            return []
            
        # Acquire a connection from the pool
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT DISTINCT
                    CAST(cb.name AS VARCHAR2(300)) AS FOURNISSEUR
                FROM 
                    M_InOut xf
                JOIN M_INOUTLINE mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
                JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
                JOIN M_PRODUCT m ON m.M_PRODUCT_id = mi.M_PRODUCT_id
                WHERE 
                    m.M_PRODUCT_ID = :product_id
                    AND xf.AD_Org_ID = 1000000
                    AND xf.C_DocType_ID IN (1000013, 1000646)
                    AND xf.M_Warehouse_ID IN (1000724, 1000000, 1000720, 1000725)
                ORDER BY 
                    FOURNISSEUR
            """

            params = {
                'product_id': product_id
            }

            # Execute the query with the provided parameters
            cursor.execute(query, params)

            # Fetch the results
            rows = cursor.fetchall()

            # Format the results into a simple list of supplier names
            suppliers = [row[0] for row in rows]

            return suppliers
    
    except Exception as e:
        logger.error(f"Error fetching suppliers by product: {e}")
        return {"error": "An error occurred while fetching suppliers by product."}

# Flask route to handle the request
@app.route('/fetchSuppliersByProduct', methods=['GET'])
def fetch_suppliers_by_product():
    product_id = request.args.get('product_id')

    # Ensure product_id is provided and not "undefined"
    if not product_id or product_id == "undefined":
        return jsonify([]), 200  # Return empty list instead of error

    # Fetch data from the database
    suppliers = fetch_fournisseur_by_product(product_id)

    # Return the result as a JSON response
    return jsonify(suppliers)




from flask import make_response
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as ReportLabTable, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from datetime import datetime
import io

# Define a color palette for years (can be extended as needed)
YEAR_COLORS = [
    colors.HexColor('#4E79A7'),  # Blue
    colors.HexColor('#F28E2B'),  # Orange
    colors.HexColor('#E15759'),  # Red
    colors.HexColor('#76B7B2'),  # Teal
    colors.HexColor('#59A14F'),  # Green
    colors.HexColor('#EDC948'),  # Yellow
    colors.HexColor('#B07AA1'),  # Purple
    colors.HexColor('#FF9DA7'),  # Pink
    colors.HexColor('#9C755F'),  # Brown
    colors.HexColor('#BAB0AC')   # Gray
]


def rot_mont_vente(year_list, fournisseurs, products=None, clients=None, zones=None, product_ids=None):
    try:
        # Acquire a connection from the pool
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            # Build the supplier condition
            supplier_condition = "1=1"  # Default to include all suppliers
            if fournisseurs:
                if isinstance(fournisseurs, str):
                    fournisseurs = [fournisseurs]

                supplier_conditions = []
                for i, f in enumerate(fournisseurs):
                    bind_var = f'fournisseur_{i}'
                    supplier_conditions.append(f"UPPER(xf.name) LIKE UPPER(:{bind_var}) || '%'")

                supplier_condition = ' OR '.join(supplier_conditions) if supplier_conditions else '1=0'

            # Build product condition for multiple products (by name)
            product_condition = "1=1"
            if products:
                if isinstance(products, str):
                    products = [products]
                product_conditions = []
                for i, p in enumerate(products):
                    bind_var = f'product_{i}'
                    product_conditions.append(f"UPPER(xf.product) LIKE UPPER(:{bind_var}) || '%'")
                product_condition = ' OR '.join(product_conditions) if product_conditions else '1=0'
            
            # Build product ID condition
            product_id_condition = "1=1"
            if product_ids:
                if isinstance(product_ids, str):
                    product_ids = [product_ids]
                product_id_conditions = []
                for i, pid in enumerate(product_ids):
                    bind_var = f'product_id_{i}'
                    product_id_conditions.append(f"xf.M_PRODUCT_ID = :{bind_var}")
                product_id_condition = ' OR '.join(product_id_conditions) if product_id_conditions else '1=0'

            # Build client condition for multiple clients
            client_condition = "1=1"
            if clients:
                if isinstance(clients, str):
                    clients = [clients]
                client_conditions = []
                for i, c in enumerate(clients):
                    bind_var = f'client_{i}'
                    client_conditions.append(f"UPPER(cb.name) LIKE UPPER(:{bind_var}) || '%'")
                client_condition = ' OR '.join(client_conditions) if client_conditions else '1=0'

            # Build zone condition for multiple zones
            zone_condition = "1=1"
            if zones:
                if isinstance(zones, str):
                    zones = [zones]
                zone_conditions = []
                for i, z in enumerate(zones):
                    bind_var = f'zone_{i}'
                    zone_conditions.append(f"UPPER(sr.name) LIKE UPPER(:{bind_var}) || '%'")
                zone_condition = ' OR '.join(zone_conditions) if zone_conditions else '1=0'

            # Create the years parameter list
            year_placeholders = ', '.join([f':year_{i}' for i in range(len(year_list))])
            
            query = f"""
                WITH monthly_data AS (
                    SELECT 
                        TO_CHAR(xf.MOVEMENTDATE, 'YYYY') AS year,
                        TO_CHAR(xf.MOVEMENTDATE, 'MM') AS month_num,
                        TO_CHAR(xf.MOVEMENTDATE, 'Month') AS month_name,
                        CAST(xf.product AS VARCHAR2(400)) AS produit,
                        xf.name AS fournisseur,
                        cb.name AS client,
                        SUM(xf.TOTALLINE) AS total,
                        SUM(xf.qtyentered) AS qty,
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS marge
                    FROM 
                        xx_ca_fournisseur xf
                        LEFT JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                        LEFT JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                        LEFT JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                        LEFT JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                        LEFT JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                        LEFT JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE 
                        TO_CHAR(xf.MOVEMENTDATE, 'YYYY') IN ({year_placeholders})
                        AND xf.AD_Org_ID = 1000000
                        AND xf.DOCSTATUS != 'RE'
                        AND ({supplier_condition})
                        AND ({product_condition})
                        AND ({product_id_condition})
                        AND ({client_condition})
                        AND ({zone_condition})
                    GROUP BY 
                        TO_CHAR(xf.MOVEMENTDATE, 'YYYY'),
                        TO_CHAR(xf.MOVEMENTDATE, 'MM'),
                        TO_CHAR(xf.MOVEMENTDATE, 'Month'),
                        xf.product,
                        xf.name,
                        cb.name
                )
                SELECT * FROM monthly_data
                ORDER BY year, month_num, produit
            """

            # Prepare parameters
            params = {}
            
            # Add supplier parameters only if suppliers are specified
            if fournisseurs:
                for i, f in enumerate(fournisseurs):
                    params[f'fournisseur_{i}'] = f

            # Add product parameters
            if products:
                for i, p in enumerate(products):
                    params[f'product_{i}'] = p
                    
            # Add product ID parameters
            if product_ids:
                for i, pid in enumerate(product_ids):
                    try:
                        params[f'product_id_{i}'] = int(pid)
                    except ValueError:
                        logger.warning(f"Invalid product ID format: {pid}, skipping")

            # Add client parameters
            if clients:
                for i, c in enumerate(clients):
                    params[f'client_{i}'] = c

            # Add zone parameters
            if zones:
                for i, z in enumerate(zones):
                    params[f'zone_{i}'] = z

            # Add year parameters
            for i, year in enumerate(year_list):
                params[f'year_{i}'] = str(year)

            # Execute the query with the parameters
            cursor.execute(query, params)

            # Fetch the results
            rows = cursor.fetchall()

            # Format the results into a nested dictionary structure
            result = {}
            
            # Process each row and organize by year and month
            for row in rows:
                year, month_num, month_name, produit, fournisseur, client, total, qty, marge = row
                month_num = month_num.zfill(2)  # Ensure two digits for month number
                
                if year not in result:
                    result[year] = {}
                
                if month_num not in result[year]:
                    result[year][month_num] = {
                        'month_name': month_name.strip(),
                        'details': [],
                        'total': {'QTY': 0, 'TOTAL': 0, 'MARGE': 0}
                    }
                
                # Add product details with supplier and client info
                result[year][month_num]['details'].append({
                    'PRODUIT': produit,
                    'FOURNISSEUR': fournisseur,
                    'CLIENT': client if client else '',
                    'QTY': float(qty) if qty is not None else 0,
                    'TOTAL': float(total) if total is not None else 0,
                    'MARGE': float(marge) if marge is not None else 0
                })
                
                # Update monthly totals
                result[year][month_num]['total']['QTY'] += float(qty) if qty is not None else 0
                result[year][month_num]['total']['TOTAL'] += float(total) if total is not None else 0
                # For marge, we need to recalculate it based on aggregated values
                # This is a simplified approach; you might need to adjust based on your business logic

            # Sort years and months numerically
            sorted_result = {
                str(year): dict(sorted(months.items()))
                for year, months in sorted(result.items())
            }
            return sorted_result

    except Exception as e:
        logger.error(f"Error fetching rotation monthly vente: {e}")
        return {"error": "An error occurred while fetching rotation monthly vente."}

@app.route('/rot_mont_vente', methods=['GET'])
def fetch_rotation_monthly_vente():
    years = request.args.get('years')  # Comma-separated list of years
    fournisseur_param = request.args.get('fournisseur')  # Can be comma-separated
    product_param = request.args.get('product')  # Can be comma-separated (legacy)
    product_id_param = request.args.get('product_id')  # New product_id parameter
    client_param = request.args.get('client')  # Can be comma-separated
    zone_param = request.args.get('zone')  # Can be comma-separated

    # If product_id is "undefined", set it to None
    if product_id_param == "undefined":
        product_id_param = None

    # If years is not provided, use current year
    if not years:
        current_year = datetime.now().year
        years = str(current_year)
    
    # Handle multiple suppliers (comma-separated or single)
    fournisseurs = None
    if fournisseur_param:
        fournisseurs = [f.strip() for f in fournisseur_param.split(',') if f.strip()]

    # Handle multiple products (comma-separated or single)
    products = None
    product_ids = None
    
    # Prioritize product_id over product name if both are provided
    if product_id_param:
        product_ids = [pid.strip() for pid in product_id_param.split(',') if pid.strip()]
    elif product_param:
        products = [p.strip() for p in product_param.split(',') if p.strip()]

    # Handle multiple clients (comma-separated or single)
    clients = None
    if client_param:
        clients = [c.strip() for c in client_param.split(',') if c.strip()]

    # Handle multiple zones (comma-separated or single)
    zones = None
    if zone_param:
        zones = [z.strip() for z in zone_param.split(',') if z.strip()]

    try:
        # Split the years string into a list and convert to integers
        year_list = [int(year.strip()) for year in years.split(',')]
    except ValueError:
        return jsonify({"error": "Invalid years format. Please provide comma-separated years (e.g., 2022,2023,2024)"}), 400

    # Fetch data from the database with the list of parameters
    data = rot_mont_vente(year_list, fournisseurs, products, clients, zones, product_ids)

    # Return the result as a JSON response
    return jsonify(data)



@app.route('/rotation_monthly_vente_pdf', methods=['GET'])
def download_product_vente_pdf():
    from flask import request, jsonify

    years = request.args.get('years')
    fournisseur_param = request.args.get('fournisseur')
    product_param = request.args.get('product')
    product_id_param = request.args.get('product_id')
    client_param = request.args.get('client')
    zone_param = request.args.get('zone')
    
    # If product_id is "undefined", set it to None
    if product_id_param == "undefined":
        product_id_param = None

    if not years:
        years = str(datetime.now().year)

    try:
        year_list = [int(y.strip()) for y in years.split(',')]
    except ValueError:
        return jsonify({"error": "Invalid years format"}), 400

    # Handle multiple suppliers (comma-separated or single)
    fournisseurs = None
    if fournisseur_param:
        fournisseurs = [f.strip() for f in fournisseur_param.split(',') if f.strip()]

    # Handle multiple clients (comma-separated or single)
    clients = None
    if client_param:
        clients = [c.strip() for c in client_param.split(',') if c.strip()]

    # Handle multiple zones (comma-separated or single)
    zones = None
    if zone_param:
        zones = [z.strip() for z in zone_param.split(',') if z.strip()]

    # Handle multiple products (comma-separated or single)
    products = None
    product_ids = None
    
    # Prioritize product_id over product name if both are provided
    if product_id_param:
        product_ids = [pid.strip() for pid in product_id_param.split(',') if pid.strip()]
    elif product_param:
        products = [p.strip() for p in product_param.split(',') if p.strip()]

    data = rot_mont_vente(year_list, fournisseurs, products, clients, zones, product_ids)
    if 'error' in data:
        return jsonify(data), 500

    # Create descriptive filename
    product_text = product_param or product_id_param or 'all_products'
    filename = f"vente_recap_{years.replace(',', '-')}_{product_text[:20]}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
    elements = []

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    normal_style = styles["Normal"]
    table_total_style = ParagraphStyle("row", parent=normal_style, alignment=1, textColor=colors.blue, fontName="Helvetica-Bold", fontSize=6)

    table_header_style = ParagraphStyle("header", parent=normal_style, alignment=1, textColor=colors.white, fontSize=8)
    table_row_style = ParagraphStyle("row", parent=normal_style, alignment=1, fontSize=6)

    elements.append(Paragraph("Récapitulatif des Ventes", title_style))
    
    # Create filter display text
    filter_parts = []
    if clients:
        client_display = ", ".join(clients) if len(clients) <= 3 else f"{len(clients)} clients"
        filter_parts.append(f"Clients: {client_display}")
    else:
        filter_parts.append("Clients: Tous")
    
    if fournisseurs:
        supplier_display = ", ".join(fournisseurs) if len(fournisseurs) <= 3 else f"{len(fournisseurs)} fournisseurs"
        filter_parts.append(f"Fournisseurs: {supplier_display}")
    
    if zones:
        zone_display = ", ".join(zones) if len(zones) <= 3 else f"{len(zones)} zones"
        filter_parts.append(f"Zones: {zone_display}")
    
    filter_parts.append(f"Produit: {product_param or product_id_param or 'Tous'}")
    filter_parts.append(f"Années: {years}")
    
    elements.append(Paragraph(" | ".join(filter_parts), styles["Heading2"]))
    elements.append(Spacer(1, 0.25 * inch))

    for year_idx, (year, months_data) in enumerate(data.items()):
        year_color = YEAR_COLORS[year_idx % len(YEAR_COLORS)]
        year_style = ParagraphStyle("year", parent=styles["Heading1"], textColor=colors.white, backColor=year_color)
        elements.append(Paragraph(f"Année: {year}", year_style))

        for part_label, month_range in [("Janvier à Juin", range(1, 7)), ("Juillet à Décembre", range(7, 13))]:
            # Create a combined product map organized by product and supplier (not client)
            product_suppliers = {}
            for m in month_range:
                m_str = str(m).zfill(2)
                if m_str in months_data:
                    for item in months_data[m_str]["details"]:
                        product_name = item['PRODUIT']
                        supplier_name = item.get('FOURNISSEUR', 'Unknown Supplier')
                        
                        # Create product-supplier key
                        product_key = f"{product_name} ({supplier_name})"
                        
                        if product_key not in product_suppliers:
                            product_suppliers[product_key] = {}
                        
                        # Aggregate data for the same product-supplier-month combination
                        if m not in product_suppliers[product_key]:
                            product_suppliers[product_key][m] = {
                                'QTY': 0,
                                'TOTAL': 0,
                                'CONSOMATION': 0
                            }
                        
                        product_suppliers[product_key][m]['QTY'] += item['QTY']
                        product_suppliers[product_key][m]['TOTAL'] += item['TOTAL']
                        # We need consomation to calculate margin properly
                        # For now, we'll calculate margin from the data we have
                        item_marge = item['MARGE'] if item['MARGE'] is not None else 0
                        consomation = item['TOTAL'] / (1 + item_marge) if item_marge > 0 else item['TOTAL']
                        product_suppliers[product_key][m]['CONSOMATION'] += consomation

            if not product_suppliers:
                continue

            # Create table for this period
            elements.append(Paragraph(f"{part_label}", styles["Heading3"]))
            
            # Header row - single row with month names
            header_row = [Paragraph("Produit (Fournisseur)", table_header_style)]
            for m in month_range:
                m_str = str(m).zfill(2)
                month_name = months_data.get(m_str, {}).get("month_name", datetime(2025, m, 1).strftime("%b")).capitalize()
                header_row.append(Paragraph(month_name, table_header_style))

            table_data = [header_row]

            # Product-supplier rows
            period_total_qty = {m: 0 for m in month_range}
            period_total_amt = {m: 0 for m in month_range}
            period_total_consomation = {m: 0 for m in month_range}
            
            for product_key in sorted(product_suppliers.keys()):
                row = [Paragraph(product_key, table_row_style)]
                for m in month_range:
                    detail = product_suppliers.get(product_key, {}).get(m)
                    if detail:
                        qty = f"{detail['QTY']:,.0f}"
                        total = f"{detail['TOTAL']:,.2f}"
                        # Calculate margin percentage
                        if detail['CONSOMATION'] > 0:
                            marge_pct = ((detail['TOTAL'] - detail['CONSOMATION']) / detail['CONSOMATION']) * 100
                            marge = f"{marge_pct:.2f}%"
                        else:
                            marge = "0.00%"
                        
                        # Create vertical layout like the frontend with better formatting
                        month_data_parts = [
                            f"<b>Qty:</b> {qty}",
                            f"<b>Total:</b> {total}",
                            f"<b>Marge:</b> {marge}"
                        ]
                        month_data = "<br/>".join(month_data_parts)
                        
                        period_total_qty[m] += detail['QTY']
                        period_total_amt[m] += detail['TOTAL']
                        period_total_consomation[m] += detail['CONSOMATION']
                    else:
                        month_data = "<b>Qty:</b> -<br/><b>Total:</b> -<br/><b>Marge:</b> -"
                    
                    row.append(Paragraph(month_data, table_row_style))
                table_data.append(row)

            # Period total row
            total_row = [Paragraph("TOTAL", table_total_style)]
            for m in month_range:
                qty = f"{period_total_qty[m]:,.0f}" if period_total_qty[m] > 0 else "0"
                amt = f"{period_total_amt[m]:,.2f}" if period_total_amt[m] > 0 else "0"
                
                if period_total_consomation[m] > 0:
                    total_marge_pct = ((period_total_amt[m] - period_total_consomation[m]) / period_total_consomation[m]) * 100
                    marge = f"{total_marge_pct:.2f}%"
                else:
                    marge = "0.00%"
                
                # Create vertical layout for totals too with better formatting
                total_data_parts = [
                    f"<b>Qty:</b> {qty}",
                    f"<b>Total:</b> {amt}",
                    f"<b>Marge:</b> {marge}"
                ]
                total_data = "<br/>".join(total_data_parts)
                total_row.append(Paragraph(total_data, table_total_style))
            table_data.append(total_row)

            # Table construction - single column per month now
            col_widths = [2.0 * inch] + [1.0 * inch] * len(month_range)
            table = ReportLabTable(table_data, colWidths=col_widths, repeatRows=1)

            # Simpler style without span since we have single column per month
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), year_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Top alignment for better readability
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),  # Header font size
                ('FONTSIZE', (0, 1), (-1, -2), 7),  # Data rows font size
                ('FONTSIZE', (0, -1), (-1, -1), 8),  # Total row font size
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.whitesmoke, colors.white]),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),  # Total row background
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 1), (-1, -1), 6),    # More padding for data rows
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('LEADING', (0, 1), (-1, -1), 12),       # Line spacing
            ]))

            elements.append(table)
            elements.append(Spacer(1, 0.2 * inch))

    doc.build(elements)
    buffer.seek(0)
    response = make_response(buffer.getvalue())
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"

    return response


@app.route('/fetchZoneClients', methods=['GET'])
def fetch_zone_clients():
    zone = request.args.get('zone')

    if not zone:
        return jsonify({"error": "Missing zone parameter"}), 400

    data = fetch_clients_by_zone(zone)
    return jsonify(data)


def fetch_clients_by_zone(zone):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT DISTINCT
                    cb.C_BPartner_ID AS "CLIENT_ID",
                    cb.name AS "CLIENT_NAME",
                    sr.name AS "ZONE"
                FROM C_SalesRegion sr
                JOIN C_BPartner_Location bpl ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                JOIN C_BPartner cb ON cb.C_BPartner_ID = bpl.C_BPartner_ID
                JOIN xx_ca_fournisseur xf ON bpl.C_BPartner_ID = xf.CLIENTID
                WHERE UPPER(sr.name) = UPPER(:zone)
                    AND xf.AD_Org_ID = 1000000
                    AND xf.DOCSTATUS != 'RE'
                ORDER BY cb.name
            """
            
            params = {
                'zone': zone
            }
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching clients by zone: {e}")
        return {"error": "An error occurred while fetching clients by zone."}




@app.route('/listfournisseur_etat')
def list_fournisseur_etat():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT cb.name, cb.C_BPartner_ID
                FROM C_BPartner cb
                WHERE cb.AD_Client_ID = 1000000
                  AND cb.ISVENDOR = 'Y'
                  AND cb.ISACTIVE = 'Y'
                ORDER BY cb.name
            """
            cursor.execute(query)
            # Return both name and ID as a list of dictionaries
            result = [{"name": row[0], "id": row[1]} for row in cursor.fetchall()]
            return jsonify(result)
    except Exception as e:
        logger.error(f"Error fetching fournisseurs: {e}")
        return jsonify({"error": "Could not fetch fournisseur list"}), 500

def fetch_etat_f(date1, date2, c_bpartner_id=None, ispaid=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT ispaid, ad_org_id, c_bpartner_id, dateinvoiced, nbl, nfact, documentno, grandtotal, verse_fact, 
                       verse_cheque, cheque, name, dateversement, nb, region,  
                       SUM(grandtotal/nb) OVER(PARTITION BY c_bpartner_id) AS bp_chiffre,
                       SUM(verse_fact/nb) OVER(PARTITION BY c_bpartner_id) AS verse_tot,
                       orgname, phone, phone2, fax,
                       address1, address2, address3, address4, city, postal
                FROM 
                (SELECT cs.ispaid, cs.ad_org_id, cs.c_bpartner_id, cs.dateinvoiced, cs.nbl, cs.nfact, cs.documentno AS documentno, 
                        cs.grandtotal, cs.verse_fact, 
                        cs.verse_cheque, cs.cheque, cs.name, cs.dateversement, cs.region,  
                        COUNT(cs.c_invoice_id) OVER(PARTITION BY cs.c_invoice_id) AS nb,
                        org.name as orgname, oi.phone, oi.phone2, oi.fax,
                        loc.address1, loc.address2, loc.address3, loc.address4, loc.city, loc.postal
                 FROM xx_vendor_status cs  
                 INNER JOIN ad_org org ON (cs.ad_org_id = org.ad_org_id)
                 INNER JOIN ad_orginfo oi ON (org.ad_org_id = oi.ad_org_id)
                 INNER JOIN c_location loc ON (oi.c_location_id = loc.c_location_id)
                 WHERE cs.AD_Client_ID = 1000000
                 AND cs.AD_Org_ID = 1000000
                 AND (:c_bpartner_id IS NULL OR cs.C_BPartner_ID = :c_bpartner_id)
                 AND cs.dateinvoiced BETWEEN TO_DATE(:date1, 'DD/MM/YYYY') AND TO_DATE(:date2, 'DD/MM/YYYY')
                 AND (:ispaid IS NULL OR :ispaid = '' OR cs.ispaid = :ispaid)
                 ORDER BY cs.dateinvoiced, cs.nbl, cs.nfact)
                ORDER BY dateinvoiced, nbl, nfact
            """
            
            params = {
                'date1': date1,
                'date2': date2,
                'c_bpartner_id': c_bpartner_id,
                'ispaid': ispaid
            }
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            
            data = [dict(zip(columns, row)) for row in rows]
            return data
            
    except Exception as e:
        logger.error(f"Error fetching etat_f data: {e}")
        return {"error": "An error occurred while fetching vendor status data."}

@app.route('/etat_f', methods=['GET'])
def etat_f():
    try:
        date1 = request.args.get('date1')
        date2 = request.args.get('date2')
        c_bpartner_id = request.args.get('c_bpartner_id')
        ispaid = request.args.get('ispaid')
        
        # Validate required parameters
        if not date1 or not date2:
            return jsonify({"error": "date1 and date2 are required parameters"}), 400
        
        # Convert c_bpartner_id to int if provided
        if c_bpartner_id:
            try:
                c_bpartner_id = int(c_bpartner_id)
            except ValueError:
                return jsonify({"error": "c_bpartner_id must be a valid integer"}), 400
        
        data = fetch_etat_f(date1, date2, c_bpartner_id, ispaid)
        
        if isinstance(data, dict) and "error" in data:
            return jsonify(data), 500
            
        return jsonify(data)
        
    except Exception as e:
        logger.error(f"Error in etat_f route: {e}")
        return jsonify({"error": "An error occurred while processing the request"}), 500


@app.route('/fetch_etat_fournisseur_cumule')
def fetch_etat_fournisseur_cumule():
    try:
        # Get parameters from request
        c_bpartner_id = request.args.get('c_bpartner_id', type=int)
        start_date = request.args.get('start_date', '01-01-2025')
        end_date = request.args.get('end_date', '01-06-2025')
        
        # Validate required parameters
        if not c_bpartner_id:
            return jsonify({"error": "c_bpartner_id parameter is required"}), 400
        
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            
            # SQL query based on your provided SQL with character set fixes
            query = """
                SELECT inv.DATEINVOICED AS DateTrx ,
                inv.DOCUMENTNO AS DOC_ID,
                inv.POREFERENCE as N_BL,
                doc.PrintName as DOC_TYPE,
                CASE
                  WHEN inv.DESCRIPTION IS NULL
                  THEN
                    (SELECT pro.NAME
                    FROM M_PRODUCT pro,
                      C_INVOICELINE il
                    WHERE inv.C_INVOICE_ID=il.C_INVOICE_ID
                    AND il.M_PRODUCT_ID   =pro.M_PRODUCT_ID
                    AND rownum=1
                    )
                  ELSE inv.DESCRIPTION
                END AS DESCRIPTION,
                CASE
                  WHEN (doc.docbasetype IN ('APC') OR doc.C_DocType_ID =1001510)
                  THEN inv.GRANDTOTAL   *-1
                  ELSE inv.GRANDTOTAL
                END AS "MONTANT"
              FROM C_INVOICE inv,
                C_DOCTYPE doc
              WHERE inv.C_DOCTYPE_ID =doc.C_DOCTYPE_ID
              AND inv.DOCSTATUS     IN ('CO','CL')
              AND doc.docbasetype   IN ('API','APC')
              AND inv.C_BPARTNER_ID  = :c_bpartner_id
              AND inv.AD_Client_ID = 1000000
              AND inv.AD_Org_ID = 1000000
              AND (inv.DATEINVOICED >= TO_DATE(:start_date, 'DD-MM-YYYY')
              AND inv.DATEINVOICED  <= TO_DATE(:end_date, 'DD-MM-YYYY') )
              UNION
              SELECT pa.DATETRX AS DateTrx,
                pa.DOCUMENTNO,
                NULL as POREFERENCE,
                N'Discount' AS NAME,
                N''         AS DESCRIPTION,
                pa.discountamt * -1
              FROM C_PAYMENT pa ,
                C_DOCTYPE doc
              WHERE pa.C_DOCTYPE_ID=doc.C_DOCTYPE_ID
              AND pa.DOCSTATUS    IN ('CO','CL')
              AND doc.docbasetype IN ('APP')
              AND pa.C_BPARTNER_ID = :c_bpartner_id
              AND pa.AD_Client_ID = 1000000
              AND pa.AD_Org_ID = 1000000
              AND (pa.DATETRX     >= TO_DATE(:start_date, 'DD-MM-YYYY')
              AND pa.DATETRX      <= TO_DATE(:end_date, 'DD-MM-YYYY') )
              AND pa.discountamt   > 0
              UNION
              SELECT pa.DATETRX AS DateTrx,
                pa.DOCUMENTNO,
                NULL as POREFERENCE,
                doc.Printname as NAME,
                CASE
                  WHEN pa.DESCRIPTION IS NOT NULL
                  THEN pa.DESCRIPTION
                  WHEN pa.C_INVOICE_ID IS NOT NULL
                  THEN
                    (SELECT inv.DOCUMENTNO
                    FROM C_INVOICE inv
                    WHERE inv.C_INVOICE_ID=pa.C_INVOICE_ID
                    AND rownum=1
                    )
                  ELSE N''
                END AS DESCRIPTION,
                (pa.PAYAMT * -1)
              FROM C_PAYMENT pa ,
                C_DOCTYPE doc
              WHERE pa.C_DOCTYPE_ID=doc.C_DOCTYPE_ID
              AND pa.DOCSTATUS    IN ('CO','CL')
              AND doc.docbasetype IN ('APP')
              AND pa.C_BPARTNER_ID = :c_bpartner_id
              AND pa.AD_Client_ID = 1000000
              AND pa.AD_Org_ID = 1000000
              AND (pa.DATETRX     >= TO_DATE(:start_date, 'DD-MM-YYYY')
              AND pa.DATETRX      <= TO_DATE(:end_date, 'DD-MM-YYYY') )

              UNION
              select DATETRX as DateTrx , 
              DOCUMENTNO as DOCUMENTNO  ,
              NULL as POREFERENCE,
              doc.PrintName  as  name , translate ('DIFFÉRENCE_' using nchar_cs) as DESCRIPTION ,  COALESCE( (select  sum(al.writeoffamt* -1) from C_ALLOCATIONLINE al where (al.C_PAYMENT_ID = par.C_PAYMENT_ID and par.C_BPartner_ID = :c_bpartner_id) ),0) 
                            
                            from C_PAYMENT par , c_doctype doc
                            where doc.docbasetype IN ('APP')
                            and par.DOCSTATUS IN ('CO','CL')
                            and  par.C_BPARTNER_ID = :c_bpartner_id
                            and par.AD_Client_ID = 1000000
                            and par.AD_Org_ID = 1000000
                            AND par.C_DOCTYPE_ID=doc.C_DOCTYPE_ID
                            AND (par.DATETRX     >= TO_DATE(:start_date, 'DD-MM-YYYY')
                    AND par.DATETRX      <= TO_DATE(:end_date, 'DD-MM-YYYY'))
                            and COALESCE( (select  sum(al.writeoffamt* -1) from C_ALLOCATIONLINE al where (al.C_PAYMENT_ID = par.C_PAYMENT_ID and par.C_BPartner_ID = :c_bpartner_id) ),0) <> 0              
                            

              UNION
              SELECT c.StatementDATE AS DateTrx,
                c.Name As DOCUMENTNO,
                i.POREFERENCE as POREFERENCE,
                N'Facture sur Caisse' as NAME,
              CASE
                  WHEN cl.DESCRIPTION IS NOT NULL
                  THEN cl.DESCRIPTION
                  WHEN cl.C_INVOICE_ID IS NOT NULL
                  THEN
                    i.DOCUMENTNO
                  ELSE N''
                END  AS DESCRIPTION,
                cl.Amount * -1
              FROM C_CashLine cl
              INNER JOIN  C_Cash c ON (cl.C_Cash_ID=c.C_Cash_ID)
              INNER JOIN C_Invoice i ON (cl.C_Invoice_ID=i.C_Invoice_ID)
              WHERE 
              c.DOCSTATUS    IN ('CO','CL')
              AND i.ispaid='Y'
              AND cl.isactive='Y'
              AND i.C_BPARTNER_ID = :c_bpartner_id
              AND i.AD_Client_ID = 1000000
              AND i.AD_Org_ID = 1000000
              AND (c.StatementDATE     >= TO_DATE(:start_date, 'DD-MM-YYYY')
              AND c.StatementDATE      <= TO_DATE(:end_date, 'DD-MM-YYYY') )
              ORDER BY DateTrx
            """
            
            # Execute query with parameters
            cursor.execute(query, {
                'c_bpartner_id': c_bpartner_id,
                'start_date': start_date,
                'end_date': end_date
            })
            
            # Fetch results and convert to list of dictionaries
            columns = [desc[0] for desc in cursor.description]
            result = []
            for row in cursor.fetchall():
                row_dict = {}
                for i, value in enumerate(row):
                    # Convert Oracle date objects to string format
                    if hasattr(value, 'strftime'):
                        row_dict[columns[i]] = value.strftime('%Y-%m-%d')
                    elif value is None:
                        row_dict[columns[i]] = None
                    else:
                        row_dict[columns[i]] = value
                result.append(row_dict)
            
            return jsonify(result)
            
    except Exception as e:
        logger.error(f"Error fetching supplier cumulative statement: {e}")
        return jsonify({"error": f"Could not fetch supplier cumulative statement: {str(e)}"}), 500



@app.route('/fetch_etat_fournisseur_cumule_paiment')
def fetch_etat_fournisseur_cumule_paiment():
    try:
        # Get parameters from request
        c_bpartner_id = request.args.get('c_bpartner_id', type=int)
        start_date = request.args.get('start_date', '01-01-2025')
        end_date = request.args.get('end_date', '01-06-2025')
        
        if not c_bpartner_id:
            return jsonify({"error": "c_bpartner_id parameter is required"}), 400

        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            
            query = """
                SELECT inv.DATEINVOICED AS DateTrx,
                       inv.DOCUMENTNO AS DOC_ID,
                       inv.POREFERENCE as N_BL,
                       doc.PrintName as DOC_TYPE,
                       CASE
                         WHEN inv.DESCRIPTION IS NULL THEN
                           (SELECT pro.NAME
                              FROM M_PRODUCT pro, C_INVOICELINE il
                             WHERE inv.C_INVOICE_ID = il.C_INVOICE_ID
                               AND il.M_PRODUCT_ID = pro.M_PRODUCT_ID
                               AND rownum = 1)
                         ELSE inv.DESCRIPTION
                       END AS DESCRIPTION,
                       CASE
                         WHEN (doc.docbasetype IN ('APC'))
                         THEN inv.GRANDTOTAL * -1
                         ELSE inv.GRANDTOTAL
                       END AS "MONTANT"
                FROM C_INVOICE inv,
                     C_DOCTYPE doc
                WHERE inv.C_DOCTYPE_ID = doc.C_DOCTYPE_ID
                  AND inv.DOCSTATUS IN ('CO','CL')
                  AND doc.docbasetype IN ('API','APC')
                  AND doc.C_DocType_ID NOT IN (1001510, 1001509, 1002841)
                  AND inv.C_BPARTNER_ID = :c_bpartner_id
                  AND inv.AD_Client_ID = 1000000
                  AND inv.AD_Org_ID = 1000000
                  AND inv.DATEINVOICED BETWEEN TO_DATE(:start_date, 'DD-MM-YYYY') AND TO_DATE(:end_date, 'DD-MM-YYYY')

                UNION

                SELECT pa.DATETRX AS DateTrx,
                       pa.DOCUMENTNO,
                       NULL as POREFERENCE,
                       N'Discount' AS NAME,
                       N'' AS DESCRIPTION,
                       pa.discountamt * -1
                FROM C_PAYMENT pa,
                     C_DOCTYPE doc
                WHERE pa.C_DOCTYPE_ID = doc.C_DOCTYPE_ID
                  AND pa.DOCSTATUS IN ('CO','CL')
                  AND doc.docbasetype IN ('APP')
                  AND doc.C_DocType_ID NOT IN (1001510, 1001509, 1002841)
                  AND pa.C_BPARTNER_ID = :c_bpartner_id
                  AND pa.AD_Client_ID = 1000000
                  AND pa.AD_Org_ID = 1000000
                  AND pa.DATETRX BETWEEN TO_DATE(:start_date, 'DD-MM-YYYY') AND TO_DATE(:end_date, 'DD-MM-YYYY')
                  AND pa.discountamt > 0

                UNION

                SELECT pa.DATETRX AS DateTrx,
                       pa.DOCUMENTNO,
                       NULL as POREFERENCE,
                       doc.Printname as NAME,
                       CASE
                         WHEN pa.DESCRIPTION IS NOT NULL THEN pa.DESCRIPTION
                         WHEN pa.C_INVOICE_ID IS NOT NULL THEN
                           (SELECT inv.DOCUMENTNO
                              FROM C_INVOICE inv
                             WHERE inv.C_INVOICE_ID = pa.C_INVOICE_ID
                               AND rownum = 1)
                         ELSE N''
                       END AS DESCRIPTION,
                       (pa.PAYAMT * -1)
                FROM C_PAYMENT pa,
                     C_DOCTYPE doc
                WHERE pa.C_DOCTYPE_ID = doc.C_DOCTYPE_ID
                  AND pa.DOCSTATUS IN ('CO','CL')
                  AND doc.docbasetype IN ('APP')
                  AND doc.C_DocType_ID NOT IN (1001510, 1001509, 1002841)
                  AND pa.C_BPARTNER_ID = :c_bpartner_id
                  AND pa.AD_Client_ID = 1000000
                  AND pa.AD_Org_ID = 1000000
                  AND pa.DATETRX BETWEEN TO_DATE(:start_date, 'DD-MM-YYYY') AND TO_DATE(:end_date, 'DD-MM-YYYY')

                UNION

                SELECT DATETRX AS DateTrx,
                       DOCUMENTNO,
                       NULL AS POREFERENCE,
                       doc.PrintName AS NAME,
                       TRANSLATE('DIFFÉRENCE_' USING NCHAR_CS) AS DESCRIPTION,
                       COALESCE(
                         (SELECT SUM(al.writeoffamt * -1)
                          FROM C_ALLOCATIONLINE al
                          WHERE al.C_PAYMENT_ID = par.C_PAYMENT_ID
                            AND par.C_BPartner_ID = :c_bpartner_id),
                         0) AS MONTANT
                FROM C_PAYMENT par,
                     C_DOCTYPE doc
                WHERE par.C_DOCTYPE_ID = doc.C_DOCTYPE_ID
                  AND doc.docbasetype IN ('APP')
                  AND doc.C_DocType_ID NOT IN (1001510, 1001509, 1002841)
                  AND par.DOCSTATUS IN ('CO','CL')
                  AND par.C_BPARTNER_ID = :c_bpartner_id
                  AND par.AD_Client_ID = 1000000
                  AND par.AD_Org_ID = 1000000
                  AND par.DATETRX BETWEEN TO_DATE(:start_date, 'DD-MM-YYYY') AND TO_DATE(:end_date, 'DD-MM-YYYY')
                  AND COALESCE(
                        (SELECT SUM(al.writeoffamt * -1)
                         FROM C_ALLOCATIONLINE al
                         WHERE al.C_PAYMENT_ID = par.C_PAYMENT_ID
                           AND par.C_BPartner_ID = :c_bpartner_id),
                        0) <> 0

                UNION

                SELECT c.StatementDATE AS DateTrx,
                       c.Name AS DOCUMENTNO,
                       i.POREFERENCE,
                       N'Facture sur Caisse' AS NAME,
                       CASE
                         WHEN cl.DESCRIPTION IS NOT NULL THEN cl.DESCRIPTION
                         WHEN cl.C_INVOICE_ID IS NOT NULL THEN i.DOCUMENTNO
                         ELSE N''
                       END AS DESCRIPTION,
                       cl.Amount * -1 AS MONTANT
                FROM C_CashLine cl
                INNER JOIN C_Cash c ON cl.C_Cash_ID = c.C_Cash_ID
                INNER JOIN C_Invoice i ON cl.C_Invoice_ID = i.C_Invoice_ID
                WHERE c.DOCSTATUS IN ('CO','CL')
                  AND i.ispaid = 'Y'
                  AND cl.isactive = 'Y'
                  AND i.C_BPARTNER_ID = :c_bpartner_id
                  AND i.AD_Client_ID = 1000000
                  AND i.AD_Org_ID = 1000000
                  AND c.StatementDATE BETWEEN TO_DATE(:start_date, 'DD-MM-YYYY') AND TO_DATE(:end_date, 'DD-MM-YYYY')
                
                ORDER BY DateTrx
            """
            
            cursor.execute(query, {
                'c_bpartner_id': c_bpartner_id,
                'start_date': start_date,
                'end_date': end_date
            })
            
            columns = [desc[0] for desc in cursor.description]
            result = []
            for row in cursor.fetchall():
                row_dict = {}
                for i, value in enumerate(row):
                    if hasattr(value, 'strftime'):
                        row_dict[columns[i]] = value.strftime('%Y-%m-%d')
                    elif value is None:
                        row_dict[columns[i]] = None
                    else:
                        row_dict[columns[i]] = value
                result.append(row_dict)
            
            return jsonify(result)
            
    except Exception as e:
        logger.error(f"Error fetching supplier cumulative statement: {e}")
        return jsonify({"error": f"Could not fetch supplier cumulative statement: {str(e)}"}), 500


@app.route('/sold_initial_etat_cum')
def sold_initial_etat_cum():
    try:
        # Get parameters from request
        c_bpartner_id = request.args.get('c_bpartner_id', type=int)
        start_date = request.args.get('start_date', '01-01-2025')
        
        # Validate required parameters
        if not c_bpartner_id:
            return jsonify({"error": "c_bpartner_id parameter is required"}), 400
        
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            
            # SQL query to calculate opening balance
            query = """
                SELECT (
                    (SELECT COALESCE(SUM(iar.GRANDTOTAL), 0) 
                     FROM C_INVOICE_V iar
                     WHERE iar.DOCSTATUS IN ('CO','CL') 
                     AND iar.ISSOTRX = 'N'
                     AND iar.C_BPARTNER_ID = :c_bpartner_id
                     AND iar.DATEINVOICED < TO_DATE(:start_date, 'DD-MM-YYYY'))
                     
                    -
                    
                    (SELECT COALESCE(SUM(par.PAYAMT) + SUM(par.discountamt), 0)
                     FROM C_PAYMENT par
                     WHERE par.DOCSTATUS IN ('CO','CL') 
                     AND par.ISRECEIPT = 'N'
                     AND par.C_BPARTNER_ID = :c_bpartner_id
                     AND par.DATETRX < TO_DATE(:start_date, 'DD-MM-YYYY'))
                     
                    -
                    
                    (SELECT COALESCE(SUM(cl.Amount), 0)
                     FROM C_CashLine cl
                     INNER JOIN C_Cash c ON (cl.C_Cash_ID = c.C_Cash_ID)
                     INNER JOIN C_Invoice i ON (cl.C_Invoice_ID = i.C_Invoice_ID)
                     WHERE c.DOCSTATUS IN ('CO','CL')
                     AND i.ispaid = 'Y'
                     AND cl.isactive = 'Y'
                     AND i.C_BPARTNER_ID = :c_bpartner_id
                     AND c.StatementDATE < TO_DATE(:start_date, 'DD-MM-YYYY'))
                ) AS OpeningBal 
                FROM DUAL
            """
            
            # Execute query with parameters
            cursor.execute(query, {
                'c_bpartner_id': c_bpartner_id,
                'start_date': start_date
            })
            
            # Fetch result
            result = cursor.fetchone()
            opening_balance = result[0] if result else 0
            
            return jsonify({
                "c_bpartner_id": c_bpartner_id,
                "start_date": start_date,
                "opening_balance": float(opening_balance) if opening_balance else 0.0
            })
            
    except Exception as e:
        logger.error(f"Error fetching opening balance: {e}")
        return jsonify({"error": f"Could not fetch opening balance: {str(e)}"}), 500







def fetch_reversed_voided_stock_movements(product_id):
    """
    Fetch stock movements for a specific product where docstatus is 'RE' (Reversed) or 'VO' (Voided)
    Calculate the difference between entries and exits
    """
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            
            query = """
            SELECT
                t.MovementDate AS MovementDate,
                nvl(nvl(io.documentno,inv.documentno),m.documentno) as documentno,
                nvl(bp.name, nvl(inv.description,m.description)) as name,
                p.name AS productname,
                CASE WHEN t.movementqty > 0 then t.movementqty else 0 end as ENTREE,
                CASE WHEN t.movementqty < 0 then ABS(t.movementqty) else 0 end as SORTIE,
                asi.lot,
                l.value AS locator,
                COALESCE(io.docstatus, m.docstatus, inv.docstatus, 'N/A') AS docstatus
            FROM M_Transaction t
            INNER JOIN M_Locator l ON (t.M_Locator_ID=l.M_Locator_ID)
            INNER JOIN M_Product p ON (t.M_Product_ID=p.M_Product_ID)
            LEFT OUTER JOIN M_InventoryLine il ON (t.M_InventoryLine_ID=il.M_InventoryLine_ID)
            LEFT OUTER JOIN M_Inventory inv ON (inv.m_inventory_id = il.m_inventory_id)
            LEFT OUTER JOIN M_MovementLine ml ON (t.M_MovementLine_ID=ml.M_MovementLine_ID)
            LEFT OUTER JOIN M_Movement m ON (m.M_Movement_ID=ml.M_Movement_ID)
            LEFT OUTER JOIN M_InOutLine iol ON (t.M_InOutLine_ID=iol.M_InOutLine_ID)
            LEFT OUTER JOIN M_Inout io ON (iol.M_InOut_ID=io.M_InOut_ID)
            LEFT OUTER JOIN C_BPartner bp ON (bp.C_BPartner_ID = io.C_BPartner_ID)
            INNER JOIN M_attributesetinstance asi on t.m_attributesetinstance_id = asi.m_attributesetinstance_id
            INNER JOIN M_attributeinstance att on (att.m_attributesetinstance_id = asi.m_attributesetinstance_id)
            WHERE
            att.m_attribute_id = 1000508
            AND p.M_Product_ID = :product_id
            AND COALESCE(io.docstatus, m.docstatus, inv.docstatus) IN ('RE', 'VO')
            AND NOT (t.movementqty = 0)
            AND t.AD_Client_ID = 1000000
            ORDER BY t.MovementDate DESC
            """
            
            cursor.execute(query, {'product_id': product_id})
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            movements = [dict(zip(columns, row)) for row in rows]
            
            # Calculate totals
            total_entree = sum(float(row.get('ENTREE', 0) or 0) for row in movements)
            total_sortie = sum(float(row.get('SORTIE', 0) or 0) for row in movements)
            difference = total_entree - total_sortie
            
            return {
                'movements': movements,
                'summary': {
                    'total_entree': total_entree,
                    'total_sortie': total_sortie,
                    'difference': difference,
                    'product_id': product_id,
                    'total_movements': len(movements)
                }
            }
            
    except Exception as e:
        logger.error(f"Database error in fetch_reversed_voided_stock_movements: {e}")
        return {"error": "An error occurred while fetching reversed/voided stock movements."}

@app.route('/fetch-reversed-voided-stock-movements', methods=['GET'])
def fetch_reversed_voided_movements():
    """
    API endpoint to get stock movements for a product with docstatus 'RE' or 'VO'
    Expected parameter: product_id
    """
    try:
        product_id = request.args.get("product_id", None)
        
        if not product_id:
            return jsonify({"error": "product_id parameter is required"}), 400
        
        try:
            product_id = int(product_id)
        except ValueError:
            return jsonify({"error": "product_id must be a valid integer"}), 400
        
        data = fetch_reversed_voided_stock_movements(product_id)
        return jsonify(data)
        
    except Exception as e:
        logger.error(f"Error in fetch_reversed_voided_movements endpoint: {e}")
        return jsonify({"error": "An error occurred while processing the request"}), 500





@app.route('/inventory-products', methods=['GET'])
def inventory_products():
    try:
        product_id = request.args.get("product_id", None)
        
        if not product_id:
            return jsonify({"error": "Product ID is required"}), 400

        data = fetch_inventory_products_data(product_id)
        return jsonify(data)

    except Exception as e:
        logger.error(f"Error fetching inventory products: {e}")
        return jsonify({"error": "Failed to fetch inventory products"}), 500


def fetch_inventory_products_data(product_id):
    """
    Fetch inventory product information - returns Product, Lot, PPA, QTY_DISPO, Guarantee Date
    """
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            
            query = """
            SELECT
                p.name AS PRODUCT,
                (
                    SELECT
                        lot
                    FROM
                        m_attributesetinstance
                    WHERE
                        m_attributesetinstance_id = mst.m_attributesetinstance_id
                ) AS LOT,
                (
                    SELECT
                        valuenumber
                    FROM
                        m_attributeinstance
                    WHERE
                        m_attributesetinstance_id = mst.m_attributesetinstance_id
                        AND m_attribute_id = 1000503
                ) AS PPA,
                (mst.qtyonhand - mst.QTYRESERVED) AS QTY_DISPO,
                mats.guaranteedate AS GUARANTEEDATE
            FROM
                m_product p
                INNER JOIN m_storage mst ON p.m_product_id = mst.m_product_id
                INNER JOIN m_attributesetinstance mats ON mst.m_attributesetinstance_id = mats.m_attributesetinstance_id
            WHERE
                p.m_product_id = :product_id
                AND mst.m_locator_id IN (1001135, 1000614, 1001128, 1001136, 1001020)
                AND mst.qtyonhand != 0
            ORDER BY
                p.name, mats.guaranteedate
            """

            cursor.execute(query, {"product_id": product_id})
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]

            return data

    except Exception as e:
        logger.error(f"Error fetching inventory products: {e}")
        return {"error": "An error occurred while fetching inventory products."}

@app.route('/listproduct_inv', methods=['GET'])
def listproduct_inv():
    """
    Returns list of products with both ID and name for inventory management
    """
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
            SELECT M_product_id, name FROM M_PRODUCT
            WHERE AD_Client_ID = 1000000
            AND AD_Org_ID = 1000000
            AND ISACTIVE = 'Y'
            ORDER BY name
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            # Return array of objects with id and name
            products = [{"id": row[0], "name": row[1]} for row in rows]
            return jsonify(products)
    except Exception as e:
        logger.error(f"Error fetching product list: {e}")
        return jsonify({"error": "Could not fetch products list"}), 500



def fetch_simulation():

    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT * FROM (
                    SELECT 
                        CAST(org.name AS VARCHAR2(300)) AS organisation,
                        CAST(co.documentno AS VARCHAR2(50)) AS ndocument,
                        CAST(cb.name AS VARCHAR2(300)) AS tier,
                        co.dateordered AS datecommande,
                        CAST(us.name AS VARCHAR2(100)) AS vendeur,
                        ROUND(((co.totallines / (SELECT SUM(mat.valuenumber * li.qtyentered) 
                             FROM c_orderline li 
                             INNER JOIN m_attributeinstance mat ON mat.m_attributesetinstance_id = li.m_attributesetinstance_id
                             WHERE mat.m_attribute_id = 1000504 
                               AND li.c_order_id = co.c_order_id 
                               AND li.qtyentered > 0 
                             GROUP BY li.c_order_id)) - 1) * 100, 2) AS marge,
                        ROUND(co.totallines, 2) AS montant,
                        1 AS sort_order
                    FROM 
                        c_order co
                    INNER JOIN ad_org org ON co.ad_org_id = org.ad_org_id
                    INNER JOIN c_bpartner cb ON co.c_bpartner_id = cb.c_bpartner_id
                    INNER JOIN ad_user us ON co.salesrep_id = us.ad_user_id
                    WHERE 
                         co.docaction = 'PR'
                        AND co.ad_org_id = 1000000
                        AND docstatus = 'IP'
                        AND issotrx = 'Y'
                    
                    UNION ALL
                    
                    SELECT 
                        CAST('Total' AS VARCHAR2(300)) AS organisation,
                        CAST(NULL AS VARCHAR2(50)) AS ndocument,
                        CAST(NULL AS VARCHAR2(300)) AS tier,
                        NULL AS datecommande,
                        CAST(NULL AS VARCHAR2(100)) AS vendeur,
                        ROUND(AVG(ROUND(((co.totallines / (SELECT SUM(mat.valuenumber * li.qtyentered) 
                             FROM c_orderline li 
                             INNER JOIN m_attributeinstance mat ON mat.m_attributesetinstance_id = li.m_attributesetinstance_id
                             WHERE mat.m_attribute_id = 1000504 
                               AND li.c_order_id = co.c_order_id 
                               AND li.qtyentered > 0 
                             GROUP BY li.c_order_id)) - 1) * 100, 2)), 2) AS marge,
                        ROUND(SUM(co.totallines), 2) AS montant,
                        0 AS sort_order
                    FROM 
                        c_order co
                    INNER JOIN ad_org org ON co.ad_org_id = org.ad_org_id
                    INNER JOIN c_bpartner cb ON co.c_bpartner_id = cb.c_bpartner_id
                    INNER JOIN ad_user us ON co.salesrep_id = us.ad_user_id
                    WHERE 
                         co.docaction = 'PR'
                        AND co.ad_org_id = 1000000
                        AND docstatus = 'IP'
                        AND issotrx = 'Y'
                )
                ORDER BY sort_order, montant DESC
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            result = [dict(zip(columns, row)) for row in rows]
            return result
    except Exception as e:
        logging.error(f"Error fetching simulation data: {e}")
        return {"error": "An error occurred while fetching simulation data."}

# New: fetch simulation by ndocument
def fetch_simulation_by_ndocument(ndocument):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT 
                    CAST(org.name AS VARCHAR2(300)) AS organisation,
                    CAST(co.documentno AS VARCHAR2(50)) AS ndocument,
                    CAST(cb.name AS VARCHAR2(300)) AS tier,
                    co.dateordered AS datecommande,
                    CAST(us.name AS VARCHAR2(100)) AS vendeur,
                    ROUND(((co.totallines / (SELECT SUM(mat.valuenumber * li.qtyentered) 
                         FROM c_orderline li 
                         INNER JOIN m_attributeinstance mat ON mat.m_attributesetinstance_id = li.m_attributesetinstance_id
                         WHERE mat.m_attribute_id = 1000504 
                           AND li.c_order_id = co.c_order_id 
                           AND li.qtyentered > 0 
                         GROUP BY li.c_order_id)) - 1) * 100, 2) AS marge,
                    ROUND(co.totallines, 2) AS montant
                FROM 
                    c_order co
                INNER JOIN ad_org org ON co.ad_org_id = org.ad_org_id
                INNER JOIN c_bpartner cb ON co.c_bpartner_id = cb.c_bpartner_id
                INNER JOIN ad_user us ON co.salesrep_id = us.ad_user_id
                WHERE 
                     co.ad_org_id = 1000000
                    AND issotrx = 'Y'
                    AND co.documentno = :ndocument
            """
            cursor.execute(query, {'ndocument': ndocument})
            row = cursor.fetchone()
            if row:
                columns = [col[0] for col in cursor.description]
                return dict(zip(columns, row))
            else:
                return None
    except Exception as e:
        logging.error(f"Error fetching simulation data by ndocument: {e}")
        return {"error": "An error occurred while fetching simulation data by ndocument."}





@app.route('/simulation', methods=['GET'])
def get_simulation():
    result = fetch_simulation()
    return jsonify(result)

# New endpoint: /simulation_all?ndocument=xxx
from flask import request

@app.route('/simulation_all', methods=['GET'])
def get_simulation_all():
    ndocument = request.args.get('ndocument')
    if not ndocument:
        return jsonify({'error': 'Missing ndocument parameter'}), 400
    result = fetch_simulation_by_ndocument(ndocument)
    if result:
        return jsonify(result)
    else:
        return jsonify({'error': 'No data found for the given ndocument'}), 404







def fetch_charges_dashboard(date_debut=None, date_fin=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            
            # Set default dates if not provided
            if not date_debut:
                date_debut = '2024-01-01'  # Default start date
            if not date_fin:
                date_fin = datetime.now().strftime('%Y-%m-%d')  # Default end date (today)
            
            # First query to get charge types and charges with totals
            summary_query = """
            SELECT 
                ct.C_ChargeType_ID,
                ct.name AS type_de_charge,
                c.C_Charge_ID,
                c.name AS charge_name,
                COALESCE(charge_amounts.total_amount, 0) AS montant,
                CASE 
                    WHEN type_totals.type_total > 0 THEN 
                        ROUND((COALESCE(charge_amounts.total_amount, 0) / type_totals.type_total) * 100, 3)
                    ELSE 0 
                END AS pourcentage
            FROM C_ChargeType ct
            INNER JOIN C_Charge c ON ct.C_ChargeType_ID = c.C_ChargeType_ID
            LEFT JOIN (
                SELECT 
                    c.C_Charge_ID,
                    SUM(il.LineNetAmt) AS total_amount
                FROM C_Charge c
                INNER JOIN C_InvoiceLine il ON c.C_Charge_ID = il.C_Charge_ID
                INNER JOIN C_Invoice i ON il.C_Invoice_ID = i.C_Invoice_ID
                WHERE c.IsActive = 'Y' 
                  AND c.AD_CLIENT_ID = 1000000
                  AND i.IsActive = 'Y'
                  AND i.DocStatus IN ('CO', 'CL')
                  AND i.DateInvoiced >= TO_DATE(:date_debut, 'YYYY-MM-DD')
                  AND i.DateInvoiced <= TO_DATE(:date_fin, 'YYYY-MM-DD')
                GROUP BY c.C_Charge_ID
            ) charge_amounts ON c.C_Charge_ID = charge_amounts.C_Charge_ID
            LEFT JOIN (
                SELECT 
                    ct.C_ChargeType_ID,
                    SUM(il.LineNetAmt) AS type_total
                FROM C_ChargeType ct
                INNER JOIN C_Charge c ON ct.C_ChargeType_ID = c.C_ChargeType_ID
                INNER JOIN C_InvoiceLine il ON c.C_Charge_ID = il.C_Charge_ID
                INNER JOIN C_Invoice i ON il.C_Invoice_ID = i.C_Invoice_ID
                WHERE ct.IsActive = 'Y' 
                  AND ct.AD_CLIENT_ID = 1000000
                  AND c.IsActive = 'Y'
                  AND i.IsActive = 'Y'
                  AND i.DocStatus IN ('CO', 'CL')
                  AND i.DateInvoiced >= TO_DATE(:date_debut, 'YYYY-MM-DD')
                  AND i.DateInvoiced <= TO_DATE(:date_fin, 'YYYY-MM-DD')
                GROUP BY ct.C_ChargeType_ID
            ) type_totals ON ct.C_ChargeType_ID = type_totals.C_ChargeType_ID
            WHERE ct.IsActive = 'Y' 
              AND ct.AD_CLIENT_ID = 1000000
              AND c.IsActive = 'Y'
            ORDER BY ct.name, c.name
            """
            
            cursor.execute(summary_query, {'date_debut': date_debut, 'date_fin': date_fin})
            summary_rows = cursor.fetchall()
            summary_columns = [col[0] for col in cursor.description]
            
            # Second query to get invoice details for each charge with line details
            details_query = """
            SELECT 
                c.C_Charge_ID,
                i.C_Invoice_ID,
                i.DocumentNo AS invoice_number,
                i.DateInvoiced,
                bp.name AS bpartner_name,
                il.LineNetAmt AS line_amount,
                i.GrandTotal AS invoice_total,
                il.C_InvoiceLine_ID,
                il.Description AS line_description,
                il.PriceEntered AS line_total_amount,
                il.QtyEntered AS line_qty_entered
            FROM C_Charge c
            INNER JOIN C_InvoiceLine il ON c.C_Charge_ID = il.C_Charge_ID
            INNER JOIN C_Invoice i ON il.C_Invoice_ID = i.C_Invoice_ID
            INNER JOIN C_BPartner bp ON i.C_BPartner_ID = bp.C_BPartner_ID
            WHERE c.IsActive = 'Y' 
              AND c.AD_CLIENT_ID = 1000000
              AND i.IsActive = 'Y'
              AND i.DocStatus IN ('CO', 'CL')
              AND i.DateInvoiced >= TO_DATE(:date_debut, 'YYYY-MM-DD')
              AND i.DateInvoiced <= TO_DATE(:date_fin, 'YYYY-MM-DD')
            ORDER BY c.C_Charge_ID, i.DateInvoiced DESC, il.C_InvoiceLine_ID
            """
            
            cursor.execute(details_query, {'date_debut': date_debut, 'date_fin': date_fin})
            details_rows = cursor.fetchall()
            details_columns = [col[0] for col in cursor.description]
            
            # Organize invoice details by charge ID
            invoice_details = {}
            for row in details_rows:
                detail_data = dict(zip(details_columns, row))
                charge_id = detail_data['C_CHARGE_ID']
                invoice_id = detail_data['C_INVOICE_ID']
                
                if charge_id not in invoice_details:
                    invoice_details[charge_id] = {}
                
                # Group by invoice ID to avoid duplicating invoice info
                if invoice_id not in invoice_details[charge_id]:
                    invoice_details[charge_id][invoice_id] = {
                        'invoice_id': detail_data['C_INVOICE_ID'],
                        'invoice_number': detail_data['INVOICE_NUMBER'],
                        'date_invoiced': detail_data['DATEINVOICED'].strftime('%Y-%m-%d') if detail_data['DATEINVOICED'] else None,
                        'bpartner_name': detail_data['BPARTNER_NAME'],
                        'invoice_total': -float(detail_data['INVOICE_TOTAL']) if detail_data['INVOICE_TOTAL'] else 0,  # Make negative
                        'invoice_lines': []
                    }
                
                # Add line details to the invoice
                invoice_details[charge_id][invoice_id]['invoice_lines'].append({
                    'line_id': detail_data['C_INVOICELINE_ID'],
                    'line_description': detail_data['LINE_DESCRIPTION'] or 'No description',
                    'line_net_amount': -float(detail_data['LINE_AMOUNT']) if detail_data['LINE_AMOUNT'] else 0,  # Make negative
                    'line_total_amount': -float(detail_data['LINE_TOTAL_AMOUNT']) if detail_data['LINE_TOTAL_AMOUNT'] else 0,  # Make negative
                    'line_qty_entered': float(detail_data['LINE_QTY_ENTERED']) if detail_data['LINE_QTY_ENTERED'] else 0
                })
            
            # Convert nested invoice_details structure to a flat list for easier processing
            flattened_invoice_details = {}
            for charge_id, invoices in invoice_details.items():
                flattened_invoice_details[charge_id] = list(invoices.values())
            
            # Organize data by charge type
            charges_data = {}
            total_all_charges = 0
            
            for row in summary_rows:
                data = dict(zip(summary_columns, row))
                charge_type_id = data['C_CHARGETYPE_ID']
                charge_type_name = data['TYPE_DE_CHARGE']
                charge_id = data['C_CHARGE_ID']
                
                if charge_type_id not in charges_data:
                    charges_data[charge_type_id] = {
                        'type_name': charge_type_name,
                        'type_total': 0,
                        'charges': []
                    }
                
                # Make charge amount negative (expense)
                charge_amount = -float(data['MONTANT']) if data['MONTANT'] else 0
                charges_data[charge_type_id]['type_total'] += charge_amount
                total_all_charges += charge_amount
                
                # Get invoice details for this charge
                charge_invoice_details = flattened_invoice_details.get(charge_id, [])
                
                charges_data[charge_type_id]['charges'].append({
                    'charge_id': charge_id,
                    'charge_name': data['CHARGE_NAME'],
                    'montant': charge_amount,
                    'pourcentage': float(data['POURCENTAGE']) if data['POURCENTAGE'] else 0,
                    'invoice_details': charge_invoice_details
                })
            
            # Calculate percentage for each charge type
            for charge_type_id in charges_data:
                if total_all_charges != 0:  # Changed from > 0 to != 0 since values are negative
                    charges_data[charge_type_id]['type_percentage'] = round(
                        (charges_data[charge_type_id]['type_total'] / total_all_charges) * 100, 3
                    )
                else:
                    charges_data[charge_type_id]['type_percentage'] = 0
            
            return {
                'charges_by_type': charges_data,
                'total_all_charges': total_all_charges,
                'date_debut': date_debut,
                'date_fin': date_fin
            }
            
    except Exception as e:
        logger.error(f"Error fetching charges dashboard: {e}")
        return {"error": "An error occurred while fetching charges dashboard."}

@app.route('/fetch-charges-dashboard', methods=['GET'])
def fetch_charges_dashboard_endpoint():
    if not test_db_connection():
        return jsonify({"error": "Database connection failed"}), 500
    
    # Get date parameters from query string
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')
    
    data = fetch_charges_dashboard(date_debut, date_fin)
    return jsonify(data)

@app.route('/download-charges-dashboard-excel', methods=['GET'])
def download_charges_dashboard_excel():
    # Get date parameters from query string
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')
    
    data = fetch_charges_dashboard(date_debut, date_fin)
    
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 404
    
    # Flatten the data for Excel export with invoice and line details
    excel_data = []
    for charge_type_id, charge_type_data in data['charges_by_type'].items():
        # Only add charge type row if total is not zero
        if charge_type_data['type_total'] != 0:
            excel_data.append({
                'Type de Charge': charge_type_data['type_name'],
                'Charge': 'TOTAL',
                'Montant': charge_type_data['type_total'],
                'Pourcentage': charge_type_data['type_percentage'],
                'Facture N°': '',
                'Date Facture': '',
                'Fournisseur': '',
                'Description Ligne': '',
                'Montant Net Ligne': '',
                'Montant Total Ligne': '',
                'Quantité': ''
            })
        
        # Add individual charges with invoice and line details
        for charge in charge_type_data['charges']:
            # Only process charges with non-zero amounts
            if charge['montant'] != 0:
                if charge['invoice_details']:
                    # Add each invoice with its lines
                    for invoice in charge['invoice_details']:
                        if invoice['invoice_lines']:
                            # Add each line as a separate row
                            for line in invoice['invoice_lines']:
                                # Only add lines with non-zero amounts
                                if line['line_net_amount'] != 0:
                                    excel_data.append({
                                        'Type de Charge': charge_type_data['type_name'],
                                        'Charge': charge['charge_name'],
                                        'Montant': charge['montant'],
                                        'Pourcentage': charge['pourcentage'],
                                        'Facture N°': invoice['invoice_number'],
                                        'Date Facture': invoice['date_invoiced'],
                                        'Fournisseur': invoice['bpartner_name'],
                                        'Description Ligne': line['line_description'],
                                        'Montant Net Ligne': line['line_net_amount'],
                                        'Montant Total Ligne': line['line_total_amount'],
                                        'Quantité': line['line_qty_entered']
                                    })
                        else:
                            # Invoice without lines
                            excel_data.append({
                                'Type de Charge': charge_type_data['type_name'],
                                'Charge': charge['charge_name'],
                                'Montant': charge['montant'],
                                'Pourcentage': charge['pourcentage'],
                                'Facture N°': invoice['invoice_number'],
                                'Date Facture': invoice['date_invoiced'],
                                'Fournisseur': invoice['bpartner_name'],
                                'Description Ligne': 'Aucune ligne de détail',
                                'Montant Net Ligne': '',
                                'Montant Total Ligne': '',
                                'Quantité': ''
                            })
                else:
                    # Add charge without invoice details
                    excel_data.append({
                        'Type de Charge': charge_type_data['type_name'],
                        'Charge': charge['charge_name'],
                        'Montant': charge['montant'],
                        'Pourcentage': charge['pourcentage'],
                        'Facture N°': 'Aucune facture',
                        'Date Facture': '',
                        'Fournisseur': '',
                        'Description Ligne': '',
                        'Montant Net Ligne': '',
                        'Montant Total Ligne': '',
                        'Quantité': ''
                    })
    
    # Add "Total des Charges" row at the end
    if data['total_all_charges'] != 0:
        excel_data.append({
            'Type de Charge': 'TOTAL DES CHARGES',
            'Charge': '',
            'Montant': data['total_all_charges'],
            'Pourcentage': 100.0,
            'Facture N°': '',
            'Date Facture': '',
            'Fournisseur': '',
            'Description Ligne': '',
            'Montant Net Ligne': '',
            'Montant Total Ligne': '',
            'Quantité': ''
        })
    
    # Generate filename with date range
    filename = f"Charges_Dashboard_Details_{data['date_debut']}_to_{data['date_fin']}.xlsx"
    
    return generate_excel_charges_dashboard(excel_data, filename)

def generate_excel_charges_dashboard(data, filename):
    if not data:
        return jsonify({"error": "No data to export"}), 404

    df = pd.DataFrame(data)
    wb = Workbook()
    ws = wb.active
    ws.title = "Charges Dashboard Details"

    # Apply header styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    ws.append(df.columns.tolist())
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Append data with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Highlight total rows (both charge type totals and total des charges)
        if row[1] == 'TOTAL' or row[0] == 'TOTAL DES CHARGES':  # Charge column or Type de Charge column
            for cell in ws[row_idx]:
                cell.font = Font(bold=True)
                if row[0] == 'TOTAL DES CHARGES':
                    # Special highlighting for the grand total
                    cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # Highlight rows with no invoice details
        elif row[4] == 'Aucune facture':  # Facture N° column
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # Highlight rows with no line details
        elif row[7] == 'Aucune ligne de détail':  # Description Ligne column
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="FFF0E6", end_color="FFF0E6", fill_type="solid")

    # Add table
    table = Table(displayName="ChargesDashboardDetailsTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        # Check if the first cell is a merged cell
        if isinstance(column[0], MergedCell):
            continue
        column_letter = column[0].column_letter
        for cell in column:
            if isinstance(cell, MergedCell):
                continue
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, 
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")




@app.route('/retour_documents', methods=['GET'])
def retour_documents():
    """
    Fetch sales documents with detailed information including margin calculation.
    Excludes documents with docstatus = 'DR' (Drafted).
    Can filter by:
    - Date range (start_date and end_date, format: YYYY-MM-DD)
    - Specific document number (ndocument)
    Returns: JSON list of documents with organization, document no, partner, etc.
    """
    from flask import request, jsonify
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    ndocument = request.args.get('ndocument')
    
    if not (start_date and end_date) and not ndocument:
        return jsonify({'error': 'Missing parameters - either start_date/end_date or ndocument required'}), 400
    
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = '''
                SELECT 
                    CAST(org.name AS VARCHAR2(300)) AS organisation,
                    CAST(co.documentno AS VARCHAR2(50)) AS ndocument,
                    CAST(cb.name AS VARCHAR2(300)) AS tier,
                    co.dateordered AS datecommande,
                    CAST(us.name AS VARCHAR2(100)) AS vendeur,
                    CAST(co.description AS VARCHAR2(255)) AS description,
                    ROUND(((co.totallines / (SELECT SUM(mat.valuenumber * li.qtyentered) 
                         FROM c_orderline li 
                         INNER JOIN m_attributeinstance mat ON mat.m_attributesetinstance_id = li.m_attributesetinstance_id
                         WHERE mat.m_attribute_id = 1000504 
                           AND li.c_order_id = co.c_order_id 
                           AND li.qtyentered > 0 
                         GROUP BY li.c_order_id)) - 1) * 100, 2) AS marge,
                    ROUND(co.totallines, 2) AS montant
                FROM 
                    c_order co
                INNER JOIN ad_org org ON co.ad_org_id = org.ad_org_id
                INNER JOIN c_bpartner cb ON co.c_bpartner_id = cb.c_bpartner_id
                INNER JOIN ad_user us ON co.salesrep_id = us.ad_user_id
                WHERE 
                    co.ad_org_id = 1000000
                    AND issotrx = 'Y'
                    AND C_DOCTYPETARGET_ID = 1001408
                    AND docstatus != 'DR'  -- NEW: Exclude drafted documents
                    AND (
                        (:ndocument IS NULL AND co.dateordered BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD'))
                        OR 
                        (co.documentno = :ndocument)
                    )
                ORDER BY co.dateordered DESC, co.documentno
            '''
            params = {
                'start_date': start_date if start_date else '1900-01-01',
                'end_date': end_date if end_date else '2999-12-31',
                'ndocument': ndocument
            }
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            return jsonify(data)
    except Exception as e:
        import logging
        logging.error(f"Error in retour_documents: {e}")
        return jsonify({'error': str(e)}), 500


def fetch_facture_recap_achat(start_date, end_date, partner_name):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT 
                    i.documentno, 
                    i.totallines, 
                    i.description, 
                    i.dateinvoiced, 
                    i.c_bpartner_id, 
                    i.C_Invoice_ID
                FROM C_Invoice i
                JOIN C_BPartner cb ON i.c_bpartner_id = cb.c_bpartner_id
                WHERE (:partner_name IS NULL OR UPPER(cb.name) LIKE UPPER(:partner_name) || '%')
                  AND i.dateinvoiced BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
                and i.docstatus IN ('CO', 'CL')
                AND i.issotrx = 'N'
                ORDER BY i.dateinvoiced DESC
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
                'partner_name': partner_name or None
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()

            data = [
                {
                    "documentno": row[0],
                    "totallines": row[1],
                    "description": row[2],
                    "dateinvoiced": row[3],
                    "c_bpartner_id": row[4],
                    "c_invoice_id": row[5]
                }
                for row in rows
            ]

            return data

    except Exception as e:
        logger.error(f"Error fetching facture recap achat: {e}")
        return {"error": "An error occurred while fetching facture recap achat."}

# Helper function to generate Excel file for invoices
def generate_excel_invoices(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # Define header styling
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Define column headers
    headers = ["Document No", "Total", "Description", "Date Invoiced"]
    ws.append(headers)

    # Apply header styling
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows
    for row_idx, row_data in enumerate(data, start=2):
        # Format date
        date_formatted = ""
        if row_data.get("dateinvoiced"):
            try:
                if isinstance(row_data["dateinvoiced"], str):
                    date_formatted = datetime.strptime(row_data["dateinvoiced"], "%Y-%m-%d").strftime("%Y-%m-%d")
                else:
                    date_formatted = row_data["dateinvoiced"].strftime("%Y-%m-%d")
            except:
                date_formatted = str(row_data["dateinvoiced"])

        row = [
            row_data.get("documentno", ""),
            row_data.get("totallines", 0),
            row_data.get("description", ""),
            date_formatted
        ]
        ws.append(row)

        # Apply alternating row colors
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Auto-fit columns
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        # Check if the first cell is a merged cell
        if isinstance(column[0], MergedCell):
            continue
        for cell in column:
            if isinstance(cell, MergedCell):
                continue
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Create Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response

# Helper function to generate Excel file for invoice lines
def generate_excel_invoice_lines(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Lines"

    # Define header styling
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Define column headers
    headers = ["Product Name", "Quantity", "Line Amount"]
    ws.append(headers)

    # Apply header styling
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows
    for row_idx, row_data in enumerate(data, start=2):
        row = [
            row_data.get("product_name", ""),
            row_data.get("qtyentered", 0),
            row_data.get("linenetamt", 0)
        ]
        ws.append(row)

        # Apply alternating row colors
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Auto-fit columns
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        # Check if the first cell is a merged cell
        if isinstance(column[0], MergedCell):
            continue
        for cell in column:
            if isinstance(cell, MergedCell):
                continue
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Create Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response

# Route to download Excel for invoices
@app.route('/download-invoices-excel', methods=['GET'])
def download_invoices_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    partner_name = request.args.get('partner_name')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    if not partner_name:
        return jsonify({"error": "Missing partner_name parameter"}), 400

    # Fetch invoice data
    data = fetch_facture_recap_achat(start_date, end_date, partner_name)
    
    if "error" in data:
        return jsonify(data), 500

    # Generate filename
    filename = f"Invoices_{partner_name}_{start_date}_to_{end_date}.xlsx"
    
    return generate_excel_invoices(data, filename)

# Route to download Excel for invoice lines
@app.route('/download-invoice-lines-excel', methods=['GET'])
def download_invoice_lines_excel():
    invoice_id = request.args.get('invoice_id')

    if not invoice_id:
        return jsonify({"error": "Missing invoice_id parameter"}), 400

    # Fetch invoice lines data
    data = BCF_product(invoice_id)
    
    if "error" in data:
        return jsonify(data), 500

    # Generate filename
    filename = f"Invoice_Lines_{invoice_id}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    
    return generate_excel_invoice_lines(data, filename)


@app.route('/fetchFactureRecapAchat', methods=['GET'])
def fetch_facture_achat():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    partner_name = request.args.get('partner_name')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    data = fetch_facture_recap_achat(start_date, end_date, partner_name)
    return jsonify(data)


def BCF_product(invoice_id):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT 
                    il.qtyentered, 
                    p.name AS product_name, 
                    il.linenetamt
                FROM C_InvoiceLine il
                JOIN M_Product p ON il.m_product_id = p.m_product_id
                WHERE il.c_invoice_id = :invoice_id
            """

            params = {'invoice_id': invoice_id}

            cursor.execute(query, params)
            rows = cursor.fetchall()

            data = [
                {
                    "qtyentered": row[0],
                    "product_name": row[1],
                    "linenetamt": row[2]
                }
                for row in rows
            ]

            return data

    except Exception as e:
        logger.error(f"Error fetching invoice lines: {e}")
        return {"error": "An error occurred while fetching invoice line details."}


@app.route('/fetchBCFProduct', methods=['GET'])
def fetch_bcf_product():
    invoice_id = request.args.get('invoice_id')

    if not invoice_id:
        return jsonify({"error": "Missing invoice_id parameter"}), 400

    data = BCF_product(invoice_id)
    return jsonify(data)





# recap vente part



def fetch_fournisseur_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT * FROM (
                    SELECT 
                        CAST(xf.name AS VARCHAR2(300)) AS FOURNISSEUR,   
                        SUM(xf.TOTALLINE) AS total, 
                        SUM(xf.qtyentered) AS QTY,
                        ROUND(
                            CASE 
                                WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                                WHEN SUM(xf.CONSOMATION) < 0 THEN ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION) * -1) * 100
                                ELSE ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION)) * 100
                            END, 4) AS marge,
                        0 AS sort_order
                    FROM xx_ca_fournisseur xf
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER C ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
                        AND xf.AD_Org_ID = :ad_org_id
                        AND xf.DOCSTATUS != 'RE'
                        AND (:fournisseur IS NULL OR xf.name LIKE :fournisseur || '%')
                        AND (:product IS NULL OR xf.product LIKE :product || '%')
                        AND (:client IS NULL OR cb.name LIKE :client || '%')
                        AND (:operateur IS NULL OR au.name LIKE :operateur || '%')
                        AND (:bccb IS NULL OR C.DOCUMENTNO LIKE :bccb || '%')
                        AND (:zone IS NULL OR sr.name LIKE :zone || '%')
                        AND (
                            :group_label IS NULL OR
                            (CASE 
                                WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                ELSE 'autre'
                            END = :group_label)
                        )
                    GROUP BY xf.name
                    UNION ALL
                    SELECT 
                        CAST('Total' AS VARCHAR2(300)) AS name, 
                        SUM(xf.TOTALLINE) AS total, 
                        SUM(xf.qtyentered) AS QTY,
                        ROUND(
                            CASE 
                                WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                                WHEN SUM(xf.CONSOMATION) < 0 THEN ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION) * -1) * 100
                                ELSE ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION)) * 100
                            END, 4) AS marge,
                        1 AS sort_order
                    FROM xx_ca_fournisseur xf
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER C ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
                        AND xf.AD_Org_ID = :ad_org_id
                        AND xf.DOCSTATUS != 'RE'
                        AND (:fournisseur IS NULL OR xf.name LIKE :fournisseur || '%')
                        AND (:product IS NULL OR xf.product LIKE :product || '%')
                        AND (:client IS NULL OR cb.name LIKE :client || '%')
                        AND (:operateur IS NULL OR au.name LIKE :operateur || '%')
                        AND (:bccb IS NULL OR C.DOCUMENTNO LIKE :bccb || '%')
                        AND (:zone IS NULL OR sr.name LIKE :zone || '%')
                        AND (
                            :group_label IS NULL OR
                            (CASE 
                                WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                ELSE 'autre'
                            END = :group_label)
                        )
                )
                ORDER BY sort_order, total DESC
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'ad_org_id': ad_org_id,
                'group_label': group_label
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching fournisseur data: {e}")
        return {"error": "An error occurred while fetching fournisseur data."}

@app.route('/fetchFournisseurData', methods=['GET'])
def fetch_fournisseur():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    if not ad_org_id:
        return jsonify({"error": "Missing ad_org_id parameter"}), 400  # Ensure ad_org_id is provided

    try:
        ad_org_id = int(ad_org_id)  # Convert to integer
    except ValueError:
        return jsonify({"error": "Invalid ad_org_id format"}), 400

    data = fetch_fournisseur_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)
    return jsonify(data)





def fetch_product_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT * FROM (
                    -- Main Product Breakdown
                    SELECT 
                        CAST(xf.product AS VARCHAR2(400)) AS "PRODUIT",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        0 AS "SORT_ORDER"
                    FROM xx_ca_fournisseur xf
                    LEFT JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    LEFT JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    LEFT JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    LEFT JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    LEFT JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    LEFT JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (
                              :group_label IS NULL OR
                              (CASE 
                                  WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                  WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                  ELSE 'autre'
                              END = :group_label)
                          )
                    GROUP BY xf.product

                    UNION ALL

                    -- Total Row
                    SELECT 
                        'Total' AS "PRODUIT",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        1 AS "SORT_ORDER"
                    FROM xx_ca_fournisseur xf
                    LEFT JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    LEFT JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    LEFT JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    LEFT JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    LEFT JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    LEFT JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (
                              :group_label IS NULL OR
                              (CASE 
                                  WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                  WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                  ELSE 'autre'
                              END = :group_label)
                          )
                ) ORDER BY "SORT_ORDER", "TOTAL" DESC
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'ad_org_id': ad_org_id,
                'group_label': group_label
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]

    except Exception as e:
        logger.error(f"Error fetching product data: {e}")
        return {"error": "An error occurred while fetching product data."}

@app.route('/fetchProductData', methods=['GET'])
def fetch_product():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get("product", "").strip()
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date or not ad_org_id:
        return jsonify({"error": "Missing start_date, end_date, or ad_org_id parameters"}), 400

    data = fetch_product_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)
    return jsonify(data)






# Fetch total recap data
def fetch_rcap_data(start_date, end_date, ad_org_id, group_label=None):

    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT 
                    SUM(xf.TOTALLINE) AS CHIFFRE, 
                    SUM(xf.qtyentered) AS QTY,
                    SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION) AS MARGE,
                    SUM(xf.CONSOMATION) AS CONSOMATION,
                    CASE 
                        WHEN SUM(xf.CONSOMATION) < 0 
                        THEN ROUND(((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / (SUM(xf.CONSOMATION) * -1)), 4)
                        ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                    END AS POURCENTAGE
                FROM xx_ca_fournisseur xf
                JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.CLIENTID
                WHERE 
                    xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                    AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND xf.AD_Org_ID = :ad_org_id
                    AND xf.DOCSTATUS != 'RE'
                    AND cb.ISCUSTOMER = 'Y'
                    AND cb.AD_Client_ID = 1000000
                    AND cb.AD_Org_ID = 1000000
                    AND (
                        :group_label IS NULL OR
                        (CASE 
                            WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                            WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                            ELSE 'autre'
                        END = :group_label)
                    )
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'ad_org_id': ad_org_id,
                'group_label': group_label
            }
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching data: {e}")
        return {"error": "An error occurred while fetching data."}
    





@app.route('/fetchTotalrecapData', methods=['GET'])
def fetch_recap():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    ad_org_id = request.args.get('ad_org_id')  # Get ad_org_id from request
    group_label = request.args.get('group_label')  # Optional group_label (para, potentiel, autre)

    if not start_date or not end_date or not ad_org_id:
        return jsonify({"error": "Missing start_date, end_date, or ad_org_id parameters"}), 400

    try:
        ad_org_id = int(ad_org_id)  # Convert to integer
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    data = fetch_rcap_data(start_date, end_date, ad_org_id, group_label)
    return jsonify(data)





@app.route('/fetchZoneRecap', methods=['GET'])
def fetch_zone():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400
    if not ad_org_id:
        return jsonify({"error": "Missing ad_org_id parameter"}), 400

    data = fetch_zone_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)
    return jsonify(data)


def fetch_zone_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT * FROM (
                    -- Recap by Zone
                    SELECT 
                        CAST(sr.name AS VARCHAR2(100)) AS "ZONE",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        0 AS "SORT_ORDER"
                    FROM C_SalesRegion sr
                    JOIN C_BPartner_Location bpl ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN xx_ca_fournisseur xf ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (
                              :group_label IS NULL OR
                              (CASE 
                                  WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                  WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                  ELSE 'autre'
                              END = :group_label)
                          )
                    GROUP BY sr.name

                    UNION ALL

                    -- Total Row
                    SELECT 
                        'Total' AS "ZONE",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        1 AS "SORT_ORDER"
                    FROM C_SalesRegion sr
                    JOIN C_BPartner_Location bpl ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN xx_ca_fournisseur xf ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (
                              :group_label IS NULL OR
                              (CASE 
                                  WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                  WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                  ELSE 'autre'
                              END = :group_label)
                          )
                ) ORDER BY "SORT_ORDER", "TOTAL" DESC
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'ad_org_id': ad_org_id,
                'group_label': group_label
            }
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching zone recap: {e}")
        return {"error": "An error occurred while fetching zone recap."}



def fetch_client_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT * FROM (
                    -- Recap by Client
                    SELECT 
                        CAST(cb.name AS VARCHAR2(100)) AS "CLIENT",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        0 AS "SORT_ORDER"
                    FROM C_BPartner cb
                    JOIN xx_ca_fournisseur xf ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (
                              :group_label IS NULL OR
                              (CASE 
                                  WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                  WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                  ELSE 'autre'
                              END = :group_label)
                          )
                    GROUP BY cb.name

                    UNION ALL

                    -- Total Row
                    SELECT 
                        'Total' AS "CLIENT",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        1 AS "SORT_ORDER"
                    FROM C_BPartner cb
                    JOIN xx_ca_fournisseur xf ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (
                              :group_label IS NULL OR
                              (CASE 
                                  WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                  WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                  ELSE 'autre'
                              END = :group_label)
                          )
                ) ORDER BY "SORT_ORDER", "TOTAL" DESC
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'ad_org_id': ad_org_id,
                'group_label': group_label
            }
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching client recap: {e}")
        return {"error": "An error occurred while fetching client recap."}


@app.route('/fetchClientRecap', methods=['GET'])
def fetch_client():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400
    if not ad_org_id:
        return jsonify({"error": "Missing ad_org_id parameter"}), 400

    data = fetch_client_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)
    return jsonify(data)

def fetch_operator_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT * FROM (
                    SELECT CAST(au.name AS VARCHAR2(100)) AS "OPERATEUR", 
                           SUM(xf.TOTALLINE) AS "TOTAL", 
                           SUM(xf.qtyentered) AS "QTY",
                           CASE 
                               WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                               WHEN SUM(xf.CONSOMATION) < 0 THEN ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION) * -1)
                               ELSE (SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION)
                           END AS "MARGE",
                           0 AS "SORT_ORDER"
                    FROM AD_User au
                    JOIN xx_ca_fournisseur xf ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (
                              :group_label IS NULL OR
                              (CASE 
                                  WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                  WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                  ELSE 'autre'
                              END = :group_label)
                          )
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                    GROUP BY au.name

                    UNION ALL

                    SELECT 'Total' AS "OPERATEUR", 
                           SUM(xf.TOTALLINE) AS "TOTAL", 
                           SUM(xf.qtyentered) AS "QTY",
                           CASE 
                               WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                               WHEN SUM(xf.CONSOMATION) < 0 THEN ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION) * -1)
                               ELSE (SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION)
                           END AS "MARGE",
                           1 AS "SORT_ORDER"
                    FROM AD_User au
                    JOIN xx_ca_fournisseur xf ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (
                              :group_label IS NULL OR
                              (CASE 
                                  WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                  WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                  ELSE 'autre'
                              END = :group_label)
                          )
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                ) ORDER BY "SORT_ORDER", "TOTAL" DESC
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'ad_org_id': ad_org_id,
                'group_label': group_label
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching operator recap: {e}")
        return {"error": "An error occurred while fetching operator recap."}


@app.route('/fetchOperatorRecap', methods=['GET'])
def fetch_operator():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400
    if not ad_org_id:
        return jsonify({"error": "Missing ad_org_id parameter"}), 400

    data = fetch_operator_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)
    return jsonify(data)




def fetch_bccb_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, group_label=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT DOCUMENTNO, DATEORDERED, GRANDTOTAL, 
                       ROUND(AVG(marge * 100), 2) AS marge
                FROM (
                    SELECT det.*
                    FROM (
                        SELECT lot.*,  
                               (lot.priceentered - ((lot.priceentered * NVL(lot.remise_vente, 0)) / 100)) / (1 + (lot.bonus_vente / 100)) AS ventef
                        FROM (
                            SELECT 
                                ol.priceentered, 
                                (SELECT valuenumber 
                                 FROM m_attributeinstance 
                                 WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                       AND m_attribute_id = 1000504) AS p_revient,
                                (SELECT valuenumber 
                                 FROM m_attributeinstance 
                                 WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                       AND m_attribute_id = 1000908) AS bonus_vente,
                                (SELECT valuenumber 
                                 FROM m_attributeinstance 
                                 WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                       AND m_attribute_id = 1001408) AS remise_vente,
                                c.DOCUMENTNO, 
                                c.DATEORDERED, 
                                c.GRANDTOTAL,
                                CASE 
                                    WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                                    WHEN SUM(xf.CONSOMATION) < 0 THEN ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION) * -1)
                                    ELSE (SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION)
                                END AS marge
                            FROM C_Order c
                            JOIN C_OrderLine ol ON c.C_Order_ID = ol.C_Order_ID 
                            JOIN M_InOut mi ON mi.C_Order_ID = c.C_Order_ID
                            JOIN xx_ca_fournisseur xf ON xf.DOCUMENTNO = mi.DOCUMENTNO
                            JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                            JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                            JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                            JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                            WHERE c.AD_Org_ID = 1000000 
                                  AND ol.qtyentered > 0
                                  AND xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                      AND TO_DATE(:end_date, 'YYYY-MM-DD')
                                  AND (c.DOCSTATUS = 'CL' OR c.DOCSTATUS = 'CO')
                                  AND c.C_DocType_ID IN (1000539, 1001408)
                                  AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                                  AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                                  AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                                  AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                                  AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                                  AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                                  AND (
                                      :group_label IS NULL OR
                                      (CASE 
                                          WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                          WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                          ELSE 'autre'
                                      END = :group_label)
                                  )
                            GROUP BY c.DOCUMENTNO, c.DATEORDERED, c.GRANDTOTAL, ol.priceentered, ol.m_attributesetinstance_id
                            ORDER BY c.DOCUMENTNO
                        ) lot
                    ) det
                )
                GROUP BY DOCUMENTNO, DATEORDERED, GRANDTOTAL
                ORDER BY DOCUMENTNO
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'group_label': group_label
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching BCCB recap: {e}")
        return {"error": "An error occurred while fetching BCCB recap."}


@app.route('/fetchBCCBRecap', methods=['GET']) 
def fetch_bccb():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    data = fetch_bccb_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, group_label)
    return jsonify(data)




@app.route('/download-bccb-product-excel', methods=['GET'])
def download_bccb_product_excel():
    bccb = request.args.get('bccb')
    ad_org_id = request.args.get('ad_org_id', '1000012')

    if not bccb:
        return jsonify({"error": "Missing BCCB parameter"}), 400

    try:
        ad_org_id = int(ad_org_id)
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_bccb_product(bccb, ad_org_id)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"BCCB_Product_{bccb}_{today_date}.xlsx"

    return generate_excel_bccb_product(data, filename)


def generate_excel_bccb_product(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "BCCB Product Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="BCCBProductTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")






@app.route('/download-fournisseur-excel', methods=['GET'])
def download_fournisseur_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_fournisseur_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id and group_label
    today_date = datetime.now().strftime("%d-%m-%Y")
    group_label_part = f"_{group_label}" if group_label else ""
    if ad_org_id == 1000012:
        filename = f"FournisseurRecap_facturation{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"FournisseurRecap{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_fournisseur(data, filename)


def generate_excel_fournisseur(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Fournisseur Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="FournisseurRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")






@app.route('/download-product-excel', methods=['GET'])
def download_product_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_product_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id and group_label
    today_date = datetime.now().strftime("%d-%m-%Y")
    group_label_part = f"_{group_label}" if group_label else ""
    if ad_org_id == 1000012:
        filename = f"ProductRecap_facturation{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"ProductRecap{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_product(data, filename)


def generate_excel_product(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Product Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="ProductRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")





@app.route('/download-totalrecap-excel', methods=['GET'])
def download_totalrecap_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date or not ad_org_id:
        return jsonify({"error": "Missing start_date, end_date, or ad_org_id parameters"}), 400

    try:
        ad_org_id = int(ad_org_id)  # Convert to integer
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data from the database
    data = fetch_rcap_data(start_date, end_date, ad_org_id, group_label)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id and group_label
    today_date = datetime.now().strftime("%d-%m-%Y")
    group_label_part = f"_{group_label}" if group_label else ""
    if ad_org_id == 1000012:
        filename = f"TotalRecap_facturation{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"TotalRecap{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_totalrecap(data, filename)


def generate_excel_totalrecap(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Total Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="TotalRecapTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")






 



@app.route('/download-zone-excel', methods=['GET'])
def download_zone_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_zone_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id and group_label
    today_date = datetime.now().strftime("%d-%m-%Y")
    group_label_part = f"_{group_label}" if group_label else ""
    if ad_org_id == 1000012:
        filename = f"ZoneRecap_facturation{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"ZoneRecap{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_zone(data, filename)


def generate_excel_zone(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Zone Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="ZoneRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/download-client-excel', methods=['GET'])
def download_client_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_client_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id and group_label
    today_date = datetime.now().strftime("%d-%m-%Y")
    group_label_part = f"_{group_label}" if group_label else ""
    if ad_org_id == 1000012:
        filename = f"ClientRecap_facturation{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"ClientRecap{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_client(data, filename)


def generate_excel_client(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Client Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="ClientRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route('/download-operator-excel', methods=['GET'])
def download_operator_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_operator_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id and group_label
    today_date = datetime.now().strftime("%d-%m-%Y")
    group_label_part = f"_{group_label}" if group_label else ""
    if ad_org_id == 1000012:
        filename = f"OperatorRecap_facturation{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"OperatorRecap{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_operator(data, filename)


def generate_excel_operator(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Operator Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="OperatorRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")








@app.route('/download-bccb-excel', methods=['GET'])
def download_bccb_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data
    data = fetch_bccb_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"BCCBRecap_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_bccb(data, filename)

def generate_excel_bccb(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "BCCB Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="BCCBRecapTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



 
@app.route('/fetchBCCBProduct', methods=['GET'])
def fetch_bccb_p():
    bccb = request.args.get('bccb')
    ad_org_id = request.args.get('ad_org_id')

    if not ad_org_id:
        return jsonify({"error": "Missing ad_org_id parameter"}), 400

    data = fetch_bccb_product(bccb, ad_org_id)
    return jsonify(data)



def fetch_bccb_product(bccb, ad_org_id):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT product, qty, remise, marge 
                FROM (
                    SELECT det.*, 
                           ROUND((det.ventef - det.p_revient) / det.p_revient * 100, 2) AS marge 
                    FROM (
                        SELECT lot.*, 
                               (lot.priceentered - ((lot.priceentered * NVL(lot.remise_vente, 0)) / 100)) / 
                               (1 + (lot.bonus_vente / 100)) AS ventef 
                        FROM (
                            SELECT ol.priceentered AS priceentered, 
                                   ol.qtyentered AS qty, 
                                   mp.name AS product, 
                                   ol.discount / 100 AS remise, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1000504) AS p_revient, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1000908) AS bonus_vente, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1001408) AS remise_vente 
                            FROM c_orderline ol 
                            INNER JOIN c_order o ON o.c_order_id = ol.c_order_id 
                            INNER JOIN m_product mp ON ol.m_product_id = mp.m_product_id 
                            WHERE ol.qtyentered > 0 
                                  AND (:bccb IS NULL OR UPPER(o.documentno) LIKE UPPER(:bccb) || '%')
                                  AND o.AD_Org_ID = :ad_org_id
                        ) lot
                    ) det
                )
            """
            
            params = {
                'bccb': bccb or None,
                'ad_org_id': ad_org_id
            }
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching BCCB product data: {e}")
        return {"error": "An error occurred while fetching BCCB product data."}



# Stock Inventory Analysis Endpoints

@app.route('/stock-manque', methods=['GET'])
def get_stock_manque():
    """Get stock manque data for date range"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not start_date or not end_date:
            return jsonify({"success": False, "error": "start_date and end_date are required"}), 400
        
        data = fetch_stock_manque_data(start_date, end_date)
        return jsonify(data)
    
    except Exception as e:
        logger.error(f"Error fetching stock manque data: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/stock-casse', methods=['GET'])
def get_stock_casse():
    """Get stock casse data for date range (Business Partner 1126375)"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not start_date or not end_date:
            return jsonify({"success": False, "error": "start_date and end_date are required"}), 400
        
        data = fetch_stock_casse_data(start_date, end_date)
        return jsonify(data)
    
    except Exception as e:
        logger.error(f"Error fetching stock casse data: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


def fetch_stock_manque_data(start_date, end_date):
    """
    Query to calculate the sum of inventory differences based on Prix Revient
    Links M_Inventory with M_InventoryLine and gets Prix Revient directly from attribute ID 1000504
    Filters for inventories with "manque" in description and within specified date range
    """
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            
            query = """
            WITH inventory_attributes AS (
                SELECT 
                    il.M_INVENTORYLINE_ID,
                    il.M_INVENTORY_ID,
                    il.QTYCOUNT,
                    il.QTYBOOK,
                    MAX(CASE WHEN ai.M_ATTRIBUTE_ID = 1000504 THEN ai.VALUENUMBER END) as PRIX_REVIENT
                FROM M_INVENTORYLINE il
                LEFT JOIN M_ATTRIBUTESETINSTANCE asi ON il.M_ATTRIBUTESETINSTANCE_ID = asi.M_ATTRIBUTESETINSTANCE_ID
                LEFT JOIN M_ATTRIBUTEINSTANCE ai ON asi.M_ATTRIBUTESETINSTANCE_ID = ai.M_ATTRIBUTESETINSTANCE_ID
                WHERE asi.ISACTIVE = 'Y'
                    AND ai.ISACTIVE = 'Y'
                    AND ai.M_ATTRIBUTE_ID = 1000504  -- Prix Revient attribute
                GROUP BY 
                    il.M_INVENTORYLINE_ID,
                    il.M_INVENTORY_ID,
                    il.QTYCOUNT,
                    il.QTYBOOK
            )
            SELECT 
                inv.M_INVENTORY_ID,
                inv.DOCUMENTNO,
                inv.DESCRIPTION AS INVENTORY_DESCRIPTION,
                inv.MOVEMENTDATE,
                COUNT(*) AS NUMBER_OF_LINES,
                SUM(
                    (ia.QTYCOUNT - ia.QTYBOOK) * COALESCE(ia.PRIX_REVIENT, 0)
                ) AS TOTAL_DIFFERENCE_AMOUNT
            FROM M_INVENTORY inv
            INNER JOIN inventory_attributes ia ON inv.M_INVENTORY_ID = ia.M_INVENTORY_ID
            WHERE inv.DOCSTATUS IN ('CO', 'CL')  -- Only completed or closed documents
                AND inv.ISACTIVE = 'Y'
                AND (ia.QTYCOUNT - ia.QTYBOOK) != 0  -- Only lines with differences
                AND UPPER(inv.DESCRIPTION) LIKE UPPER('%manque%')  -- Only inventories with "manque" in description
                AND inv.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')  -- Date range filter
            GROUP BY 
                inv.M_INVENTORY_ID,
                inv.DOCUMENTNO,
                inv.DESCRIPTION,
                inv.MOVEMENTDATE
            ORDER BY inv.MOVEMENTDATE DESC, inv.DOCUMENTNO
            """
            
            cursor.execute(query, {
                'start_date': start_date,
                'end_date': end_date
            })
            
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            
            return {"success": True, "data": data}
            
    except Exception as e:
        logger.error(f"Error fetching stock manque data: {e}")
        return {"success": False, "error": str(e)}


def fetch_stock_casse_data(start_date, end_date):
    """
    Query to get invoices for specific business partner (C_BPartner_ID=1126375)
    Gets C_Invoice data with document number, description, and total lines
    Filters for DOCSTATUS in ('CO', 'CL') and includes date range filter
    """
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            
            query = """
            SELECT 
                inv.DOCUMENTNO,
                inv.DESCRIPTION,
                inv.DATEINVOICED,
                inv.DOCSTATUS,
                COUNT(il.C_INVOICELINE_ID) AS TOTAL_LINES,
                SUM(il.LINENETAMT) AS TOTAL_NET_AMOUNT,
                inv.GRANDTOTAL
            FROM C_INVOICE inv
            LEFT JOIN C_INVOICELINE il ON inv.C_INVOICE_ID = il.C_INVOICE_ID 
                AND il.ISACTIVE = 'Y'
            WHERE inv.C_BPARTNER_ID = 1126375  -- Specific business partner
                AND inv.DOCSTATUS IN ('CO', 'CL')  -- Only completed or closed documents
                AND inv.ISACTIVE = 'Y'
                AND inv.DATEINVOICED BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')  -- Date range filter
            GROUP BY 
                inv.C_INVOICE_ID,
                inv.DOCUMENTNO,
                inv.DESCRIPTION,
                inv.DATEINVOICED,
                inv.DOCSTATUS,
                inv.GRANDTOTAL
            ORDER BY inv.DATEINVOICED DESC, inv.DOCUMENTNO
            """
            
            cursor.execute(query, {
                'start_date': start_date,
                'end_date': end_date
            })
            
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            
            return {"success": True, "data": data}
            
    except Exception as e:
        logger.error(f"Error fetching stock casse data: {e}")
        return {"success": False, "error": str(e)}


@app.route('/stock-manque/excel', methods=['GET'])
def download_stock_manque_excel():
    """Download stock manque data as Excel file"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not start_date or not end_date:
            return jsonify({"success": False, "error": "start_date and end_date are required"}), 400
        
        data = fetch_stock_manque_data(start_date, end_date)
        
        if not data.get('success') or not data.get('data'):
            return jsonify({"success": False, "error": "No data found for the specified date range"}), 404
        
        # Generate Excel file
        excel_file = generate_manque_excel(data['data'], start_date, end_date)
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=f'manque_stock_{start_date}_{end_date}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        logger.error(f"Error generating manque Excel: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/stock-casse/excel', methods=['GET'])
def download_stock_casse_excel():
    """Download stock casse data as Excel file"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not start_date or not end_date:
            return jsonify({"success": False, "error": "start_date and end_date are required"}), 400
        
        data = fetch_stock_casse_data(start_date, end_date)
        
        if not data.get('success') or not data.get('data'):
            return jsonify({"success": False, "error": "No data found for the specified date range"}), 404
        
        # Generate Excel file
        excel_file = generate_casse_excel(data['data'], start_date, end_date)
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=f'casse_stock_{start_date}_{end_date}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        logger.error(f"Error generating casse Excel: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


def generate_manque_excel(data, start_date, end_date):
    """Generate Excel file for manque stock data"""
    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Manque Stock"
        
        # Set headers
        headers = [
            "Document No",
            "Description",
            "Date Mouvement",
            "Nombre de Lignes",
            "Montant Différence (DA)"
        ]
        
        # Add title row
        title = f"Rapport Manque de Stock - Période: {start_date} au {end_date}"
        ws.merge_cells('A1:E1')
        ws['A1'] = title
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Add headers to row 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # Add data rows
        for row_idx, item in enumerate(data, start=4):
            ws.cell(row=row_idx, column=1, value=item.get('DOCUMENTNO', ''))
            ws.cell(row=row_idx, column=2, value=item.get('INVENTORY_DESCRIPTION', ''))
            ws.cell(row=row_idx, column=3, value=item.get('MOVEMENTDATE', ''))
            ws.cell(row=row_idx, column=4, value=item.get('NUMBER_OF_LINES', 0))
            ws.cell(row=row_idx, column=5, value=float(item.get('TOTAL_DIFFERENCE_AMOUNT', 0)))
        
        # Calculate totals
        if data:
            total_amount = sum(float(item.get('TOTAL_DIFFERENCE_AMOUNT', 0)) for item in data)
            total_documents = len(data)
            total_lines = sum(int(item.get('NUMBER_OF_LINES', 0)) for item in data)
            
            # Add totals row
            total_row = len(data) + 5
            ws.cell(row=total_row, column=1, value="TOTAUX:")
            ws.cell(row=total_row, column=1).font = Font(bold=True)
            ws.cell(row=total_row, column=2, value=f"{total_documents} documents")
            ws.cell(row=total_row, column=4, value=total_lines)
            ws.cell(row=total_row, column=5, value=total_amount)
            
            # Style totals row
            for col in range(1, 6):
                ws.cell(row=total_row, column=col).fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                ws.cell(row=total_row, column=col).font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            # Check if the first cell is a merged cell
            if isinstance(column[0], MergedCell):
                continue
            column_letter = column[0].column_letter
            for cell in column:
                if isinstance(cell, MergedCell):
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create table for better formatting
        if data:
            table_range = f"A3:E{len(data) + 3}"
            table = Table(displayName="ManqueStockTable", ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium9", 
                showFirstColumn=False,
                showLastColumn=False, 
                showRowStripes=True, 
                showColumnStripes=True
            )
            table.tableStyleInfo = style
            ws.add_table(table)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    except Exception as e:
        logger.error(f"Error creating manque Excel file: {e}")
        raise


def generate_casse_excel(data, start_date, end_date):
    """Generate Excel file for casse stock data"""
    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Casse Stock"
        
        # Set headers
        headers = [
            "Document No",
            "Description",
            "Date Facture",
            "Statut",
            "Total Lignes",
            "Montant Net (DA)",
            "Grand Total (DA)"
        ]
        
        # Add title row
        title = f"Rapport Casse de Stock (BP: 1126375) - Période: {start_date} au {end_date}"
        ws.merge_cells('A1:G1')
        ws['A1'] = title
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Add headers to row 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # Add data rows
        for row_idx, item in enumerate(data, start=4):
            ws.cell(row=row_idx, column=1, value=item.get('DOCUMENTNO', ''))
            ws.cell(row=row_idx, column=2, value=item.get('DESCRIPTION', ''))
            ws.cell(row=row_idx, column=3, value=item.get('DATEINVOICED', ''))
            ws.cell(row=row_idx, column=4, value=item.get('DOCSTATUS', ''))
            ws.cell(row=row_idx, column=5, value=int(item.get('TOTAL_LINES', 0)))
            ws.cell(row=row_idx, column=6, value=float(item.get('TOTAL_NET_AMOUNT', 0)))
            ws.cell(row=row_idx, column=7, value=float(item.get('GRANDTOTAL', 0)))
        
        # Calculate totals
        if data:
            total_net_amount = sum(float(item.get('TOTAL_NET_AMOUNT', 0)) for item in data)
            total_grand_total = sum(float(item.get('GRANDTOTAL', 0)) for item in data)
            total_documents = len(data)
            total_lines = sum(int(item.get('TOTAL_LINES', 0)) for item in data)
            
            # Add totals row
            total_row = len(data) + 5
            ws.cell(row=total_row, column=1, value="TOTAUX:")
            ws.cell(row=total_row, column=1).font = Font(bold=True)
            ws.cell(row=total_row, column=2, value=f"{total_documents} documents")
            ws.cell(row=total_row, column=5, value=total_lines)
            ws.cell(row=total_row, column=6, value=total_net_amount)
            ws.cell(row=total_row, column=7, value=total_grand_total)
            
            # Style totals row
            for col in range(1, 8):
                ws.cell(row=total_row, column=col).fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                ws.cell(row=total_row, column=col).font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            # Check if the first cell is a merged cell
            if isinstance(column[0], MergedCell):
                continue
            column_letter = column[0].column_letter
            for cell in column:
                if isinstance(cell, MergedCell):
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create table for better formatting
        if data:
            table_range = f"A3:G{len(data) + 3}"
            table = Table(displayName="CasseStockTable", ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium2", 
                showFirstColumn=False,
                showLastColumn=False, 
                showRowStripes=True, 
                showColumnStripes=True
            )
            table.tableStyleInfo = style
            ws.add_table(table)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    except Exception as e:
        logger.error(f"Error creating casse Excel file: {e}")
        raise




#---------------------------wifi part---------------------------------------
from cryptography.fernet import Fernet
import mysql.connector

# === Encryption Key ===
# !! In production, store this key in a secure location (env variable or secrets manager)
ENCRYPTION_KEY = b'hukIqOCK3RGhnVUpZRN5qdZCe-Tu4wS5QDXAVo8Wick='  # replace with your own key if needed
cipher = Fernet(ENCRYPTION_KEY)


# === Encrypt ===
def encrypt(text):
    return cipher.encrypt(text.encode()).decode()

# === Decrypt ===
def decrypt(token):
    return cipher.decrypt(token.encode()).decode()

# === Get List ===
@app.route("/list", methods=["GET"])
def list_wifi():
    conn = get_localdb_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id, name, password, modempasswd, ip FROM wifi_passwords ORDER BY id DESC")
    result = cur.fetchall()
    conn.close()

    # Decrypt passwords before sending
    for row in result:
        try:
            row['password'] = decrypt(row['password'])
            row['modempasswd'] = decrypt(row['modempasswd'])
        except Exception as e:
            row['password'] = "[decryption error]"
            row['modempasswd'] = "[decryption error]"
            logger.error(f"Decryption error: {e}")

    return jsonify(result)

# === Add WiFi ===
@app.route("/add", methods=["POST"])
def add_wifi():
    data = request.get_json()
    conn = get_localdb_connection()
    cur = conn.cursor()

    try:
        encrypted_password = encrypt(data['password'])
        encrypted_modem = encrypt(data['modempasswd'])

        cur.execute("""
            INSERT INTO wifi_passwords (name, password, ip, modempasswd, created_by)
            VALUES (%s, %s, %s, %s, 'admin')
        """, (data['name'], encrypted_password, data['ip'], encrypted_modem))

        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"Add WiFi error: {e}")
        return jsonify({"success": False, "error": str(e)})
    finally:
        conn.close()

# === Update WiFi ===
@app.route("/update/<int:id>", methods=["POST"])
def update_wifi(id):
    data = request.get_json()
    conn = get_localdb_connection()
    cur = conn.cursor()

    try:
        encrypted_password = encrypt(data['password'])
        encrypted_modem = encrypt(data['modempasswd'])

        cur.execute("""
            UPDATE wifi_passwords
            SET name = %s,
                password = %s,
                ip = %s,
                modempasswd = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (data['name'], encrypted_password, data['ip'], encrypted_modem, id))

        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"Update WiFi error: {e}")
        return jsonify({"success": False, "error": str(e)})
    finally:
        conn.close()

# === Delete WiFi ===
@app.route("/delete/<int:id>", methods=["POST"])
def delete_wifi(id):
    conn = get_localdb_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM wifi_passwords WHERE id = %s", (id,))
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"Delete error: {e}")
        return jsonify({"success": False, "error": str(e)})
    finally:
        conn.close()











@app.route('/recieved_products_bydate')
def recieved_products_bydate():
    try:
        # Get date parameters from request
        start_date = request.args.get('start_date', datetime.now().strftime('%Y-%m-%d'))
        end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))

        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT 
                  bp.NAME AS SUPPLIER_NAME,
                  p.NAME AS PRODUCT_NAME,
                  sc.name AS BONUS,
                  MAX((
                      SELECT ai.VALUENUMBER
                      FROM M_ATTRIBUTEINSTANCE ai
                      WHERE ai.M_ATTRIBUTESETINSTANCE_ID = iol.M_ATTRIBUTESETINSTANCE_ID
                        AND ai.M_ATTRIBUTE_ID = 1001408
                        AND ai.ISACTIVE = 'Y'
                  )) AS REM_VENTE,
                  MAX((
                      SELECT ai.VALUENUMBER
                      FROM M_ATTRIBUTEINSTANCE ai
                      WHERE ai.M_ATTRIBUTESETINSTANCE_ID = iol.M_ATTRIBUTESETINSTANCE_ID
                        AND ai.M_ATTRIBUTE_ID = 1000908
                        AND ai.ISACTIVE = 'Y'
                  )) AS BON_VENTE,
                  (SELECT SUM(ms.qtyonhand - ms.qtyreserved)
                      FROM m_storage ms
                      WHERE ms.m_product_id = p.m_product_id
                        AND ms.ad_client_id = 1000000
                        AND ms.qtyonhand > 0
                        AND ms.m_locator_id = 1000614
                  ) AS qty_dispo,
                  MAX(dsp.NAME)   AS REM_POT,
                  MAX(dspara.NAME) AS REM_PARA
              FROM M_INOUT io
              INNER JOIN M_INOUTLINE iol ON io.M_INOUT_ID = iol.M_INOUT_ID
              INNER JOIN M_PRODUCT p ON iol.M_PRODUCT_ID = p.M_PRODUCT_ID
              INNER JOIN C_BPARTNER bp ON io.C_BPARTNER_ID = bp.C_BPARTNER_ID
              JOIN XX_SALESCONTEXT sc ON p.XX_SALESCONTEXT_ID = sc.XX_SALESCONTEXT_ID
              LEFT JOIN (
                  SELECT M_PRODUCT_ID, M_DISCOUNTSCHEMA_ID 
                  FROM C_BPartner_Product 
                  WHERE ISACTIVE = 'Y' AND C_BP_Group_ID = 1001330
              ) bppot ON bppot.M_PRODUCT_ID = p.M_PRODUCT_ID
              LEFT JOIN M_DiscountSchema dsp ON dsp.M_DiscountSchema_ID = bppot.M_DiscountSchema_ID
              LEFT JOIN (
                  SELECT M_PRODUCT_ID, M_DISCOUNTSCHEMA_ID 
                  FROM C_BPartner_Product 
                  WHERE ISACTIVE = 'Y' AND C_BP_Group_ID = 1000003
              ) bppara ON bppara.M_PRODUCT_ID = p.M_PRODUCT_ID
              LEFT JOIN M_DiscountSchema dspara ON dspara.M_DiscountSchema_ID = bppara.M_DiscountSchema_ID
              WHERE io.DOCSTATUS IN ('CO', 'CL')
                AND io.C_DOCTYPE_ID = 1000013
                AND io.AD_CLIENT_ID = 1000000
                AND io.ISACTIVE = 'Y'
                AND iol.ISACTIVE = 'Y'
                AND p.ISACTIVE = 'Y'
                AND bp.ISACTIVE = 'Y'
                AND iol.M_PRODUCT_ID IS NOT NULL
                AND io.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
              GROUP BY bp.NAME, p.NAME, sc.name, p.m_product_id
              ORDER BY bp.NAME, p.NAME
            """
            cursor.execute(query, {
                'start_date': start_date,
                'end_date': end_date
            })
            columns = [desc[0] for desc in cursor.description]
            grouped = {}
            for row in cursor.fetchall():
                row_dict = {columns[i]: row[i] for i in range(len(columns))}
                supplier = row_dict.pop('SUPPLIER_NAME')
                if supplier not in grouped:
                    grouped[supplier] = []
                grouped[supplier].append(row_dict)
            # Convert to list of dicts for JSON
            reception_data = [{"SUPPLIER_NAME": k, "PRODUCTS": v} for k, v in grouped.items()]

        # Also get fake reception data by calling product_details_by_id logic
        fake_reception_data = []
        try:
            # Get all records from fake_reception table
            mysql_conn = get_localdb_connection()
            if mysql_conn:
                try:
                    mysql_cursor = mysql_conn.cursor(dictionary=True)
                    
                    # Query to get all product details from fake_reception
                    fake_reception_query = """
                        SELECT 
                            id,
                            product_id,
                            m_attributesetinstance_id,
                            product_name
                        FROM fake_reception
                        ORDER BY created_at DESC
                    """
                    
                    mysql_cursor.execute(fake_reception_query)
                    fake_reception_records = mysql_cursor.fetchall()
                    
                    if fake_reception_records:
                        # Process fake reception data
                        fake_grouped = {}
                        
                        with DB_POOL.acquire() as fake_connection:
                            fake_cursor = fake_connection.cursor()
                            
                            for record in fake_reception_records:
                                product_id = record['product_id']
                                attributesetinstance_id = record['m_attributesetinstance_id']
                                fake_reception_id = record['id']
                                
                                # Query to get data using the values from fake_reception
                                fake_query = """
                                    SELECT 
                                      (SELECT ai.VALUE
                                       FROM M_ATTRIBUTEINSTANCE ai
                                       WHERE ai.M_ATTRIBUTESETINSTANCE_ID = :attributesetinstance_id
                                         AND ai.M_ATTRIBUTE_ID = 1000508
                                         AND ai.ISACTIVE = 'Y'
                                         AND ROWNUM = 1
                                      ) AS SUPPLIER_NAME,
                                      p.NAME AS PRODUCT_NAME,
                                      sc.name AS BONUS,
                                      NVL((SELECT ai.VALUENUMBER
                                           FROM M_ATTRIBUTEINSTANCE ai
                                           WHERE ai.M_ATTRIBUTESETINSTANCE_ID = :attributesetinstance_id
                                             AND ai.M_ATTRIBUTE_ID = 1001408
                                             AND ai.ISACTIVE = 'Y'
                                             AND ROWNUM = 1
                                          ), 0) AS REM_VENTE,
                                      NVL((SELECT ai.VALUENUMBER
                                           FROM M_ATTRIBUTEINSTANCE ai
                                           WHERE ai.M_ATTRIBUTESETINSTANCE_ID = :attributesetinstance_id
                                             AND ai.M_ATTRIBUTE_ID = 1000908
                                             AND ai.ISACTIVE = 'Y'
                                             AND ROWNUM = 1
                                          ), 0) AS BON_VENTE,
                                      NVL((SELECT SUM(ms.qtyonhand - ms.qtyreserved)
                                           FROM m_storage ms
                                           WHERE ms.m_product_id = :product_id
                                             AND ms.ad_client_id = 1000000
                                             AND ms.qtyonhand > 0
                                             AND ms.m_locator_id = 1000614
                                          ), 0) AS qty_dispo,
                                      (SELECT ds.NAME 
                                       FROM C_BPartner_Product bp, M_DiscountSchema ds
                                       WHERE bp.M_PRODUCT_ID = :product_id
                                         AND bp.ISACTIVE = 'Y' 
                                         AND bp.C_BP_Group_ID = 1001330
                                         AND ds.M_DiscountSchema_ID = bp.M_DISCOUNTSCHEMA_ID
                                         AND ROWNUM = 1
                                      ) AS REM_POT,
                                      (SELECT ds.NAME 
                                       FROM C_BPartner_Product bp, M_DiscountSchema ds
                                       WHERE bp.M_PRODUCT_ID = :product_id
                                         AND bp.ISACTIVE = 'Y' 
                                         AND bp.C_BP_Group_ID = 1000003
                                         AND ds.M_DiscountSchema_ID = bp.M_DISCOUNTSCHEMA_ID
                                         AND ROWNUM = 1
                                      ) AS REM_PARA,
                                      :fake_reception_id AS FAKE_RECEPTION_ID,
                                      :product_id AS ORIGINAL_PRODUCT_ID,
                                      :attributesetinstance_id AS ORIGINAL_ATTRIBUTESETINSTANCE_ID
                                  FROM M_PRODUCT p
                                  JOIN XX_SALESCONTEXT sc ON p.XX_SALESCONTEXT_ID = sc.XX_SALESCONTEXT_ID
                                  WHERE p.M_PRODUCT_ID = :product_id
                                    AND p.ISACTIVE = 'Y'
                                    AND p.AD_CLIENT_ID = 1000000
                                """
                                
                                # Execute query with parameters from fake_reception
                                fake_cursor.execute(fake_query, {
                                    'product_id': product_id,
                                    'attributesetinstance_id': attributesetinstance_id,
                                    'fake_reception_id': fake_reception_id
                                })
                                
                                # Fetch results and group by supplier
                                fake_columns = [desc[0] for desc in fake_cursor.description]
                                for fake_row in fake_cursor.fetchall():
                                    fake_row_dict = {fake_columns[i]: fake_row[i] for i in range(len(fake_columns))}
                                    fake_supplier = fake_row_dict.pop('SUPPLIER_NAME')
                                    
                                    # Handle case where supplier name might be None
                                    if fake_supplier is None:
                                        fake_supplier = "Unknown Supplier"
                                    
                                    if fake_supplier not in fake_grouped:
                                        fake_grouped[fake_supplier] = []
                                    fake_grouped[fake_supplier].append(fake_row_dict)
                            
                            # Convert to list of dicts for JSON (same structure as recieved_products_bydate)
                            fake_reception_data = [{"SUPPLIER_NAME": k, "PRODUCTS": v} for k, v in fake_grouped.items()]
                            
                except mysql.connector.Error as e:
                    logger.error(f"Error fetching fake reception data: {e}")
                finally:
                    mysql_cursor.close()
                    mysql_conn.close()
        except Exception as e:
            logger.error(f"Error processing fake reception data: {e}")

        # Return both datasets
        # Merge suppliers that exist in both datasets
        merged_data = {}
        
        # First, add all real reception data
        for supplier_data in reception_data:
            supplier_name = supplier_data["SUPPLIER_NAME"]
            merged_data[supplier_name] = {
                "SUPPLIER_NAME": supplier_name,
                "PRODUCTS": supplier_data["PRODUCTS"].copy()  # Copy real reception products
            }
        
        # Then, merge fake reception data
        for fake_supplier_data in fake_reception_data:
            fake_supplier_name = fake_supplier_data["SUPPLIER_NAME"]
            
            if fake_supplier_name in merged_data:
                # Supplier exists in both - merge products
                merged_data[fake_supplier_name]["PRODUCTS"].extend(fake_supplier_data["PRODUCTS"])
            else:
                # Supplier only exists in fake data - add as new supplier
                merged_data[fake_supplier_name] = {
                    "SUPPLIER_NAME": fake_supplier_name,
                    "PRODUCTS": fake_supplier_data["PRODUCTS"].copy()
                }
        
        # Convert back to list format
        combined_data = list(merged_data.values())
        
        return jsonify({
            "reception_data": combined_data,  # Now contains merged data
            "fake_reception_data": []  # Empty since data is now merged
        })

    except Exception as e:
        logger.error(f"Error fetching recieved products by date: {e}")
        return jsonify({"error": f"Could not fetch recieved products: {str(e)}"}), 500

# ...existing code...



@app.route('/details-products', methods=['GET'])
def details_products():
    try:
        product_id = request.args.get("product_id", None)
        category = request.args.get("category", "all")  # Default to "all" if no category provided
        
        if not product_id:
            return jsonify({"error": "Product ID is required"}), 400

        data = details_products_data(product_id, category)
        return jsonify(data)

    except Exception as e:
        logger.error(f"Error fetching updated inventory products: {e}")
        return jsonify({"error": "Failed to fetch updated inventory products"}), 500
    



def details_products_data(product_id, category="all"):
    """
    Fetch inventory product data where:
    - ANY of these is non-zero: QTY_ONHAND, QTY_RESERVED, QTY_DISPO, QTYORDERED
    - AND GUARANTEEDATE is NOT NULL
    """
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            
            # Define locator IDs for different categories
            locator_groups = {
                "all": "(1001135, 1000614, 1001128, 1001136, 1001020, 1000314, 1000210, 1000211, 1000109, 1000209, 1000213, 1000214, 1000414, 1000817, 1001129)",
                "preparation": "(1001135, 1000614, 1001128, 1001136, 1001020)",
                "tempo": "(1000314, 1000210, 1000211, 1000109, 1000209, 1000213, 1000214, 1000414, 1000817, 1001129)"
            }
            
            locator_list = locator_groups.get(category, locator_groups["all"])
            
            logger.info(f"Fetching inventory data for product_id: {product_id}, category: {category}")
            
            query = f"""
            SELECT
                p.name AS PRODUCT,
                (SELECT lot FROM m_attributesetinstance WHERE m_attributesetinstance_id = mst.m_attributesetinstance_id) AS LOT,
                (SELECT description FROM m_attributesetinstance WHERE m_attributesetinstance_id = mst.m_attributesetinstance_id) AS DESCRIPTION,
                (SELECT valuenumber FROM m_attributeinstance WHERE m_attributesetinstance_id = mst.m_attributesetinstance_id AND m_attribute_id = 1000503) AS PPA,
                (mst.qtyonhand - mst.QTYRESERVED) AS QTY_DISPO,
                mst.m_attributesetinstance_id as M_ATTRIBUTESSETINSTANCE_ID,
                mst.qtyonhand AS QTY_ONHAND,
                mst.QTYRESERVED AS QTY_RESERVED,
                mst.QTYORDERED AS QTYORDERED,
                mats.guaranteedate AS GUARANTEEDATE,
                ROUND(
                    (
                        (
                            (SELECT valuenumber FROM m_attributeinstance WHERE m_attributesetinstance_id = mst.m_attributesetinstance_id AND m_attribute_id = 1000501)
                            - 
                            ((SELECT valuenumber FROM m_attributeinstance WHERE m_attributesetinstance_id = mst.m_attributesetinstance_id AND m_attribute_id = 1000501) 
                             * 
                             (SELECT NVL(valuenumber, 0) FROM m_attributeinstance WHERE m_attributesetinstance_id = mst.m_attributesetinstance_id AND m_attribute_id = 1001009) / 100)
                        ) 
                        / 
                        (1 + (SELECT NVL(valuenumber, 0) FROM m_attributeinstance WHERE m_attributesetinstance_id = mst.m_attributesetinstance_id AND m_attribute_id = 1000808) / 100)
                    ), 2
                ) AS P_REVIENT,
                (
                    SELECT mt.movementtype 
                    FROM m_transaction mt 
                    WHERE mt.m_product_id = mst.m_product_id 
                    AND mt.m_attributesetinstance_id = mst.m_attributesetinstance_id 
                    AND mt.m_locator_id = mst.m_locator_id
                    ORDER BY mt.created DESC 
                    FETCH FIRST 1 ROW ONLY
                ) AS LTS
            FROM
                m_product p
                INNER JOIN m_storage mst ON p.m_product_id = mst.m_product_id
                INNER JOIN m_attributesetinstance mats ON mst.m_attributesetinstance_id = mats.m_attributesetinstance_id
            WHERE
                p.m_product_id = :product_id
                AND mst.m_locator_id IN {locator_list}
                AND (
                    mst.qtyonhand != 0 
                    OR mst.QTYRESERVED != 0 
                    OR (mst.qtyonhand - mst.QTYRESERVED) != 0
                    OR mst.QTYORDERED != 0
                )
                AND mats.guaranteedate IS NOT NULL  -- NEW: Exclude NULL guarantee dates
            ORDER BY
                p.name, mats.guaranteedate
            """
            
            cursor.execute(query, {"product_id": product_id})
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]

            return data

    except Exception as e:
        logger.error(f"Error fetching inventory products: {e}")
        return {"error": "An error occurred while fetching inventory products."}



# MySQL connection for bank data
def get_localdb_connection():
    try:
        return mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="bnm",
            charset="utf8",
            use_unicode=True,
            autocommit=False
        )
    except mysql.connector.Error as err:
        logger.error(f"Error connecting to MySQL database: {err}")
        return None


@app.route('/insert_fake_reception', methods=['POST'])
def insert_fake_reception():
    """
    Simple function to insert data into fake_reception table
    """
    try:
        # Get parameters from request
        data = request.get_json()
        m_attributesetinstance_id = data.get('m_attributesetinstance_id')
        product_id = data.get('product_id')
        product_name = data.get('product_name')
        
        # Validate required parameters
        if not product_id or not product_name:
            return jsonify({"error": "product_id and product_name are required"}), 400
        
        # Connect to MySQL
        mysql_conn = get_localdb_connection()
        if not mysql_conn:
            return jsonify({"error": "Could not connect to MySQL database"}), 500
        
        try:
            mysql_cursor = mysql_conn.cursor()
            
            # Insert query (explicitly set created_at for old MySQL versions)
            insert_query = """
                INSERT INTO fake_reception (m_attributesetinstance_id, product_id, product_name, created_at)
                VALUES (%s, %s, %s, NOW())
            """
            
            mysql_cursor.execute(insert_query, (
                m_attributesetinstance_id,
                product_id,
                product_name
            ))
            
            mysql_conn.commit()
            
            return jsonify({
                "success": True,
                "message": "Data inserted successfully",
                "inserted_id": mysql_cursor.lastrowid
            }), 200
            
        except mysql.connector.Error as e:
            mysql_conn.rollback()
            logger.error(f"Error inserting data: {e}")
            return jsonify({"error": f"Could not insert data: {str(e)}"}), 500
        finally:
            mysql_cursor.close()
            mysql_conn.close()
            
    except Exception as e:
        logger.error(f"Error in insert_fake_reception: {e}")
        return jsonify({"error": f"Could not process request: {str(e)}"}), 500


@app.route('/get_fake_reception', methods=['GET'])
def get_fake_reception():
    """
    Get all records from fake_reception table
    """
    try:
        mysql_conn = get_localdb_connection()
        if not mysql_conn:
            return jsonify({"error": "Could not connect to MySQL database"}), 500
        
        try:
            mysql_cursor = mysql_conn.cursor(dictionary=True)
            
            # Query to get all records
            query = """
                SELECT 
                    id,
                    m_attributesetinstance_id,
                    product_id,
                    product_name,
                    created_at
                FROM fake_reception
                ORDER BY created_at DESC
            """
            
            mysql_cursor.execute(query)
            results = mysql_cursor.fetchall()
            
            # Convert datetime objects to strings for JSON serialization
            for record in results:
                if record.get('created_at'):
                    record['created_at'] = record['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            
            return jsonify({
                "data": results,
                "count": len(results)
            }), 200
            
        except mysql.connector.Error as e:
            logger.error(f"Error fetching data from MySQL: {e}")
            return jsonify({"error": f"Could not fetch data: {str(e)}"}), 500
        finally:
            mysql_cursor.close()
            mysql_conn.close()
            
    except Exception as e:
        logger.error(f"Error in get_fake_reception: {e}")
        return jsonify({"error": f"Could not process request: {str(e)}"}), 500


@app.route('/delete_fake_reception/<int:record_id>', methods=['DELETE'])
def delete_fake_reception(record_id):
    """
    Delete a specific record from fake_reception table
    """
    try:
        mysql_conn = get_localdb_connection()
        if not mysql_conn:
            return jsonify({"error": "Could not connect to MySQL database"}), 500
        
        try:
            mysql_cursor = mysql_conn.cursor()
            
            # First check if record exists
            check_query = "SELECT id, product_name FROM fake_reception WHERE id = %s"
            mysql_cursor.execute(check_query, (record_id,))
            existing_record = mysql_cursor.fetchone()
            
            if not existing_record:
                return jsonify({"error": "Record not found"}), 404
            
            # Delete the record
            delete_query = "DELETE FROM fake_reception WHERE id = %s"
            mysql_cursor.execute(delete_query, (record_id,))
            mysql_conn.commit()
            
            if mysql_cursor.rowcount > 0:
                return jsonify({
                    "success": True,
                    "message": f"Record deleted successfully",
                    "deleted_id": record_id
                }), 200
            else:
                return jsonify({"error": "No record was deleted"}), 400
            
        except mysql.connector.Error as e:
            mysql_conn.rollback()
            logger.error(f"Error deleting data from MySQL: {e}")
            return jsonify({"error": f"Could not delete data: {str(e)}"}), 500
        finally:
            mysql_cursor.close()
            mysql_conn.close()
            
    except Exception as e:
        logger.error(f"Error in delete_fake_reception: {e}")
        return jsonify({"error": f"Could not process request: {str(e)}"}), 500


# ...existing code...

# simulation part

@app.route('/simulation_fetchBCCBProduct', methods=['GET'])
def simulation_fetch_bccb_p():
    bccb = request.args.get('bccb')
    ad_org_id = request.args.get('ad_org_id')

    data = simulation_fetch_bccb_product(bccb, ad_org_id)
    return jsonify(data)



def simulation_fetch_bccb_product(bccb, ad_org_id):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT product, qty, remise, marge, priceentered, remise_vente, bonus_vente, p_revient,ventef, pricelist
                FROM (
                    SELECT det.*, 
                           ROUND((det.ventef - det.p_revient) / det.p_revient * 100, 2) AS marge 
                    FROM (
                        SELECT lot.*, 
                               (lot.priceentered - ((lot.priceentered * NVL(lot.remise_vente, 0)) / 100)) / 
                               (1 + (lot.bonus_vente / 100)) AS ventef 
                        FROM (
                            SELECT ol.priceentered AS priceentered, 
                                   ol.qtyentered AS qty, 
                                   ol.pricelist AS pricelist,
                                   mp.name AS product, 
                                   ol.discount / 100 AS remise, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1000504) AS p_revient, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1000908) AS bonus_vente, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1001408) AS remise_vente 
                            FROM c_orderline ol 
                            INNER JOIN c_order o ON o.c_order_id = ol.c_order_id 
                            INNER JOIN m_product mp ON ol.m_product_id = mp.m_product_id 
                            WHERE ol.qtyentered > 0 
                                  AND (:bccb IS NULL OR UPPER(o.documentno) LIKE UPPER(:bccb) || '%')
                                  AND o.AD_Org_ID = 1000000
                        ) lot
                    ) det
                )
            """
            
            params = {
                'bccb': bccb or None
            }
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching BCCB product data: {e}")
        return {"error": "An error occurred while fetching BCCB product data."}



@app.route('/real_simulation_all', methods=['GET'])
def simulation_get_simulation_all():
    ndocument = request.args.get('ndocument')
    if not ndocument:
        return jsonify({'error': 'Missing ndocument parameter'}), 400
    result = simulation_fetch_simulation_by_ndocument(ndocument)
    if result:
        return jsonify(result)
    else:
        return jsonify({'error': 'No data found for the given ndocument'}), 404





# New: fetch simulation by ndocument
def simulation_fetch_simulation_by_ndocument(ndocument):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT 
                    CAST(org.name AS VARCHAR2(300)) AS organisation,
                    CAST(co.documentno AS VARCHAR2(50)) AS ndocument,
                    CAST(cb.name AS VARCHAR2(300)) AS tier,
                    co.dateordered AS datecommande,
                    CAST(us.name AS VARCHAR2(100)) AS vendeur,
                    ROUND(((co.totallines / (SELECT SUM(mat.valuenumber * li.qtyentered) 
                         FROM c_orderline li 
                         INNER JOIN m_attributeinstance mat ON mat.m_attributesetinstance_id = li.m_attributesetinstance_id
                         WHERE mat.m_attribute_id = 1000504 
                           AND li.c_order_id = co.c_order_id 
                           AND li.qtyentered > 0 
                         GROUP BY li.c_order_id)) - 1) * 100, 2) AS marge,
                    ROUND(co.totallines, 2) AS montant,
                    (SELECT mat.valuenumber 
                     FROM c_orderline li 
                     INNER JOIN m_attributeinstance mat ON mat.m_attributesetinstance_id = li.m_attributesetinstance_id
                     WHERE mat.m_attribute_id = 1000504 
                       AND li.c_order_id = co.c_order_id 
                       AND li.qtyentered > 0 
                       AND ROWNUM = 1) AS valuenumber
                FROM 
                    c_order co
                INNER JOIN ad_org org ON co.ad_org_id = org.ad_org_id
                INNER JOIN c_bpartner cb ON co.c_bpartner_id = cb.c_bpartner_id
                INNER JOIN ad_user us ON co.salesrep_id = us.ad_user_id
                WHERE 
                     co.ad_org_id = 1000000
                    AND issotrx = 'Y'
                    AND co.documentno = :ndocument
            """
            cursor.execute(query, {'ndocument': ndocument})
            row = cursor.fetchone()
            if row:
                columns = [col[0] for col in cursor.description]
                return dict(zip(columns, row))
            else:
                return None
    except Exception as e:
        logging.error(f"Error fetching simulation data by ndocument: {e}")
        return {"error": "An error occurred while fetching simulation data by ndocument."}




@app.route('/simulation_fetch-product-details', methods=['GET'])
def simulation_fetch_product_details():
    try:
        product_name = request.args.get("product_name", None)
        
        if not product_name:
            return jsonify({"error": "Product name is required"}), 400

        data = simulation_fetch_product_details_data(product_name)
        return jsonify(data)

    except Exception as e:
        logger.error(f"Error fetching product details: {e}")
        return jsonify({"error": "Failed to fetch product details"}), 500




 
def simulation_fetch_product_details_data(product_name):
    """
    Fetch detailed product information similar to the marge data structure
    """
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            
            query = """
            SELECT
                *
            FROM
                (
                    SELECT
                        "source"."FOURNISSEUR" "FOURNISSEUR",
                        "source"."PRODUCT" "PRODUCT",
                        "source"."P_ACHAT" "P_ACHAT",
                        "source"."P_VENTE" "P_VENTE",
                        "source"."REM_ACHAT" "REM_ACHAT",
                        "source"."REM_VENTE" "REM_VENTE",
                        "source"."BON_ACHAT" "BON_ACHAT",
                        "source"."BON_VENTE" "BON_VENTE",
                        "source"."REMISE_AUTO" "REMISE_AUTO",
                        "source"."BONUS_AUTO" "BONUS_AUTO",
                        "source"."P_REVIENT" "P_REVIENT",
                        "source"."MARGE" "MARGE",
                        "source"."LABO" "LABO",
                        "source"."LOT" "LOT",
                        "source"."LOT_ACTIVE" "LOT_ACTIVE",

                        "source"."QTY" "QTY",
                        "source"."QTY_DISPO" "QTY_DISPO",
                        "source"."GUARANTEEDATE" "GUARANTEEDATE",
                        "source"."PPA" "PPA",
                        "source"."LOCATION" "LOCATION"

                    FROM
                        (
                            SELECT
                                DISTINCT fournisseur,
                                product,
                                p_achat,
                                p_vente,
                                round(rem_achat, 2) AS rem_achat,
                                rem_vente,
                                round(bon_achat, 2) AS bon_achat,
                                bon_vente,
                                remise_auto,
                                bonus_auto,
                                round(p_revient, 2) AS p_revient,
                                LEAST(round((marge), 2), 100) AS marge,
                                labo,
                                lot,
                                lot_active,
                                qty,
                                qty_dispo,
                                guaranteedate,
                                ppa,
                                CASE 
                                    WHEN m_locator_id = 1000614 THEN 'Préparation'
                                    WHEN m_locator_id = 1001135 THEN 'HANGAR'
                                    WHEN m_locator_id = 1001128 THEN 'Dépot_réserve'
                                    WHEN m_locator_id = 1001136 THEN 'HANGAR_'
                                    WHEN m_locator_id = 1001020 THEN 'Depot_Vente'
                                END AS location
                            FROM
                                (
                                    SELECT
                                        d.*,
                                        LEAST(
                                            round(
                                                (((ventef - ((ventef * nvl(rma, 0)) / 100))) - p_revient) / p_revient * 100,
                                                2
                                            ), 
                                            100
                                        ) AS marge
                                    FROM
                                        (
                                            SELECT
                                                det.*,
                                                (det.p_achat - ((det.p_achat * det.rem_achat) / 100)) / (1 + (det.bon_achat / 100)) p_revient,
                                                (
                                                    det.p_vente - ((det.p_vente * nvl(det.rem_vente, 0)) / 100)
                                                ) / (
                                                    1 + (
                                                        CASE
                                                            WHEN det.bna > 0 THEN det.bna
                                                            ELSE det.bon_vente
                                                        END / 100
                                                    )
                                                ) ventef
                                            FROM
                                                (
                                                    SELECT
                                                        p.name product,
                                                        (
                                                            SELECT
                                                                NAME
                                                            FROM
                                                                XX_Laboratory
                                                            WHERE
                                                                XX_Laboratory_id = p.XX_Laboratory_id
                                                        ) labo,
                                                        mst.qtyonhand qty,
                                                        (mst.qtyonhand - mst.QTYRESERVED) qty_dispo,
                                                        mst.m_locator_id,
                                                        mati.value fournisseur,
                                                        mats.guaranteedate,
                                                        md.name remise_auto,
                                                        sal.description bonus_auto,
                                                        md.flatdiscount rma,
                                                        TO_NUMBER(
                                                            CASE
                                                                WHEN REGEXP_LIKE(sal.name, '^[0-9]+$') THEN sal.name
                                                                ELSE NULL
                                                            END
                                                        ) AS bna,
                                                        (
                                                            SELECT
                                                                valuenumber
                                                            FROM
                                                                m_attributeinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                                AND m_attribute_id = 1000501
                                                        ) p_achat,
                                                        (
                                                            SELECT
                                                                valuenumber
                                                            FROM
                                                                m_attributeinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                                AND m_attribute_id = 1001009
                                                        ) rem_achat,
                                                        (
                                                            SELECT
                                                                valuenumber
                                                            FROM
                                                                m_attributeinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                                AND m_attribute_id = 1000808
                                                        ) bon_achat,
                                                        (
                                                            SELECT
                                                                valuenumber
                                                            FROM
                                                                m_attributeinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                                AND m_attribute_id = 1000502
                                                        ) p_vente,
                                                        (
                                                            SELECT
                                                                valuenumber
                                                            FROM
                                                                m_attributeinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                                AND m_attribute_id = 1001408
                                                        ) rem_vente,
                                                        (
                                                            SELECT
                                                                valuenumber
                                                            FROM
                                                                m_attributeinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                                AND m_attribute_id = 1000908
                                                        ) bon_vente,
                                                        (
                                                            SELECT
                                                                lot
                                                            FROM
                                                                m_attributesetinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                        ) lot,
                                                        (
                                                            SELECT
                                                                isactive
                                                            FROM
                                                                m_attributesetinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                        ) lot_active,
                                                        (
                                                            SELECT
                                                                valuenumber
                                                            FROM
                                                                m_attributeinstance
                                                            WHERE
                                                                m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                                AND m_attribute_id = 1000503
                                                        ) ppa
                                                    FROM
                                                        m_product p
                                                        INNER JOIN m_storage mst ON p.m_product_id = mst.m_product_id
                                                        INNER JOIN m_attributeinstance mati ON mst.m_attributesetinstance_id = mati.m_attributesetinstance_id
                                                        INNER JOIN m_attributesetinstance mats ON mst.m_attributesetinstance_id = mats.m_attributesetinstance_id
                                                        LEFT JOIN C_BPartner_Product cp ON cp.m_product_id = p.m_product_id
                                                            OR cp.C_BPartner_Product_id IS NULL
                                                        LEFT JOIN M_DiscountSchema md ON cp.M_DiscountSchema_id = md.M_DiscountSchema_id
                                                        LEFT JOIN XX_SalesContext sal ON p.XX_SalesContext_ID = sal.XX_SalesContext_ID
                                                    WHERE
                                                        mati.m_attribute_id = 1000508
                                                        AND mst.m_locator_id IN (1001135, 1000614, 1001128, 1001136, 1001020)
                                                        AND mst.qtyonhand != 0
                                                        AND p.name = :product_name
                                                    ORDER BY
                                                        p.name
                                                ) det
                                            WHERE
                                                det.rem_achat < 200
                                        ) d
                                )
                            GROUP BY
                                fournisseur,
                                product,
                                p_achat,
                                p_vente,
                                rem_achat,
                                rem_vente,
                                bon_achat,
                                bon_vente,
                                remise_auto,
                                bonus_auto,
                                p_revient,
                                marge,
                                labo,
                                lot,
                                lot_active,

                                qty,
                                qty_dispo,
                                guaranteedate,
                                ppa,
                                m_locator_id
                            ORDER BY
                                fournisseur
                        ) "source"
                )
            WHERE
                rownum <= 1048575
            """

            cursor.execute(query, {"product_name": product_name})
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]

            return data

    except Exception as e:
        logger.error(f"Error fetching product details: {e}")
        return {"error": "An error occurred while fetching product details."}


@app.route('/simulation_listproduct')
def simulation_listproduct():
    try:
        data = simulation_fetch_all_products()
        return jsonify(data)
    except Exception as e:
        logger.error(f"Error fetching products list: {e}")
        return jsonify({"error": "An error occurred while fetching products list."}), 500

def simulation_fetch_all_products():
    """
    Fetch all products with basic information for search functionality
    """
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT DISTINCT
                    mp.name AS NAME,
                    mp.value AS CODE,
                    mp.m_product_id AS PRODUCT_ID
                FROM m_product mp
                WHERE mp.isactive = 'Y'
                AND mp.issold = 'Y'
                ORDER BY mp.name
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching all products: {e}")
        return []
    



if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000, debug=True)