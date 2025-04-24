
import oracledb
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import logging
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

app = Flask(__name__)
CORS(app, supports_credentials=True)  # Important!
# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure Oracle database connection pool
DB_POOL = oracledb.create_pool(
    user="compiere",
    password="compiere",
    dsn="192.168.1.213/compiere",
    min=2,
    max=10,
    increment=1
)


# Test database connection
def test_db_connection():
    try:
        with DB_POOL.acquire() as connection:
            logger.info("Database connection successful.")
        return True
    except Exception as e:
        logger.error(f"Database connection failed: {str(e)}")
        return False
    










@app.route('/download-quota-excel', methods=['GET'])
def download_quota_excel():
    data = fetch_quota_product()  # Fetch all data
    produit_filter = request.args.get("produit", "").strip()

    # Apply filter if provided (already applied inside fetch_quota_product, but double check here)
    if produit_filter:
        data = [row for row in data if produit_filter.lower() in row["NAME"].lower()]

    return generate_excel_quota(data, produit_filter)

#----------------------------
def generate_excel_quota(data, produit_filter):
    if not data:
        return {"error": "No data available"}, 400

    df = pd.DataFrame(data)
    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered Quota"

    # Apply header styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(df.columns.tolist())
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Append data with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Generate filename dynamically based on filter
    today_date = datetime.now().strftime("%d-%m-%Y")
    filter_text = f"produit-{produit_filter}" if produit_filter else ""
    filename = f"Quota_{filter_text}_{today_date}.xlsx" if filter_text else f"Quota_{today_date}.xlsx"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Fetch product quota data from Oracle DB
def fetch_quota_product():
    try:
        produit = request.args.get('produit', '')  # Get 'produit' param from URL, default to empty string

        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            # Base SQL with optional filtering
            query = """
                SELECT m.name, SUM(s.QTYONHAND) AS QTY
                FROM m_product m
                JOIN m_storage s ON s.M_PRODUCT_ID = m.M_PRODUCT_ID
                WHERE m.XX_SalesContext_ID = 1000100
                    AND m.AD_Client_ID = 1000000
                    AND s.QTYONHAND > 0
                    AND s.M_Locator_ID = 1000614
            """

            # If produit filter is provided, add it to the query
            if produit:
                query += " AND LOWER(m.name) LIKE :produit"

            query += """
                GROUP BY m.name
                ORDER BY m.name
            """

            # Bind parameter if filtering
            if produit:
                cursor.execute(query, {"produit": produit.lower() + '%'})
            else:
                cursor.execute(query)

            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            return data

    except Exception as e:
        logger.error(f"Error fetching product quota data: {e}")
        return {"error": "An error occurred while fetching product quota data."}





@app.route('/quota-product', methods=['GET'])
def quota_product():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_quota_product()
    return jsonify(data)



def fetch_quota_operator():
    try:
        produit = request.args.get('produit', '')

        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            # SQL query with optional filtering
            query = """
                SELECT a.name, SUM(q.quantity - q.QTYORDERED) AS qty
                FROM AD_USER a
                JOIN XX_Quota q ON q.AD_USER_id = a.AD_USER_id
                JOIN M_PRODUCT m ON m.M_PRODUCT_id = q.M_PRODUCT_ID
                WHERE q.M_Warehouse_ID = 1000000
            """

            # Apply filter if produit is provided
            if produit:
                query += " AND LOWER(m.name) LIKE :produit"

            query += """
                GROUP BY a.name
                ORDER BY qty DESC
            """

            # Execute with binding
            if produit:
                cursor.execute(query, {"produit": f"%{produit.lower()}%"})
            else:
                cursor.execute(query)

            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            return data

    except Exception as e:
        logger.error(f"Error fetching operator quota data: {e}")
        return {"error": "An error occurred while fetching operator quota data."}

@app.route('/quota-operator', methods=['GET'])
def quota_operator():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_quota_operator()
    return jsonify(data)


if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000)