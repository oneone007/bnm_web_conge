
import oracledb
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import logging
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure Oracle database connection pool
DB_POOL = oracledb.create_pool(
    user="compiere",
    password="compiere",
    dsn="192.168.1.213/compiere",
    min=2,
    max=10,
    increment=1
)

# Fetch reserved data from Oracle DB
def fetch_reserved_data():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
            SELECT 
                us.name AS OPERATEUR,
                co.documentno AS NDOCUMENT,
                m.name AS PRODUCT,
                co.dateordered AS DATECOMMANDE,
                s.qtyreserved AS TOTALRESERVE,
                cl.qtyreserved AS QTYRESERVE,
                l.name AS NAME,
                CASE
                    WHEN co.docstatus = 'CO' THEN 'prepared'
                    ELSE 'not prepared'
                END AS STATUS
            FROM
                m_storage s
            INNER JOIN c_orderline cl
                ON cl.m_attributesetinstance_id = s.m_attributesetinstance_id
            INNER JOIN c_order co
                ON co.c_order_id = cl.c_order_id
            INNER JOIN m_product m
                ON m.m_product_id = s.m_product_id
            INNER JOIN ad_user us
                ON us.ad_user_id = co.salesrep_id
            JOIN XX_Laboratory l
                ON l.XX_Laboratory_ID = m.XX_Laboratory_ID
            WHERE
                cl.qtyreserved != 0
                AND co.c_doctypetarget_id = 1000539
                AND cl.m_product_id = s.m_product_id
                AND co.ad_org_id = 1000000
                AND s.qtyonhand > 0
                AND s.qtyreserved > 0
            FETCH FIRST 1048575 ROWS ONLY
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching reserved data: {e}")
        return {"error": "An error occurred while fetching reserved data."}

# Fetch remise data from Oracle DB
def fetch_remise_data():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
            WITH LastFournisseur AS (
                SELECT
                    cf.m_product_id,
                    cf.c_bpartner_id,
                    ROW_NUMBER() OVER (
                        PARTITION BY cf.m_product_id
                        ORDER BY cf.dateinvoiced DESC
                    ) AS rn
                FROM
                    xx_ca_fournisseur_facture cf
            )
            SELECT
                mp.name AS product,
                md.name AS reward,
                l.name AS laboratory_name,
                cb.name AS fournisseur,
                CBG.NAME AS Type_Client
            FROM
                m_product mp
                INNER JOIN C_BPartner_Product cbp ON mp.m_product_id = cbp.m_product_id
                INNER JOIN  C_BP_Group CBG  ON CBG.C_BP_Group_ID = CBP.C_BP_Group_ID
                JOIN XX_Laboratory l ON l.XX_Laboratory_ID = mp.XX_Laboratory_ID
                JOIN LastFournisseur lf ON mp.m_product_id = lf.m_product_id AND lf.rn = 1
                JOIN c_bpartner cb ON cb.c_bpartner_id = lf.c_bpartner_id
                INNER JOIN M_DiscountSchema md ON cbp.M_DiscountSchema_id = md.M_DiscountSchema_id
            WHERE
                cbp.C_BPartner_Product_id IS NOT NULL
                AND cbp.m_discountschema_id IS NOT NULL
            ORDER BY
                mp.name
            FETCH FIRST 1048575 ROWS ONLY
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching remise data: {e}")
        return {"error": "An error occurred while fetching remise data."}

# Fetch marge data from Oracle DB
def fetch_marge_data():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
SELECT
    *
FROM
    (
        SELECT
            "source"."FOURNISSEUR" "FOURNISSEUR",
            "source"."PRODUCT" "PRODUCT",
            "source"."P_ACHAT" "P_ACHAT",
            "source"."P_VENTE" "P_VENTE",
            "source"."REM_ACHAT" "REM_ACHAT",
            "source"."REM_VENTE" "REM_VENTE",
            "source"."BON_ACHAT" "BON_ACHAT",
            "source"."BON_VENTE" "BON_VENTE",
            "source"."REMISE_AUTO" "REMISE_AUTO",
            "source"."BONUS_AUTO" "BONUS_AUTO",
            "source"."P_REVIENT" "P_REVIENT",
            "source"."MARGE" "MARGE",
            "source"."LABO" "LABO",
            "source"."LOT" "LOT",
            "source"."QTY" "QTY",
            "source"."GUARANTEEDATE" "GUARANTEEDATE",  -- Added the GUARANTEEDATE column
            "source"."LOCATION" "LOCATION"
        FROM
            (
                SELECT
                    DISTINCT fournisseur,
                    product,
                    p_achat,
                    p_vente,
                    round(rem_achat, 2) AS rem_achat,
                    rem_vente,
                    round(bon_achat, 2) AS bon_achat,
                    bon_vente,
                    remise_auto,
                    bonus_auto,
                    round(p_revient, 2) AS p_revient,
                    LEAST(round((marge), 2), 100) AS marge,
                    labo,
                    lot,
                    qty,
                    guaranteedate,  -- Added the GUARANTEEDATE column
                    CASE 
                        WHEN m_locator_id = 1000614 THEN 'Préparation'
                        WHEN m_locator_id = 1001135 THEN 'HANGAR'
                        WHEN m_locator_id = 1001128 THEN 'Dépot_réserve'
                        WHEN m_locator_id = 1001136 THEN 'HANGAR_'
                        WHEN m_locator_id = 1001020 THEN 'Depot_Vente'
                    END AS location
                FROM
                    (
                        SELECT
                            d.*,
                            LEAST(
                                round(
                                    (((ventef - ((ventef * nvl(rma, 0)) / 100))) - p_revient) / p_revient * 100,
                                    2
                                ), 
                                100
                            ) AS marge
                        FROM
                            (
                                SELECT
                                    det.*,
                                    (det.p_achat - ((det.p_achat * det.rem_achat) / 100)) / (1 + (det.bon_achat / 100)) p_revient,
                                    (
                                        det.p_vente - ((det.p_vente * nvl(det.rem_vente, 0)) / 100)
                                    ) / (
                                        1 + (
                                            CASE
                                                WHEN det.bna > 0 THEN det.bna
                                                ELSE det.bon_vente
                                            END / 100
                                        )
                                    ) ventef
                                FROM
                                    (
                                        SELECT
                                            p.name product,
                                            (
                                                SELECT
                                                    NAME
                                                FROM
                                                    XX_Laboratory
                                                WHERE
                                                    XX_Laboratory_id = p.XX_Laboratory_id
                                            ) labo,
                                            mst.qtyonhand qty,
                                            mst.m_locator_id,
                                            mati.value fournisseur,
                                            mats.guaranteedate,
                                            md.name remise_auto,
                                            sal.description bonus_auto,
                                            md.flatdiscount rma,
                                            TO_NUMBER(
                                                CASE
                                                    WHEN REGEXP_LIKE(sal.name, '^[0-9]+$') THEN sal.name
                                                    ELSE NULL
                                                END
                                            ) AS bna,
                                            (
                                                SELECT
                                                    valuenumber
                                                FROM
                                                    m_attributeinstance
                                                WHERE
                                                    m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                    AND m_attribute_id = 1000501
                                            ) p_achat,
                                            (
                                                SELECT
                                                    valuenumber
                                                FROM
                                                    m_attributeinstance
                                                WHERE
                                                    m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                    AND m_attribute_id = 1001009
                                            ) rem_achat,
                                            (
                                                SELECT
                                                    valuenumber
                                                FROM
                                                    m_attributeinstance
                                                WHERE
                                                    m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                    AND m_attribute_id = 1000808
                                            ) bon_achat,
                                            (
                                                SELECT
                                                    valuenumber
                                                FROM
                                                    m_attributeinstance
                                                WHERE
                                                    m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                    AND m_attribute_id = 1000502
                                            ) p_vente,
                                            (
                                                SELECT
                                                    valuenumber
                                                FROM
                                                    m_attributeinstance
                                                WHERE
                                                    m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                    AND m_attribute_id = 1001408
                                            ) rem_vente,
                                            (
                                                SELECT
                                                    valuenumber
                                                FROM
                                                    m_attributeinstance
                                                WHERE
                                                    m_attributesetinstance_id = mst.m_attributesetinstance_id
                                                    AND m_attribute_id = 1000908
                                            ) bon_vente,
                                            (
                                                SELECT
                                                    lot
                                                FROM
                                                    m_attributesetinstance
                                                WHERE
                                                    m_attributesetinstance_id = mst.m_attributesetinstance_id
                                            ) lot
                                        FROM
                                            m_product p
                                            INNER JOIN m_storage mst ON p.m_product_id = mst.m_product_id
                                            INNER JOIN m_attributeinstance mati ON mst.m_attributesetinstance_id = mati.m_attributesetinstance_id
                                            INNER JOIN m_attributesetinstance mats ON mst.m_attributesetinstance_id = mats.m_attributesetinstance_id
                                            LEFT JOIN C_BPartner_Product cp ON cp.m_product_id = p.m_product_id
                                                OR cp.C_BPartner_Product_id IS NULL
                                            LEFT JOIN M_DiscountSchema md ON cp.M_DiscountSchema_id = md.M_DiscountSchema_id
                                            LEFT JOIN XX_SalesContext sal ON p.XX_SalesContext_ID = sal.XX_SalesContext_ID
                                        WHERE
                                            mati.m_attribute_id = 1000508
                                            AND mst.m_locator_id IN (1001135, 1000614, 1001128, 1001136, 1001020)
                                            AND mst.qtyonhand != 0
                                        ORDER BY
                                            p.name
                                    ) det
                                WHERE
                                    det.rem_achat < 200
                            ) d
                    )
                GROUP BY
                    fournisseur,
                    product,
                    p_achat,
                    p_vente,
                    rem_achat,
                    rem_vente,
                    bon_achat,
                    bon_vente,
                    remise_auto,
                    bonus_auto,
                    p_revient,
                    marge,
                    labo,
                    lot,
                    qty,
                    guaranteedate,  -- Added to GROUP BY
                    m_locator_id
                ORDER BY
                    fournisseur
            ) "source"
    )
WHERE
    rownum <= 1048575
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching marge data: {e}")
        return {"error": "An error occurred while fetching marge data."}


# Fetch bonus data from Oracle DB
def fetch_bonus_data():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
            SELECT DISTINCT 
                mp.name AS product,
                (SELECT description 
                 FROM XX_SalesContext xsc 
                 WHERE mp.XX_SalesContext_id = xsc.XX_SalesContext_id) AS bonus,
                l.name AS laboratory_name,
                cbp.name AS fournisseur
            FROM 
                m_product mp
            JOIN XX_Laboratory l ON l.XX_Laboratory_ID = mp.XX_Laboratory_ID
            JOIN m_storage s ON s.m_product_id = mp.m_product_id
            JOIN M_ATTRIBUTEINSTANCE asi ON s.M_ATTRIBUTESETINSTANCE_ID = asi.M_ATTRIBUTESETINSTANCE_ID
            JOIN c_bpartner cbp ON cbp.c_bpartner_id = ValueNUMBER_of_ASI('Fournisseur', asi.m_attributesetinstance_id)
            WHERE 
                mp.xx_salescontext_id NOT IN (1000000, 1000100)
                AND mp.ad_org_id = 1000000
            ORDER BY 
                mp.name
            FETCH FIRST 1048575 ROWS ONLY
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching bonus data: {e}")
        return {"error": "An error occurred while fetching bonus data."}



# Test database connection
def test_db_connection():
    try:
        with DB_POOL.acquire() as connection:
            logger.info("Database connection successful.")
        return True
    except Exception as e:
        logger.error(f"Database connection failed: {str(e)}")
        return False
    



# Route to fetch reserved data
@app.route('/fetch-reserved-data', methods=['GET'])
def fetch_reserved():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_reserved_data()
    return jsonify(data)

# Route to fetch remise data
@app.route('/fetch-remise-data', methods=['GET'])
def fetch_remise():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_remise_data()
    return jsonify(data)

# Route to fetch marge data
@app.route('/fetch-data', methods=['GET'])
def fetch_marge():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_marge_data()
    return jsonify(data)


# Route to fetch bonus data
@app.route('/fetch-bonus-data', methods=['GET'])
def fetch_bonus():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_bonus_data()
    return jsonify(data)

# Route to download Excel for reserved data
@app.route('/download-reserved-excel', methods=['GET'])
def download_reserved_excel():
    reserve_value = request.args.get("reserve", "RESERVE")
    data = fetch_reserved_data()
    return generate_excel(data, reserve_value)

# Route to download Excel for remise data
@app.route('/download-remise-excel', methods=['GET'])
def download_remise_excel():
    remise_value = request.args.get("remise", "REMISE")
    data = fetch_remise_data()
    return generate_excel(data, remise_value)


# Route to download Excel for bonus data
@app.route('/download-bonus-excel', methods=['GET'])
def download_bonus_excel():
    bonus_value = request.args.get("bonus", "BONUS")
    data = fetch_bonus_data()
    return generate_excel(data, bonus_value)


# Route to download Excel for marge data
@app.route('/download-marge-excel', methods=['GET'])
def download_marge_excel():
    data = fetch_marge_data()  # Fetch all data
    filters = {
        "fournisseur": request.args.get("fournisseur", "").strip(),
        "product": request.args.get("product", "").strip(),
        "marge": request.args.get("marge", "").strip(),
    }

    # Apply filters if provided
    if filters["fournisseur"]:
        data = [row for row in data if filters["fournisseur"].lower() in row["FOURNISSEUR"].lower()]
    if filters["product"]:
        data = [row for row in data if filters["product"].lower() in row["PRODUCT"].lower()]
    if filters["marge"]:
        try:
            marge_value = float(filters["marge"])
            data = [row for row in data if float(row["MARGE"]) >= marge_value]
        except ValueError:
            return jsonify({"error": "Invalid marge value"}), 400

    return generate_excel_marge(data, filters)

#----------------------------
def generate_excel_marge(data, filters):
    if not data:
        return {"error": "No data available"}, 400

    df = pd.DataFrame(data)
    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered Data"

    # Apply header styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(df.columns.tolist())
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Append data with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Generate filename dynamically based on filters (include values)
    today_date = datetime.now().strftime("%d-%m-%Y")
    filter_text = "_".join([f"{k}-{v}" for k, v in filters.items() if v])  # Include values
    filename = f"Marge_{filter_text}_{today_date}.xlsx" if filter_text else f"Marge_{today_date}.xlsx"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# Helper function to generate Excel file
def generate_excel(data, value):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "All Data"

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    table = Table(displayName="DataTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"{value}_{today_date}.xlsx"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def fetch_stock_data_from_db(fournisseur=None, magasin=None, emplacement=None, name=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
            SELECT 
                mati.value AS fournisseur, 
                m.name,  
                SUM(m_storage.qtyonhand) AS qty,
                SUM(M_ATTRIBUTEINSTANCE.valuenumber * m_storage.qtyonhand) AS prix,
                SUM(m_storage.qtyonhand - m_storage.QTYRESERVED) AS qty_dispo, 
                SUM(M_ATTRIBUTEINSTANCE.valuenumber * (m_storage.qtyonhand - m_storage.QTYRESERVED)) AS prix_dispo,
                ml.M_Locator_ID AS locatorid,
                m.m_product_id AS productid,
                1 AS sort_order,
                CASE 
                    WHEN ml.M_Locator_ID = 1000614 THEN 'Préparation'
                    WHEN ml.M_Locator_ID = 1001135 THEN 'HANGAR'
                    WHEN ml.M_Locator_ID = 1001128 THEN 'Dépot_réserve'
                    WHEN ml.M_Locator_ID = 1001136 THEN 'HANGAR_'
                    WHEN ml.M_Locator_ID = 1001020 THEN 'Depot_Vente'
                    ELSE 'Unknown' 
                END AS place
            FROM 
                M_ATTRIBUTEINSTANCE
            JOIN 
                m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
            JOIN 
                M_PRODUCT m ON m.M_PRODUCT_id = m_storage.M_PRODUCT_id
            JOIN 
                M_Locator ml ON ml.M_Locator_ID = m_storage.M_Locator_ID
            INNER JOIN 
                m_attributeinstance mati ON m_storage.m_attributesetinstance_id = mati.m_attributesetinstance_id
            WHERE 
                M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                AND m_storage.qtyonhand > 0
                AND mati.m_attribute_id = 1000508
                AND m_storage.AD_Client_ID = 1000000
            """

            params = {}

            if fournisseur:
                query += " AND mati.value LIKE :fournisseur || '%'"
                params["fournisseur"] = fournisseur

            if name:
                query += " AND m.name LIKE :name || '%'"
                params["name"] = name

            if magasin:
                query += """
                AND m_storage.M_Locator_ID IN (
                    SELECT M_Locator_ID 
                    FROM M_Locator 
                    WHERE M_Warehouse_ID IN (
                        SELECT M_Warehouse_ID 
                        FROM M_Warehouse 
                        WHERE VALUE LIKE :magasin || '%'
                    )
                )
                """
                params["magasin"] = magasin
            else:
                query += """
                AND m_storage.M_Locator_ID IN (
                    SELECT M_Locator_ID 
                    FROM M_Locator 
                    WHERE M_Warehouse_ID IN (
                        SELECT M_Warehouse_ID 
                        FROM M_Warehouse 
                        WHERE VALUE IN ('HANGAR', '1-Dépôt Principal', '8-Dépot réserve', '88-Dépot Hangar réserve')
                    )
                )
                """

            if emplacement:
                query += """
                AND (
                    (M_Warehouse_ID != 1000000 AND 
                     m_storage.M_Locator_ID IN (
                         SELECT M_Locator_ID 
                         FROM M_Locator 
                         WHERE value LIKE :emplacement || '%'
                     ))
                    OR 
                    (M_Warehouse_ID = 1000000 AND 
                     m.M_Locator_ID IN (
                         SELECT M_Locator_ID 
                         FROM M_Locator 
                         WHERE value LIKE :emplacement || '%'
                     ))
                )
                """
                params["emplacement"] = emplacement

            query += """
            GROUP BY m.name, mati.value, m.m_product_id, ml.M_Locator_ID

            UNION ALL

            SELECT 
                CAST('Total' AS NVARCHAR2(300)) AS fournisseur, 
                CAST('' AS NVARCHAR2(300)) AS name, 
                SUM(m_storage.qtyonhand) AS qty,
                SUM(M_ATTRIBUTEINSTANCE.valuenumber * m_storage.qtyonhand) AS prix,
                SUM(m_storage.qtyonhand - m_storage.QTYRESERVED) AS qty_dispo, 
                SUM(M_ATTRIBUTEINSTANCE.valuenumber * (m_storage.qtyonhand - m_storage.QTYRESERVED)) AS prix_dispo,
                NULL AS locatorid,
                NULL AS productid,
                0 AS sort_order,
                NULL AS place
            FROM 
                M_ATTRIBUTEINSTANCE
            JOIN 
                m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
            JOIN 
                M_PRODUCT m ON m.M_PRODUCT_ID = m_storage.M_PRODUCT_ID
            JOIN 
                M_Locator ml ON ml.M_Locator_ID = m_storage.M_Locator_ID
            INNER JOIN 
                m_attributeinstance mati ON m_storage.m_attributesetinstance_id = mati.m_attributesetinstance_id
            WHERE 
                M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                AND m_storage.qtyonhand > 0
                AND mati.m_attribute_id = 1000508
                AND m_storage.AD_Client_ID = 1000000
            """

            if fournisseur:
                query += " AND mati.value LIKE :fournisseur || '%'"

            if name:
                query += " AND m.name LIKE :name || '%'"

            if magasin:
                query += """
                AND m_storage.M_Locator_ID IN (
                    SELECT M_Locator_ID 
                    FROM M_Locator 
                    WHERE M_Warehouse_ID IN (
                        SELECT M_Warehouse_ID 
                        FROM M_Warehouse 
                        WHERE VALUE LIKE :magasin || '%'
                    )
                )
                """
            else:
                query += """
                AND m_storage.M_Locator_ID IN (
                    SELECT M_Locator_ID 
                    FROM M_Locator 
                    WHERE M_Warehouse_ID IN (
                        SELECT M_Warehouse_ID 
                        FROM M_Warehouse 
                        WHERE VALUE IN ('HANGAR', '1-Dépôt Principal', '8-Dépot réserve', '88-Dépot Hangar réserve')
                    )
                )
                """

            if emplacement:
                query += """
                AND (
                    (M_Warehouse_ID != 1000000 AND 
                     m_storage.M_Locator_ID IN (
                         SELECT M_Locator_ID 
                         FROM M_Locator 
                         WHERE value LIKE :emplacement || '%'
                     ))
                    OR 
                    (M_Warehouse_ID = 1000000 AND 
                     m.M_Locator_ID IN (
                         SELECT M_Locator_ID 
                         FROM M_Locator 
                         WHERE value LIKE :emplacement || '%'
                     ))
                )
                """

            query += """
            ORDER BY sort_order, fournisseur, name
            """

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]

            return data

    except Exception as e:
        logger.error(f"Database error: {e}")
        return {"error": "An error occurred while fetching stock data."}

# Flask route
@app.route('/fetch-stock-data', methods=['GET'])
def fetch_stock_data():
    try:
        fournisseur = request.args.get("fournisseur", None)
        magasin = request.args.get("magasin", None)
        emplacement = request.args.get("emplacement", None)
        name = request.args.get("name", None)

        data = fetch_stock_data_from_db(fournisseur, magasin, emplacement, name)
        return jsonify(data)

    except Exception as e:
        logger.error(f"Error fetching stock data: {e}")
        return jsonify({"error": "Failed to fetch stock data"}), 500


def generate_excel_stock(data):
    # Create a DataFrame from the data
    df = pd.DataFrame(data)

    # Create workbook and worksheet using openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Data"

    # Define header styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Write headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Apply auto filter (using the dimensions of the sheet)
    ws.auto_filter.ref = ws.dimensions

    # Write data rows with alternating row styling for readability
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Optionally, create an Excel table
    table = Table(displayName="DataTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save workbook to a BytesIO stream
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route('/download-stock-excel', methods=['GET'])
def download_stock_excel():
    try:
        # Get query parameters
        fournisseur = request.args.get("fournisseur", None)
        magasin = request.args.get("magasin", None)
        emplacement = request.args.get("emplacement", None)

        # Fetch data from the database
        data = fetch_stock_data_from_db(fournisseur, magasin, emplacement)

        # Check if data contains an error
        if "error" in data:
            return jsonify({"error": data["error"]}), 500

        # Check if data is empty
        if not data:
            return jsonify({"error": "No data available to generate Excel"}), 400

        # Generate Excel file in memory
        excel_output = generate_excel_stock(data)

        # Generate filename based on parameters
        today_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"stock_data_{today_date}.xlsx"

        if fournisseur:
            filename = f"stock_data_fournisseur_{fournisseur}_{today_date}.xlsx"
        if magasin:
            filename = f"stock_data_magasin_{magasin}_{today_date}.xlsx"
        if emplacement:
            filename = f"stock_data_emplacement_{emplacement}_{today_date}.xlsx"
        if fournisseur and magasin:
            filename = f"stock_data_fournisseur_{fournisseur}_magasin_{magasin}_{today_date}.xlsx"
        if fournisseur and emplacement:
            filename = f"stock_data_fournisseur_{fournisseur}_emplacement_{emplacement}_{today_date}.xlsx"
        if magasin and emplacement:
            filename = f"stock_data_magasin_{magasin}_emplacement_{emplacement}_{today_date}.xlsx"
        if fournisseur and magasin and emplacement:
            filename = f"stock_data_fournisseur_{fournisseur}_magasin_{magasin}_emplacement_{emplacement}_{today_date}.xlsx"

        # Send file as a download response
        return send_file(
            excel_output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"Error generating Excel: {e}")
        return jsonify({"error": "Failed to generate Excel file"}), 500


    
# Flask route for generating Excel from fetched stock data

def fetch_magasins_from_db(magasin=None, emplacement=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT DISTINCT m.value AS MAGASIN
                FROM M_Locator ml
                JOIN M_Warehouse m ON m.M_WAREHOUSE_ID = ml.M_WAREHOUSE_ID
                WHERE m.ISACTIVE = 'Y'
                  AND m.AD_Client_ID = 1000000
                  AND ml.ISACTIVE = 'Y'
                  AND ml.AD_Client_ID = 1000000
            """

            params = {}

            # Dynamically add filters
            if magasin:
                query += " AND m.value LIKE :magasin || '%'"
                params["magasin"] = magasin

            if emplacement:
                query += " AND ml.value LIKE :emplacement || '%'"
                params["emplacement"] = emplacement

            query += " ORDER BY m.value"

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]

            return data

    except Exception as e:
        logger.error(f"Error fetching magasins: {e}")
        return {"error": "An error occurred while fetching magasins."}


def fetch_emplacements_from_db(magasin=None, emplacement=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT ml.value AS EMPLACEMENT
                FROM M_Locator ml
                JOIN M_Warehouse m ON m.M_WAREHOUSE_ID = ml.M_WAREHOUSE_ID
                WHERE m.ISACTIVE = 'Y'
                  AND m.AD_Client_ID = 1000000
                  AND ml.ISACTIVE = 'Y'
                  AND ml.AD_Client_ID = 1000000
            """

            params = {}

            # Dynamically add filters
            if magasin:
                query += " AND m.value LIKE :magasin || '%'"
                params["magasin"] = magasin

            if emplacement:
                query += " AND ml.value LIKE :emplacement || '%'"
                params["emplacement"] = emplacement

            query += " ORDER BY m.value"

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]

            return data

    except Exception as e:
        logger.error(f"Error fetching emplacements: {e}")
        return {"error": "An error occurred while fetching emplacements."}


@app.route('/fetch-magasins', methods=['GET'])
def fetch_magasins():
    try:
        magasin = request.args.get("magasin", None)
        emplacement = request.args.get("emplacement", None)

        data = fetch_magasins_from_db(magasin, emplacement)
        return jsonify(data)

    except Exception as e:
        logger.error(f"Error fetching magasins: {e}")
        return jsonify({"error": "Failed to fetch magasins"}), 500

@app.route('/fetch-emplacements', methods=['GET'])
def fetch_emplacements():
    try:
        magasin = request.args.get("magasin", None)
        emplacement = request.args.get("emplacement", None)

        data = fetch_emplacements_from_db(magasin, emplacement)
        return jsonify(data)

    except Exception as e:
        logger.error(f"Error fetching emplacements: {e}")
        return jsonify({"error": "Failed to fetch emplacements"}), 500
    





# Helper function to generate Excel file

# Fetch total recap data
def fetch_rcap_data(start_date, end_date, ad_org_id):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT 
                    SUM(xf.TOTALLINE) AS CHIFFRE, 
                    SUM(xf.qtyentered) AS QTY,
                    SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION) AS MARGE,
                    SUM(xf.CONSOMATION) AS CONSOMATION,
                    CASE 
                        WHEN SUM(xf.CONSOMATION) < 0 
                        THEN ROUND(((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / (SUM(xf.CONSOMATION) * -1)), 4)
                        ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                    END AS POURCENTAGE
                FROM xx_ca_fournisseur xf
                WHERE 
                    xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                    AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND xf.AD_Org_ID = :ad_org_id
                    AND xf.DOCSTATUS != 'RE'
            """
            cursor.execute(query, {'start_date': start_date, 'end_date': end_date, 'ad_org_id': ad_org_id})
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching data: {e}")
        return {"error": "An error occurred while fetching data."}

# Fetch total recap data

# Fetch fournisseur data



def fetch_bccb_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT DOCUMENTNO, DATEORDERED, GRANDTOTAL, 
                       ROUND(AVG(marge * 100), 2) AS marge
                FROM (
                    SELECT det.*
                    FROM (
                        SELECT lot.*,  
                               (lot.priceentered - ((lot.priceentered * NVL(lot.remise_vente, 0)) / 100)) / (1 + (lot.bonus_vente / 100)) AS ventef
                        FROM (
                            SELECT 
                                ol.priceentered, 
                                (SELECT valuenumber 
                                 FROM m_attributeinstance 
                                 WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                       AND m_attribute_id = 1000504) AS p_revient,
                                (SELECT valuenumber 
                                 FROM m_attributeinstance 
                                 WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                       AND m_attribute_id = 1000908) AS bonus_vente,
                                (SELECT valuenumber 
                                 FROM m_attributeinstance 
                                 WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                       AND m_attribute_id = 1001408) AS remise_vente,
                                c.DOCUMENTNO, 
                                c.DATEORDERED, 
                                c.GRANDTOTAL,
                                CASE 
                                    WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                                    WHEN SUM(xf.CONSOMATION) < 0 THEN ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION) * -1)
                                    ELSE (SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION)
                                END AS marge
                            FROM C_Order c
                            JOIN C_OrderLine ol ON c.C_Order_ID = ol.C_Order_ID 
                            JOIN M_InOut mi ON mi.C_Order_ID = c.C_Order_ID
                            JOIN xx_ca_fournisseur xf ON xf.DOCUMENTNO = mi.DOCUMENTNO
                            JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                            JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                            JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                            JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                            WHERE c.AD_Org_ID = 1000000 
                                  AND ol.qtyentered > 0
                                  AND xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                      AND TO_DATE(:end_date, 'YYYY-MM-DD')
                                  AND (c.DOCSTATUS = 'CL' OR c.DOCSTATUS = 'CO')
                                  AND c.C_DocType_ID IN (1000539, 1001408)
                                  AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                                  AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                                  AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                                  AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                                  AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                                  AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                            GROUP BY c.DOCUMENTNO, c.DATEORDERED, c.GRANDTOTAL, ol.priceentered, ol.m_attributesetinstance_id
                            ORDER BY c.DOCUMENTNO
                        ) lot
                    ) det
                )
                GROUP BY DOCUMENTNO, DATEORDERED, GRANDTOTAL
                ORDER BY DOCUMENTNO
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching BCCB recap: {e}")
        return {"error": "An error occurred while fetching BCCB recap."}


@app.route('/fetchBCCBRecap', methods=['GET']) 
def fetch_bccb():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    data = fetch_bccb_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone)
    return jsonify(data)

 


def fetch_bccb_recap_fact(start_date, end_date, fournisseur, product, client, operateur, bccb, zone):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT DOCUMENTNO, DATEORDERED, GRANDTOTAL, 
      round(avg( CASE WHEN marge < 0 THEN 0 ELSE marge END),2) AS marge
FROM (  
    SELECT det.*, 
           ROUND((det.ventef - det.p_revient) / det.p_revient * 100, 2) AS marge
    FROM (  
        SELECT lot.*,  
               (lot.priceentered - ((lot.priceentered * NVL(lot.remise_vente, 0)) / 100)) / (1 + (lot.bonus_vente / 100)) AS ventef
        FROM (  
            SELECT ol.priceentered, 
                   (SELECT valuenumber 
                    FROM m_attributeinstance 
                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                          AND m_attribute_id = 1001519) AS p_revient,
                   
                   (SELECT valuenumber 
                    FROM m_attributeinstance 
                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                          AND m_attribute_id = 1001523) AS bonus_vente,
                   
                   (SELECT valuenumber 
                    FROM m_attributeinstance 
                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                          AND m_attribute_id = 1001532) AS remise_vente,
                   
                   c.DOCUMENTNO, 
                   c.DATEORDERED, 
                   c.GRANDTOTAL
            FROM C_Order c
            JOIN C_OrderLine ol ON c.C_Order_ID = ol.C_Order_ID 
            JOIN M_InOut mi ON mi.C_Order_ID = c.C_Order_ID
            JOIN xx_ca_fournisseur xf ON xf.DOCUMENTNO = mi.DOCUMENTNO
            JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
            JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
            JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
            JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
            WHERE c.AD_Org_ID = 1000012 
                  AND ol.qtyentered > 0
                  AND xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                      AND TO_DATE(:end_date, 'YYYY-MM-DD')
                  AND (c.DOCSTATUS = 'CL' OR c.DOCSTATUS = 'CO')
                  AND c.C_DocType_ID IN (1002845, 1002846)
                  AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                  AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                  AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                  AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                  AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                  AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
            GROUP BY c.DOCUMENTNO, c.DATEORDERED, c.GRANDTOTAL, ol.priceentered, ol.m_attributesetinstance_id
            ORDER BY c.DOCUMENTNO
        )  lot  
    )  det  
)  
group by DOCUMENTNO, DATEORDERED, GRANDTOTAL
order by DOCUMENTNO

            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None
            }
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching BCCB recap: {e}")
        return {"error": "An error occurred while fetching BCCB recap."}

@app.route('/fetchBCCBRecapfact', methods=['GET'])
def fetch_bccb_fact():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    data = fetch_bccb_recap_fact(start_date, end_date, fournisseur, product, client, operateur, bccb, zone)
    return jsonify(data)




@app.route('/fetchZoneRecap', methods=['GET'])
def fetch_zone():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400
    if not ad_org_id:
        return jsonify({"error": "Missing ad_org_id parameter"}), 400

    data = fetch_zone_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id)
    return jsonify(data)


def fetch_zone_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT * FROM (
                    -- Recap by Zone
                    SELECT 
                        CAST(sr.name AS VARCHAR2(100)) AS "ZONE",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        0 AS "SORT_ORDER"
                    FROM C_SalesRegion sr
                    JOIN C_BPartner_Location bpl ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN xx_ca_fournisseur xf ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                    GROUP BY sr.name

                    UNION ALL

                    -- Total Row
                    SELECT 
                        'Total' AS "ZONE",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        1 AS "SORT_ORDER"
                    FROM C_SalesRegion sr
                    JOIN C_BPartner_Location bpl ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN xx_ca_fournisseur xf ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                ) ORDER BY "SORT_ORDER", "TOTAL" DESC
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'ad_org_id': ad_org_id
            }
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching zone recap: {e}")
        return {"error": "An error occurred while fetching zone recap."}


def fetch_client_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT * FROM (
                    -- Recap by Client
                    SELECT 
                        CAST(cb.name AS VARCHAR2(100)) AS "CLIENT",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        0 AS "SORT_ORDER"
                    FROM C_BPartner cb
                    JOIN xx_ca_fournisseur xf ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                    GROUP BY cb.name

                    UNION ALL

                    -- Total Row
                    SELECT 
                        'Total' AS "CLIENT",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        1 AS "SORT_ORDER"
                    FROM C_BPartner cb
                    JOIN xx_ca_fournisseur xf ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                ) ORDER BY "SORT_ORDER", "TOTAL" DESC
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'ad_org_id': ad_org_id
            }
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching client recap: {e}")
        return {"error": "An error occurred while fetching client recap."}


@app.route('/fetchClientRecap', methods=['GET'])
def fetch_client():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400
    if not ad_org_id:
        return jsonify({"error": "Missing ad_org_id parameter"}), 400

    data = fetch_client_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id)
    return jsonify(data)

def fetch_operator_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT * FROM (
                    SELECT CAST(au.name AS VARCHAR2(100)) AS "OPERATEUR", 
                           SUM(xf.TOTALLINE) AS "TOTAL", 
                           SUM(xf.qtyentered) AS "QTY",
                           CASE 
                               WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                               WHEN SUM(xf.CONSOMATION) < 0 THEN ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION) * -1)
                               ELSE (SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION)
                           END AS "MARGE",
                           0 AS "SORT_ORDER"
                    FROM AD_User au
                    JOIN xx_ca_fournisseur xf ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                    GROUP BY au.name

                    UNION ALL

                    SELECT 'Total' AS "OPERATEUR", 
                           SUM(xf.TOTALLINE) AS "TOTAL", 
                           SUM(xf.qtyentered) AS "QTY",
                           CASE 
                               WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                               WHEN SUM(xf.CONSOMATION) < 0 THEN ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION) * -1)
                               ELSE (SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION)
                           END AS "MARGE",
                           1 AS "SORT_ORDER"
                    FROM AD_User au
                    JOIN xx_ca_fournisseur xf ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                ) ORDER BY "SORT_ORDER", "TOTAL" DESC
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'ad_org_id': ad_org_id
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching operator recap: {e}")
        return {"error": "An error occurred while fetching operator recap."}


@app.route('/fetchOperatorRecap', methods=['GET'])
def fetch_operator():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400
    if not ad_org_id:
        return jsonify({"error": "Missing ad_org_id parameter"}), 400

    data = fetch_operator_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id)
    return jsonify(data)










@app.route('/fetchTotalrecapData', methods=['GET'])
def fetch_recap():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    ad_org_id = request.args.get('ad_org_id')  # Get ad_org_id from request

    if not start_date or not end_date or not ad_org_id:
        return jsonify({"error": "Missing start_date, end_date, or ad_org_id parameters"}), 400

    try:
        ad_org_id = int(ad_org_id)  # Convert to integer
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    data = fetch_rcap_data(start_date, end_date, ad_org_id)
    return jsonify(data)






@app.route('/download-totalrecap-excel', methods=['GET'])
def download_totalrecap_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    ad_org_id = request.args.get('ad_org_id')

    if not start_date or not end_date or not ad_org_id:
        return jsonify({"error": "Missing start_date, end_date, or ad_org_id parameters"}), 400

    try:
        ad_org_id = int(ad_org_id)  # Convert to integer
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data from the database
    data = fetch_rcap_data(start_date, end_date, ad_org_id)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id
    today_date = datetime.now().strftime("%d-%m-%Y")
    if ad_org_id == 1000012:
        filename = f"TotalRecap_facturation_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"TotalRecap_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_totalrecap(data, filename)


def generate_excel_totalrecap(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Total Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="TotalRecapTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def generate_excel_totalrecap_facturation(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Total Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="TotalRecapTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")




def fetch_fournisseur_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT * FROM (
                    SELECT 
                        CAST(xf.name AS VARCHAR2(300)) AS FOURNISSEUR,   
                        SUM(xf.TOTALLINE) AS total, 
                        SUM(xf.qtyentered) AS QTY,
                        ROUND(
                            CASE 
                                WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                                WHEN SUM(xf.CONSOMATION) < 0 THEN ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION) * -1) * 100
                                ELSE ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION)) * 100
                            END, 4) AS marge,
                        0 AS sort_order
                    FROM xx_ca_fournisseur xf
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER C ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
                        AND xf.AD_Org_ID = :ad_org_id
                        AND xf.DOCSTATUS != 'RE'
                        AND (:fournisseur IS NULL OR xf.name LIKE :fournisseur || '%')
                        AND (:product IS NULL OR xf.product LIKE :product || '%')
                        AND (:client IS NULL OR cb.name LIKE :client || '%')
                        AND (:operateur IS NULL OR au.name LIKE :operateur || '%')
                        AND (:bccb IS NULL OR C.DOCUMENTNO LIKE :bccb || '%')
                        AND (:zone IS NULL OR sr.name LIKE :zone || '%')
                    GROUP BY xf.name
                    UNION ALL
                    SELECT 
                        CAST('Total' AS VARCHAR2(300)) AS name, 
                        SUM(xf.TOTALLINE) AS total, 
                        SUM(xf.qtyentered) AS QTY,
                        ROUND(
                            CASE 
                                WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                                WHEN SUM(xf.CONSOMATION) < 0 THEN ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION) * -1) * 100
                                ELSE ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION)) * 100
                            END, 4) AS marge,
                        1 AS sort_order
                    FROM xx_ca_fournisseur xf
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER C ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
                        AND xf.AD_Org_ID = :ad_org_id
                        AND xf.DOCSTATUS != 'RE'
                        AND (:fournisseur IS NULL OR xf.name LIKE :fournisseur || '%')
                        AND (:product IS NULL OR xf.product LIKE :product || '%')
                        AND (:client IS NULL OR cb.name LIKE :client || '%')
                        AND (:operateur IS NULL OR au.name LIKE :operateur || '%')
                        AND (:bccb IS NULL OR C.DOCUMENTNO LIKE :bccb || '%')
                        AND (:zone IS NULL OR sr.name LIKE :zone || '%')
                )
                ORDER BY sort_order, total DESC
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'ad_org_id': ad_org_id  # New parameter
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching fournisseur data: {e}")
        return {"error": "An error occurred while fetching fournisseur data."}

@app.route('/fetchFournisseurData', methods=['GET'])
def fetch_fournisseur():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')  # New parameter

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    if not ad_org_id:
        return jsonify({"error": "Missing ad_org_id parameter"}), 400  # Ensure ad_org_id is provided

    try:
        ad_org_id = int(ad_org_id)  # Convert to integer
    except ValueError:
        return jsonify({"error": "Invalid ad_org_id format"}), 400

    data = fetch_fournisseur_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id)
    return jsonify(data)


@app.route('/download-fournisseur-excel', methods=['GET'])
def download_fournisseur_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_fournisseur_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id
    today_date = datetime.now().strftime("%d-%m-%Y")
    if ad_org_id == 1000012:
        filename = f"FournisseurRecap_facturation_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"FournisseurRecap_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_fournisseur(data, filename)


def generate_excel_fournisseur(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Fournisseur Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="FournisseurRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


 

def fetch_product_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT * FROM (
                    -- Main Product Breakdown
                    SELECT 
                        CAST(xf.product AS VARCHAR2(400)) AS "PRODUIT",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        0 AS "SORT_ORDER"
                    FROM xx_ca_fournisseur xf
                    LEFT JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    LEFT JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    LEFT JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    LEFT JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    LEFT JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    LEFT JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                    GROUP BY xf.product

                    UNION ALL

                    -- Total Row
                    SELECT 
                        'Total' AS "PRODUIT",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        1 AS "SORT_ORDER"
                    FROM xx_ca_fournisseur xf
                    LEFT JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    LEFT JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    LEFT JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    LEFT JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    LEFT JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    LEFT JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                ) ORDER BY "SORT_ORDER", "TOTAL" DESC
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'ad_org_id': ad_org_id
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]

    except Exception as e:
        logger.error(f"Error fetching product data: {e}")
        return {"error": "An error occurred while fetching product data."}

@app.route('/fetchProductData', methods=['GET'])
def fetch_product():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get("product", "").strip()
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')

    if not start_date or not end_date or not ad_org_id:
        return jsonify({"error": "Missing start_date, end_date, or ad_org_id parameters"}), 400

    data = fetch_product_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id)
    return jsonify(data)




@app.route('/download-product-excel', methods=['GET'])
def download_product_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_product_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id
    today_date = datetime.now().strftime("%d-%m-%Y")
    if ad_org_id == 1000012:
        filename = f"ProductRecap_facturation_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"ProductRecap_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_product(data, filename)


def generate_excel_product(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Product Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="ProductRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

from datetime import datetime

@app.route('/download-zone-excel', methods=['GET'])
def download_zone_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_zone_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id
    today_date = datetime.now().strftime("%d-%m-%Y")
    if ad_org_id == 1000012:
        filename = f"ZoneRecap_facturation_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"ZoneRecap_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_zone(data, filename)


def generate_excel_zone(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Zone Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="ZoneRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/download-client-excel', methods=['GET'])
def download_client_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_client_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id
    today_date = datetime.now().strftime("%d-%m-%Y")
    if ad_org_id == 1000012:
        filename = f"ClientRecap_facturation_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"ClientRecap_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_client(data, filename)


def generate_excel_client(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Client Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="ClientRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route('/download-operator-excel', methods=['GET'])
def download_operator_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_operator_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id
    today_date = datetime.now().strftime("%d-%m-%Y")
    if ad_org_id == 1000012:
        filename = f"OperatorRecap_facturation_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"OperatorRecap_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_operator(data, filename)


def generate_excel_operator(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Operator Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="OperatorRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")





@app.route('/download-BCCB-excel-fac', methods=['GET'])
def download_bccb_excelf():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data
    data = fetch_bccb_recap_fact(start_date, end_date, fournisseur, product, client, operateur, bccb, zone)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"BCCBRecap_Facturation_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_bccbf(data, filename)

def generate_excel_bccbf(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "BCCB Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="BCCBRecapTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")




@app.route('/download-bccb-excel', methods=['GET'])
def download_bccb_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data
    data = fetch_bccb_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"BCCBRecap_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_bccb(data, filename)

def generate_excel_bccb(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "BCCB Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="BCCBRecapTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def fetch_total_recap_achat(start_date, end_date, fournisseur, product):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT
                    SUM(CASE 
                        WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                        ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                    END) AS chiffre
                FROM M_InOut xf
                JOIN M_InOutline mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
                JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
                LEFT JOIN C_InvoiceLine ci ON ci.M_INOUTLINE_ID = mi.M_INOUTLINE_ID
                JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID = mi.M_ATTRIBUTESETINSTANCE_ID
                JOIN M_PRODUCT m ON m.M_PRODUCT_id = mi.M_PRODUCT_id
                WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND ma.M_Attribute_ID = 1000504
                    AND xf.AD_Org_ID = 1000000
                    AND xf.C_DocType_ID IN (1000013, 1000646)
                    AND xf.M_Warehouse_ID IN (1000724, 1000000, 1000720, 1000725)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None
            }

            cursor.execute(query, params)
            row = cursor.fetchone()
            
            return {"chiffre": row[0] if row and row[0] is not None else 0}
    
    except Exception as e:
        logger.error(f"Error fetching total recap achat: {e}")
        return {"error": "An error occurred while fetching total recap achat."}
@app.route('/fetchTotalRecapAchat', methods=['GET'])
def fetch_total_achat():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    data = fetch_total_recap_achat(start_date, end_date, fournisseur, product)
    return jsonify(data)

 
# PRIX DE VENT 1001519   1000084 DOCTYPE  1002854 RETOUR
def fetch_total_recap_achat_fact(start_date, end_date, fournisseur, product):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT
                    SUM(CASE 
                        WHEN xf.C_DocType_ID = 1002854 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                        ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                    END) AS chiffre
                FROM M_InOut xf
                JOIN M_InOutline mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
                JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
                LEFT JOIN C_InvoiceLine ci ON ci.M_INOUTLINE_ID = mi.M_INOUTLINE_ID
                JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID = mi.M_ATTRIBUTESETINSTANCE_ID
                JOIN M_PRODUCT m ON m.M_PRODUCT_id = mi.M_PRODUCT_id
                WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND ma.M_Attribute_ID = 1001519 
                    AND xf.AD_Org_ID = 1000012
                    AND XF.DOCSTATUS!='RE'
                    AND xf.C_DocType_ID IN (1000084, 1002854)
                    AND xf.M_Warehouse_ID IN (1000014, 1000721)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None
            }

            cursor.execute(query, params)
            row = cursor.fetchone()
            
            return {"chiffre": row[0] if row and row[0] is not None else 0}
    
    except Exception as e:
        logger.error(f"Error fetching total recap achat: {e}")
        return {"error": "An error occurred while fetching total recap achat."}
@app.route('/fetchTotalRecapAchat_fact', methods=['GET'])
def fetch_total_achat_fact():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    data = fetch_total_recap_achat_fact(start_date, end_date, fournisseur, product)
    return jsonify(data)

#facturation
def fetch_fournisseur_recap_achat_fact(start_date, end_date, fournisseur, product):
    try:
        # Acquire a connection from the pool
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT
                    CAST(cb.name AS VARCHAR2(300)) AS FOURNISSEUR,   
                    SUM(CASE 
                        WHEN xf.C_DocType_ID = 1002854 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                        ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                    END) AS chiffre,
                    0 AS sort_order
                FROM 
                    M_InOut  xf
                JOIN M_INOUTLINE mi ON mi.M_INOUT_ID=xf.M_INOUT_ID
                JOIN C_BPartner cb ON cb.C_BPARTNER_ID=xf.C_BPARTNER_ID
                LEFT JOIN C_InvoiceLine ci ON ci.M_INOUTLINE_ID=mi.M_INOUTLINE_ID
                JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID=mi.M_ATTRIBUTESETINSTANCE_ID
                JOIN M_PRODUCT m ON m.M_PRODUCT_id=mi.M_PRODUCT_id
                WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND ma.M_Attribute_ID=1001519
                    AND xf.AD_Org_ID = 1000012
                    AND XF.DOCSTATUS!='RE'
                    AND xf.C_DocType_ID IN (1000084, 1002854)
                    AND xf.M_Warehouse_ID IN (1000014, 1000721)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
                GROUP BY 
                    cb.name

                UNION ALL

                SELECT
                    CAST('Total' AS VARCHAR2(300)) AS FOURNISSEUR, 
                    SUM(CASE 
                        WHEN xf.C_DocType_ID = 1002854 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                        ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                    END) AS chiffre,
                    1 AS sort_order
                FROM 
                    M_InOut  xf
                JOIN M_INOUTLINE mi ON mi.M_INOUT_ID=xf.M_INOUT_ID
                JOIN C_BPartner cb ON cb.C_BPARTNER_ID=xf.C_BPARTNER_ID
                LEFT JOIN C_InvoiceLine ci ON ci.M_INOUTLINE_ID=mi.M_INOUTLINE_ID
                JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID=mi.M_ATTRIBUTESETINSTANCE_ID
                JOIN M_PRODUCT m ON m.M_PRODUCT_id=mi.M_PRODUCT_id
                WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND ma.M_Attribute_ID=1001519
                    AND xf.AD_Org_ID = 1000012
                    AND XF.DOCSTATUS!='RE'
                    AND xf.C_DocType_ID IN (1000084, 1002854)
                    AND xf.M_Warehouse_ID IN (1000014, 1000721)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None
            }

            # Execute the query with the provided parameters
            cursor.execute(query, params)

            # Fetch the results
            rows = cursor.fetchall()

            # Format the results into a list of dictionaries
            data = [{"FOURNISSEUR": row[0], "CHIFFRE": row[1], "SORT_ORDER": row[2]} for row in rows]

            return data
    
    except Exception as e:
        logger.error(f"Error fetching fournisseur recap achat: {e}")
        return {"error": "An error occurred while fetching fournisseur recap achat."}

# Flask route to handle the request
@app.route('/fetchfourisseurRecapAchat_fact', methods=['GET'])
def fetch_fournisseur_achat_fact():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    # Ensure both start_date and end_date are provided
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data from the database
    data = fetch_fournisseur_recap_achat_fact(start_date, end_date, fournisseur, product)

    # Return the result as a JSON response
    return jsonify(data)


def fetch_product_achat_recap_fact(start_date, end_date, fournisseur, product):
    try:
        # Acquire a connection from the pool
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT 
                    CAST(m.name AS VARCHAR2(300)) AS produit,   
                    SUM(CASE 
                            WHEN xf.C_DocType_ID = 1002854 THEN -1 * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(mi.QTYENTERED)
                        END) AS qty,
                    
                    SUM(CASE 
                            WHEN xf.C_DocType_ID = 1002854 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                        END) AS chiffre,
                    0 AS sort_order
                FROM 
                    M_InOut xf
                    JOIN M_INOUTLINE mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
                    JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID = mi.M_ATTRIBUTESETINSTANCE_ID
                    JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
                    JOIN M_PRODUCT m ON m.M_PRODUCT_id = mi.M_PRODUCT_id
                WHERE 
                    xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND xf.AD_Org_ID = 1000012
                    AND XF.DOCSTATUS!='RE'
                    AND xf.C_DocType_ID IN (1000084, 1002854)
                    AND ma.M_Attribute_ID = 1001519
                    AND xf.DOCSTATUS = 'CO'
                    AND xf.M_Warehouse_ID IN (1000014, 1000721)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
                GROUP BY 
                    m.name

                UNION ALL

                SELECT
                    CAST('Total' AS VARCHAR2(300)) AS produit, 
                    SUM(CASE 
                            WHEN xf.C_DocType_ID = 1002854 THEN -1 * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(mi.QTYENTERED)
                        END) AS qty,
                    SUM(CASE 
                            WHEN xf.C_DocType_ID = 1002854 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                        END) AS chiffre,
                    1 AS sort_order
                FROM 
                    M_InOut xf
                    JOIN M_INOUTLINE mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
                    JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID = mi.M_ATTRIBUTESETINSTANCE_ID
                    JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
                    JOIN M_PRODUCT m ON m.M_PRODUCT_id = mi.M_PRODUCT_id
                WHERE 
                    xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND xf.AD_Org_ID = 1000012
                    AND XF.DOCSTATUS!='RE'
                    AND xf.C_DocType_ID IN (1000084, 1002854)
                    AND ma.M_Attribute_ID = 1001519
                    AND xf.DOCSTATUS = 'CO'
                    AND xf.M_Warehouse_ID IN (1000014, 1000721)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None
            }

            # Execute the query with the provided parameters
            cursor.execute(query, params)

            # Fetch the results
            rows = cursor.fetchall()

            # Format the results into a list of dictionaries
            data = [{"PRODUIT": row[0], "QTY": row[1], "CHIFFRE": row[2], "SORT_ORDER": row[3]} for row in rows]

            return data
    
    except Exception as e:
        logger.error(f"Error fetching product recap achat: {e}")
        return {"error": "An error occurred while fetching product recap achat."}

@app.route('/fetchProductRecapAchat_fact', methods=['GET'])
def fetch_product_achat_fact():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    # Ensure both start_date and end_date are provided
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data from the database
    data = fetch_product_achat_recap_fact(start_date, end_date, fournisseur, product)

    # Return the result as a JSON response
    return jsonify(data)



def fetch_fournisseur_recap_achat(start_date, end_date, fournisseur, product):
    try:
        # Acquire a connection from the pool
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT
                    CAST(cb.name AS VARCHAR2(300)) AS FOURNISSEUR,   
                    SUM(CASE 
                        WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                        ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                    END) AS chiffre,
                    0 AS sort_order
                FROM 
                    M_InOut  xf
                JOIN M_INOUTLINE mi ON mi.M_INOUT_ID=xf.M_INOUT_ID
                JOIN C_BPartner cb ON cb.C_BPARTNER_ID=xf.C_BPARTNER_ID
                LEFT JOIN C_InvoiceLine ci ON ci.M_INOUTLINE_ID=mi.M_INOUTLINE_ID
                JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID=mi.M_ATTRIBUTESETINSTANCE_ID
                JOIN M_PRODUCT m ON m.M_PRODUCT_id=mi.M_PRODUCT_id
                WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND ma.M_Attribute_ID=1000504
                    AND xf.AD_Org_ID = 1000000
                    AND xf.C_DocType_ID IN (1000013, 1000646)
                    AND xf.M_Warehouse_ID IN (1000724, 1000000, 1000720, 1000725)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
                GROUP BY 
                    cb.name

                UNION ALL

                SELECT
                    CAST('Total' AS VARCHAR2(300)) AS FOURNISSEUR, 
                    SUM(CASE 
                        WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                        ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED) 
                    END) AS chiffre,
                    1 AS sort_order
                FROM 
                    M_InOut  xf
                JOIN M_INOUTLINE mi ON mi.M_INOUT_ID=xf.M_INOUT_ID
                JOIN C_BPartner cb ON cb.C_BPARTNER_ID=xf.C_BPARTNER_ID
                LEFT JOIN C_InvoiceLine ci ON ci.M_INOUTLINE_ID=mi.M_INOUTLINE_ID
                JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID=mi.M_ATTRIBUTESETINSTANCE_ID
                JOIN M_PRODUCT m ON m.M_PRODUCT_id=mi.M_PRODUCT_id
                WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND ma.M_Attribute_ID=1000504
                    AND xf.AD_Org_ID = 1000000
                    AND xf.C_DocType_ID IN (1000013, 1000646)
                    AND xf.M_Warehouse_ID IN (1000724, 1000000, 1000720, 1000725)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None
            }

            # Execute the query with the provided parameters
            cursor.execute(query, params)

            # Fetch the results
            rows = cursor.fetchall()

            # Format the results into a list of dictionaries
            data = [{"FOURNISSEUR": row[0], "CHIFFRE": row[1], "SORT_ORDER": row[2]} for row in rows]

            return data
    
    except Exception as e:
        logger.error(f"Error fetching fournisseur recap achat: {e}")
        return {"error": "An error occurred while fetching fournisseur recap achat."}

# Flask route to handle the request
@app.route('/fetchfourisseurRecapAchat', methods=['GET'])
def fetch_fournisseur_achat():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    # Ensure both start_date and end_date are provided
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data from the database
    data = fetch_fournisseur_recap_achat(start_date, end_date, fournisseur, product)

    # Return the result as a JSON response
    return jsonify(data)


def fetch_product_achat_recap(start_date, end_date, fournisseur, product):
    try:
        # Acquire a connection from the pool
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT 
                    CAST(m.name AS VARCHAR2(300)) AS produit,   
                    SUM(CASE 
                            WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(mi.QTYENTERED)
                        END) AS qty,
                    
                    SUM(CASE 
                            WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                        END) AS chiffre,
                    0 AS sort_order
                FROM 
                    M_InOut xf
                    JOIN M_INOUTLINE mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
                    JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID = mi.M_ATTRIBUTESETINSTANCE_ID
                    JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
                    JOIN M_PRODUCT m ON m.M_PRODUCT_id = mi.M_PRODUCT_id
                WHERE 
                    xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND xf.AD_Org_ID = 1000000
                    AND xf.C_DocType_ID IN (1000013, 1000646)
                    AND ma.M_Attribute_ID = 1000504
                    AND xf.DOCSTATUS = 'CO'
                    AND xf.M_Warehouse_ID IN (1000724, 1000000, 1000720, 1000725)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
                GROUP BY 
                    m.name

                UNION ALL

                SELECT
                    CAST('Total' AS VARCHAR2(300)) AS produit, 
                    SUM(CASE 
                            WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(mi.QTYENTERED)
                        END) AS qty,
                    SUM(CASE 
                            WHEN xf.C_DocType_ID = 1000646 THEN -1 * TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                            ELSE TO_NUMBER(ma.valuenumber) * TO_NUMBER(mi.QTYENTERED)
                        END) AS chiffre,
                    1 AS sort_order
                FROM 
                    M_InOut xf
                    JOIN M_INOUTLINE mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
                    JOIN M_ATTRIBUTEINSTANCE ma ON ma.M_ATTRIBUTESETINSTANCE_ID = mi.M_ATTRIBUTESETINSTANCE_ID
                    JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
                    JOIN M_PRODUCT m ON m.M_PRODUCT_id = mi.M_PRODUCT_id
                WHERE 
                    xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND xf.AD_Org_ID = 1000000
                    AND xf.C_DocType_ID IN (1000013, 1000646)
                    AND ma.M_Attribute_ID = 1000504
                    AND xf.DOCSTATUS = 'CO'
                    AND xf.M_Warehouse_ID IN (1000724, 1000000, 1000720, 1000725)
                    AND (:fournisseur IS NULL OR UPPER(cb.name) LIKE UPPER(:fournisseur) || '%')
                    AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None
            }

            # Execute the query with the provided parameters
            cursor.execute(query, params)

            # Fetch the results
            rows = cursor.fetchall()

            # Format the results into a list of dictionaries
            data = [{"PRODUIT": row[0], "QTY": row[1], "CHIFFRE": row[2], "SORT_ORDER": row[3]} for row in rows]

            return data
    
    except Exception as e:
        logger.error(f"Error fetching product recap achat: {e}")
        return {"error": "An error occurred while fetching product recap achat."}

@app.route('/fetchProductRecapAchat', methods=['GET'])
def fetch_product_achat():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    # Ensure both start_date and end_date are provided
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data from the database
    data = fetch_product_achat_recap(start_date, end_date, fournisseur, product)

    # Return the result as a JSON response
    return jsonify(data)


@app.route('/download-recap-product-achat-excel', methods=['GET'])
def download_recap_product_achat_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
 

    # Ensure both start and end dates are provided
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch the data using the fetch_product_achat_recap function
    data = fetch_product_achat_recap(start_date, end_date, fournisseur, product)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"ProductRecapAchat_{start_date}_to_{end_date}_{today_date}.xlsx"

    # Generate and return the Excel file
    return generate_excel_product_achat(data, filename)

def generate_excel_product_achat(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Convert the data into a DataFrame
    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    # Create an Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers to the sheet
    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table for better presentation
    table = Table(displayName="ProductRecapTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send the file to the client for download
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

  
@app.route('/download-recap-fournisseur-achat-excel', methods=['GET'])
def download_recap_fournisseur_achat_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    # Ensure both start and end dates are provided
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data using the fetch_fournisseur_recap_achat function
    data = fetch_fournisseur_recap_achat(start_date, end_date, fournisseur, product)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"FournisseurRecapAchat_{start_date}_to_{end_date}_{today_date}.xlsx"

    # Generate and return the Excel file
    return generate_excel_fournisseur_achat(data, filename)

def generate_excel_fournisseur_achat(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Convert the data into a DataFrame
    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    # Create an Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Fournisseur Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers to the sheet
    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table for better presentation
    table = Table(displayName="FournisseurRecapAchatTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send the file to the client for download
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


#download achat fact

@app.route('/download-recap-fournisseur-achat_facturation-excel', methods=['GET'])
def download_recap_fournisseur_achatf_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    # Ensure both start_date and end_date are provided
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data using the updated fetch function
    data = fetch_fournisseur_recap_achat_fact(start_date, end_date, fournisseur, product)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"FournisseurRecapAchat_Facturation_{start_date}_to_{end_date}_{today_date}.xlsx"

    # Generate and return the Excel file
    return generate_excel_fournisseur_achatf(data, filename)

def generate_excel_fournisseur_achatf(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Convert the data into a DataFrame
    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    # Create an Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Fournisseur Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers to the sheet
    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table for better presentation
    table = Table(displayName="FournisseurRecapAchatTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send the file to the client for download
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



@app.route('/download-recap-product-achat_facturation-excel', methods=['GET'])
def download_recap_product_achatf_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')

    # Ensure both start_date and end_date are provided
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data using the updated fetch function
    data = fetch_product_achat_recap_fact(start_date, end_date, fournisseur, product)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"Product_RecapAchat_Facturation_{start_date}_to_{end_date}_{today_date}.xlsx"

    # Generate and return the Excel file
    return generate_excel_prdct_achatf(data, filename)

def generate_excel_prdct_achatf(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Convert the data into a DataFrame
    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    # Create an Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Fournisseur Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers to the sheet
    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table for better presentation
    table = Table(displayName="FournisseurRecapAchatTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send the file to the client for download
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")




def fetch_bccb_product(bccb, ad_org_id):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT product, qty, remise, marge 
                FROM (
                    SELECT det.*, 
                           ROUND((det.ventef - det.p_revient) / det.p_revient * 100, 2) AS marge 
                    FROM (
                        SELECT lot.*, 
                               (lot.priceentered - ((lot.priceentered * NVL(lot.remise_vente, 0)) / 100)) / 
                               (1 + (lot.bonus_vente / 100)) AS ventef 
                        FROM (
                            SELECT ol.priceentered AS priceentered, 
                                   ol.qtyentered AS qty, 
                                   mp.name AS product, 
                                   ol.discount / 100 AS remise, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1000504) AS p_revient, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1000908) AS bonus_vente, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1001408) AS remise_vente 
                            FROM c_orderline ol 
                            INNER JOIN c_order o ON o.c_order_id = ol.c_order_id 
                            INNER JOIN m_product mp ON ol.m_product_id = mp.m_product_id 
                            WHERE ol.qtyentered > 0 
                                  AND (:bccb IS NULL OR UPPER(o.documentno) LIKE UPPER(:bccb) || '%')
                                  AND o.AD_Org_ID = :ad_org_id
                        ) lot
                    ) det
                )
            """
            
            params = {
                'bccb': bccb or None,
                'ad_org_id': ad_org_id
            }
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching BCCB product data: {e}")
        return {"error": "An error occurred while fetching BCCB product data."}

def fetch_bccb_productfact(bccb, ad_org_id):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT product, qty, remise, 
                       CASE WHEN marge < 0 THEN 0 ELSE marge END AS marge 
                FROM (
                    SELECT det.*, 
                           ROUND((det.ventef - det.p_revient) / det.p_revient * 100, 2) AS marge 
                    FROM (
                        SELECT lot.*, 
                               (lot.priceentered - ((lot.priceentered * NVL(lot.remise_vente, 0)) / 100)) / 
                               (1 + (lot.bonus_vente / 100)) AS ventef 
                        FROM (
                            SELECT ol.priceentered AS priceentered, 
                                   ol.qtyentered AS qty, 
                                   mp.name AS product, 
                                   ol.discount / 100 AS remise, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1001519) AS p_revient, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1001523) AS bonus_vente, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1001532) AS remise_vente 
                            FROM c_orderline ol 
                            INNER JOIN c_order o ON o.c_order_id = ol.c_order_id 
                            INNER JOIN m_product mp ON ol.m_product_id = mp.m_product_id 
                            WHERE ol.qtyentered > 0 
                                  AND (:bccb IS NULL OR UPPER(o.documentno) LIKE UPPER(:bccb) || '%')
                                  AND o.AD_Org_ID = :ad_org_id
                        ) lot
                    ) det
                )
            """
            
            params = {
                'bccb': bccb or None,
                'ad_org_id': ad_org_id
            }
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching BCCB product data: {e}")
        return {"error": "An error occurred while fetching BCCB product data."}

@app.route('/fetchBCCBProduct', methods=['GET'])
def fetch_bccb_p():
    bccb = request.args.get('bccb')
    ad_org_id = request.args.get('ad_org_id')

    if not ad_org_id:
        return jsonify({"error": "Missing ad_org_id parameter"}), 400

    data = fetch_bccb_product(bccb, ad_org_id)
    return jsonify(data)


@app.route('/fetchBCCBProductfact', methods=['GET'])
def fetch_bccb_pf():
    bccb = request.args.get('bccb')
    ad_org_id = request.args.get('ad_org_id')

    if not ad_org_id:
        return jsonify({"error": "Missing ad_org_id parameter"}), 400

    data = fetch_bccb_productfact(bccb, ad_org_id)
    return jsonify(data)

@app.route('/download-bccb-product-excel', methods=['GET'])
def download_bccb_product_excel():
    bccb = request.args.get('bccb')
    ad_org_id = request.args.get('ad_org_id', '1000012')

    if not bccb:
        return jsonify({"error": "Missing BCCB parameter"}), 400

    try:
        ad_org_id = int(ad_org_id)
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_bccb_product(bccb, ad_org_id)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"BCCB_Product_{bccb}_{today_date}.xlsx"

    return generate_excel_bccb_product(data, filename)


def generate_excel_bccb_product(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "BCCB Product Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="BCCBProductTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route('/download-bccb-product-excel-f', methods=['GET'])
def download_bccb_product_excelf():
    bccb = request.args.get('bccb')
    ad_org_id = request.args.get('ad_org_id', '1000012')

    if not bccb:
        return jsonify({"error": "Missing BCCB parameter"}), 400

    try:
        ad_org_id = int(ad_org_id)
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_bccb_productfact(bccb, ad_org_id)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"BCCB_Product_Facturation_{bccb}_{today_date}.xlsx"

    return generate_excel_bccb_productf(data, filename)


def generate_excel_bccb_productf(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "BCCB Product Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="BCCBProductTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



def fetch_rotation_product_data():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
            select name from M_PRODUCT
WHERE AD_Client_ID = 1000000
AND AD_Org_ID = 1000000
  AND ISACTIVE = 'Y'
ORDER BY name
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching remise data: {e}")
        return {"error": "An error occurred while fetching emplacement data."}
    
@app.route('/fetch-rotation-product-data', methods=['GET'])
def fetch_data():
    data = fetch_rotation_product_data()
    return jsonify(data)


def fetch_historique_rotation(product):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                WITH Latest_Purchase AS (
                    SELECT 
                        cl.M_Product_ID,
                        SUM(cl.qtyentered) AS last_purchase_qty,
                        c.dateinvoiced,
                        ROW_NUMBER() OVER (PARTITION BY cl.M_Product_ID ORDER BY c.dateinvoiced DESC) AS rn
                    FROM 
                        C_InvoiceLine cl
                    JOIN C_Invoice c ON c.C_Invoice_id = cl.C_Invoice_id
                    JOIN M_INOUTLINE ml ON ml.M_INOUTLINE_id = cl.M_INOUTLINE_ID
                    WHERE 
                        c.dateinvoiced BETWEEN TO_DATE('01/01/2020', 'DD/MM/YYYY') AND SYSDATE
                        AND c.AD_Client_ID = 1000000
                        AND c.AD_Org_ID = 1000000
                        AND c.ISSOTRX = 'N'
                        AND c.DOCSTATUS in ('CO','CL')
                        AND ml.M_Locator_ID != 1001020
                    GROUP BY 
                        cl.M_Product_ID, c.dateinvoiced
                ),
                Filtered_Latest_Purchase AS (
                    SELECT 
                        M_Product_ID,
                        last_purchase_qty,
                        dateinvoiced
                    FROM 
                        Latest_Purchase
                    WHERE 
                        rn = 1
                ),
                On_Hand_Quantity AS (
                    SELECT  
                        m.M_Product_ID as midp,
                        m.name AS product_name,
                        SUM(s.QTYONHAND) - SUM(s.QTYRESERVED) AS QTYONHAND
                    FROM 
                        m_product m
                    JOIN m_storage s ON s.M_PRODUCT_ID = m.M_PRODUCT_ID
                    JOIN M_Locator ml ON ml.M_Locator_ID = s.M_Locator_ID
                    WHERE 
                        s.AD_Client_ID = 1000000
                        AND m.AD_Client_ID = 1000000
                        AND s.M_Locator_ID IN (1001135, 1000614, 1001128, 1001136)
                        AND (:product IS NULL OR UPPER(m.name) LIKE UPPER(:product) || '%')
                    GROUP BY 
                        m.M_Product_ID, m.name
                ),
                Stock_Principale AS (
                    SELECT 
                        ROUND(SUM(M_ATTRIBUTEINSTANCE.valuenumber * (m_storage.qtyonhand - m_storage.QTYRESERVED)), 2) AS stock_principale
                    FROM 
                        M_ATTRIBUTEINSTANCE
                    JOIN m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
                    WHERE 
                        M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                        AND m_storage.qtyonhand > 0
                        AND m_storage.M_Locator_ID IN (1001135, 1000614, 1001128, 1001136)
                        AND (:product IS NULL OR m_storage.M_Product_ID IN (SELECT M_Product_ID FROM M_Product WHERE UPPER(name) LIKE UPPER(:product) || '%'))
                )
                SELECT 
                    oq.midp,
                    oq.product_name,
                    oq.QTYONHAND AS "QTY DISPO",
                    COALESCE(fp.last_purchase_qty, 0) AS "DERNIER ACHAT",
                    fp.dateinvoiced AS "DATE",
                    sp.stock_principale AS "VALEUR"
                FROM 
                    On_Hand_Quantity oq
                LEFT JOIN 
                    Filtered_Latest_Purchase fp ON oq.midp = fp.M_Product_ID
                CROSS JOIN 
                    Stock_Principale sp
                ORDER BY 
                    oq.product_name
            """

            params = {
                'product': product or None
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()

            data = [
                {
                    "PRODUCT_ID": row[0],
                    "PRODUCT_NAME": row[1],
                    "QTY_DISPO": row[2],
                    "DERNIER_ACHAT": row[3],
                    "DATE": row[4],
                    "VALEUR": row[5]
                }
                for row in rows
            ]

            return data
    
    except Exception as e:
        logger.error(f"Error fetching historique rotation: {e}")
        return {"error": "An error occurred while fetching historique rotation."}

@app.route('/fetchHistoriqueRotation', methods=['GET'])
def fetch_historique():
    product = request.args.get('product')

    data = fetch_historique_rotation(product)
    return jsonify(data)




def rotation_par_mois(start_date, end_date, product):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
WITH date_range AS (
    SELECT 
        CASE 
            WHEN TO_DATE(:end_date, 'YYYY-MM-DD') - TO_DATE(:start_date, 'YYYY-MM-DD') < 30 THEN 
                TO_CHAR(TO_DATE(:start_date, 'YYYY-MM-DD') + LEVEL - 1, 'YYYY-MM-DD')
            ELSE 
                TO_CHAR(ADD_MONTHS(TO_DATE(:start_date, 'YYYY-MM-DD'), LEVEL - 1), 'YYYY-MM')
        END AS invoice_period
    FROM DUAL
    CONNECT BY 
        CASE 
            WHEN TO_DATE(:end_date, 'YYYY-MM-DD') - TO_DATE(:start_date, 'YYYY-MM-DD') < 30 THEN 
                TO_DATE(:start_date, 'YYYY-MM-DD') + LEVEL - 1
            ELSE 
                ADD_MONTHS(TO_DATE(:start_date, 'YYYY-MM-DD'), LEVEL - 1)
        END <= TO_DATE(:end_date, 'YYYY-MM-DD')
),

invoice_data AS (
    SELECT
        dr.invoice_period,
        NVL(SUM(ff.QTYENTERED), 0) AS qty_vendu
    FROM
        date_range dr
    LEFT JOIN
        xx_ca_fournisseur_facture ff
    ON
        CASE 
            WHEN TO_DATE(:end_date, 'YYYY-MM-DD') - TO_DATE(:start_date, 'YYYY-MM-DD') < 30 THEN 
                TO_CHAR(ff.DATEINVOICED, 'YYYY-MM-DD')
            ELSE 
                TO_CHAR(ff.DATEINVOICED, 'YYYY-MM')
        END = dr.invoice_period
        AND ff.DATEINVOICED BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
        AND ff.AD_ORG_ID = 1000000
        AND ff.product LIKE '%' || :product || '%'
    GROUP BY
        dr.invoice_period
),

purchase_data AS (
    SELECT
        dr.invoice_period,
        SUM(
            CASE 
                WHEN xf.C_DocType_ID = 1000646 THEN -1 * mi.QTYENTERED
                ELSE mi.QTYENTERED
            END
        ) AS qty_acheté
    FROM
        date_range dr
    LEFT JOIN
        M_InOut xf ON
            CASE 
                WHEN TO_DATE(:end_date, 'YYYY-MM-DD') - TO_DATE(:start_date, 'YYYY-MM-DD') < 30 THEN 
                    TO_CHAR(xf.MOVEMENTDATE, 'YYYY-MM-DD')
                ELSE 
                    TO_CHAR(xf.MOVEMENTDATE, 'YYYY-MM')
            END = dr.invoice_period
        AND xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
        AND xf.AD_Org_ID = 1000000
        AND xf.C_DocType_ID IN (1000013, 1000646)
        AND xf.M_Warehouse_ID != 1000521
    JOIN M_INOUTLINE mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
        AND mi.AD_Org_ID = 1000000
    JOIN C_InvoiceLine ci ON ci.M_INOUTLINE_ID = mi.M_INOUTLINE_ID
    JOIN M_PRODUCT m ON m.M_PRODUCT_ID = mi.M_PRODUCT_ID
        AND m.AD_Org_ID = 1000000
        AND m.name LIKE '%' || :product || '%'
    LEFT JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
    GROUP BY
        dr.invoice_period
),

sold_totals AS (
    SELECT
        'TOTAL' AS invoice_period,
        SUM(qty_vendu) AS qty_vendu,
        NULL AS qty_acheté,
        2 AS sort_order
    FROM
        invoice_data
),
sold_averages AS (
    SELECT
        'MOYENNE' AS invoice_period,
        TRUNC(AVG(qty_vendu)) AS qty_vendu,
        NULL AS qty_acheté,
        3 AS sort_order
    FROM
        invoice_data
),

purchased_totals AS (
    SELECT
        'TOTAL' AS invoice_period,
        NULL AS qty_vendu,
        SUM(qty_acheté) AS qty_acheté,
        2 AS sort_order
    FROM
        purchase_data
),
purchased_averages AS (
    SELECT
        'MOYENNE' AS invoice_period,
        NULL AS qty_vendu,
        TRUNC(AVG(qty_acheté)) AS qty_acheté,
        3 AS sort_order
    FROM
        purchase_data
),

combined_data AS (
    SELECT
        COALESCE(id.invoice_period, pd.invoice_period) AS invoice_period,
        id.qty_vendu,
        pd.qty_acheté,
        1 AS sort_order
    FROM
        invoice_data id
    FULL OUTER JOIN
        purchase_data pd
    ON
        id.invoice_period = pd.invoice_period
),

final_results AS (
    SELECT
        invoice_period,
        COALESCE(qty_vendu, 0) AS qty_vendu,
        COALESCE(qty_acheté, 0) AS qty_acheté,
        sort_order
    FROM
        combined_data
    UNION ALL
    SELECT
        invoice_period,
        qty_vendu,
        qty_acheté,
        sort_order
    FROM
        sold_totals
    UNION ALL
    SELECT
        invoice_period,
        qty_vendu,
        qty_acheté,
        sort_order
    FROM
        sold_averages
    UNION ALL
    SELECT
        invoice_period,
        qty_vendu,
        qty_acheté,
        sort_order
    FROM
        purchased_totals
    UNION ALL
    SELECT
        invoice_period,
        qty_vendu,
        qty_acheté,
        sort_order
    FROM
        purchased_averages
)

SELECT
    invoice_period AS period,
    SUM(qty_vendu) AS qty_vendu,
    SUM(qty_acheté) AS qty_acheté
FROM
    final_results
GROUP BY
    invoice_period,
    sort_order
ORDER BY
    sort_order, invoice_period
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
                'product': product or '%'
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()

            data = [
                {
                    "PERIOD": row[0],
                    "QTY_VENDU": row[1],
                    "QTY_ACHETÉ": row[2]
                }
                for row in rows
            ]

            return data

    except Exception as e:
        logger.error(f"Error fetching rotation par mois: {e}")
        return {"error": "An error occurred while fetching rotation data."}

@app.route('/rotationParMois', methods=['GET'])
def fetch_rotation():
    product = request.args.get('product')  # Ensure product is received first
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # Debugging
    logger.info(f"Received parameters: product={product}, start_date={start_date}, end_date={end_date}")

    data = rotation_par_mois(start_date, end_date, product)
    return jsonify(data)




@app.route('/download-rotation-par-mois-excel', methods=['GET'])
def download_rotation_par_mois_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    product = request.args.get('product', 'All_Products')  # Default if no product is provided

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data
    data = rotation_par_mois(start_date, end_date, product)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename with product name, date range, and download date & time
    download_datetime = datetime.now().strftime("%d-%m-%Y_%H-%M")  # Day-Month-Year_Hour-Minute
    sanitized_product = product.replace(" ", "_").replace("/", "-")  # Replace spaces & slashes
    filename = f"Rotation_{sanitized_product}_{start_date}_to_{end_date}_{download_datetime}.xlsx"

    return generate_excel_rotation(data, filename)


def generate_excel_rotation(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Convert data to DataFrame
    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Rotation Par Mois"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="RotationTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send the file to the client
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")




def histogram(start_date, end_date, product):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
WITH 
date_range AS (
  SELECT 
    CASE 
      WHEN TO_DATE(:end_date, 'YYYY-MM-DD') - TO_DATE(:start_date, 'YYYY-MM-DD') < 30 THEN 
        TO_CHAR(TO_DATE(:start_date, 'YYYY-MM-DD') + LEVEL - 1, 'YYYY-MM-DD')
      ELSE 
        TO_CHAR(ADD_MONTHS(TO_DATE(:start_date, 'YYYY-MM-DD'), LEVEL - 1), 'YYYY-MM')
    END AS invoice_period
  FROM DUAL
  CONNECT BY 
    CASE 
      WHEN TO_DATE(:end_date, 'YYYY-MM-DD') - TO_DATE(:start_date, 'YYYY-MM-DD') < 30 THEN 
        TO_DATE(:start_date, 'YYYY-MM-DD') + LEVEL - 1
      ELSE 
        ADD_MONTHS(TO_DATE(:start_date, 'YYYY-MM-DD'), LEVEL - 1)
    END <= TO_DATE(:end_date, 'YYYY-MM-DD')
),

invoice_data AS (
  SELECT
    dr.invoice_period,
    NVL(SUM(ff.QTYENTERED), 0) AS qty_vendu
  FROM
    date_range dr
  LEFT JOIN
    xx_ca_fournisseur_facture ff
  ON
    CASE 
      WHEN TO_DATE(:end_date, 'YYYY-MM-DD') - TO_DATE(:start_date, 'YYYY-MM-DD') < 30 THEN 
        TO_CHAR(ff.DATEINVOICED, 'YYYY-MM-DD')
      ELSE 
        TO_CHAR(ff.DATEINVOICED, 'YYYY-MM')
    END = dr.invoice_period
    AND ff.DATEINVOICED BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
    AND ff.AD_ORG_ID = 1000000
    AND ff.product LIKE '%' || :product || '%'
  GROUP BY
    dr.invoice_period
),

purchase_data AS (
  SELECT
    dr.invoice_period,
    SUM(
      CASE 
        WHEN xf.C_DocType_ID = 1000646 THEN -1 * mi.QTYENTERED
        ELSE mi.QTYENTERED
      END
    ) AS qty_acheté
  FROM
    date_range dr
  LEFT JOIN
    M_InOut xf ON
      CASE 
        WHEN TO_DATE(:end_date, 'YYYY-MM-DD') - TO_DATE(:start_date, 'YYYY-MM-DD') < 30 THEN 
          TO_CHAR(xf.MOVEMENTDATE, 'YYYY-MM-DD')
        ELSE 
          TO_CHAR(xf.MOVEMENTDATE, 'YYYY-MM')
      END = dr.invoice_period
    AND xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
    AND xf.AD_Org_ID = 1000000
    AND xf.C_DocType_ID IN (1000013, 1000646)
    AND xf.M_Warehouse_ID != 1000521
  JOIN M_INOUTLINE mi ON mi.M_INOUT_ID = xf.M_INOUT_ID
    AND mi.AD_Org_ID = 1000000
  JOIN C_InvoiceLine ci ON ci.M_INOUTLINE_ID = mi.M_INOUTLINE_ID
  JOIN M_PRODUCT m ON m.M_PRODUCT_ID = mi.M_PRODUCT_ID
    AND m.AD_Org_ID = 1000000
    AND m.name LIKE '%' || :product || '%'
  LEFT JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.C_BPARTNER_ID
  GROUP BY
    dr.invoice_period
),

combined_data AS (
  SELECT
    COALESCE(id.invoice_period, pd.invoice_period) AS invoice_period,
    id.qty_vendu,
    pd.qty_acheté,
    1 AS sort_order
  FROM
    invoice_data id
  FULL OUTER JOIN
    purchase_data pd
  ON
    id.invoice_period = pd.invoice_period
)

SELECT
  invoice_period AS period,
  COALESCE(SUM(qty_acheté), 0) AS qty_acheté,
  COALESCE(SUM(qty_vendu), 0) AS qty_vendu
FROM
  combined_data
GROUP BY
  invoice_period,
  sort_order
ORDER BY
  invoice_period
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
                'product': product or '%'
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()

            data = [
                {
                    "PERIOD": row[0],
                    "QTY_ACHETÉ": row[1],
                    "QTY_VENDU": row[2]
                }
                for row in rows
            ]

            return data

    except Exception as e:
        logger.error(f"Error fetching histogram data: {e}")
        return {"error": "An error occurred while fetching histogram data."}

@app.route('/histogram', methods=['GET'])
def histogram_route():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    product = request.args.get('product', '')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date"}), 400

    data = histogram(start_date, end_date, product)
    return jsonify(data)



def journal_vente(start_date, end_date, client):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
            SELECT ci.DocumentNo,
                   ci.DateInvoiced,
                   cb.name as client,
                   CASE WHEN ci.isreturntrx = 'Y'
                        THEN -(ci.GRANDTOTAL - getinvoicetaxamt(ci.c_invoice_id)) 
                        ELSE ci.GRANDTOTAL - getinvoicetaxamt(ci.c_invoice_id) 
                   END as TotalHT,
                   CASE WHEN ci.isreturntrx = 'Y'           
                        THEN -getinvoicetaxamt(ci.c_invoice_id) 
                        ELSE getinvoicetaxamt(ci.c_invoice_id) 
                   END as TotalTVA,
                   CASE WHEN ci.isreturntrx = 'Y'
                        THEN -ci.chargeamt 
                        ELSE ci.chargeamt 
                   END as TotalDT,
                   CASE WHEN ci.isreturntrx = 'Y'
                        THEN -ci.GRANDTOTAL 
                        ELSE ci.GRANDTOTAL 
                   END as TotalTTC,
                   CASE WHEN ci.isreturntrx = 'Y'
                        THEN -ci.GRANDTOTAL - ci.chargeamt 
                        ELSE ci.GRANDTOTAL + ci.chargeamt 
                   END as NETAPAYER,   
                   sr.name as region,
                   adc.name as Entreprise,
                   COALESCE((
                        SELECT sum(tax.taxbaseamt)
                        FROM C_InvoiceTax tax 
                        WHERE tax.C_Invoice_ID = ci.C_Invoice_ID 
                        AND tax.taxamt = 0
                    ), 0) as Montant_Exonere
            FROM C_INVOICE ci
            INNER JOIN AD_CLIENT adc ON (adc.ad_client_id = ci.ad_client_id)
            INNER JOIN AD_ORG ado ON (ado.ad_org_id = ci.ad_org_id)
            INNER JOIN C_BPARTNER cb ON (cb.c_bpartner_id = ci.c_bpartner_id) 
            INNER JOIN C_BPARTNER_Location bpl ON (bpl.C_BPARTNER_Location_id = ci.C_BPARTNER_Location_id) 
            LEFT OUTER JOIN c_salesregion sr ON (bpl.C_SalesRegion_ID = sr.C_SalesRegion_ID)
            WHERE ci.dateInvoiced BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                     AND TO_DATE(:end_date, 'YYYY-MM-DD')
            AND ci.ad_Org_id = 1000012
            AND ci.ISSOTRX = 'Y' 
            AND ci.docstatus IN ('CO', 'CL')
            AND ci.c_doctype_id IN (SELECT c_doctype_id FROM c_doctype WHERE Xx_Excluejournalvente = 'N')
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
            }

            if client:
                query += " AND cb.name LIKE :client"
                params['client'] = f"%{client}%"

            query += " ORDER BY ci.DateInvoiced"

            cursor.execute(query, params)
            rows = cursor.fetchall()

            data = [
                {
                    "DocumentNo": row[0],
                    "DateInvoiced": row[1],
                    "Client": row[2],
                    "TotalHT": row[3],
                    "TotalTVA": row[4],
                    "TotalDT": row[5],
                    "TotalTTC": row[6],
                    "NETAPAYER": row[7],
                    "Region": row[8],
                    "Entreprise": row[9],
                    "Montant_Exonere": row[10]
                }
                for row in rows
            ]

            return data

    except Exception as e:
        logger.error(f"Error fetching journal de vente: {e}")
        return {"error": "An error occurred while fetching sales journal data."}



@app.route('/journalVente', methods=['GET'])
def fetch_journal():
    client = request.args.get('client')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    data = journal_vente(start_date, end_date, client)
    return jsonify(data)

def total_journal(start_date, end_date, client):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
            SELECT 
                SUM(CASE 
                    WHEN ci.isreturntrx = 'Y' THEN -(ci.GRANDTOTAL - getinvoicetaxamt(ci.c_invoice_id)) 
                    ELSE ci.GRANDTOTAL - getinvoicetaxamt(ci.c_invoice_id) 
                END) AS TotalHT,
                SUM(CASE 
                    WHEN ci.isreturntrx = 'Y' THEN -getinvoicetaxamt(ci.c_invoice_id) 
                    ELSE getinvoicetaxamt(ci.c_invoice_id) 
                END) AS TotalTVA,
                SUM(CASE 
                    WHEN ci.isreturntrx = 'Y' THEN -ci.chargeamt 
                    ELSE ci.chargeamt 
                END) AS TotalDT,
                SUM(CASE 
                    WHEN ci.isreturntrx = 'Y' THEN -ci.GRANDTOTAL  
                    ELSE ci.GRANDTOTAL 
                END) AS TotalTTC,
                SUM(CASE 
                    WHEN ci.isreturntrx = 'Y' THEN -ci.GRANDTOTAL - ci.chargeamt
                    ELSE ci.GRANDTOTAL + ci.chargeamt
                END) AS NETAPAYER
            FROM C_INVOICE ci
            INNER JOIN AD_CLIENT adc ON adc.ad_client_id = ci.ad_client_id
            INNER JOIN AD_ORG ado ON ado.ad_org_id = ci.ad_org_id
            INNER JOIN C_BPARTNER cb ON cb.c_bpartner_id = ci.c_bpartner_id 
            INNER JOIN C_BPARTNER_Location bpl ON bpl.C_BPARTNER_Location_id = ci.C_BPARTNER_Location_id 
            LEFT OUTER JOIN C_SALESREGION sr ON bpl.C_SalesRegion_ID = sr.C_SalesRegion_ID
            WHERE ci.dateInvoiced BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                     AND TO_DATE(:end_date, 'YYYY-MM-DD')
            AND ci.ad_Org_id = 1000012
            AND ci.ISSOTRX = 'Y' 
            AND ci.docstatus IN ('CO', 'CL')
            AND ci.c_doctype_id IN (SELECT c_doctype_id FROM c_doctype WHERE Xx_Excluejournalvente = 'N')
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
            }

            if client:
                query += " AND cb.name LIKE :client"
                params['client'] = f"%{client}%"

            cursor.execute(query, params)
            row = cursor.fetchone()

            # Assuming there's only one row of results, we return the aggregated totals
            total_data = {
                "TotalHT": row[0],
                "TotalTVA": row[1],
                "TotalDT": row[2],
                "TotalTTC": row[3],
                "NETAPAYER": row[4]
            }

            return total_data

    except Exception as e:
        logger.error(f"Error fetching total journal data: {e}")
        return {"error": "An error occurred while fetching total journal data."}

@app.route('/totalJournal', methods=['GET'])
def fetch_total_journal():
    client = request.args.get('client')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    total_data = total_journal(start_date, end_date, client)
    return jsonify(total_data)



@app.route('/download-journal-vente-excel', methods=['GET'])
def download_journal_vente_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    client = request.args.get('client', 'All_Clients')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data
    data = journal_vente(start_date, end_date, client)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    download_datetime = datetime.now().strftime("%d-%m-%Y_%H-%M")
    sanitized_client = client.replace(" ", "_").replace("/", "-")
    filename = f"JournalVente_{sanitized_client}_{start_date}_to_{end_date}_{download_datetime}.xlsx"

    return generate_excel_journal(data, filename)


def generate_excel_journal(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Journal Vente"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="JournalVenteTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send the file to the client
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")







# Assuming DB_POOL is a valid database connection pool


def fetch_etat_fournisseur():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            # Static SQL query without dynamic filtering
            query = """
                SELECT 
                    ROUND(SUM(total_echu), 2) AS "TOTAL ECHU",
                    ROUND(SUM(total_dette), 2) AS "TOTAL DETTE",
                    ROUND(SUM(STOCK), 2) AS "TOTAL STOCK"
                FROM (
                    -- TOTAL ECHU Calculation
                    SELECT  
                        SUM(COALESCE(invoiceOpen(inv.C_Invoice_ID, 0), 0)) AS total_echu,
                        0 AS total_dette,
                        0 AS STOCK
                    FROM C_Invoice inv
                    INNER JOIN c_bpartner bp ON bp.c_bpartner_id = inv.c_bpartner_id
                    LEFT OUTER JOIN C_BPARTNER_LOCATION bpl ON bp.c_bpartner_id = bpl.c_bpartner_id
                    LEFT OUTER JOIN C_SalesRegion SR ON SR.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    INNER JOIN C_PAYMENTTERM pt ON inv.C_PaymentTerm_ID = pt.C_PaymentTerm_ID
                    WHERE inv.docstatus IN ('CO', 'CL')
                      AND inv.ad_client_id = 1000000
                      AND inv.ISSOTRX = 'N'
                      AND bp.isactive = 'Y'
                      AND bp.isactive = 'Y'

                      AND bp.isvendor = 'Y'
                      AND COALESCE(invoiceOpen(inv.C_Invoice_ID, 0), 0) <> 0
                      AND inv.AD_Org_ID = 1000000
                      AND inv.AD_Client_ID = 1000000
                      AND (inv.dateinvoiced + pt.netdays) BETWEEN TO_DATE('01/01/2000', 'DD/MM/YYYY') AND SYSDATE
                      AND bp.name NOT LIKE 'solde initial%'

                    UNION ALL

                    -- TOTAL DETTE Calculation
                    SELECT  
                        0 AS total_echu,
                        SUM(cs.grandtotal - cs.verse_fact) AS total_dette,
                        0 AS STOCK
                    FROM (
                        SELECT cs.grandtotal, cs.verse_fact, cs.C_Invoice_ID,
                               ROW_NUMBER() OVER (PARTITION BY cs.name, cs.C_Invoice_ID ORDER BY cs.dateinvoiced DESC) AS rn
                        FROM xx_vendor_status cs
                        WHERE cs.AD_Client_ID = 1000000
                          AND cs.AD_Org_ID = 1000000
                          AND cs.dateinvoiced BETWEEN TO_DATE('01/01/2000', 'DD/MM/YYYY') AND TO_DATE('30/12/3000', 'DD/MM/YYYY')
                          AND cs.name NOT LIKE 'solde initial%'
                    ) cs
                    WHERE cs.rn = 1

                    UNION ALL

                    -- TOTAL STOCK Calculation
                    SELECT 
                        0 AS total_echu,
                        0 AS total_dette,
                        ROUND(SUM(asi.valuenumber * (ms.qtyonhand - ms.QTYRESERVED)), 2) AS STOCK
                    FROM 
                        M_ATTRIBUTEINSTANCE asi
                    JOIN 
                        m_storage ms ON ms.M_ATTRIBUTEsetINSTANCE_id = asi.M_ATTRIBUTEsetINSTANCE_id
                    LEFT JOIN 
                        C_bpartner bp ON bp.c_bpartner_id = ValueNUMBER_of_ASI('Fournisseur', asi.m_attributesetinstance_id)
                    WHERE 
                        asi.M_Attribute_ID = 1000504
                        AND ms.qtyonhand > 0
                        AND ms.m_locator_id IN (1000614, 1001128, 1001135, 1001136)
                        AND bp.name NOT LIKE 'solde initial%'
                ) temp
                HAVING ROUND(SUM(total_echu), 2) <> 0 OR ROUND(SUM(total_dette), 2) <> 0
                ORDER BY "TOTAL DETTE" DESC, "TOTAL ECHU" DESC, "TOTAL STOCK" DESC
            """

            # Execute the query
            cursor.execute(query)
            row = cursor.fetchone()

            # Return the results
            return {
                "TOTAL ECHU": row[0] if row else 0,
                "TOTAL DETTE": row[1] if row else 0,
                "TOTAL STOCK": row[2] if row else 0
            }

    except Exception as e:
        logging.error(f"Error fetching etat fournisseur: {e}")
        return {"error": "An error occurred while fetching etat fournisseur."}



@app.route('/etat_fournisseur', methods=['GET'])
def get_etat_fournisseur():
    result = fetch_etat_fournisseur()
    return jsonify(result)

 

def fetch_fournisseur_dette(fournisseur=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            fournisseur_filter = "AND bp.name LIKE :fournisseur || '%'" if fournisseur else ""
            fournisseur_filter_cs = "AND cs.name LIKE :fournisseur || '%'" if fournisseur else ""

            query = f"""
                SELECT 
                    fournisseur,
                    ROUND(SUM(total_echu), 2) AS "TOTAL ECHU",
                    ROUND(SUM(total_dette), 2) AS "TOTAL DETTE",
                    ROUND(SUM(STOCK), 2) AS "TOTAL STOCK"
                FROM (
                    SELECT  
                        bp.name AS fournisseur,
                        SUM(COALESCE(invoiceOpen(inv.C_Invoice_ID, 0), 0)) AS total_echu,
                        0 AS total_dette,
                        0 AS STOCK
                    FROM C_Invoice inv
                    INNER JOIN c_bpartner bp ON bp.c_bpartner_id = inv.c_bpartner_id
                    LEFT OUTER JOIN C_BPARTNER_LOCATION bpl ON bp.c_bpartner_id = bpl.c_bpartner_id
                    LEFT OUTER JOIN C_SalesRegion SR ON SR.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    INNER JOIN C_PAYMENTTERM pt ON inv.C_PaymentTerm_ID = pt.C_PaymentTerm_ID
                    WHERE inv.docstatus IN ('CO', 'CL')
                      AND inv.ad_client_id = 1000000
                      AND inv.ISSOTRX = 'N'
                      AND bp.isactive = 'Y'
                      AND bp.isvendor = 'Y'
                      AND COALESCE(invoiceOpen(inv.C_Invoice_ID, 0), 0) <> 0
                      AND inv.AD_Org_ID = 1000000
                      AND (inv.dateinvoiced + pt.netdays) BETWEEN TO_DATE('01/01/2020', 'DD/MM/YYYY') AND SYSDATE
                      AND bp.name NOT LIKE 'solde initial%'
                      {fournisseur_filter}
                    GROUP BY bp.name

                    UNION ALL

                    SELECT  
                        cs.name AS fournisseur,
                        0 AS total_echu,
                        SUM(cs.grandtotal - cs.verse_fact) AS total_dette,
                        0 AS STOCK
                    FROM (
                        SELECT cs.name, cs.grandtotal, cs.verse_fact, cs.C_Invoice_ID,
                               ROW_NUMBER() OVER (PARTITION BY cs.name, cs.C_Invoice_ID ORDER BY cs.dateinvoiced DESC) AS rn
                        FROM xx_vendor_status cs
                        WHERE cs.AD_Client_ID = 1000000
                          AND cs.AD_Org_ID = 1000000
                          AND cs.dateinvoiced BETWEEN TO_DATE('01/01/2015', 'DD/MM/YYYY') AND TO_DATE('30/12/3000', 'DD/MM/YYYY')
                          AND cs.name NOT LIKE 'solde initial%'
                          {fournisseur_filter_cs}
                    ) cs
                    WHERE cs.rn = 1
                    GROUP BY cs.name

                    UNION ALL

                    SELECT 
                        bp.name AS fournisseur,
                        0 AS total_echu,
                        0 AS total_dette,
                        ROUND(SUM(asi.valuenumber * (ms.qtyonhand - ms.QTYRESERVED)), 2) AS STOCK
                    FROM 
                        M_ATTRIBUTEINSTANCE asi
                    JOIN 
                        m_storage ms ON ms.M_ATTRIBUTEsetINSTANCE_id = asi.M_ATTRIBUTEsetINSTANCE_id
                    LEFT JOIN 
                        C_bpartner bp ON bp.c_bpartner_id = ValueNUMBER_of_ASI('Fournisseur', asi.m_attributesetinstance_id)
                    WHERE 
                        asi.M_Attribute_ID = 1000504
                        AND ms.qtyonhand > 0
                        AND ms.m_locator_id IN (1000614, 1001128, 1001135, 1001136)
                        AND bp.name NOT LIKE 'solde initial%'
                        {fournisseur_filter}
                    GROUP BY bp.name
                ) temp
                GROUP BY fournisseur
                HAVING ROUND(SUM(total_echu), 2) <> 0 OR ROUND(SUM(total_dette), 2) <> 0
                ORDER BY "TOTAL DETTE" DESC, "TOTAL ECHU" DESC, "TOTAL STOCK" DESC
            """

            params = {'fournisseur': fournisseur} if fournisseur else {}
            cursor.execute(query, params)
            rows = cursor.fetchall()

            result = []
            for row in rows:
                result.append({
                    "FOURNISSEUR": row[0],
                    "TOTAL ECHU": row[1],
                    "TOTAL DETTE": row[2],
                    "TOTAL STOCK": row[3]
                })

            return result

    except Exception as e:
        logging.error(f"Error fetching fournisseur dette: {e}")
        return {"error": "An error occurred while fetching fournisseur dette."}


@app.route('/fetchFournisseurDette', methods=['GET'])
def fetch_fournisseur_dette_api():
    fournisseur = request.args.get('fournisseur')  # Get the fournisseur parameter from the request

    data = fetch_fournisseur_dette(fournisseur)  # Call the function to get the data
    return jsonify(data)  # Return the result as JSON



def fetch_order_confirmed():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            
            query = """
                SELECT * FROM (
                    SELECT 
                        CAST(org.name AS VARCHAR2(300)) AS organisation,
                        CAST(co.documentno AS VARCHAR2(50)) AS ndocument,
                        CAST(cb.name AS VARCHAR2(300)) AS tier,
                        co.dateordered AS datecommande,
                        CAST(us.name AS VARCHAR2(100)) AS vendeur,
                        ROUND(((co.totallines / (SELECT SUM(mat.valuenumber * li.qtyentered) 
                             FROM c_orderline li 
                             INNER JOIN m_attributeinstance mat ON mat.m_attributesetinstance_id = li.m_attributesetinstance_id
                             WHERE mat.m_attribute_id = 1000504 
                               AND li.c_order_id = co.c_order_id 
                               AND li.qtyentered > 0 
                             GROUP BY li.c_order_id)) - 1) * 100, 2) AS marge,
                        ROUND(co.totallines, 2) AS montant,
                        1 AS sort_order
                    FROM 
                        c_order co
                    INNER JOIN ad_org org ON co.ad_org_id = org.ad_org_id
                    INNER JOIN c_bpartner cb ON co.c_bpartner_id = cb.c_bpartner_id
                    INNER JOIN ad_user us ON co.salesrep_id = us.ad_user_id
                    INNER JOIN c_orderconfirm conf ON co.c_order_id = conf.c_order_id
                    WHERE 
                        conf.c_orderconfirm_id IS NOT NULL 
                        AND co.docstatus = 'IP' 
                        AND co.docaction IN ('CO', 'PR') 
                        AND co.c_doctypetarget_id = 1000539 
                        AND co.ad_org_id = 1000000
                    
                    UNION ALL
                    
                    SELECT 
                        CAST('Total' AS VARCHAR2(300)) AS organisation,
                        CAST(NULL AS VARCHAR2(50)) AS ndocument,
                        CAST(NULL AS VARCHAR2(300)) AS tier,
                        NULL AS datecommande,
                        CAST(NULL AS VARCHAR2(100)) AS vendeur,
                        ROUND(AVG(ROUND(((co.totallines / (SELECT SUM(mat.valuenumber * li.qtyentered) 
                             FROM c_orderline li 
                             INNER JOIN m_attributeinstance mat ON mat.m_attributesetinstance_id = li.m_attributesetinstance_id
                             WHERE mat.m_attribute_id = 1000504 
                               AND li.c_order_id = co.c_order_id 
                               AND li.qtyentered > 0 
                             GROUP BY li.c_order_id)) - 1) * 100, 2)), 2) AS marge,
                        ROUND(SUM(co.totallines), 2) AS montant,
                        0 AS sort_order
                    FROM 
                        c_order co
                    INNER JOIN ad_org org ON co.ad_org_id = org.ad_org_id
                    INNER JOIN c_bpartner cb ON co.c_bpartner_id = cb.c_bpartner_id
                    INNER JOIN ad_user us ON co.salesrep_id = us.ad_user_id
                    INNER JOIN c_orderconfirm conf ON co.c_order_id = conf.c_order_id
                    WHERE 
                        conf.c_orderconfirm_id IS NOT NULL 
                        AND co.docstatus = 'IP' 
                        AND co.docaction IN ('CO', 'PR') 
                        AND co.c_doctypetarget_id = 1000539 
                        AND co.ad_org_id = 1000000
                )
                ORDER BY sort_order, montant DESC
            """
            
            cursor.execute(query)
            rows = cursor.fetchall()
            
            columns = [col[0] for col in cursor.description]
            result = [dict(zip(columns, row)) for row in rows]
            
            return result

    except Exception as e:
        logging.error(f"Error fetching order confirmed: {e}")
        return {"error": "An error occurred while fetching order confirmed."}

@app.route('/order_confirmed', methods=['GET'])
def get_order_confirmed():
    result = fetch_order_confirmed()
    return jsonify(result)




# Updated function that returns only the total price
def fetch_total_stock_price():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
            SELECT 
                SUM(M_ATTRIBUTEINSTANCE.valuenumber * m_storage.qtyonhand) AS prix_total
            FROM 
                M_ATTRIBUTEINSTANCE
            JOIN 
                m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
            JOIN 
                M_PRODUCT m ON m.M_PRODUCT_ID = m_storage.M_PRODUCT_ID
            JOIN 
                M_Locator ml ON ml.M_Locator_ID = m_storage.M_Locator_ID
            INNER JOIN 
                m_attributeinstance mati ON m_storage.M_ATTRIBUTEsetINSTANCE_id = mati.M_ATTRIBUTEsetINSTANCE_id
            WHERE 
                M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                AND m_storage.qtyonhand > 0
                AND mati.m_attribute_id = 1000508
                AND m_storage.AD_Client_ID = 1000000
                AND m_storage.M_Locator_ID IN (
                    SELECT M_Locator_ID 
                    FROM M_Locator 
                    WHERE M_Warehouse_ID IN (
                        SELECT M_Warehouse_ID 
                        FROM M_Warehouse 
                        WHERE VALUE IN (
                            'HANGAR', '1-Dépôt Principal', '8-Dépot réserve', '88-Dépot Hangar réserve'
                        )
                    )
                )
            """

            cursor.execute(query)
            result = cursor.fetchone()

            return {"prix_total": result[0] if result else 0}

    except Exception as e:
        logger.error(f"Database error: {e}")
        return {"error": "An error occurred while fetching total price."}



@app.route('/total-stock', methods=['GET'])
def get_total_stock_price():
    try:
        total_data = fetch_total_stock_price()
        return jsonify(total_data)
    except Exception as e:
        logger.error(f"Error fetching total stock price: {e}")
        return jsonify({"error": "Failed to fetch total price"}), 500

def fetch_total_stock_by_location():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                WITH stock_principale_data AS (
                    SELECT 
                        ROUND(SUM(M_ATTRIBUTEINSTANCE.valuenumber * m_storage.qtyonhand), 2) AS stock_principale
                    FROM 
                        M_ATTRIBUTEINSTANCE
                    JOIN 
                        m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
                    WHERE 
                        M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                        AND m_storage.qtyonhand > 0
                        AND m_storage.m_locator_id = 1000614
                ),
                hangar_data AS (
                    SELECT 
                        ROUND(SUM(M_ATTRIBUTEINSTANCE.valuenumber * m_storage.qtyonhand), 2) AS hangar
                    FROM 
                        M_ATTRIBUTEINSTANCE
                    JOIN 
                        m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
                    WHERE 
                        M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                        AND m_storage.qtyonhand > 0
                        AND m_storage.m_locator_id = 1001135
                ),
                hangarresrev_data AS (
                    SELECT 
                        ROUND(SUM(M_ATTRIBUTEINSTANCE.valuenumber * m_storage.qtyonhand), 2) AS hangarresrev
                    FROM 
                        M_ATTRIBUTEINSTANCE
                    JOIN 
                        m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
                    WHERE 
                        M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                        AND m_storage.qtyonhand > 0
                        AND m_storage.m_locator_id = 1001136
                ),
                depot_reserver_data AS (
                    SELECT 
                        ROUND(SUM(M_ATTRIBUTEINSTANCE.valuenumber * m_storage.qtyonhand), 2) AS depot_reserver
                    FROM 
                        M_ATTRIBUTEINSTANCE
                    JOIN 
                        m_storage ON m_storage.M_ATTRIBUTEsetINSTANCE_id = M_ATTRIBUTEINSTANCE.M_ATTRIBUTEsetINSTANCE_id
                    WHERE 
                        M_ATTRIBUTEINSTANCE.M_Attribute_ID = 1000504
                        AND m_storage.qtyonhand > 0
                        AND m_storage.m_locator_id = 1001128
                )

                SELECT 
                    ROUND(SUM(sp.stock_principale), 2) AS STOCK_principale,
                    ROUND(SUM(h.hangar), 2) AS hangar,
                    ROUND(SUM(hr.hangarresrev), 2) AS hangarréserve,
                    ROUND(SUM(dr.depot_reserver), 2) AS depot_reserver,
                    ROUND(
                        NVL(SUM(sp.stock_principale), 0) + 
                        NVL(SUM(h.hangar), 0) + 
                        NVL(SUM(hr.hangarresrev), 0) + 
                        NVL(SUM(dr.depot_reserver), 0), 
                        2
                    ) AS total_stock
                FROM 
                    stock_principale_data sp,
                    hangar_data h,
                    hangarresrev_data hr,
                    depot_reserver_data dr
            """

            cursor.execute(query)
            result = cursor.fetchone()

            return {
                "STOCK_principale": result[0] or 0,
                "hangar": result[1] or 0,
                "hangarréserve": result[2] or 0,
                "depot_reserver": result[3] or 0,
                "total_stock": result[4] or 0
            }

    except Exception as e:
        logger.error(f"Database error: {e}")
        return {"error": "An error occurred while fetching stock data."}



@app.route('/stock-summary', methods=['GET'])
def get_stock_summary():
    try:
        stock_data = fetch_total_stock_by_location()
        return jsonify(stock_data)
    except Exception as e:
        logger.error(f"Error fetching stock summary: {e}")
        return jsonify({"error": "Failed to fetch stock summary"}), 500





@app.route('/download-quota-excel', methods=['GET'])
def download_quota_excel():
    data = fetch_quota_product()  # Fetch all data
    produit_filter = request.args.get("produit", "").strip()

    # Apply filter if provided (already applied inside fetch_quota_product, but double check here)
    if produit_filter:
        data = [row for row in data if produit_filter.lower() in row["NAME"].lower()]

    return generate_excel_quota(data, produit_filter)

#----------------------------
def generate_excel_quota(data, produit_filter):
    if not data:
        return {"error": "No data available"}, 400

    df = pd.DataFrame(data)
    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered Quota"

    # Apply header styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(df.columns.tolist())
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Append data with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Generate filename dynamically based on filter
    today_date = datetime.now().strftime("%d-%m-%Y")
    filter_text = f"produit-{produit_filter}" if produit_filter else ""
    filename = f"Quota_{filter_text}_{today_date}.xlsx" if filter_text else f"Quota_{today_date}.xlsx"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Fetch product quota data from Oracle DB
def fetch_quota_product():
    try:
        produit = request.args.get('produit', '')  # Get 'produit' param from URL, default to empty string

        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            # Base SQL with optional filtering
            query = """
                SELECT m.name, SUM(s.QTYONHAND) AS QTY , sum(s.QTYONHAND * at.valuenumber) as Prix
                
                FROM m_product m

                JOIN m_storage s ON s.M_PRODUCT_ID = m.M_PRODUCT_ID
                INNER JOIN m_attributeinstance at
                ON at.m_attributesetinstance_id = s.m_attributesetinstance_id



                WHERE m.XX_SalesContext_ID = 1000100
                    AND m.AD_Client_ID = 1000000
                    AND s.QTYONHAND > 0
                    AND s.M_Locator_ID = 1000614
                    and at.M_Attribute_ID=1000502
            """

            # If produit filter is provided, add it to the query
            if produit:
                query += " AND LOWER(m.name) LIKE :produit"

            query += """
                GROUP BY m.name
                ORDER BY m.name
            """

            # Bind parameter if filtering
            if produit:
                cursor.execute(query, {"produit": produit.lower() + '%'})
            else:
                cursor.execute(query)

            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            return data

    except Exception as e:
        logger.error(f"Error fetching product quota data: {e}")
        return {"error": "An error occurred while fetching product quota data."}





@app.route('/quota-product', methods=['GET'])
def quota_product():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_quota_product()
    return jsonify(data)



def fetch_quota_operator():
    try:
        produit = request.args.get('produit', '')

        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            # SQL query with optional filtering
            query = """
                SELECT a.name, SUM(q.quantity - q.QTYORDERED) AS qty
                FROM AD_USER a
                JOIN XX_Quota q ON q.AD_USER_id = a.AD_USER_id
                JOIN M_PRODUCT m ON m.M_PRODUCT_id = q.M_PRODUCT_ID
                WHERE q.M_Warehouse_ID = 1000000
            """

            # Apply filter if produit is provided
            if produit:
                query += " AND LOWER(m.name) LIKE :produit"

            query += """
                GROUP BY a.name
                ORDER BY qty DESC
            """

            # Execute with binding
            if produit:
                cursor.execute(query, {"produit": f"%{produit.lower()}%"})
            else:
                cursor.execute(query)

            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            return data

    except Exception as e:
        logger.error(f"Error fetching operator quota data: {e}")
        return {"error": "An error occurred while fetching operator quota data."}

@app.route('/quota-operator', methods=['GET'])
def quota_operator():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_quota_operator()
    return jsonify(data)




def fetch_credit_client():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT 
                    ROUND(SUM(SoldeFact + SoldeBL), 2) AS credit_client
                FROM (
                    SELECT 
                        bp.c_bpartner_id, 
                        (
                            SELECT COALESCE(SUM(invoiceOpen(inv.C_Invoice_ID, 0)), 0) 
                            FROM C_Invoice inv 
                            WHERE bp.c_bpartner_id = inv.c_bpartner_id 
                            AND paymentTermDueDays(inv.C_PAYMENTTERM_ID, inv.DATEINVOICED, TO_DATE('01/01/3000', 'DD/MM/YYYY')) >= 0
                            AND inv.docstatus IN ('CO', 'CL') 
                            AND inv.AD_ORGTRX_ID = inv.ad_org_id 
                            AND inv.ad_client_id = 1000000
                            AND inv.C_PaymentTerm_ID != 1000000
                        ) AS SoldeFact,
                        (
                            SELECT COALESCE(SUM(invoiceOpen(inv.C_Invoice_ID, 0)), 0) 
                            FROM C_Invoice inv 
                            WHERE bp.c_bpartner_id = inv.c_bpartner_id 
                            AND paymentTermDueDays(inv.C_PAYMENTTERM_ID, inv.DATEINVOICED, TO_DATE('01/01/3000', 'DD/MM/YYYY')) >= 0
                            AND inv.docstatus IN ('CO', 'CL') 
                            AND inv.AD_ORGTRX_ID <> inv.ad_org_id 
                            AND inv.ad_client_id = 1000000
                            AND inv.C_PaymentTerm_ID != 1000000
                        ) AS SoldeBL
                    FROM 
                        c_bpartner bp 
                        INNER JOIN C_BPartner_Location bpl ON bp.C_BPartner_id = bpl.C_BPartner_id
                        INNER JOIN ad_user u ON bp.salesrep_id = u.ad_user_id
                        LEFT OUTER JOIN AD_User u2 ON u2.AD_User_ID = bp.XX_TempSalesRep_ID
                        LEFT OUTER JOIN C_SalesRegion sr ON sr.C_SalesRegion_id = bpl.C_SalesRegion_id
                        LEFT OUTER JOIN C_City sr2 ON sr2.C_City_id = bpl.C_City_id
                    WHERE 
                        bp.iscustomer = 'Y' 
                        AND bp.C_BP_Group_ID IN (1000003, 1000926, 1001330)
                        AND bp.isactive = 'Y' 
                        AND sr.isactive = 'Y'
                        AND bp.C_PaymentTerm_ID != 1000000
                        AND bpl.c_salesregion_id IN (
                            101, 102, 1000032, 1001776, 1001777, 1001778, 1001779, 1001780, 1001781,
                            1001782, 1001783, 1001784, 1001785, 1001786, 1001787, 1001788, 1001789,
                            1001790, 1001791, 1001792, 1001793, 1001794, 1002076, 1002077, 1002078,
                            1002079, 1002080, 1002176, 1002177, 1002178, 1002179, 1002180, 1002181,
                            1002283, 1002285, 1002286, 1002287, 1002288
                        )
                ) subquery
            """

            cursor.execute(query)
            row = cursor.fetchone()
            credit_client = row[0] if row else 0.0
            return {"credit_client": credit_client}

    except Exception as e:
        logger.error(f"Error fetching credit client data: {e}")
        return {"error": "An error occurred while fetching credit client data."}




@app.route('/credit-client', methods=['GET'])
def credit_client():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_credit_client()
    return jsonify(data)

def fetch_caisse():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT 
                    ROUND(endingbalance, 2) AS caisse
                FROM 
                    C_BankStatement
                WHERE 
                    C_BankAccount_ID = 1000205
                    AND docstatus = 'CO'
                    AND AD_Client_ID = 1000000
                ORDER BY 
                    statementdate DESC
                FETCH FIRST 1 ROW ONLY
            """

            cursor.execute(query)
            row = cursor.fetchone()
            caisse = row[0] if row else 0.0
            return {"caisse": caisse}

    except Exception as e:
        logger.error(f"Error fetching caisse data: {e}")
        return {"error": "An error occurred while fetching caisse data."}


@app.route('/caisse', methods=['GET'])
def caisse():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_caisse()
    return jsonify(data)



def fourniseurdettfond():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
                SELECT 
                    ROUND(SUM(invopenamt), 2) AS credit_fournisseur
                FROM (
                    SELECT 
                        COALESCE(invoiceOpen(inv.C_Invoice_ID, 0), 0) AS invopenamt
                    FROM
                        C_Invoice inv
                        INNER JOIN c_bpartner bp ON bp.c_bpartner_id = inv.c_bpartner_id
                        LEFT OUTER JOIN C_BPARTNER_LOCATION bpl ON bp.c_bpartner_id = bpl.c_bpartner_id
                        LEFT OUTER JOIN C_SalesRegion SR ON SR.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                        INNER JOIN C_PAYMENTTERM pt ON inv.C_PaymentTerm_ID = pt.C_PaymentTerm_ID
                        LEFT OUTER JOIN ad_user usr ON usr.ad_user_id = inv.salesrep_id
                        LEFT OUTER JOIN ad_user usr2 ON usr2.ad_user_id = bp.salesrep_id
                        LEFT OUTER JOIN C_City ct ON ct.C_City_ID = bpl.C_City_ID
                    WHERE
                        inv.docstatus IN ('CO', 'CL')
                        AND inv.ad_client_id = 1000000
                        --AND inv.ISSOTRX = 'N'
                        AND bp.isactive = 'Y'
                        AND bp.isvendor = 'Y'
                )
            """

            cursor.execute(query)
            result = cursor.fetchone()
            return result[0] if result else 0.0

    except Exception as e:
        print(f"An error occurred: {e}")
        return None


@app.route('/fourniseurdettfond', methods=['GET'])
def get_fourniseurdettfond():
    result = fourniseurdettfond()
    return jsonify({"value": result})  # Now the frontend gets a key



from flask import Flask, request, jsonify
import os
import json
from datetime import datetime


# Path to store the previous values
DATA_FILE = os.path.join(os.path.dirname(__file__), 'kpi_data.json')

# Initialize data file if it doesn't exist
if not os.path.exists(DATA_FILE):
    with open(DATA_FILE, 'w') as f:
        json.dump({"fonds_propre": []}, f)

@app.route('/previous-fonds-propre', methods=['GET'])
def get_previous_fonds_propre():
    try:
        with open(DATA_FILE, 'r') as f:
            data = json.load(f)
        
        last_value = data['fonds_propre'][-1] if data['fonds_propre'] else None
        return jsonify(last_value)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/store-fonds-propre', methods=['POST'])
def store_fonds_propre():
    try:
        with open(DATA_FILE, 'r') as f:
            data = json.load(f)
        
        new_data = request.get_json()
        value = new_data.get('value')
        date = new_data.get('date', datetime.now().isoformat())
        
        # Keep only the last 12 months of data
        if len(data['fonds_propre']) >= 12:
            data['fonds_propre'].pop(0)
        
        # Ensure that the data is stored correctly
        data['fonds_propre'].append({"value": value, "date": date})
        
        with open(DATA_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# @app.route('/fonds-propre-chart-data')
# def get_fonds_propre_chart_data():
#     try:
#         # Load the data from JSON file
#         with open('kpi_data.json', 'r') as f:
#             data = json.load(f)
        
#         # Extract and format the data
#         chart_data = []
#         for entry in data.get('fonds_propre', []):
#             # Convert date string to datetime object for sorting
#             date = datetime.strptime(entry['date'], '%Y-%m-%dT%H:%M:%S.%fZ')
#             chart_data.append({
#                 'date': date.strftime('%Y-%m-%d %H:%M:%S'),  # Format for display
#                 'timestamp': date.timestamp(),  # For sorting
#                 'value': entry['value']
#             })
        
#         # Sort data by timestamp (oldest first)
#         chart_data.sort(key=lambda x: x['timestamp'])
        
#         # Prepare response data
#         response = {
#             'labels': [entry['date'] for entry in chart_data],
#             'values': [entry['value'] for entry in chart_data],
#             'min_date': chart_data[0]['date'] if chart_data else None,
#             'max_date': chart_data[-1]['date'] if chart_data else None
#         }
        
#         return jsonify(response)
    
#     except Exception as e:
#         return jsonify({'error': str(e)}), 500


@app.route('/fonds-propre-chart-data')
def get_fonds_propre_chart_data():
    try:
        with open(DATA_FILE, 'r') as f:
            data = json.load(f)
        
        sorted_data = sorted(data['fonds_propre'], key=lambda x: x['date'])
        labels = [entry['date'] for entry in sorted_data]
        values = [entry['value'] for entry in sorted_data]
        
        return jsonify({
            "labels": labels,
            "values": values,
            "min_date": labels[0] if labels else None,
            "max_date": labels[-1] if labels else None
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/dette-fournisseur-chart-data')
def get_dette_fournisseur_chart_data():
    try:
        with open(DETTE_FILE, 'r') as f:
            data = json.load(f)
        
        sorted_data = sorted(data['dette_fournisseur'], key=lambda x: x['date'])
        labels = [entry['date'] for entry in sorted_data]
        values = [entry['value'] for entry in sorted_data]
        
        return jsonify({
            "labels": labels,
            "values": values,
            "min_date": labels[0] if labels else None,
            "max_date": labels[-1] if labels else None
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Add similar endpoints for other metrics (credit-client, caisse, tresorerie)


# Update your existing DATA_FILE path or create a new one
DETTE_FILE = 'kpi_dette.json'

if not os.path.exists(DETTE_FILE):
    with open(DETTE_FILE, 'w') as f:
        json.dump({"dette_fournisseur": []}, f)

@app.route('/previous-dette-fournisseur', methods=['GET'])
def get_previous_dette_fournisseur():
    try:
        with open(DETTE_FILE, 'r') as f:
            data = json.load(f)
        
        last_value = data['dette_fournisseur'][-1] if data['dette_fournisseur'] else None
        return jsonify(last_value)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/store-dette-fournisseur', methods=['POST'])
def store_dette_fournisseur():
    try:
        with open(DETTE_FILE, 'r') as f:
            data = json.load(f)
        
        new_data = request.get_json()
        value = new_data.get('value')
        date = new_data.get('date', datetime.now().isoformat())
        
        # Keep only the last 12 months of data
        if len(data['dette_fournisseur']) >= 12:
            data['dette_fournisseur'].pop(0)
        
        data['dette_fournisseur'].append({"value": value, "date": date})
        
        with open(DETTE_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


STOCK_FILE = 'kpi_stock.json'

if not os.path.exists(STOCK_FILE):
    with open(STOCK_FILE, 'w') as f:
        json.dump({
            "total_stock": [],
            "stock_principale": [],
            "hangar": [],
            "hangar_reserve": [],
            "depot_reserver": []
        }, f)

@app.route('/previous-stock/<stock_type>', methods=['GET'])
def get_previous_stock(stock_type):
    try:
        with open(STOCK_FILE, 'r') as f:
            data = json.load(f)
        
        key = {
            'total': 'total_stock',
            'principale': 'stock_principale',
            'hangar': 'hangar',
            'hangar_reserve': 'hangar_reserve',
            'depot': 'depot_reserver'
        }.get(stock_type)
        
        if not key:
            return jsonify({"error": "Invalid stock type"}), 400
            
        last_value = data[key][-1] if data[key] else None
        return jsonify(last_value)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/store-stock', methods=['POST'])
def store_stock():
    try:
        with open(STOCK_FILE, 'r') as f:
            data = json.load(f)
        
        new_data = request.get_json()
        
        # Store all stock values
        stock_types = ['total_stock', 'stock_principale', 'hangar', 'hangar_reserve', 'depot_reserver']
        current_date = datetime.now().isoformat()
        
        for stype in stock_types:
            value = new_data.get(stype)
            if value is not None:
                # Keep only the last 12 months of data
                if len(data[stype]) >= 12:
                    data[stype].pop(0)
                
                data[stype].append({
                    "value": value,
                    "date": current_date
                })
        
        with open(STOCK_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
        
# Add this to your Flask app
CREDIT_FILE = 'kpi_credit.json'

if not os.path.exists(CREDIT_FILE):
    with open(CREDIT_FILE, 'w') as f:
        json.dump({"credit_client": []}, f)

@app.route('/previous-credit-client', methods=['GET'])
def get_previous_credit_client():
    try:
        with open(CREDIT_FILE, 'r') as f:
            data = json.load(f)
        
        last_value = data['credit_client'][-1] if data['credit_client'] else None
        return jsonify(last_value)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/store-credit-client', methods=['POST'])
def store_credit_client():
    try:
        with open(CREDIT_FILE, 'r') as f:
            data = json.load(f)
        
        new_data = request.get_json()
        value = new_data.get('value')
        date = new_data.get('date', datetime.now().isoformat())
        
        # Keep only the last 12 months of data
        if len(data['credit_client']) >= 12:
            data['credit_client'].pop(0)
        
        data['credit_client'].append({"value": value, "date": date})
        
        with open(CREDIT_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    

CAISSE_FILE = 'kpi_caisse.json'

if not os.path.exists(CAISSE_FILE):
    with open(CAISSE_FILE, 'w') as f:
        json.dump({"caisse": []}, f)

@app.route('/previous-caisse', methods=['GET'])
def get_previous_caisse():
    try:
        with open(CAISSE_FILE, 'r') as f:
            data = json.load(f)
        
        last_value = data['caisse'][-1] if data['caisse'] else None
        return jsonify(last_value)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/store-caisse', methods=['POST'])
def store_caisse():
    try:
        with open(CAISSE_FILE, 'r') as f:
            data = json.load(f)
        
        new_data = request.get_json()
        value = new_data.get('value')
        date = new_data.get('date', datetime.now().isoformat())
        
        # Keep only the last 12 months of data
        if len(data['caisse']) >= 12:
            data['caisse'].pop(0)
        
        data['caisse'].append({"value": value, "date": date})
        
        with open(CAISSE_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/test-caisse', methods=['GET'])
def test_caisse():
    try:
        # Return a fixed value of 10000 as JSON
        return jsonify({"caisse": 10000})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/test-fournisseur-dette', methods=['GET'])
def test_fournisseur_dette():
    try:
        # Return a fixed test value of 20000 as JSON
        return jsonify({"value": 20000})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


TRESORERIE_FILE = 'kpi_tresorerie.json'

if not os.path.exists(TRESORERIE_FILE):
    with open(TRESORERIE_FILE, 'w') as f:
        json.dump({"tresorerie": []}, f)

@app.route('/previous-tresorerie', methods=['GET'])
def get_previous_tresorerie():
    try:
        with open(TRESORERIE_FILE, 'r') as f:
            data = json.load(f)
        
        last_value = data['tresorerie'][-1] if data['tresorerie'] else None
        return jsonify(last_value)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/store-tresorerie', methods=['POST'])
def store_tresorerie():
    try:
        with open(TRESORERIE_FILE, 'r') as f:
            data = json.load(f)
        
        new_data = request.get_json()
        value = new_data.get('value')
        date = new_data.get('date', datetime.now().isoformat())
        
        # Keep only the last 12 months of data
        if len(data['tresorerie']) >= 12:
            data['tresorerie'].pop(0)
        
        data['tresorerie'].append({"value": value, "date": date})
        
        with open(TRESORERIE_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/credit-client-chart-data')
def get_credit_client_chart_data():
    try:
        with open(CREDIT_FILE, 'r') as f:
            data = json.load(f)

        sorted_data = sorted(data['credit_client'], key=lambda x: x['date'])
        labels = [entry['date'] for entry in sorted_data]
        values = [entry['value'] for entry in sorted_data]

        return jsonify({
            "labels": labels,
            "values": values,
            "min_date": labels[0] if labels else None,
            "max_date": labels[-1] if labels else None
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
@app.route('/caisse-chart-data')
def get_caisse_chart_data():
    try:
        with open('kpi_caisse.json', 'r') as f:
            data = json.load(f)

        sorted_data = sorted(data['caisse'], key=lambda x: x['date'])
        labels = [entry['date'] for entry in sorted_data]
        values = [entry['value'] for entry in sorted_data]

        return jsonify({
            "labels": labels,
            "values": values,
            "min_date": labels[0] if labels else None,
            "max_date": labels[-1] if labels else None
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
@app.route('/tresorerie-chart-data')
def get_tresorerie_chart_data():
    try:
        with open('kpi_tresorerie.json', 'r') as f:
            data = json.load(f)

        sorted_data = sorted(data['tresorerie'], key=lambda x: x['date'])
        labels = [entry['date'] for entry in sorted_data]
        values = [entry['value'] for entry in sorted_data]

        return jsonify({
            "labels": labels,
            "values": values,
            "min_date": labels[0] if labels else None,
            "max_date": labels[-1] if labels else None
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/combined-stock-chart-data')
def get_combined_stock_chart_data():
    try:
        with open(STOCK_FILE, 'r') as f:
            data = json.load(f)
        
        # Get all dates from all stock types
        all_dates = set()
        for stock_type in ['total_stock', 'stock_principale', 'hangar', 'hangar_reserve', 'depot_reserver']:
            for entry in data[stock_type]:
                all_dates.add(entry['date'])
        
        sorted_dates = sorted(all_dates)
        
        # Prepare datasets for each stock type
        datasets = []
        colors = ['#4e73df', '#e74a3b', '#f6c23e', '#1cc88a', '#36b9cc']
        stock_types = [
            ('total_stock', 'Total Stock'),
            ('stock_principale', 'Stock Principale'),
            ('hangar', 'Hangar'),
            ('hangar_reserve', 'Hangar Reserve'),
            ('depot_reserver', 'Dépôt Reserve')
        ]
        
        for i, (stock_type, label) in enumerate(stock_types):
            values = []
            stock_data = {entry['date']: entry['value'] for entry in data[stock_type]}
            
            for date in sorted_dates:
                values.append(stock_data.get(date, None))
            
            datasets.append({
                "label": label,
                "data": values,
                "borderColor": colors[i],
                "backgroundColor": f"{colors[i]}20",
                "pointBackgroundColor": colors[i],
                "pointBorderColor": '#fff',
                "pointHoverRadius": 5,
                "pointHoverBackgroundColor": colors[i],
                "pointHoverBorderColor": '#fff',
                "pointHitRadius": 10,
                "pointBorderWidth": 2,
                "borderWidth": 2,
                "tension": 0.3,
                "fill": i == 0  # Only fill the first dataset for better visibility
            })
        
        return jsonify({
            "labels": sorted_dates,
            "datasets": datasets,
            "min_date": sorted_dates[0] if sorted_dates else None,
            "max_date": sorted_dates[-1] if sorted_dates else None
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def fetch_paiment():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()

            query = """
SELECT
    NVL(SUM(
        CASE 
            WHEN z.name = 'Encaiss: Espèces' THEN ROUND(ci.PayAmt, 2)
            WHEN z.name = 'Décaiss: Espèces' THEN -ROUND(ci.PayAmt, 2)
            ELSE 0 
        END
    ), 0) AS total_difference
FROM 
    C_Payment ci
    INNER JOIN ZSubPaymentRule z ON ci.ZSubPaymentRule_ID = z.ZSubPaymentRule_ID
WHERE 
    TRUNC(ci.DATETRX) = TRUNC(SYSDATE)
    AND ci.DOCACTION IN ('CO', 'CL', 'co', 'cl')
    AND ci.AD_Client_ID = 1000000
    AND z.name IN ('Encaiss: Espèces', 'Décaiss: Espèces')


            """

            cursor.execute(query)
            row = cursor.fetchone()
            paiment = row[0] if row else 0.0
            return {"Total_Paiment": paiment}

    except Exception as e:
        logger.error(f"Error fetching paiment data: {e}")
        return {"error": "An error occurred while fetching paiment data."}


@app.route('/paiement-net', methods=['GET'])
def paiment():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."}), 500
    data = fetch_paiment()
    return jsonify(data)




import datetime
from flask import request, jsonify



# @app.route('/fetchFournisseurDataByYear', methods=['GET'])
# def fetch_fournisseur_by_year():
#     fournisseur = request.args.get('fournisseur')
#     product = request.args.get('product')
#     client = request.args.get('client')
#     operateur = request.args.get('operateur')
#     bccb = request.args.get('bccb')
#     zone = request.args.get('zone')
#     year = request.args.get('year')
#     month = request.args.get('month')

#     try:
#         year = int(year) if year else None
#         month = int(month) if month else None
#     except ValueError:
#         return jsonify({"error": "Invalid year or month format"}), 400

#     data = fetch_fournisseur_data_by_year(
#         fournisseur=fournisseur,
#         product=product,
#         client=client,
#         operateur=operateur,
#         bccb=bccb,
#         zone=zone,
#         year=year,
#         month=month
#     )
#     return jsonify(data)

def fetch_fournisseur_data_by_year(fournisseur=None, product=None, client=None, operateur=None, bccb=None, zone=None, year=None, month=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            ad_org_id = 1000000

            if year and month:
                start_date = f"{year}-{month:02d}-01"
                last_day = 31 if month in [1, 3, 5, 7, 8, 10, 12] else 30
                if month == 2:
                    last_day = 29 if int(year) % 4 == 0 and (int(year) % 100 != 0 or int(year) % 400 == 0) else 28
                end_date = f"{year}-{month:02d}-{last_day}"
            elif year:
                start_date = f"{year}-01-01"
                end_date = f"{year}-12-31"
            else:
                current_year = datetime.datetime.now().year
                start_date = f"{current_year}-01-01"
                end_date = f"{current_year}-12-31"
                year = current_year

            query = """
                SELECT 
                    TO_CHAR(xf.MOVEMENTDATE, 'YYYY') AS year,
                    TO_CHAR(xf.MOVEMENTDATE, 'MM') AS month,
                    SUM(xf.TOTALLINE) AS total,
                    ROUND(
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            WHEN SUM(xf.CONSOMATION) < 0 THEN ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION) * -1) * 100
                            ELSE ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION)) * 100
                        END, 4) AS marge
                FROM xx_ca_fournisseur xf
                JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                JOIN C_ORDER C ON mi.C_ORDER_ID = c.C_ORDER_ID
                WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND xf.AD_Org_ID = :ad_org_id
                    AND xf.DOCSTATUS != 'RE'
                    AND (:fournisseur IS NULL OR xf.name LIKE :fournisseur || '%')
                    AND (:product IS NULL OR xf.product LIKE :product || '%')
                    AND (:client IS NULL OR cb.name LIKE :client || '%')
                    AND (:operateur IS NULL OR au.name LIKE :operateur || '%')
                    AND (:bccb IS NULL OR C.DOCUMENTNO LIKE :bccb || '%')
                    AND (:zone IS NULL OR sr.name LIKE :zone || '%')
                GROUP BY ROLLUP(TO_CHAR(xf.MOVEMENTDATE, 'YYYY'), TO_CHAR(xf.MOVEMENTDATE, 'MM'))
                HAVING TO_CHAR(xf.MOVEMENTDATE, 'YYYY') = :year_str OR TO_CHAR(xf.MOVEMENTDATE, 'YYYY') IS NULL
                ORDER BY year NULLS LAST, month NULLS LAST
            """

            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'ad_org_id': ad_org_id,
                'year_str': str(year)
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            return data

    except Exception as e:
        logger.error(f"Error fetching total fournisseur data: {e}")
        return {"error": "An error occurred while fetching total data."}

@app.route('/fetchFournisseurDataByYear', methods=['GET'])
def fetch_fournisseur_by_year():
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    year_param = request.args.get('year')
    month = request.args.get('month')

    try:
        year_param = int(year_param) if year_param else None
        month = int(month) if month else None
    except ValueError:
        return jsonify({"error": "Invalid year or month format"}), 400

    if year_param:
        data = fetch_fournisseur_data_by_year(
            fournisseur=fournisseur,
            product=product,
            client=client,
            operateur=operateur,
            bccb=bccb,
            zone=zone,
            year=year_param,
            month=month
        )
        return jsonify({str(year_param): data})
    else:
        current_year = datetime.datetime.now().year
        full_data = {}
        for y in range(2022, current_year + 1):
            yearly_data = fetch_fournisseur_data_by_year(
                fournisseur=fournisseur,
                product=product,
                client=client,
                operateur=operateur,
                bccb=bccb,
                zone=zone,
                year=y,
                month=None
            )
            full_data[str(y)] = yearly_data
        return jsonify(full_data)






if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000)