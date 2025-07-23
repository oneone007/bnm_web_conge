import oracledb
from flask import Flask, jsonify, request, send_file, make_response
from flask_cors import CORS
import logging
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime


from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as ReportLabTable, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from datetime import datetime
import io

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure Oracle database connection pool
DB_POOL = oracledb.create_pool(
    user="compiere",
    password="compiere",
    dsn="192.168.1.213/compiere",
    min=2,
    max=10,
    increment=1
)




# Test database connection
def test_db_connection():
    try:
        with DB_POOL.acquire() as connection:
            logger.info("Database connection successful.")
        return True
    except Exception as e:
        logger.error(f"Database connection failed: {str(e)}")
        return False
    


def fetch_charges_dashboard(date_debut=None, date_fin=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            
            # Set default dates if not provided
            if not date_debut:
                date_debut = '2024-01-01'  # Default start date
            if not date_fin:
                date_fin = datetime.now().strftime('%Y-%m-%d')  # Default end date (today)
            
            # First query to get charge types and charges with totals
            summary_query = """
            SELECT 
                ct.C_ChargeType_ID,
                ct.name AS type_de_charge,
                c.C_Charge_ID,
                c.name AS charge_name,
                COALESCE(charge_amounts.total_amount, 0) AS montant,
                CASE 
                    WHEN type_totals.type_total > 0 THEN 
                        ROUND((COALESCE(charge_amounts.total_amount, 0) / type_totals.type_total) * 100, 3)
                    ELSE 0 
                END AS pourcentage
            FROM C_ChargeType ct
            INNER JOIN C_Charge c ON ct.C_ChargeType_ID = c.C_ChargeType_ID
            LEFT JOIN (
                SELECT 
                    c.C_Charge_ID,
                    SUM(il.LineNetAmt) AS total_amount
                FROM C_Charge c
                INNER JOIN C_InvoiceLine il ON c.C_Charge_ID = il.C_Charge_ID
                INNER JOIN C_Invoice i ON il.C_Invoice_ID = i.C_Invoice_ID
                WHERE c.IsActive = 'Y' 
                  AND c.AD_CLIENT_ID = 1000000
                  AND i.IsActive = 'Y'
                  AND i.DocStatus IN ('CO', 'CL')
                  AND i.DateInvoiced >= TO_DATE(:date_debut, 'YYYY-MM-DD')
                  AND i.DateInvoiced <= TO_DATE(:date_fin, 'YYYY-MM-DD')
                GROUP BY c.C_Charge_ID
            ) charge_amounts ON c.C_Charge_ID = charge_amounts.C_Charge_ID
            LEFT JOIN (
                SELECT 
                    ct.C_ChargeType_ID,
                    SUM(il.LineNetAmt) AS type_total
                FROM C_ChargeType ct
                INNER JOIN C_Charge c ON ct.C_ChargeType_ID = c.C_ChargeType_ID
                INNER JOIN C_InvoiceLine il ON c.C_Charge_ID = il.C_Charge_ID
                INNER JOIN C_Invoice i ON il.C_Invoice_ID = i.C_Invoice_ID
                WHERE ct.IsActive = 'Y' 
                  AND ct.AD_CLIENT_ID = 1000000
                  AND c.IsActive = 'Y'
                  AND i.IsActive = 'Y'
                  AND i.DocStatus IN ('CO', 'CL')
                  AND i.DateInvoiced >= TO_DATE(:date_debut, 'YYYY-MM-DD')
                  AND i.DateInvoiced <= TO_DATE(:date_fin, 'YYYY-MM-DD')
                GROUP BY ct.C_ChargeType_ID
            ) type_totals ON ct.C_ChargeType_ID = type_totals.C_ChargeType_ID
            WHERE ct.IsActive = 'Y' 
              AND ct.AD_CLIENT_ID = 1000000
              AND c.IsActive = 'Y'
            ORDER BY ct.name, c.name
            """
            
            cursor.execute(summary_query, {'date_debut': date_debut, 'date_fin': date_fin})
            summary_rows = cursor.fetchall()
            summary_columns = [col[0] for col in cursor.description]
            
            # Second query to get invoice details for each charge with line details
            details_query = """
            SELECT 
                c.C_Charge_ID,
                i.C_Invoice_ID,
                i.DocumentNo AS invoice_number,
                i.DateInvoiced,
                bp.name AS bpartner_name,
                il.LineNetAmt AS line_amount,
                i.GrandTotal AS invoice_total,
                il.C_InvoiceLine_ID,
                il.Description AS line_description,
                il.PriceEntered AS line_total_amount,
                il.QtyEntered AS line_qty_entered
            FROM C_Charge c
            INNER JOIN C_InvoiceLine il ON c.C_Charge_ID = il.C_Charge_ID
            INNER JOIN C_Invoice i ON il.C_Invoice_ID = i.C_Invoice_ID
            INNER JOIN C_BPartner bp ON i.C_BPartner_ID = bp.C_BPartner_ID
            WHERE c.IsActive = 'Y' 
              AND c.AD_CLIENT_ID = 1000000
              AND i.IsActive = 'Y'
              AND i.DocStatus IN ('CO', 'CL')
              AND i.DateInvoiced >= TO_DATE(:date_debut, 'YYYY-MM-DD')
              AND i.DateInvoiced <= TO_DATE(:date_fin, 'YYYY-MM-DD')
            ORDER BY c.C_Charge_ID, i.DateInvoiced DESC, il.C_InvoiceLine_ID
            """
            
            cursor.execute(details_query, {'date_debut': date_debut, 'date_fin': date_fin})
            details_rows = cursor.fetchall()
            details_columns = [col[0] for col in cursor.description]
            
            # Organize invoice details by charge ID
            invoice_details = {}
            for row in details_rows:
                detail_data = dict(zip(details_columns, row))
                charge_id = detail_data['C_CHARGE_ID']
                invoice_id = detail_data['C_INVOICE_ID']
                
                if charge_id not in invoice_details:
                    invoice_details[charge_id] = {}
                
                # Group by invoice ID to avoid duplicating invoice info
                if invoice_id not in invoice_details[charge_id]:
                    invoice_details[charge_id][invoice_id] = {
                        'invoice_id': detail_data['C_INVOICE_ID'],
                        'invoice_number': detail_data['INVOICE_NUMBER'],
                        'date_invoiced': detail_data['DATEINVOICED'].strftime('%Y-%m-%d') if detail_data['DATEINVOICED'] else None,
                        'bpartner_name': detail_data['BPARTNER_NAME'],
                        'invoice_total': -float(detail_data['INVOICE_TOTAL']) if detail_data['INVOICE_TOTAL'] else 0,  # Make negative
                        'invoice_lines': []
                    }
                
                # Add line details to the invoice
                invoice_details[charge_id][invoice_id]['invoice_lines'].append({
                    'line_id': detail_data['C_INVOICELINE_ID'],
                    'line_description': detail_data['LINE_DESCRIPTION'] or 'No description',
                    'line_net_amount': -float(detail_data['LINE_AMOUNT']) if detail_data['LINE_AMOUNT'] else 0,  # Make negative
                    'line_total_amount': -float(detail_data['LINE_TOTAL_AMOUNT']) if detail_data['LINE_TOTAL_AMOUNT'] else 0,  # Make negative
                    'line_qty_entered': float(detail_data['LINE_QTY_ENTERED']) if detail_data['LINE_QTY_ENTERED'] else 0
                })
            
            # Convert nested invoice_details structure to a flat list for easier processing
            flattened_invoice_details = {}
            for charge_id, invoices in invoice_details.items():
                flattened_invoice_details[charge_id] = list(invoices.values())
            
            # Organize data by charge type
            charges_data = {}
            total_all_charges = 0
            
            for row in summary_rows:
                data = dict(zip(summary_columns, row))
                charge_type_id = data['C_CHARGETYPE_ID']
                charge_type_name = data['TYPE_DE_CHARGE']
                charge_id = data['C_CHARGE_ID']
                
                if charge_type_id not in charges_data:
                    charges_data[charge_type_id] = {
                        'type_name': charge_type_name,
                        'type_total': 0,
                        'charges': []
                    }
                
                # Make charge amount negative (expense)
                charge_amount = -float(data['MONTANT']) if data['MONTANT'] else 0
                charges_data[charge_type_id]['type_total'] += charge_amount
                total_all_charges += charge_amount
                
                # Get invoice details for this charge
                charge_invoice_details = flattened_invoice_details.get(charge_id, [])
                
                charges_data[charge_type_id]['charges'].append({
                    'charge_id': charge_id,
                    'charge_name': data['CHARGE_NAME'],
                    'montant': charge_amount,
                    'pourcentage': float(data['POURCENTAGE']) if data['POURCENTAGE'] else 0,
                    'invoice_details': charge_invoice_details
                })
            
            # Calculate percentage for each charge type
            for charge_type_id in charges_data:
                if total_all_charges != 0:  # Changed from > 0 to != 0 since values are negative
                    charges_data[charge_type_id]['type_percentage'] = round(
                        (charges_data[charge_type_id]['type_total'] / total_all_charges) * 100, 3
                    )
                else:
                    charges_data[charge_type_id]['type_percentage'] = 0
            
            return {
                'charges_by_type': charges_data,
                'total_all_charges': total_all_charges,
                'date_debut': date_debut,
                'date_fin': date_fin
            }
            
    except Exception as e:
        logger.error(f"Error fetching charges dashboard: {e}")
        return {"error": "An error occurred while fetching charges dashboard."}

@app.route('/fetch-charges-dashboard', methods=['GET'])
def fetch_charges_dashboard_endpoint():
    if not test_db_connection():
        return jsonify({"error": "Database connection failed"}), 500
    
    # Get date parameters from query string
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')
    
    data = fetch_charges_dashboard(date_debut, date_fin)
    return jsonify(data)

@app.route('/download-charges-dashboard-excel', methods=['GET'])
def download_charges_dashboard_excel():
    # Get date parameters from query string
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')
    
    data = fetch_charges_dashboard(date_debut, date_fin)
    
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 404
    
    # Flatten the data for Excel export with invoice and line details
    excel_data = []
    for charge_type_id, charge_type_data in data['charges_by_type'].items():
        # Only add charge type row if total is not zero
        if charge_type_data['type_total'] != 0:
            excel_data.append({
                'Type de Charge': charge_type_data['type_name'],
                'Charge': 'TOTAL',
                'Montant': charge_type_data['type_total'],
                'Pourcentage': charge_type_data['type_percentage'],
                'Facture N°': '',
                'Date Facture': '',
                'Fournisseur': '',
                'Description Ligne': '',
                'Montant Net Ligne': '',
                'Montant Total Ligne': '',
                'Quantité': ''
            })
        
        # Add individual charges with invoice and line details
        for charge in charge_type_data['charges']:
            # Only process charges with non-zero amounts
            if charge['montant'] != 0:
                if charge['invoice_details']:
                    # Add each invoice with its lines
                    for invoice in charge['invoice_details']:
                        if invoice['invoice_lines']:
                            # Add each line as a separate row
                            for line in invoice['invoice_lines']:
                                # Only add lines with non-zero amounts
                                if line['line_net_amount'] != 0:
                                    excel_data.append({
                                        'Type de Charge': charge_type_data['type_name'],
                                        'Charge': charge['charge_name'],
                                        'Montant': charge['montant'],
                                        'Pourcentage': charge['pourcentage'],
                                        'Facture N°': invoice['invoice_number'],
                                        'Date Facture': invoice['date_invoiced'],
                                        'Fournisseur': invoice['bpartner_name'],
                                        'Description Ligne': line['line_description'],
                                        'Montant Net Ligne': line['line_net_amount'],
                                        'Montant Total Ligne': line['line_total_amount'],
                                        'Quantité': line['line_qty_entered']
                                    })
                        else:
                            # Invoice without lines
                            excel_data.append({
                                'Type de Charge': charge_type_data['type_name'],
                                'Charge': charge['charge_name'],
                                'Montant': charge['montant'],
                                'Pourcentage': charge['pourcentage'],
                                'Facture N°': invoice['invoice_number'],
                                'Date Facture': invoice['date_invoiced'],
                                'Fournisseur': invoice['bpartner_name'],
                                'Description Ligne': 'Aucune ligne de détail',
                                'Montant Net Ligne': '',
                                'Montant Total Ligne': '',
                                'Quantité': ''
                            })
                else:
                    # Add charge without invoice details
                    excel_data.append({
                        'Type de Charge': charge_type_data['type_name'],
                        'Charge': charge['charge_name'],
                        'Montant': charge['montant'],
                        'Pourcentage': charge['pourcentage'],
                        'Facture N°': 'Aucune facture',
                        'Date Facture': '',
                        'Fournisseur': '',
                        'Description Ligne': '',
                        'Montant Net Ligne': '',
                        'Montant Total Ligne': '',
                        'Quantité': ''
                    })
    
    # Add "Total des Charges" row at the end
    if data['total_all_charges'] != 0:
        excel_data.append({
            'Type de Charge': 'TOTAL DES CHARGES',
            'Charge': '',
            'Montant': data['total_all_charges'],
            'Pourcentage': 100.0,
            'Facture N°': '',
            'Date Facture': '',
            'Fournisseur': '',
            'Description Ligne': '',
            'Montant Net Ligne': '',
            'Montant Total Ligne': '',
            'Quantité': ''
        })
    
    # Generate filename with date range
    filename = f"Charges_Dashboard_Details_{data['date_debut']}_to_{data['date_fin']}.xlsx"
    
    return generate_excel_charges_dashboard(excel_data, filename)

def generate_excel_charges_dashboard(data, filename):
    if not data:
        return jsonify({"error": "No data to export"}), 404

    df = pd.DataFrame(data)
    wb = Workbook()
    ws = wb.active
    ws.title = "Charges Dashboard Details"

    # Apply header styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    ws.append(df.columns.tolist())
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Append data with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Highlight total rows (both charge type totals and total des charges)
        if row[1] == 'TOTAL' or row[0] == 'TOTAL DES CHARGES':  # Charge column or Type de Charge column
            for cell in ws[row_idx]:
                cell.font = Font(bold=True)
                if row[0] == 'TOTAL DES CHARGES':
                    # Special highlighting for the grand total
                    cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # Highlight rows with no invoice details
        elif row[4] == 'Aucune facture':  # Facture N° column
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # Highlight rows with no line details
        elif row[7] == 'Aucune ligne de détail':  # Description Ligne column
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="FFF0E6", end_color="FFF0E6", fill_type="solid")

    # Add table
    table = Table(displayName="ChargesDashboardDetailsTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, 
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5003)