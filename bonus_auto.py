import oracledb
from flask import Flask, jsonify, request
from flask_cors import CORS  # Import CORS
import logging
from flask import Flask, request, send_file
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure Oracle database connection pool
DB_POOL = oracledb.create_pool(
    user="compiere",
    password="compiere",
    dsn="192.168.1.213/compiere",
    min=2,
    max=10,
    increment=1
)

# Fetch data from Oracle DB
def fetch_data_from_db():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            # Define your SQL query here
            query = """
       SELECT DISTINCT 
    mp.name AS product,
    (SELECT description 
     FROM XX_SalesContext xsc 
     WHERE mp.XX_SalesContext_id = xsc.XX_SalesContext_id) AS bonus,
    l.name AS laboratory_name,
    cbp.name AS fournisseur
FROM 
    m_product mp
JOIN XX_Laboratory l ON l.XX_Laboratory_ID = mp.XX_Laboratory_ID
JOIN m_storage s ON s.m_product_id = mp.m_product_id
JOIN M_ATTRIBUTEINSTANCE asi ON s.M_ATTRIBUTESETINSTANCE_ID = asi.M_ATTRIBUTESETINSTANCE_ID
JOIN c_bpartner cbp ON cbp.c_bpartner_id = ValueNUMBER_of_ASI('Fournisseur', asi.m_attributesetinstance_id)
WHERE 
    mp.xx_salescontext_id NOT IN (1000000, 1000100)
    AND mp.ad_org_id = 1000000
ORDER BY 
    mp.name
FETCH FIRST 1048575 ROWS ONLY

            """
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching data: {e}")
        return {"error": "An error occurred while fetching data."}
def test_db_connection():
    try:
        with DB_POOL.acquire() as connection:
            logger.info("Database connection successful.")
        return True
    except Exception as e:
        logger.error(f"Database connection failed: {str(e)}")
        return False



@app.route('/fetch-bonus-data', methods=['GET'])
def fetch_remise_data():
    if not test_db_connection():
        return jsonify({"error": "Failed to connect to the database."})
     
    data = fetch_data_from_db()
    return jsonify(data)

@app.route('/download-excel', methods=['GET'])
def download_excel():
    # Get MARGE value from request (if available)
    bonus_value = request.args.get("bonus", "BONUS")

    # Fetch real data from Oracle DB
    data = fetch_data_from_db()

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400  # Handle empty dataset

    # Convert to Pandas DataFrame
    df = pd.DataFrame(data)

    if df.empty:
        return jsonify({"error": "No data available"}), 400

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "All Data"

    # Write headers with formatting
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(df.columns.tolist())  # Add header row

    for col_idx, cell in enumerate(ws[1], 1):  # Format header row
        cell.fill = header_fill
        cell.font = header_font
        ws.auto_filter.ref = ws.dimensions  # Enable filter & sort

    # Write data with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)  # Add data row
        if row_idx % 2 == 0:  # Apply alternating row colors
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Create an Excel table with formatting
    table = Table(displayName="DataTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Generate a dynamic filename
    today_date = datetime.now().strftime("%d-%m-%Y")  # Format: day-month-year
    filename = f"{bonus_value}_{today_date}.xlsx"

    # Save to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    app.run(debug=True, port=5002)
