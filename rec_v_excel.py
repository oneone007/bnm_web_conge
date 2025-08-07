

import oracledb
from flask import Flask, jsonify, request, send_file, make_response
from flask_cors import CORS
import logging
import pandas as pd
from io import BytesIO
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import os
import json

import calendar
import mysql.connector




app = Flask(__name__)
CORS(app)  # Allow your DDNS domains and ports

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure Oracle database connection pool
DB_POOL = oracledb.create_pool(
    user="compiere",
    password="compiere",
    dsn="192.168.1.213/compiere",
    min=2,
    max=10,
    increment=1
)


# MySQL connection for recouvrement data
def get_localdb_connection():
    try:
        return mysql.connector.connect(
            host="localhost",
            user="bmk",
            password="",
            database="bnm_web",
            charset="utf8",
            use_unicode=True,
            autocommit=False
        )
    except mysql.connector.Error as err:
        logger.error(f"Error connecting to MySQL database: {err}")
        return None




@app.route('/download-bccb-product-excel', methods=['GET'])
def download_bccb_product_excel():
    bccb = request.args.get('bccb')
    ad_org_id = request.args.get('ad_org_id', '1000012')

    if not bccb:
        return jsonify({"error": "Missing BCCB parameter"}), 400

    try:
        ad_org_id = int(ad_org_id)
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_bccb_product(bccb, ad_org_id)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"BCCB_Product_{bccb}_{today_date}.xlsx"

    return generate_excel_bccb_product(data, filename)


def generate_excel_bccb_product(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "BCCB Product Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="BCCBProductTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")






@app.route('/download-fournisseur-excel', methods=['GET'])
def download_fournisseur_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_fournisseur_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id
    today_date = datetime.now().strftime("%d-%m-%Y")
    if ad_org_id == 1000012:
        filename = f"FournisseurRecap_facturation_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"FournisseurRecap_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_fournisseur(data, filename)


def generate_excel_fournisseur(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Fournisseur Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="FournisseurRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")






@app.route('/download-product-excel', methods=['GET'])
def download_product_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_product_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id
    today_date = datetime.now().strftime("%d-%m-%Y")
    if ad_org_id == 1000012:
        filename = f"ProductRecap_facturation_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"ProductRecap_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_product(data, filename)


def generate_excel_product(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Product Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="ProductRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")





@app.route('/download-totalrecap-excel', methods=['GET'])
def download_totalrecap_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    ad_org_id = request.args.get('ad_org_id')

    if not start_date or not end_date or not ad_org_id:
        return jsonify({"error": "Missing start_date, end_date, or ad_org_id parameters"}), 400

    try:
        ad_org_id = int(ad_org_id)  # Convert to integer
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data from the database
    data = fetch_rcap_data(start_date, end_date, ad_org_id)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id
    today_date = datetime.now().strftime("%d-%m-%Y")
    if ad_org_id == 1000012:
        filename = f"TotalRecap_facturation_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"TotalRecap_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_totalrecap(data, filename)


def generate_excel_totalrecap(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Total Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="TotalRecapTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")






 



@app.route('/download-zone-excel', methods=['GET'])
def download_zone_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_zone_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id
    today_date = datetime.now().strftime("%d-%m-%Y")
    if ad_org_id == 1000012:
        filename = f"ZoneRecap_facturation_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"ZoneRecap_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_zone(data, filename)


def generate_excel_zone(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Zone Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="ZoneRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/download-client-excel', methods=['GET'])
def download_client_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_client_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id
    today_date = datetime.now().strftime("%d-%m-%Y")
    if ad_org_id == 1000012:
        filename = f"ClientRecap_facturation_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"ClientRecap_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_client(data, filename)


def generate_excel_client(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Client Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="ClientRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route('/download-operator-excel', methods=['GET'])
def download_operator_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_operator_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id
    today_date = datetime.now().strftime("%d-%m-%Y")
    if ad_org_id == 1000012:
        filename = f"OperatorRecap_facturation_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"OperatorRecap_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_operator(data, filename)


def generate_excel_operator(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Operator Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="OperatorRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")








@app.route('/download-bccb-excel', methods=['GET'])
def download_bccb_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data
    data = fetch_bccb_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"BCCBRecap_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_bccb(data, filename)

def generate_excel_bccb(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "BCCB Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="BCCBRecapTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")




if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000)
