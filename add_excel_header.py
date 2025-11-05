#!/usr/bin/env python3
"""
add_excel_header.py

Usage:
  python add_excel_header.py input.xls output.xlsx --logo logo.png

This script reads an Excel file (xls or xlsx), inserts a multi-line header at the top of the first sheet
matching the provided design (company name, contact info, report title and period, page indicator).
It uses pandas to read the input and openpyxl to write the header and preserve the data.

Features:
- Accepts input Excel file path (xls/xlsx).
- Optional logo image to embed at top-left.
- Writes output as .xlsx (openpyxl).
- Simple formatting: merged cells, font sizes, bold where appropriate, thin bottom border.

Dependencies: pandas, openpyxl, pillow

Note: This is intentionally conservative about styling so it works without advanced fonts.
"""
import sys
import argparse
from pathlib import Path
import pandas as pd

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, Border, Side
except Exception:
    print("Missing dependency: install requirements in requirements.txt")
    raise


HEADER_LINES = [
    # kept for reference if needed; the code below composes the header explicitly
]


def insert_header(ws, logo_path: Path | None = None, data_columns: int = 14):
    """Insert a multi-row header into worksheet `ws`.

    This function composes the header in multiple cells (not a single long line), sets wrap_text
    where needed, and leaves the worksheet ready for writing data starting at row `start_row`.
    """
    header_height = 9

    # If logo provided, insert image at A1 area (cols A-B)
    if logo_path is None:
        default_logo = Path(__file__).with_name('log.png')
        if default_logo.exists():
            logo_path = default_logo

    if logo_path and logo_path.exists():
        try:
            img = XLImage(str(logo_path))
            # anchor top-left and scale to approximately fit into columns A-B (width ~160px, height ~80px)
            img.anchor = 'A1'
            # try to set width/height to fit nicely
            img.width = 160
            img.height = 80
            ws.add_image(img)
        except Exception:
            pass

    thin = Side(border_style="thin", color="000000")
    border_bottom = Border(bottom=thin)

    # Set a reasonable column width map up to data_columns
    default_widths = [18, 10, 12, 12, 14, 14, 10, 10, 12, 12, 12, 10, 8, 8, 8, 8]
    for c in range(1, data_columns + 1):
        w = default_widths[c - 1] if c - 1 < len(default_widths) else 10
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w

    # Left block (company short name under logo) placed in cols 3..min(5,data_columns)
    left_end = min(5, data_columns)
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=left_end)
    c1 = ws.cell(row=1, column=3)
    c1.value = "BNM PARAPHARM"
    c1.font = Font(size=14, bold=True)
    c1.alignment = Alignment(horizontal="left", vertical="top")

    # Center-left block: address and contact (rows 1..3, columns 6..center_right)
    center_left_start = 6
    center_left_end = min(9, max(6, data_columns - 4))
    ws.merge_cells(start_row=1, start_column=center_left_start, end_row=3, end_column=center_left_end)
    leftinfo = ws.cell(row=1, column=center_left_start)
    leftinfo.value = (
        "RTE NATIONALE N 05 N 46 Ain\n"
        "Constantine\n"
        "Tél : 031 97 55 93\n"
        "Fax : 031 97 55 94\n"
        "Email : bnm.parapharm@gmail.com\n"
        "Site : www.bnm.com"
    )
    leftinfo.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")

    # Right block: registration info - use last columns so it's aligned with the table width
    right_start = max(center_left_end + 1, data_columns - 4)
    right_start = min(right_start, data_columns)
    right_end = data_columns
    if right_start < 1:
        right_start = 1
    ws.merge_cells(start_row=1, start_column=right_start, end_row=3, end_column=right_end)
    reg = ws.cell(row=1, column=right_start)
    reg.value = (
        "RC : 25/00-0071142 B 15\n"
        "NIF : 001525007114238\n"
        "NAI : 25100124031\n"
        "NIS : 001525100039948\n"
        "RIB : 001008500300000097130\n"
        "Capital : 12 300 000,00 DA"
    )
    reg.alignment = Alignment(wrap_text=True, horizontal="right", vertical="top")


    # Separator line (thin bottom border) across columns 1..data_columns
    sep_row = 4
    for col in range(1, data_columns + 1):
        ws.cell(row=sep_row, column=col).border = border_bottom

    # Title (center) across the data columns except the last column which we'll reserve for page info
    title_row = 5
    title_end = max(1, data_columns - 1)
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=title_end)
    title = ws.cell(row=title_row, column=1)
    title.value = "JOURNAL DES VENTES GLOBAL"
    title.font = Font(size=16, bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Period on next row across the data columns (leave last column for page info)
    period_row = 6
    period_end = max(1, data_columns - 1)
    ws.merge_cells(start_row=period_row, start_column=1, end_row=period_row, end_column=period_end)
    period = ws.cell(row=period_row, column=1)
    period.value = "PERIODE    01/10/2025    AU    21/10/2025"
    period.alignment = Alignment(horizontal="center")

    # Page indicator on the same horizontal band, right side (last column)
    page_cell = ws.cell(row=period_row, column=data_columns)
    page_cell.value = "Page : 1  /  3"
    page_cell.alignment = Alignment(horizontal="right")

    # Organisation line at bottom of header
    org_row = 8
    ws.merge_cells(start_row=org_row, start_column=1, end_row=org_row, end_column=min(3, data_columns))
    org_label = ws.cell(row=org_row, column=1)
    org_label.value = "Organisation :"
    org_label.font = Font(bold=True)
    ws.merge_cells(start_row=org_row, start_column=4, end_row=org_row, end_column=min(12, data_columns))
    org_val = ws.cell(row=org_row, column=4)
    org_val.value = "BNM PARAPHARM"
    org_val.alignment = Alignment(horizontal="left")

    # Make sure we left enough space; we'll write data starting at row header_height+1



def process_file(input_path: Path, output_path: Path, logo: Path | None = None):
    # Read first sheet into DataFrame
    try:
        df = pd.read_excel(input_path, sheet_name=0, engine='xlrd' if input_path.suffix in ('.xls',) else None)
    except ValueError:
        # fallback without specifying engine
        df = pd.read_excel(input_path, sheet_name=0)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Insert header with width matching data columns
    data_columns = max(1, len(df.columns))
    insert_header(ws, logo, data_columns)

    # Write DataFrame starting at row after header (row 9)
    start_row = 9
    # write header
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=str(col))
    # write rows
    for i, row in enumerate(df.itertuples(index=False, name=None), start=start_row+1):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    # Save
    wb.save(output_path)


def main(argv):
    p = argparse.ArgumentParser(description="Add header to Excel file and save as xlsx")
    p.add_argument('input', help='Input Excel file (.xls or .xlsx)')
    p.add_argument('output', nargs='?', help='Output XLSX path (default: input_header.xlsx)')
    p.add_argument('--logo', help='Optional logo image file to embed (png/jpg)')
    args = p.parse_args(argv)

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Input file not found: {input_path}")
        sys.exit(2)

    output_path = Path(args.output) if args.output else input_path.with_name(input_path.stem + '_with_header.xlsx')
    logo_path = Path(args.logo) if args.logo else None

    process_file(input_path, output_path, logo_path)
    print(f"Wrote: {output_path}")


if __name__ == '__main__':
    main(sys.argv[1:])
