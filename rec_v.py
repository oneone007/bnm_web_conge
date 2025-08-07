

import oracledb
from flask import Flask, jsonify, request, send_file, make_response
from flask_cors import CORS
import logging
import pandas as pd
from io import BytesIO
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import os
import json

import calendar
import mysql.connector




app = Flask(__name__)
CORS(app)  # Allow your DDNS domains and ports

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure Oracle database connection pool
DB_POOL = oracledb.create_pool(
    user="compiere",
    password="compiere",
    dsn="192.168.1.213/compiere",
    min=2,
    max=10,
    increment=1
)





def fetch_fournisseur_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT * FROM (
                    SELECT 
                        CAST(xf.name AS VARCHAR2(300)) AS FOURNISSEUR,   
                        SUM(xf.TOTALLINE) AS total, 
                        SUM(xf.qtyentered) AS QTY,
                        ROUND(
                            CASE 
                                WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                                WHEN SUM(xf.CONSOMATION) < 0 THEN ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION) * -1) * 100
                                ELSE ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION)) * 100
                            END, 4) AS marge,
                        0 AS sort_order
                    FROM xx_ca_fournisseur xf
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER C ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
                        AND xf.AD_Org_ID = :ad_org_id
                        AND xf.DOCSTATUS != 'RE'
                        AND (:fournisseur IS NULL OR xf.name LIKE :fournisseur || '%')
                        AND (:product IS NULL OR xf.product LIKE :product || '%')
                        AND (:client IS NULL OR cb.name LIKE :client || '%')
                        AND (:operateur IS NULL OR au.name LIKE :operateur || '%')
                        AND (:bccb IS NULL OR C.DOCUMENTNO LIKE :bccb || '%')
                        AND (:zone IS NULL OR sr.name LIKE :zone || '%')
                        AND (
                            :group_label IS NULL OR
                            (CASE 
                                WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                ELSE 'autre'
                            END = :group_label)
                        )
                    GROUP BY xf.name
                    UNION ALL
                    SELECT 
                        CAST('Total' AS VARCHAR2(300)) AS name, 
                        SUM(xf.TOTALLINE) AS total, 
                        SUM(xf.qtyentered) AS QTY,
                        ROUND(
                            CASE 
                                WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                                WHEN SUM(xf.CONSOMATION) < 0 THEN ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION) * -1) * 100
                                ELSE ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION)) * 100
                            END, 4) AS marge,
                        1 AS sort_order
                    FROM xx_ca_fournisseur xf
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER C ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') AND TO_DATE(:end_date, 'YYYY-MM-DD')
                        AND xf.AD_Org_ID = :ad_org_id
                        AND xf.DOCSTATUS != 'RE'
                        AND (:fournisseur IS NULL OR xf.name LIKE :fournisseur || '%')
                        AND (:product IS NULL OR xf.product LIKE :product || '%')
                        AND (:client IS NULL OR cb.name LIKE :client || '%')
                        AND (:operateur IS NULL OR au.name LIKE :operateur || '%')
                        AND (:bccb IS NULL OR C.DOCUMENTNO LIKE :bccb || '%')
                        AND (:zone IS NULL OR sr.name LIKE :zone || '%')
                        AND (
                            :group_label IS NULL OR
                            (CASE 
                                WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                ELSE 'autre'
                            END = :group_label)
                        )
                )
                ORDER BY sort_order, total DESC
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'ad_org_id': ad_org_id,
                'group_label': group_label
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching fournisseur data: {e}")
        return {"error": "An error occurred while fetching fournisseur data."}

@app.route('/fetchFournisseurData', methods=['GET'])
def fetch_fournisseur():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    if not ad_org_id:
        return jsonify({"error": "Missing ad_org_id parameter"}), 400  # Ensure ad_org_id is provided

    try:
        ad_org_id = int(ad_org_id)  # Convert to integer
    except ValueError:
        return jsonify({"error": "Invalid ad_org_id format"}), 400

    data = fetch_fournisseur_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)
    return jsonify(data)





def fetch_product_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT * FROM (
                    -- Main Product Breakdown
                    SELECT 
                        CAST(xf.product AS VARCHAR2(400)) AS "PRODUIT",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        0 AS "SORT_ORDER"
                    FROM xx_ca_fournisseur xf
                    LEFT JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    LEFT JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    LEFT JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    LEFT JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    LEFT JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    LEFT JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (
                              :group_label IS NULL OR
                              (CASE 
                                  WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                  WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                  ELSE 'autre'
                              END = :group_label)
                          )
                    GROUP BY xf.product

                    UNION ALL

                    -- Total Row
                    SELECT 
                        'Total' AS "PRODUIT",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        1 AS "SORT_ORDER"
                    FROM xx_ca_fournisseur xf
                    LEFT JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    LEFT JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    LEFT JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    LEFT JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    LEFT JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    LEFT JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (
                              :group_label IS NULL OR
                              (CASE 
                                  WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                  WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                  ELSE 'autre'
                              END = :group_label)
                          )
                ) ORDER BY "SORT_ORDER", "TOTAL" DESC
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'ad_org_id': ad_org_id,
                'group_label': group_label
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]

    except Exception as e:
        logger.error(f"Error fetching product data: {e}")
        return {"error": "An error occurred while fetching product data."}

@app.route('/fetchProductData', methods=['GET'])
def fetch_product():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get("product", "").strip()
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date or not ad_org_id:
        return jsonify({"error": "Missing start_date, end_date, or ad_org_id parameters"}), 400

    data = fetch_product_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)
    return jsonify(data)






# Fetch total recap data
def fetch_rcap_data(start_date, end_date, ad_org_id, group_label=None):

    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT 
                    SUM(xf.TOTALLINE) AS CHIFFRE, 
                    SUM(xf.qtyentered) AS QTY,
                    SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION) AS MARGE,
                    SUM(xf.CONSOMATION) AS CONSOMATION,
                    CASE 
                        WHEN SUM(xf.CONSOMATION) < 0 
                        THEN ROUND(((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / (SUM(xf.CONSOMATION) * -1)), 4)
                        ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                    END AS POURCENTAGE
                FROM xx_ca_fournisseur xf
                JOIN C_BPartner cb ON cb.C_BPARTNER_ID = xf.CLIENTID
                WHERE 
                    xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                    AND TO_DATE(:end_date, 'YYYY-MM-DD')
                    AND xf.AD_Org_ID = :ad_org_id
                    AND xf.DOCSTATUS != 'RE'
                    AND cb.ISCUSTOMER = 'Y'
                    AND cb.AD_Client_ID = 1000000
                    AND cb.AD_Org_ID = 1000000
                    AND (
                        :group_label IS NULL OR
                        (CASE 
                            WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                            WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                            ELSE 'autre'
                        END = :group_label)
                    )
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'ad_org_id': ad_org_id,
                'group_label': group_label
            }
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching data: {e}")
        return {"error": "An error occurred while fetching data."}
    





@app.route('/fetchTotalrecapData', methods=['GET'])
def fetch_recap():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    ad_org_id = request.args.get('ad_org_id')  # Get ad_org_id from request
    group_label = request.args.get('group_label')  # Optional group_label (para, potentiel, autre)

    if not start_date or not end_date or not ad_org_id:
        return jsonify({"error": "Missing start_date, end_date, or ad_org_id parameters"}), 400

    try:
        ad_org_id = int(ad_org_id)  # Convert to integer
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    data = fetch_rcap_data(start_date, end_date, ad_org_id, group_label)
    return jsonify(data)





@app.route('/fetchZoneRecap', methods=['GET'])
def fetch_zone():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400
    if not ad_org_id:
        return jsonify({"error": "Missing ad_org_id parameter"}), 400

    data = fetch_zone_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)
    return jsonify(data)


def fetch_zone_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT * FROM (
                    -- Recap by Zone
                    SELECT 
                        CAST(sr.name AS VARCHAR2(100)) AS "ZONE",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        0 AS "SORT_ORDER"
                    FROM C_SalesRegion sr
                    JOIN C_BPartner_Location bpl ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN xx_ca_fournisseur xf ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (
                              :group_label IS NULL OR
                              (CASE 
                                  WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                  WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                  ELSE 'autre'
                              END = :group_label)
                          )
                    GROUP BY sr.name

                    UNION ALL

                    -- Total Row
                    SELECT 
                        'Total' AS "ZONE",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        1 AS "SORT_ORDER"
                    FROM C_SalesRegion sr
                    JOIN C_BPartner_Location bpl ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN xx_ca_fournisseur xf ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (
                              :group_label IS NULL OR
                              (CASE 
                                  WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                  WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                  ELSE 'autre'
                              END = :group_label)
                          )
                ) ORDER BY "SORT_ORDER", "TOTAL" DESC
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'ad_org_id': ad_org_id,
                'group_label': group_label
            }
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching zone recap: {e}")
        return {"error": "An error occurred while fetching zone recap."}



def fetch_client_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT * FROM (
                    -- Recap by Client
                    SELECT 
                        CAST(cb.name AS VARCHAR2(100)) AS "CLIENT",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        0 AS "SORT_ORDER"
                    FROM C_BPartner cb
                    JOIN xx_ca_fournisseur xf ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (
                              :group_label IS NULL OR
                              (CASE 
                                  WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                  WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                  ELSE 'autre'
                              END = :group_label)
                          )
                    GROUP BY cb.name

                    UNION ALL

                    -- Total Row
                    SELECT 
                        'Total' AS "CLIENT",
                        SUM(xf.TOTALLINE) AS "TOTAL",
                        SUM(xf.qtyentered) AS "QTY",
                        CASE 
                            WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                            ELSE ROUND((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / NULLIF(SUM(xf.CONSOMATION), 0), 4)
                        END AS "MARGE",
                        1 AS "SORT_ORDER"
                    FROM C_BPartner cb
                    JOIN xx_ca_fournisseur xf ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (
                              :group_label IS NULL OR
                              (CASE 
                                  WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                  WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                  ELSE 'autre'
                              END = :group_label)
                          )
                ) ORDER BY "SORT_ORDER", "TOTAL" DESC
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'ad_org_id': ad_org_id,
                'group_label': group_label
            }
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching client recap: {e}")
        return {"error": "An error occurred while fetching client recap."}


@app.route('/fetchClientRecap', methods=['GET'])
def fetch_client():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')
    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400
    if not ad_org_id:
        return jsonify({"error": "Missing ad_org_id parameter"}), 400

    data = fetch_client_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)
    return jsonify(data)

def fetch_operator_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT * FROM (
                    SELECT CAST(au.name AS VARCHAR2(100)) AS "OPERATEUR", 
                           SUM(xf.TOTALLINE) AS "TOTAL", 
                           SUM(xf.qtyentered) AS "QTY",
                           CASE 
                               WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                               WHEN SUM(xf.CONSOMATION) < 0 THEN ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION) * -1)
                               ELSE (SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION)
                           END AS "MARGE",
                           0 AS "SORT_ORDER"
                    FROM AD_User au
                    JOIN xx_ca_fournisseur xf ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (
                              :group_label IS NULL OR
                              (CASE 
                                  WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                  WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                  ELSE 'autre'
                              END = :group_label)
                          )
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                    GROUP BY au.name

                    UNION ALL

                    SELECT 'Total' AS "OPERATEUR", 
                           SUM(xf.TOTALLINE) AS "TOTAL", 
                           SUM(xf.qtyentered) AS "QTY",
                           CASE 
                               WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                               WHEN SUM(xf.CONSOMATION) < 0 THEN ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION) * -1)
                               ELSE (SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION)
                           END AS "MARGE",
                           1 AS "SORT_ORDER"
                    FROM AD_User au
                    JOIN xx_ca_fournisseur xf ON au.AD_User_ID = xf.SALESREP_ID
                    JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                    JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                    JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                    JOIN M_InOut mi ON xf.DOCUMENTNO = mi.DOCUMENTNO
                    JOIN C_ORDER c ON mi.C_ORDER_ID = c.C_ORDER_ID
                    WHERE xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                          AND TO_DATE(:end_date, 'YYYY-MM-DD')
                          AND xf.AD_Org_ID = :ad_org_id
                          AND xf.DOCSTATUS != 'RE'
                          AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                          AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                          AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                          AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                          AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                          AND (
                              :group_label IS NULL OR
                              (CASE 
                                  WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                  WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                  ELSE 'autre'
                              END = :group_label)
                          )
                          AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                ) ORDER BY "SORT_ORDER", "TOTAL" DESC
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'ad_org_id': ad_org_id,
                'group_label': group_label
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching operator recap: {e}")
        return {"error": "An error occurred while fetching operator recap."}


@app.route('/fetchOperatorRecap', methods=['GET'])
def fetch_operator():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400
    if not ad_org_id:
        return jsonify({"error": "Missing ad_org_id parameter"}), 400

    data = fetch_operator_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)
    return jsonify(data)




def fetch_bccb_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, group_label=None):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT DOCUMENTNO, DATEORDERED, GRANDTOTAL, 
                       ROUND(AVG(marge * 100), 2) AS marge
                FROM (
                    SELECT det.*
                    FROM (
                        SELECT lot.*,  
                               (lot.priceentered - ((lot.priceentered * NVL(lot.remise_vente, 0)) / 100)) / (1 + (lot.bonus_vente / 100)) AS ventef
                        FROM (
                            SELECT 
                                ol.priceentered, 
                                (SELECT valuenumber 
                                 FROM m_attributeinstance 
                                 WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                       AND m_attribute_id = 1000504) AS p_revient,
                                (SELECT valuenumber 
                                 FROM m_attributeinstance 
                                 WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                       AND m_attribute_id = 1000908) AS bonus_vente,
                                (SELECT valuenumber 
                                 FROM m_attributeinstance 
                                 WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                       AND m_attribute_id = 1001408) AS remise_vente,
                                c.DOCUMENTNO, 
                                c.DATEORDERED, 
                                c.GRANDTOTAL,
                                CASE 
                                    WHEN SUM(xf.CONSOMATION) = 0 THEN 0
                                    WHEN SUM(xf.CONSOMATION) < 0 THEN ((SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION) * -1)
                                    ELSE (SUM(xf.TOTALLINE) - SUM(xf.CONSOMATION)) / SUM(xf.CONSOMATION)
                                END AS marge
                            FROM C_Order c
                            JOIN C_OrderLine ol ON c.C_Order_ID = ol.C_Order_ID 
                            JOIN M_InOut mi ON mi.C_Order_ID = c.C_Order_ID
                            JOIN xx_ca_fournisseur xf ON xf.DOCUMENTNO = mi.DOCUMENTNO
                            JOIN C_BPartner cb ON cb.C_BPartner_ID = xf.CLIENTID
                            JOIN AD_User au ON au.AD_User_ID = xf.SALESREP_ID
                            JOIN C_BPartner_Location bpl ON bpl.C_BPartner_ID = xf.CLIENTID
                            JOIN C_SalesRegion sr ON sr.C_SalesRegion_ID = bpl.C_SalesRegion_ID
                            WHERE c.AD_Org_ID = 1000000 
                                  AND ol.qtyentered > 0
                                  AND xf.MOVEMENTDATE BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                                      AND TO_DATE(:end_date, 'YYYY-MM-DD')
                                  AND (c.DOCSTATUS = 'CL' OR c.DOCSTATUS = 'CO')
                                  AND c.C_DocType_ID IN (1000539, 1001408)
                                  AND (:bccb IS NULL OR UPPER(c.DOCUMENTNO) LIKE UPPER(:bccb) || '%')
                                  AND (:fournisseur IS NULL OR UPPER(xf.name) LIKE UPPER(:fournisseur) || '%')
                                  AND (:product IS NULL OR UPPER(xf.product) LIKE UPPER(:product) || '%')
                                  AND (:client IS NULL OR UPPER(cb.name) LIKE UPPER(:client) || '%')
                                  AND (:operateur IS NULL OR UPPER(au.name) LIKE UPPER(:operateur) || '%')
                                  AND (:zone IS NULL OR UPPER(sr.name) LIKE UPPER(:zone) || '%')
                                  AND (
                                      :group_label IS NULL OR
                                      (CASE 
                                          WHEN cb.c_bp_group_id = 1000003 THEN 'para'
                                          WHEN cb.c_bp_group_id = 1001330 THEN 'potentiel'
                                          ELSE 'autre'
                                      END = :group_label)
                                  )
                            GROUP BY c.DOCUMENTNO, c.DATEORDERED, c.GRANDTOTAL, ol.priceentered, ol.m_attributesetinstance_id
                            ORDER BY c.DOCUMENTNO
                        ) lot
                    ) det
                )
                GROUP BY DOCUMENTNO, DATEORDERED, GRANDTOTAL
                ORDER BY DOCUMENTNO
            """
            params = {
                'start_date': start_date,
                'end_date': end_date,
                'fournisseur': fournisseur or None,
                'product': product or None,
                'client': client or None,
                'operateur': operateur or None,
                'bccb': bccb or None,
                'zone': zone or None,
                'group_label': group_label
            }

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching BCCB recap: {e}")
        return {"error": "An error occurred while fetching BCCB recap."}


@app.route('/fetchBCCBRecap', methods=['GET']) 
def fetch_bccb():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    data = fetch_bccb_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, group_label)
    return jsonify(data)




@app.route('/download-bccb-product-excel', methods=['GET'])
def download_bccb_product_excel():
    bccb = request.args.get('bccb')
    ad_org_id = request.args.get('ad_org_id', '1000012')

    if not bccb:
        return jsonify({"error": "Missing BCCB parameter"}), 400

    try:
        ad_org_id = int(ad_org_id)
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_bccb_product(bccb, ad_org_id)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"BCCB_Product_{bccb}_{today_date}.xlsx"

    return generate_excel_bccb_product(data, filename)


def generate_excel_bccb_product(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "BCCB Product Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="BCCBProductTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")






@app.route('/download-fournisseur-excel', methods=['GET'])
def download_fournisseur_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_fournisseur_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id and group_label
    today_date = datetime.now().strftime("%d-%m-%Y")
    group_label_part = f"_{group_label}" if group_label else ""
    if ad_org_id == 1000012:
        filename = f"FournisseurRecap_facturation{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"FournisseurRecap{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_fournisseur(data, filename)


def generate_excel_fournisseur(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Fournisseur Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="FournisseurRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")






@app.route('/download-product-excel', methods=['GET'])
def download_product_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_product_data(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id and group_label
    today_date = datetime.now().strftime("%d-%m-%Y")
    group_label_part = f"_{group_label}" if group_label else ""
    if ad_org_id == 1000012:
        filename = f"ProductRecap_facturation{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"ProductRecap{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_product(data, filename)


def generate_excel_product(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Product Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="ProductRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")





@app.route('/download-totalrecap-excel', methods=['GET'])
def download_totalrecap_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date or not ad_org_id:
        return jsonify({"error": "Missing start_date, end_date, or ad_org_id parameters"}), 400

    try:
        ad_org_id = int(ad_org_id)  # Convert to integer
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data from the database
    data = fetch_rcap_data(start_date, end_date, ad_org_id, group_label)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id and group_label
    today_date = datetime.now().strftime("%d-%m-%Y")
    group_label_part = f"_{group_label}" if group_label else ""
    if ad_org_id == 1000012:
        filename = f"TotalRecap_facturation{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"TotalRecap{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_totalrecap(data, filename)


def generate_excel_totalrecap(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Total Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="TotalRecapTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")






 



@app.route('/download-zone-excel', methods=['GET'])
def download_zone_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_zone_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id and group_label
    today_date = datetime.now().strftime("%d-%m-%Y")
    group_label_part = f"_{group_label}" if group_label else ""
    if ad_org_id == 1000012:
        filename = f"ZoneRecap_facturation{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"ZoneRecap{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_zone(data, filename)


def generate_excel_zone(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Zone Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="ZoneRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/download-client-excel', methods=['GET'])
def download_client_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_client_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id and group_label
    today_date = datetime.now().strftime("%d-%m-%Y")
    group_label_part = f"_{group_label}" if group_label else ""
    if ad_org_id == 1000012:
        filename = f"ClientRecap_facturation{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"ClientRecap{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_client(data, filename)


def generate_excel_client(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Client Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="ClientRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route('/download-operator-excel', methods=['GET'])
def download_operator_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')
    ad_org_id = request.args.get('ad_org_id')
    group_label = request.args.get('group_label')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    try:
        ad_org_id = int(ad_org_id) if ad_org_id else None  # Convert if provided
    except ValueError:
        return jsonify({"error": "ad_org_id must be an integer"}), 400

    # Fetch data
    data = fetch_operator_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone, ad_org_id, group_label)

    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename based on ad_org_id and group_label
    today_date = datetime.now().strftime("%d-%m-%Y")
    group_label_part = f"_{group_label}" if group_label else ""
    if ad_org_id == 1000012:
        filename = f"OperatorRecap_facturation{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"
    else:
        filename = f"OperatorRecap{group_label_part}_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_operator(data, filename)


def generate_excel_operator(data, filename):
    if not data or isinstance(data, dict) and "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Operator Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Add headers
    ws.append(df.columns.tolist())
    for cell in ws[1]:  # Style header row
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))  # Convert tuple to list before appending
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="OperatorRecapTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")








@app.route('/download-bccb-excel', methods=['GET'])
def download_bccb_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    fournisseur = request.args.get('fournisseur')
    product = request.args.get('product')
    client = request.args.get('client')
    operateur = request.args.get('operateur')
    bccb = request.args.get('bccb')
    zone = request.args.get('zone')

    if not start_date or not end_date:
        return jsonify({"error": "Missing start_date or end_date parameters"}), 400

    # Fetch data
    data = fetch_bccb_recap(start_date, end_date, fournisseur, product, client, operateur, bccb, zone)

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    # Generate filename
    today_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"BCCBRecap_{start_date}_to_{end_date}_{today_date}.xlsx"

    return generate_excel_bccb(data, filename)

def generate_excel_bccb(data, filename):
    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400

    df = pd.DataFrame(data)
    if df.empty:
        return jsonify({"error": "No data available"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "BCCB Recap"

    # Formatting headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(df.columns.tolist())
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
    ws.auto_filter.ref = ws.dimensions

    # Add data rows with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    # Add an Excel table
    table = Table(displayName="BCCBRecapTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



 
@app.route('/fetchBCCBProduct', methods=['GET'])
def fetch_bccb_p():
    bccb = request.args.get('bccb')
    ad_org_id = request.args.get('ad_org_id')

    if not ad_org_id:
        return jsonify({"error": "Missing ad_org_id parameter"}), 400

    data = fetch_bccb_product(bccb, ad_org_id)
    return jsonify(data)



def fetch_bccb_product(bccb, ad_org_id):
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            query = """
                SELECT product, qty, remise, marge 
                FROM (
                    SELECT det.*, 
                           ROUND((det.ventef - det.p_revient) / det.p_revient * 100, 2) AS marge 
                    FROM (
                        SELECT lot.*, 
                               (lot.priceentered - ((lot.priceentered * NVL(lot.remise_vente, 0)) / 100)) / 
                               (1 + (lot.bonus_vente / 100)) AS ventef 
                        FROM (
                            SELECT ol.priceentered AS priceentered, 
                                   ol.qtyentered AS qty, 
                                   mp.name AS product, 
                                   ol.discount / 100 AS remise, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1000504) AS p_revient, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1000908) AS bonus_vente, 
                                   (SELECT valuenumber FROM m_attributeinstance 
                                    WHERE m_attributesetinstance_id = ol.m_attributesetinstance_id 
                                          AND m_attribute_id = 1001408) AS remise_vente 
                            FROM c_orderline ol 
                            INNER JOIN c_order o ON o.c_order_id = ol.c_order_id 
                            INNER JOIN m_product mp ON ol.m_product_id = mp.m_product_id 
                            WHERE ol.qtyentered > 0 
                                  AND (:bccb IS NULL OR UPPER(o.documentno) LIKE UPPER(:bccb) || '%')
                                  AND o.AD_Org_ID = :ad_org_id
                        ) lot
                    ) det
                )
            """
            
            params = {
                'bccb': bccb or None,
                'ad_org_id': ad_org_id
            }
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching BCCB product data: {e}")
        return {"error": "An error occurred while fetching BCCB product data."}







if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000)
