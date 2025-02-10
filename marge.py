import oracledb
from flask import Flask, jsonify, request
from flask_cors import CORS  # Import CORS
import logging
from flask import Flask, request, send_file
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure Oracle database connection pool
DB_POOL = oracledb.create_pool(
    user="compiere",
    password="compiere",
    dsn="192.168.1.213/compiere",
    min=2,
    max=10,
    increment=1
)

# Fetch data from Oracle DB
def fetch_data_from_db():
    try:
        with DB_POOL.acquire() as connection:
            cursor = connection.cursor()
            # Define your SQL query here
            query = """
            SELECT
              *
            FROM
              (
                SELECT
                  "source"."FOURNISSEUR" "FOURNISSEUR",
                  "source"."PRODUCT" "PRODUCT",
                  "source"."P_ACHAT" "P_ACHAT",
                  "source"."P_VENTE" "P_VENTE",
                  "source"."REM_ACHAT" "REM_ACHAT",
                  "source"."REM_VENTE" "REM_VENTE",
                  "source"."BON_ACHAT" "BON_ACHAT",
                  "source"."BON_VENTE" "BON_VENTE",
                  "source"."REMISE_AUTO" "REMISE_AUTO",
                  "source"."BONUS_AUTO" "BONUS_AUTO",
                  "source"."P_REVIENT" "P_REVIENT",
                  "source"."MARGE" "MARGE",
                  "source"."LABO" "LABO",
                  "source"."LOT" "LOT",
                  "source"."QTY" "QTY"
                FROM
                  (
                    SELECT
                      DISTINCT fournisseur,
                      product,
                      p_achat,
                      p_vente,
                      round(rem_achat, 2) AS rem_achat,
                      rem_vente,
                      round(bon_achat, 2) AS bon_achat,
                      bon_vente,
                      remise_auto,
                      bonus_auto,
                      round(p_revient, 2) AS p_revient,
                      round((marge), 2) marge,
                      labo,
                      lot,
                      qty
                    FROM
                      (
                        SELECT
                          d.*,
                          round(
                            (((ventef - ((ventef * nvl(rma, 0)) / 100))) - p_revient) / p_revient * 100,
                            2
                          ) marge
                        FROM
                          (
                            SELECT
                              det.*,
                              (det.p_achat - ((det.p_achat * det.rem_achat) / 100)) / (1 + (det.bon_achat / 100)) p_revient,
                              (
                                det.p_vente - ((det.p_vente * nvl(det.rem_vente, 0)) / 100)
                              ) / (
                                1 + (
                                  CASE
                                    WHEN det.bna > 0 THEN det.bna
                                    ELSE det.bon_vente
                                  END / 100
                                )
                              ) ventef
                            FROM
                              (
                                SELECT
                                  p.name product,
                                  (
                                    SELECT
                                      NAME
                                    FROM
                                      XX_Laboratory
                                    WHERE
                                      XX_Laboratory_id = p.XX_Laboratory_id
                                  ) labo,
                                  mst.qtyonhand qty,
                                  mati.value fournisseur,
                                  mats.guaranteedate,
                                  md.name remise_auto,
                                  sal.description bonus_auto,
                                  md.flatdiscount rma,
                                  TO_NUMBER(
                                    CASE
                                      WHEN REGEXP_LIKE(sal.name, '^[0-9]+$') THEN sal.name
                                      ELSE NULL
                                    END
                                  ) AS bna,
                                  (
                                    SELECT
                                      valuenumber
                                    FROM
                                      m_attributeinstance
                                    WHERE
                                      m_attributesetinstance_id = mst.m_attributesetinstance_id
                                      AND m_attribute_id = 1000501
                                  ) p_achat,
                                  (
                                    SELECT
                                      valuenumber
                                    FROM
                                      m_attributeinstance
                                    WHERE
                                      m_attributesetinstance_id = mst.m_attributesetinstance_id
                                      AND m_attribute_id = 1001009
                                  ) rem_achat,
                                  (
                                    SELECT
                                      valuenumber
                                    FROM
                                      m_attributeinstance
                                    WHERE
                                      m_attributesetinstance_id = mst.m_attributesetinstance_id
                                      AND m_attribute_id = 1000808
                                  ) bon_achat,
                                  (
                                    SELECT
                                      valuenumber
                                    FROM
                                      m_attributeinstance
                                    WHERE
                                      m_attributesetinstance_id = mst.m_attributesetinstance_id
                                      AND m_attribute_id = 1000502
                                  ) p_vente,
                                  (
                                    SELECT
                                      valuenumber
                                    FROM
                                      m_attributeinstance
                                    WHERE
                                      m_attributesetinstance_id = mst.m_attributesetinstance_id
                                      AND m_attribute_id = 1001408
                                  ) rem_vente,
                                  (
                                    SELECT
                                      valuenumber
                                    FROM
                                      m_attributeinstance
                                    WHERE
                                      m_attributesetinstance_id = mst.m_attributesetinstance_id
                                      AND m_attribute_id = 1000908
                                  ) bon_vente,
                                  (
                                    SELECT
                                      lot
                                    FROM
                                      m_attributesetinstance
                                    WHERE
                                      m_attributesetinstance_id = mst.m_attributesetinstance_id
                                  ) lot
                                FROM
                                  m_product p
                                  INNER JOIN m_storage mst ON p.m_product_id = mst.m_product_id
                                  INNER JOIN m_attributeinstance mati ON mst.m_attributesetinstance_id = mati.m_attributesetinstance_id
                                  INNER JOIN m_attributesetinstance mats ON mst.m_attributesetinstance_id = mats.m_attributesetinstance_id
                                  LEFT JOIN C_BPartner_Product cp ON cp.m_product_id = p.m_product_id
                                    OR cp.C_BPartner_Product_id IS NULL
                                  LEFT JOIN M_DiscountSchema md ON cp.M_DiscountSchema_id = md.M_DiscountSchema_id
                                  LEFT JOIN XX_SalesContext sal ON p.XX_SalesContext_ID = sal.XX_SalesContext_ID
                                WHERE
                                  mati.m_attribute_id = 1000508
                                  AND mst.m_locator_id IN (1000614, 1001135)
                                  AND mst.qtyonhand != 0
                                ORDER BY
                                  p.name
                              ) det
                            WHERE
                              det.rem_achat < 200
                          ) d
                      )
                    GROUP BY
                      fournisseur,
                      product,
                      p_achat,
                      p_vente,
                      rem_achat,
                      rem_vente,
                      bon_achat,
                      bon_vente,
                      remise_auto,
                      bonus_auto,
                      p_revient,
                      marge,
                      labo,
                      lot,
                      qty
                    ORDER BY
                      fournisseur
                  ) "source"
              )
            WHERE
              rownum <= 1048575
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]  # Get column names
            data = [dict(zip(columns, row)) for row in rows]
            return data
    except Exception as e:
        logger.error(f"Error fetching data: {e}")
        return {"error": "An error occurred while fetching data."}

@app.route('/fetch-data', methods=['GET'])
def fetch_data():
    data = fetch_data_from_db()
    return jsonify(data)



@app.route('/download-excel', methods=['GET'])
def download_excel():
    # Get MARGE value from request (if available)
    marge_value = request.args.get("marge", "MARGE")

    # Fetch real data from Oracle DB
    data = fetch_data_from_db()

    if not data or "error" in data:
        return jsonify({"error": "No data available"}), 400  # Handle empty dataset

    # Convert to Pandas DataFrame
    df = pd.DataFrame(data)

    if df.empty:
        return jsonify({"error": "No data available"}), 400

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "All Data"

    # Write headers with formatting
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(df.columns.tolist())  # Add header row

    for col_idx, cell in enumerate(ws[1], 1):  # Format header row
        cell.fill = header_fill
        cell.font = header_font
        ws.auto_filter.ref = ws.dimensions  # Enable filter & sort

    # Write data with alternating row colors
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)  # Add data row
        if row_idx % 2 == 0:  # Apply alternating row colors
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Create an Excel table with formatting
    table = Table(displayName="DataTable", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Generate a dynamic filename
    today_date = datetime.now().strftime("%d-%m-%Y")  # Format: day-month-year
    filename = f"{marge_value}_{today_date}.xlsx"

    # Save to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    app.run(debug=True)